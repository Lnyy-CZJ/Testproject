"""历史 JSONL/Excel 导入与统一存储测试。"""

from __future__ import annotations

import json
import tempfile
import time
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from analysis_service import (
    ActiveRunError,
    AnalysisService,
    _candidate_identity_rule,
    _field_comparison_scores_v3,
    DEFAULT_FIELD_DEFINITIONS_V3,
    DEFAULT_FIELD_DEFINITIONS,
    DEFAULT_FIELD_SCHEMA_VERSION,
    DEFAULT_FIELD_SCHEMA_V3_VERSION,
    DuplicateImportError,
    FIELD_PROCESSING_RULE_VERSION,
    FieldSchemaValidationError,
    ImportValidationError,
    METRICS_RULE_VERSION,
    PersonLinkValidationError,
    REPORT_MODEL_VERSION,
    REPORT_V5_CORE_METRIC_DEFINITIONS,
    ReviewValidationError,
    V5_REPORT_MODEL_VERSION,
    COMPARE_REPORT_MODEL_VERSION,
    V5_FIELD_PROCESSING_RULE_VERSION,
    extract_source_path,
    extract_profile_item,
    normalize_field_value,
    validate_field_definitions,
)
from analysis_store import AnalysisStore
from search_tool import FlowError


PROJECT_ROOT = Path(__file__).resolve().parents[1]
IMPORT_WORKBOOK = (
    PROJECT_ROOT / "tests" / "fixtures" / "v1_3_import" / "import_sources.xlsx"
)
E2E_FIXTURE = PROJECT_ROOT / "tests" / "fixtures" / "v1_3_e2e"
DATA_PROCESSING_PHASE0_FIXTURE = (
    PROJECT_ROOT
    / "tests"
    / "fixtures"
    / "v1_3_data_processing_phase0"
    / "problem_contract.json"
)
FIELD_CONFIGURATION_PHASE0_FIXTURE = (
    PROJECT_ROOT
    / "tests"
    / "fixtures"
    / "v1_3_field_configuration_phase0_contract.json"
)

# 报告优化阶段 0 冻结的最小 v5 契约。阶段 1 必须产出这些数据，但不能
# 通过在模板中临时查询 Process/Candidate 来绕过不可变报告快照。
REPORT_MODEL_V5_REQUIRED_TOP_LEVEL_KEYS = frozenset({
    "metadata",
    "summary",
    "overview",
    "processing_scope",
    "query_stage_metrics",
    "query_explorer",
    "diagnostics",
})
REPORT_MODEL_V5_REQUIRED_QUERY_KEYS = frozenset({
    "person_id",
    "query_stage",
    "candidate_run",
})
REPORT_MODEL_V5_REQUIRED_CANDIDATE_KEYS = frozenset({
    "candidate_pk",
    "candidate_id",
    "candidate_rank",
    "detail_status",
    "identity",
    "metrics",
})


def write_jsonl(path: Path, records: list[dict]) -> None:
    """写入测试用脱敏 JSONL，每个对象占一行。"""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(
            json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n"
            for record in records
        ),
        encoding="utf-8",
    )


def api_body(data: dict) -> dict:
    """构造采集服务测试使用的成功业务响应。"""

    return {
        "code": 0,
        "message": "ok",
        "responses": [
            {
                "id": "req_0",
                "success": True,
                "code": 0,
                "message": "ok",
                "data": data,
            }
        ],
    }


class FakeExecutionClient:
    """按既定顺序返回业务响应，避免阶段3测试调用真实接口。"""

    def __init__(self, responses: list[dict | Exception]) -> None:
        """保存响应队列和 process_one 依赖的轮询配置。"""

        self.responses = list(responses)
        self.calls: list[tuple[str, dict]] = []
        self.config = SimpleNamespace(
            max_poll_count=3,
            poll_interval_seconds=0.001,
        )

    def call(self, stage: str, params: dict) -> dict:
        """返回下一条响应；异常对象用于模拟接口流程失败。"""

        self.calls.append((stage, params))
        if not self.responses:
            raise AssertionError("出现了未预期的接口调用")
        response = self.responses.pop(0)
        if isinstance(response, Exception):
            raise response
        return response


class AnalysisServiceTests(unittest.TestCase):
    """验证导入校验、统一模型、Raw 标记、归档和重复保护。"""

    def setUp(self):
        """为每个测试创建隔离数据库和数据目录。"""

        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.store = AnalysisStore(self.root / "data" / "searchtool.db")
        self.store.initialize()
        self.store.create_evaluation("eval-import", "导入测试")
        self.service = AnalysisService(self.store, self.root / "data")

    def tearDown(self):
        """清理当前测试创建的临时数据库和归档。"""

        self.temp_dir.cleanup()

    def test_compare_report_delta_direction_respects_metric_meaning(self):
        """非命中相似度越低越好，资料完整度变化只做观察不误标提升。"""

        baseline = {
            key: {"value": 0.5}
            for key in REPORT_V5_CORE_METRIC_DEFINITIONS
        }
        candidate = {
            key: {"value": 0.6}
            for key in REPORT_V5_CORE_METRIC_DEFINITIONS
        }
        candidate["nonmatched_baseline_overlap"] = {"value": 0.4}
        rows = {
            item["key"]: item
            for item in self.service._report_compare_metric_rows(
                baseline, candidate
            )
        }
        self.assertEqual(
            "IMPROVED",
            rows["nonmatched_baseline_overlap"]["delta_status"],
        )
        self.assertEqual(
            "OBSERVED",
            rows["nonmatched_data_completeness"]["delta_status"],
        )

    def test_identity_rule_uses_social_priority_then_photo_fallback(self):
        """Social 冲突优先，照片仅在 Social 无结论时作为自动终判兜底。"""

        judgement, reason, _ = _candidate_identity_rule(
            ["https://linkedin.com/in/target"],
            [
                "https://linkedin.com/in/target",
                "https://linkedin.com/in/another-person",
            ],
            99,
        )
        self.assertEqual(("NOT_HIT", "SOCIAL_CONFLICT"), (judgement, reason))

        judgement, reason, _ = _candidate_identity_rule(
            ["https://twitter.com/Target/"],
            ["https://x.com/target?utm_source=test"],
            None,
        )
        self.assertEqual(("HIT", "SOCIAL_MATCH"), (judgement, reason))

        # Summary Social Links 的接口值是对象数组，也必须取其中 url 参与强绑定。
        judgement, reason, _ = _candidate_identity_rule(
            ["https://linkedin.com/in/target"],
            [{"platform": "linkedin", "url": "https://www.linkedin.com/in/target/"}],
            None,
        )
        self.assertEqual(("HIT", "SOCIAL_MATCH"), (judgement, reason))

        judgement, reason, _ = _candidate_identity_rule(
            ["https://linkedin.com/in/target"],
            [],
            0.8,
        )
        self.assertEqual(("HIT", "PHOTO_MATCH"), (judgement, reason))

        judgement, reason, _ = _candidate_identity_rule([], [], None)
        self.assertEqual(("SUSPECTED", "NO_STRONG_FIELD"), (judgement, reason))

        judgement, reason, _ = _candidate_identity_rule(
            ["https://linkedin.com/in/target"],
            [],
            79.9,
        )
        self.assertEqual(
            ("NOT_HIT", "PHOTO_BELOW_THRESHOLD"),
            (judgement, reason),
        )

    def _build_data_processing_phase0_scenario(
        self,
        force_rule_version: str | None = None,
    ) -> dict:
        """构造数据处理优化阶段0的脱敏问题场景。

        功能说明:
            根据冻结契约生成10条未关联人物的 Query、45名候选人和10名
            Baseline Person，用于稳定复现 v2 的作用域、模块判空和指标状态
            问题。全部数据只写入当前测试临时目录，不访问外部接口。

        返回值:
            dict: 包含冻结契约、Process 指标和数据库计数的测试上下文。

        异常说明:
            夹具损坏、导入失败或处理失败时直接让测试失败，避免预期失败
            装饰器掩盖测试前置条件错误。
        """

        contract = json.loads(
            DATA_PROCESSING_PHASE0_FIXTURE.read_text(encoding="utf-8")
        )
        schema_version = self.service.ensure_default_field_schema()
        available_fields = contract["baseline_available_fields"]
        baseline_path = self.root / "phase0-baseline.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": f"person-phase0-{index:02d}",
                    "display_name": f"Phase0 Person {index:02d}",
                    "fields": {
                        "social_urls": [
                            f"https://social.example.test/phase0-{index:02d}"
                        ]
                    },
                    "baseline_available_fields": available_fields,
                }
                for index in range(1, 11)
            ],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="数据处理阶段0脱敏基准",
            baseline_version="baseline-data-processing-phase0",
        )

        candidate_index = 0
        result_records = []
        for query_index, candidate_count in enumerate(
            contract["candidate_count_by_query"],
            start=1,
        ):
            query_id = f"phase0-query-{query_index:02d}"
            candidates = []
            for candidate_rank in range(1, candidate_count + 1):
                candidate_index += 1
                has_insights = candidate_index <= 2
                has_social = candidate_index <= 34
                ui_sections = {
                    "insights": {
                        "status": "data" if has_insights else "empty",
                        "data": {
                            "count": 1 if has_insights else 0,
                            "items": (
                                [
                                    {
                                        "description": (
                                            f"Phase0 insight {candidate_index}"
                                        ),
                                        "links": [
                                            "https://example.test/insight"
                                        ],
                                    }
                                ]
                                if has_insights
                                else []
                            ),
                        },
                    },
                    "photos": {
                        "status": "empty",
                        "data": {"count": 0, "image_urls": []},
                    },
                    "profile": {
                        "status": "data",
                        "data": {
                            "sections": {
                                "person_biography": (
                                    f"Phase0 biography {candidate_index}"
                                )
                            }
                        },
                    },
                    "social": {
                        "status": "data" if has_social else "empty",
                        "data": {
                            "profiles": (
                                [
                                    {
                                        "display_handle": (
                                            f"phase0-{candidate_index}"
                                        ),
                                        "platform": "example",
                                        "url": (
                                            "https://social.example.test/"
                                            f"candidate-{candidate_index}"
                                        ),
                                    }
                                ]
                                if has_social
                                else []
                            )
                        },
                    },
                    "summary": {
                        "status": "data",
                        "data": {
                            "display_name": (
                                f"Phase0 Candidate {candidate_index}"
                            ),
                            "confidence_level": "HIGH",
                        },
                    },
                }
                candidate_id = f"phase0-candidate-{candidate_index:02d}"
                candidates.append(
                    {
                        "candidate_rank": candidate_rank,
                        "candidate_id": candidate_id,
                        "rank_score": 1 - candidate_rank / 100,
                        "detail_status": "SUCCESS",
                        "detail_error": "",
                        "list_item_raw": {"candidate_id": candidate_id},
                        "detail_data_raw": {"ui_sections": ui_sections},
                        "ui_sections": ui_sections,
                    }
                )
            result_records.append(
                {
                    "result_schema_version": "1.3.1",
                    "input_id": query_id,
                    "person_id": None,
                    "query_stage": "FULL_NAME",
                    "task_id": f"phase0-task-{query_index:02d}",
                    "query_status": "SUCCESS",
                    "result_status": "HAS_CANDIDATES",
                    "candidate_count_total": candidate_count,
                    "results": candidates,
                }
            )

        results_path = self.root / "phase0-results.jsonl"
        write_jsonl(results_path, result_records)
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="数据处理阶段0问题复现",
            system_version="phase0-v2",
            run_id="run-data-processing-phase0",
        )
        process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
            baseline_version="baseline-data-processing-phase0",
            process_id="process-data-processing-phase0",
        )
        if force_rule_version:
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    UPDATE process_runs SET rule_version = ?
                    WHERE process_id = ?
                    """,
                    (force_rule_version, process.process_id),
                )
        metrics = self.service.calculate_process_metrics(process.process_id)
        counts = {
            "query_count": self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM run_queries
                WHERE run_id = ?
                """,
                (imported.object_id,),
            )["count"],
            "missing_person_id_count": self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM run_queries
                WHERE run_id = ?
                  AND (person_id IS NULL OR person_id = '')
                """,
                (imported.object_id,),
            )["count"],
            "candidate_count": self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM candidates
                WHERE run_id = ?
                """,
                (imported.object_id,),
            )["count"],
            "pending_review_count": self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM reviews
                WHERE process_id = ? AND judgement = 'PENDING_REVIEW'
                """,
                (process.process_id,),
            )["count"],
            "query_task_id_count": self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM processed_queries
                WHERE process_id = ?
                  AND json_extract(fields_json, '$.task_id') IS NOT NULL
                """,
                (process.process_id,),
            )["count"],
        }
        return {
            "contract": contract,
            "process_id": process.process_id,
            "metrics": metrics,
            "counts": counts,
        }

    def test_data_processing_phase0_v2_problem_snapshot_is_reproducible(self):
        """v2 的六类真实问题可在脱敏数据中稳定复现。"""

        scenario = self._build_data_processing_phase0_scenario(
            "field-processing-v2"
        )
        contract = scenario["contract"]
        metrics = scenario["metrics"]
        counts = scenario["counts"]

        self.assertEqual(contract["query_count"], counts["query_count"])
        self.assertEqual(
            contract["missing_person_id_count"],
            counts["missing_person_id_count"],
        )
        self.assertEqual(contract["candidate_count"], counts["candidate_count"])
        self.assertEqual(
            contract["pending_review_count"],
            counts["pending_review_count"],
        )
        self.assertEqual(
            contract["query_task_id_count"],
            counts["query_task_id_count"],
        )
        self.assertEqual("metrics-v2", metrics["metrics_rule_version"])
        self.assertEqual(
            contract["v2_snapshot"]["task_id"],
            {
                "returned_count": metrics["field_metrics"]["task_id"][
                    "returned_count"
                ],
                "candidate_count": metrics["field_metrics"]["task_id"][
                    "candidate_count"
                ],
            },
        )
        self.assertEqual(
            contract["v2_snapshot"]["insights_return_rate"],
            metrics["module_metrics"]["Insights"]["return_rate"],
        )
        self.assertEqual(
            contract["v2_snapshot"]["photos_return_rate"],
            metrics["module_metrics"]["Photos"]["return_rate"],
        )
        self.assertEqual(
            "NOT_APPLICABLE",
            metrics["matched_completeness"]["status"],
        )
        self.assertEqual(
            contract["baseline_available_field_count"],
            len(contract["baseline_available_fields"]),
        )
        self.assertEqual(
            contract["v2_snapshot"]["completeness_field_keys"],
            [
                definition["field_key"]
                for definition in DEFAULT_FIELD_DEFINITIONS
                if "completeness" in definition["scoring_role"]
            ],
        )

    def test_data_processing_phase0_v3_query_scope_contract(self):
        """v3 应按10条 Query 统计 task_id，而不是按45名候选人统计。"""

        metrics = self._build_data_processing_phase0_scenario()["metrics"]
        self.assertEqual(
            {
                "value_scope": "QUERY",
                "entity_count": 10,
                "returned_count": 10,
                "return_rate": 1.0,
            },
            {
                key: metrics["field_metrics"]["task_id"][key]
                for key in (
                    "value_scope",
                    "entity_count",
                    "returned_count",
                    "return_rate",
                )
            },
        )

    def test_data_processing_phase0_v3_module_empty_contract(self):
        """v3 应按模块语义判空，空容器不能算作有数据。"""

        metrics = self._build_data_processing_phase0_scenario()["metrics"]
        self.assertEqual(
            {
                "data_count": 2,
                "empty_count": 43,
                "unknown_count": 0,
                "candidate_count": 45,
                "data_rate": 2 / 45,
            },
            {
                key: metrics["module_metrics"]["Insights"][key]
                for key in (
                    "data_count",
                    "empty_count",
                    "unknown_count",
                    "candidate_count",
                    "data_rate",
                )
            },
        )
        self.assertEqual(
            0,
            metrics["module_metrics"]["Photos"]["data_count"],
        )
        self.assertEqual(
            45,
            metrics["module_metrics"]["Profile"]["data_count"],
        )
        self.assertEqual(
            34,
            metrics["module_metrics"]["Social"]["data_count"],
        )

    def test_data_processing_phase0_v3_identity_pending_reason_contract(self):
        """v3 应把未关联和待归类识别为 NOT_READY 并返回原因码。"""

        metrics = self._build_data_processing_phase0_scenario()["metrics"]
        self.assertEqual(
            "NOT_READY",
            metrics["matched_completeness"]["status"],
        )
        self.assertEqual(
            {"BASELINE_NOT_LINKED", "IDENTITY_PENDING"},
            set(metrics["matched_completeness"]["reason_codes"]),
        )

    def test_data_processing_phase0_v3_candidate_return_rate_contract(self):
        """v3 候选人字段返回率不依赖身份归类，应在全 PENDING 时就绪。"""

        metrics = self._build_data_processing_phase0_scenario()["metrics"]
        social_urls = metrics["field_metrics"]["social_urls"]
        self.assertEqual("READY", social_urls["status"])
        self.assertEqual("CANDIDATE", social_urls["value_scope"])
        self.assertEqual(45, social_urls["entity_count"])
        self.assertEqual(34, social_urls["returned_count"])
        self.assertEqual(34 / 45, social_urls["return_rate"])

    def test_data_processing_phase0_v3_field_matrix_conflict_contract(self):
        """v3 字段矩阵应明确暴露15个基准字段与1个评分字段的冲突。"""

        scenario = self._build_data_processing_phase0_scenario()
        matrix = self.service.get_field_comparison_matrix(
            schema_version=DEFAULT_FIELD_SCHEMA_VERSION,
            baseline_version="baseline-data-processing-phase0",
            process_id=scenario["process_id"],
        )
        self.assertEqual(15, matrix["baseline_available_field_count"])
        self.assertEqual(1, matrix["completeness_field_count"])
        self.assertEqual(14, matrix["baseline_only_conflict_count"])

    def test_stage4_metrics_v3_and_report_model_v3_explain_not_ready(self):
        """v3 字段返回率保持可用，质量指标和报告给出明确未就绪原因。"""

        scenario = self._build_data_processing_phase0_scenario()
        metrics = scenario["metrics"]
        self.assertEqual("metrics-v3", metrics["metrics_rule_version"])
        self.assertEqual(
            {
                "value_scope": "QUERY",
                "entity_count": 10,
                "returned_count": 10,
                "return_rate": 1.0,
                "status": "READY",
            },
            {
                key: metrics["field_metrics"]["task_id"][key]
                for key in (
                    "value_scope",
                    "entity_count",
                    "returned_count",
                    "return_rate",
                    "status",
                )
            },
        )
        self.assertEqual(
            {"BASELINE_NOT_LINKED", "IDENTITY_PENDING"},
            set(metrics["matched_completeness"]["reason_codes"]),
        )
        model = self.service.build_report_model(
            report_id="report-stage4-not-ready",
            candidate_process_id=scenario["process_id"],
        )
        self.assertEqual(
            "report-model-v3",
            model["metadata"]["report_model_version"],
        )
        self.assertEqual(
            "metrics-v3",
            model["metadata"]["metrics_rule_version"],
        )
        self.assertTrue(model["not_ready_reasons"])
        self.assertIn(
            "IDENTITY_PENDING",
            {
                item["reason_code"]
                for item in model["not_ready_reasons"]
            },
        )
        self.assertIn(
            "FIELD_NOT_CONNECTED",
            {
                item["reason_code"]
                for item in model["not_ready_reasons"]
            },
        )

    def test_field_configuration_phase0_default_schema_contract_is_frozen(self):
        """阶段0冻结默认 FieldSchema 的数量、模块与评分职责。

        功能说明：验证后续 FieldSchema v3 开发前，默认 v2 配置仍可发布、
        读取，并保持当前 33 个字段与既有评分职责不变。
        返回值：无；断言失败表示旧配置兼容性被破坏。
        异常说明：夹具缺失或字段定义变化时由测试框架直接报告差异。
        """

        contract = json.loads(
            FIELD_CONFIGURATION_PHASE0_FIXTURE.read_text(encoding="utf-8")
        )
        schema_version = self.service.ensure_default_field_schema()
        schema = self.store.fetch_one(
            "SELECT definitions_json FROM field_schemas WHERE schema_version = ?",
            (schema_version,),
        )
        self.assertIsNotNone(schema)
        definitions = validate_field_definitions(
            json.loads(schema["definitions_json"])
        )

        self.assertEqual(
            contract["versions"]["default_field_schema_version"],
            schema_version,
        )
        self.assertEqual(contract["default_schema"]["field_count"], len(definitions))
        self.assertEqual(
            contract["default_schema"]["module_counts"],
            {
                module: sum(
                    1 for definition in definitions
                    if definition["module"] == module
                )
                for module in contract["default_schema"]["module_counts"]
            },
        )
        for role, expected_keys in contract["default_schema"][
            "scoring_role_field_keys"
        ].items():
            self.assertEqual(
                expected_keys,
                [
                    definition["field_key"]
                    for definition in definitions
                    if role in definition["scoring_role"]
                ],
            )

    def test_field_configuration_phase0_v4_v3_routing_contract_is_frozen(self):
        """阶段0冻结 v4 处理对 metrics-v3/report-model-v3 的兼容路由。

        功能说明：使用既有脱敏 10 Query、45 Candidate 场景，确保最新 v4
        Process 不会误路由到新 metrics，且报告继续使用 v3 模型。
        返回值：无；断言通过代表版本路由和核心统计口径可重复。
        异常说明：未知规则或报告构建失败会由服务层或测试框架直接抛出。
        """

        contract = json.loads(
            FIELD_CONFIGURATION_PHASE0_FIXTURE.read_text(encoding="utf-8")
        )
        scenario = self._build_data_processing_phase0_scenario()
        metrics = scenario["metrics"]
        model = self.service.build_report_model(
            report_id="report-field-configuration-phase0",
            candidate_process_id=scenario["process_id"],
        )

        self.assertEqual(
            contract["versions"]["field_processing_rule_version"],
            FIELD_PROCESSING_RULE_VERSION,
        )
        self.assertEqual("metrics-v3", contract["versions"]["metrics_rule_version"])
        self.assertEqual("report-model-v3", contract["versions"]["report_model_version"])
        self.assertEqual("metrics-v3", metrics["metrics_rule_version"])
        self.assertEqual(
            contract["regression_scenario"]["query_count"],
            scenario["counts"]["query_count"],
        )
        self.assertEqual(
            contract["regression_scenario"]["candidate_count"],
            scenario["counts"]["candidate_count"],
        )
        self.assertEqual(
            contract["regression_scenario"]["module_data_count"],
            {
                module: metrics["module_metrics"][module]["data_count"]
                for module in contract["regression_scenario"]["module_data_count"]
            },
        )
        self.assertEqual(
            "report-model-v3",
            model["metadata"]["report_model_version"],
        )
        self.assertEqual(
            "metrics-v3",
            model["metadata"]["metrics_rule_version"],
        )

    def test_report_optimization_phase0_v4_gap_is_reproducible(self):
        """冻结当前报告缺少完整 Query/Candidate 快照的基线事实。

        功能说明：使用现有脱敏 10 Query、45 Candidate 场景构建当前报告，
        确认 v4 只保存聚合指标和案例分组，尚未具备 v5 Query Explorer。
        返回值：无；断言通过说明阶段 1 的目标缺口仍可稳定复现。
        异常说明：场景无法构建或报告结构意外变化时直接失败。
        """

        scenario = self._build_data_processing_phase0_scenario()
        model = self.service.build_report_model(
            report_id="report-optimization-phase0-gap",
            candidate_process_id=scenario["process_id"],
        )

        self.assertEqual(10, scenario["counts"]["query_count"])
        self.assertEqual(45, scenario["counts"]["candidate_count"])
        self.assertNotEqual(
            V5_REPORT_MODEL_VERSION,
            model["metadata"]["report_model_version"],
        )
        self.assertFalse(
            REPORT_MODEL_V5_REQUIRED_TOP_LEVEL_KEYS.issubset(model)
        )
        self.assertIn("case_groups", model)
        self.assertNotIn("query_explorer", model)

    def test_report_optimization_phase0_v5_version_contract_is_frozen(self):
        """冻结 v5 标识与最小快照键，供阶段 1 路由和结构测试复用。

        功能说明：验证阶段 0 只登记目标版本和不可变快照的最小契约，
        不提前把尚未完成的 v4 模型伪装为 v5。
        返回值：无；断言通过表示版本常量和测试契约可被后续阶段使用。
        异常说明：常量或契约被意外改动时直接失败。
        """

        self.assertEqual("report-model-v5", V5_REPORT_MODEL_VERSION)
        self.assertEqual(
            {
                "metadata",
                "summary",
                "overview",
                "processing_scope",
                "query_stage_metrics",
                "query_explorer",
                "diagnostics",
            },
            set(REPORT_MODEL_V5_REQUIRED_TOP_LEVEL_KEYS),
        )

    def test_field_schema_v3_catalog_and_legacy_definition_are_compatible(self):
        """v3 目录完整发布，旧 v2 定义仅在内存中补齐新属性。"""

        self.service.ensure_default_field_schema()
        schema_version = self.service.ensure_default_field_schema_v3()
        schema = self.store.fetch_one(
            "SELECT definitions_json FROM field_schemas WHERE schema_version = ?",
            (schema_version,),
        )
        definitions = validate_field_definitions(
            json.loads(schema["definitions_json"])
        )
        by_key = {item["field_key"]: item for item in definitions}
        legacy = validate_field_definitions([DEFAULT_FIELD_DEFINITIONS[0]])[0]

        self.assertEqual(DEFAULT_FIELD_SCHEMA_V3_VERSION, schema_version)
        self.assertEqual(87, len(definitions))
        self.assertTrue(by_key["summary_web_links"]["display_enabled"])
        self.assertFalse(by_key["llm_cost"]["enabled"])
        self.assertIn("pdl_person_identify_call_count", by_key)
        self.assertIn("pdl_provider_cost_status", by_key)
        self.assertIn("google_vision_cost_status", by_key)
        self.assertIn("tool_usage_summary", by_key)
        self.assertIn("cost_by_provider", by_key)
        self.assertTrue(by_key["social_urls"]["identity_enabled"])
        self.assertTrue(by_key["social_urls"]["baseline_compare_enabled"])
        self.assertEqual("PATH", legacy["source_type"])
        self.assertEqual({}, legacy["source_options"])
        self.assertTrue(legacy["display_enabled"])

    def test_field_schema_v3_profile_item_selector_and_validation(self):
        """PROFILE_ITEM 只按 section + label 提取，并拒绝缺少选择器的配置。"""

        source = {
            "ui_sections": {
                "profile": {
                    "data": {
                        "sections": [
                            {
                                "title": "Identity",
                                "items": [
                                    {"label": "Full Name", "value": "Alice"},
                                    {"label": "Full Name", "value": "A. Example"},
                                ],
                            }
                        ]
                    }
                }
            }
        }
        value, duplicated = extract_profile_item(
            source,
            "ui_sections.profile.data.sections",
            {"section": " identity ", "label": "full name"},
        )
        profile_definition = next(
            dict(item)
            for item in DEFAULT_FIELD_DEFINITIONS_V3
            if item["field_key"] == "profile_sections"
        )
        profile_definition.update(
            {
                "field_key": "profile_identity_full_name",
                "display_name": "Identity / Full Name",
                "data_type": "array",
                "array_mode": "preserve",
                "normalizer": "string_list",
                "source_type": "PROFILE_ITEM",
                "source_options": {
                    "section": "Identity", "label": "Full Name",
                },
            }
        )
        invalid_definition = dict(profile_definition)
        invalid_definition["source_options"] = {"section": "Identity"}

        self.assertEqual(["Alice", "A. Example"], value)
        self.assertTrue(duplicated)
        self.assertEqual(
            "PROFILE_ITEM",
            validate_field_definitions([profile_definition])[0]["source_type"],
        )
        profile_full_name = next(
            item
            for item in DEFAULT_FIELD_DEFINITIONS_V3
            if item["field_key"] == "profile_full_name"
        )
        self.assertEqual(
            ["Alice", "A. Example"],
            extract_profile_item(
                source,
                profile_full_name["source_path"],
                profile_full_name["source_options"],
            )[0],
        )
        self.assertTrue(profile_full_name["baseline_compare_enabled"])
        self.assertTrue(profile_full_name["completeness_enabled"])
        profile_age = next(
            item
            for item in DEFAULT_FIELD_DEFINITIONS_V3
            if item["field_key"] == "profile_age"
        )
        self.assertEqual(
            38.0,
            normalize_field_value("38", profile_age["normalizer"]),
        )
        current_and_other_roles, duplicated_roles = extract_profile_item(
            {
                "ui_sections": {
                    "profile": {
                        "data": {
                            "sections": [{
                                "title": "Career",
                                "items": [
                                    {"label": "Current Role", "value": "Engineer"},
                                    {"label": "Other Roles", "value": ["Founder"]},
                                ],
                            }],
                        },
                    },
                },
            },
            "ui_sections.profile.data.sections",
            {"section": "Career", "labels": ["Current Role", "Other Roles"]},
        )
        self.assertEqual(["Engineer", ["Founder"]], current_and_other_roles)
        self.assertFalse(duplicated_roles)
        with self.assertRaises(FieldSchemaValidationError):
            validate_field_definitions([invalid_definition])

    def test_stage2_v5_field_comparison_and_cost_free_reprocess(self):
        """v3 Schema 按映射生成字段对比，重处理只读取已入库数据。"""

        definitions = []
        for field_key in ("social_urls", "summary_location"):
            definition = next(
                dict(item)
                for item in DEFAULT_FIELD_DEFINITIONS_V3
                if item["field_key"] == field_key
            )
            definition.update(
                {
                    "baseline_field_key": field_key,
                    "baseline_compare_enabled": True,
                    "completeness_enabled": True,
                    "accuracy_enabled": True,
                    "scoring_role": ["completeness", "accuracy"],
                }
            )
            if field_key == "social_urls":
                definition["identity_enabled"] = True
                definition["scoring_role"] = [
                    "identity", "completeness", "accuracy",
                ]
            definitions.append(definition)
        profile_name = next(
            dict(item)
            for item in DEFAULT_FIELD_DEFINITIONS_V3
            if item["field_key"] == "profile_full_name"
        )
        profile_name.update({
            "baseline_field_key": "profile_full_name",
            "baseline_compare_enabled": True,
            "completeness_enabled": True,
            # 姓名参与资料完整度，但不得进入非命中资料相似度或准确度。
            "accuracy_enabled": False,
            "scoring_role": ["completeness"],
        })
        definitions.append(profile_name)
        schema_version = self.service.publish_field_schema(
            name="阶段2字段对比配置",
            definitions=definitions,
            schema_version="field-schema-stage2-v3",
        )
        baseline_path = self.root / "stage2-v3-baseline.jsonl"
        write_jsonl(
            baseline_path,
            [{
                "person_id": "person-stage2-v3",
                "display_name": "Stage2 V3 Person",
                "fields": {
                    "social_urls": ["https://linkedin.com/in/stage2"],
                    "summary_location": "Redmond, Washington, USA",
                    "profile_full_name": "Stage2 V3 Person",
                },
                "baseline_available_fields": [
                    "social_urls", "summary_location", "profile_full_name",
                ],
            }],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段2 V3 基准",
            baseline_version="baseline-stage2-v3",
        )
        results_path = self.root / "stage2-v3-results.jsonl"
        ui_sections = {
            "social": {
                "status": "data",
                "data": {"profiles": [{
                    "url": "https://www.linkedin.com/in/stage2/",
                }]},
            },
            "summary": {
                "status": "data",
                "data": {
                    "location": "Redmond, Washington, United States",
                },
            },
            "profile": {
                "status": "data",
                "data": {
                    "sections": [{
                        "title": "Identity",
                        "items": [{
                            "label": "Full Name",
                            "value": "Stage2 V3 Person",
                        }],
                    }],
                },
            },
        }
        write_jsonl(
            results_path,
            [{
                "result_schema_version": "1.3",
                "input_id": "query-stage2-v3",
                "person_id": "person-stage2-v3",
                "query_stage": "FULL_NAME_SOCIAL",
                "task_id": "task-stage2-v3",
                "query_status": "SUCCESS",
                "results": [
                    {
                        "candidate_rank": 1,
                        "rank_score": 0.9,
                        "candidate_id": "candidate-stage2-v3",
                        "detail_status": "SUCCESS",
                        "detail_error": "",
                        "list_item_raw": {
                            "candidate_id": "candidate-stage2-v3",
                            "rank_score": 0.9,
                        },
                        "detail_data_raw": {"ui_sections": ui_sections},
                        "ui_sections": ui_sections,
                    },
                    {
                        # Social 同平台 URL 不一致：自动归类为 NOT_HIT。
                        "candidate_rank": 2,
                        "rank_score": 0.3,
                        "candidate_id": "candidate-stage2-v3-non-hit",
                        "detail_status": "SUCCESS",
                        "detail_error": "",
                        "list_item_raw": {
                            "candidate_id": "candidate-stage2-v3-non-hit",
                            "rank_score": 0.3,
                        },
                        "detail_data_raw": {"ui_sections": {
                            **ui_sections,
                            "social": {"status": "data", "data": {
                                "profiles": [{"url": "https://linkedin.com/in/other"}],
                            }},
                        }},
                        "ui_sections": {
                            **ui_sections,
                            "social": {"status": "data", "data": {
                                "profiles": [{"url": "https://linkedin.com/in/other"}],
                            }},
                        },
                    },
                ],
            }],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="stage2-v3",
            system_version="v3",
            run_id="run-stage2-v3",
        )
        raw_count_before = self.store.fetch_one(
            "SELECT COUNT(*) AS count FROM raw_records WHERE run_id = ?",
            (imported.object_id,),
        )["count"]
        process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
            baseline_version="baseline-stage2-v3",
            process_id="process-stage2-v3",
        )
        review = self.store.fetch_one(
            "SELECT field_scores_json FROM reviews WHERE process_id = ?",
            (process.process_id,),
        )
        scores = json.loads(review["field_scores_json"])
        location = scores["summary_location"]

        self.assertEqual(
            V5_FIELD_PROCESSING_RULE_VERSION,
            self.store.fetch_one(
                "SELECT rule_version FROM process_runs WHERE process_id = ?",
                (process.process_id,),
            )["rule_version"],
        )
        self.assertEqual(1.0, scores["social_urls"]["completeness_score"])
        self.assertEqual(1.0, scores["social_urls"]["data_completeness_score"])
        self.assertEqual(1.0, scores["social_urls"]["baseline_coverage_score"])
        self.assertEqual(1.0, scores["social_urls"]["accuracy_score"])
        self.assertEqual("summary_location", location["baseline_field_key"])
        self.assertEqual(1.0, location["completeness_score"])
        # 地点返回了真实资料，即使语义准确度未达阈值，资料完整度也必须为 1。
        self.assertEqual(1.0, location["data_completeness_score"])
        self.assertEqual("WARNING", location["comparison_status"])
        self.assertEqual("BELOW_SIMILARITY_THRESHOLD", location["reason_code"])

        metrics_v4 = self.service.calculate_process_metrics(process.process_id)
        report_v4 = self.service.build_report_model(
            report_id="report-stage4-v4",
            candidate_process_id=process.process_id,
        )
        self.assertEqual("metrics-v4", metrics_v4["metrics_rule_version"])
        self.assertEqual(
            1,
            metrics_v4["baseline_quality_metrics"]["primary_hit_query_count"],
        )
        self.assertEqual(
            1.0,
            metrics_v4["baseline_quality_metrics"]["modules"]["Social"]
            ["completeness"]["value"],
        )
        self.assertEqual(
            "NOT_APPLICABLE",
            metrics_v4["baseline_quality_metrics"]["modules"]["Insights"]
            ["accuracy"]["status"],
        )
        self.assertEqual(
            "report-model-v5",
            report_v4["metadata"]["report_model_version"],
        )
        self.assertIn("non_hit_data_return", report_v4)
        self.assertIn("regression_metrics", report_v4)
        self.assertTrue(
            REPORT_MODEL_V5_REQUIRED_TOP_LEVEL_KEYS.issubset(report_v4)
        )
        explorer = report_v4["query_explorer"]
        self.assertEqual(1, explorer["total_query_count"])
        self.assertEqual(2, explorer["total_candidate_count"])
        self.assertEqual(5, explorer["initial_query_count"])
        self.assertEqual(10, explorer["load_more_query_count"])
        self.assertEqual(
            "衡量系统能否在全部有效 Query 中找到目标人物",
            report_v4["overview"]["core_metrics"]["retrieval_success"][
                "purpose"
            ],
        )
        self.assertEqual(
            1,
            report_v4["overview"]["core_metrics"]["retrieval_success"][
                "breakdown_count"
            ],
        )
        processing_fields = {
            item["field_key"] for item in report_v4["processing_scope"]["fields"]
        }
        self.assertIn("social_urls", processing_fields)
        self.assertIn("summary_location", processing_fields)
        self.assertIn("profile_full_name", processing_fields)
        economics = report_v4["execution_economics"]
        self.assertEqual(1, economics["query_count"])
        self.assertEqual(
            [
                "pdl_person_identify", "pdl_person_search", "llm_search",
                "wiki", "google_lens", "google_vision",
                "social_profile_extraction",
            ],
            [item["key"] for item in economics["tool_rows"]],
        )
        module_overview = report_v4["module_return_overview"]
        modules_by_name = {
            item["module"]: item for item in module_overview["modules"]
        }
        self.assertEqual(
            ["Insights", "Photos", "Profile", "Social", "Summary"],
            module_overview["module_order"],
        )
        self.assertEqual(
            1,
            module_overview["populations"]["hit"]["detail_success_count"],
        )
        self.assertEqual(
            1,
            module_overview["populations"]["nonmatched"][
                "detail_success_count"
            ],
        )
        self.assertEqual(
            ["profile_full_name"],
            modules_by_name["Profile"]["field_keys"],
        )
        self.assertEqual(
            1.0,
            modules_by_name["Profile"]["populations"]["hit"][
                "module_return_rate"
            ],
        )
        self.assertEqual(
            1.0,
            modules_by_name["Social"]["populations"]["nonmatched"][
                "field_completeness"
            ],
        )
        query_item = explorer["items"][0]
        self.assertTrue(
            REPORT_MODEL_V5_REQUIRED_QUERY_KEYS.issubset(query_item)
        )
        candidate_item = query_item["candidate_run"]["candidates"][0]
        self.assertTrue(
            REPORT_MODEL_V5_REQUIRED_CANDIDATE_KEYS.issubset(
                candidate_item
            )
        )
        self.assertEqual(
            "candidate-stage2-v3",
            candidate_item["candidate_id"],
        )
        # Social 完全一致、地点为语义近似，详情快照必须复用既有 0.7 准确度，
        # 不能在报告层重新计算为 1.0。
        self.assertEqual(0.7, candidate_item["metrics"]["matched_accuracy"]["value"])
        self.assertEqual(
            1.0,
            candidate_item["metrics"]["matched_completeness"]["value"],
        )
        quality = report_v4["quality_metrics"]
        self.assertEqual(1.0, quality["candidate_return_rate"]["value"])
        self.assertEqual(1.0, quality["retrieval_success"]["value"])
        self.assertEqual(1.0, quality["conditional_hit_rate"]["value"])
        self.assertEqual(3.0, quality["matched_completeness"]["numerator"])
        self.assertEqual(3, quality["matched_completeness"]["denominator"])
        completeness_detail = report_v4["overview"]["core_metrics"][
            "matched_completeness"
        ]["breakdown"][0]
        self.assertEqual(3.0, completeness_detail["numerator"])
        self.assertEqual(3, completeness_detail["denominator"])
        completeness_groups = report_v4["overview"]["core_metrics"][
            "matched_completeness"
        ]["population_groups"]
        self.assertEqual(
            ["all_hit", "primary_hit", "secondary_hit"],
            [item["key"] for item in completeness_groups],
        )
        self.assertEqual(1, completeness_groups[0]["candidate_count"])
        self.assertTrue(completeness_groups[1]["formal"])
        self.assertEqual(0, completeness_groups[2]["candidate_count"])
        accuracy_detail = report_v4["overview"]["core_metrics"][
            "matched_accuracy"
        ]
        self.assertIn("70.0000%", accuracy_detail["calculation_expression"])
        self.assertAlmostEqual(
            0.7,
            accuracy_detail["auxiliary_calculation"]["value"],
        )
        accuracy_primary_group = next(
            item
            for item in accuracy_detail["population_groups"]
            if item["key"] == "primary_hit"
        )
        self.assertEqual("候选人宏平均", accuracy_primary_group["calculation_label"])
        self.assertAlmostEqual(
            0.7,
            accuracy_primary_group["auxiliary_calculation"]["value"],
        )
        self.assertEqual(
            1,
            report_v4["confidence_metrics"]["all_hit"]["UNKNOWN"],
        )
        non_hit_candidate = next(
            item for item in query_item["candidate_run"]["candidates"]
            if item["identity"]["judgement"] == "NOT_HIT"
        )
        # 非命中资料完整度仍统计姓名是否返回；候选人资料相似度排除
        # profile_full_name，只保留 Social 冲突 0 分与地点字段 1 分。
        self.assertEqual(
            1.0,
            non_hit_candidate["metrics"]["nonmatched_data_completeness"]["value"],
        )
        self.assertEqual(
            3.0,
            report_v4["quality_metrics"]["nonmatched_data_completeness"][
                "numerator"
            ],
        )
        self.assertEqual(
            3,
            report_v4["quality_metrics"]["nonmatched_data_completeness"][
                "denominator"
            ],
        )
        self.assertEqual(
            0.5,
            non_hit_candidate["metrics"]["nonmatched_baseline_overlap"]["value"],
        )
        self.assertEqual(
            1.0,
            report_v4["quality_metrics"]["nonmatched_baseline_overlap"][
                "numerator"
            ],
        )
        self.assertEqual(
            2,
            report_v4["quality_metrics"]["nonmatched_baseline_overlap"][
                "denominator"
            ],
        )
        similarity_components = report_v4["overview"]["core_metrics"][
            "nonmatched_baseline_overlap"
        ]["breakdown"][0]["field_components"]
        self.assertNotIn(
            "profile_full_name",
            {item["field_key"] for item in similarity_components},
        )
        self.assertIsNone(report_v4["threshold_assessment"])

        original_execute = self.service.execute_run
        self.service.execute_run = lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("无成本重处理不得调用检索接口")
        )
        try:
            reprocessed = self.service.reprocess_existing_run(
                run_id=imported.object_id,
                schema_version=schema_version,
                baseline_version="baseline-stage2-v3",
                process_id="process-stage2-v3-reprocessed",
            )
        finally:
            self.service.execute_run = original_execute
        self.assertEqual(2, reprocessed.candidate_count)
        self.assertEqual(
            raw_count_before,
            self.store.fetch_one(
                "SELECT COUNT(*) AS count FROM raw_records WHERE run_id = ?",
                (imported.object_id,),
            )["count"],
        )
        # 报告创建后，完整明细必须直接固化到 reports.metrics_json；不依赖
        # 后续页面查询，也不能把原始接口 payload 写入报告快照。
        single_report = self.service.create_report(
            candidate_process_id=process.process_id,
            data_marker="MOCK",
            report_id="report-stage2-v5-single",
        )
        stored_row = self.store.fetch_one(
            "SELECT metrics_json FROM reports WHERE report_id = ?",
            (single_report.report_id,),
        )
        stored_model = json.loads(stored_row["metrics_json"])
        stored_candidate = stored_model["query_explorer"]["items"][0][
            "candidate_run"
        ]["candidates"][0]
        self.assertEqual(
            V5_REPORT_MODEL_VERSION,
            stored_model["metadata"]["report_model_version"],
        )
        self.assertTrue(stored_candidate["references"]["raw_ids"])
        self.assertNotIn("detail_data_raw", stored_candidate)

        # 双 Process 报告必须使用独立对比模型，不能继续套用单 Run v5 主体。
        compare_report = self.service.create_report(
            candidate_process_id=reprocessed.process_id,
            baseline_process_id=process.process_id,
            data_marker="MOCK",
            report_id="report-stage2-v6-compare",
        )
        compare_model = compare_report.model
        self.assertEqual(
            COMPARE_REPORT_MODEL_VERSION,
            compare_model["metadata"]["report_model_version"],
        )
        comparison_report = compare_model["comparison_report"]
        self.assertIn("core_metric_changes", comparison_report)
        self.assertIn("execution_cost_comparison", comparison_report)
        self.assertEqual(
            ["Insights", "Photos", "Profile", "Social", "Summary"],
            [item["module"] for item in comparison_report["module_changes"]],
        )
        self.assertIn("confidence_comparison", comparison_report)
        self.assertIn("tool_call_comparison", comparison_report)
        self.assertIn(
            "execution_economics_comparison",
            comparison_report,
        )
        self.assertIn("appendix", comparison_report)

    def test_v5_completeness_does_not_require_baseline_compare(self):
        """完整度开关独立于基准对比，避免字段出现“完整度未生成”。

        功能说明：复现 Summary Web Links 已提取、完整度开启、基准对比关闭
        的配置组合。候选人有值时完整度为 1，空值时为 0；准确度保持不适用，
        不会因为没有 Baseline 对比而遗漏整个字段评分。

        返回值：无；断言通过表示两个配置开关的职责被明确隔离。
        异常说明：字段定义无法通过 v3 校验或评分结构变化时测试失败。
        """

        definition = next(
            dict(item)
            for item in DEFAULT_FIELD_DEFINITIONS_V3
            if item["field_key"] == "summary_web_links"
        )
        definition.update({
            "baseline_compare_enabled": False,
            "completeness_enabled": True,
            "accuracy_enabled": True,
            "scoring_role": ["completeness", "accuracy"],
        })
        definitions = validate_field_definitions([definition])
        present = _field_comparison_scores_v3(
            definitions,
            {"summary_web_links": ["https://example.test/profile"]},
            {"summary_web_links": False},
            {"summary_web_links": ["https://baseline.test/profile"]},
            {"summary_web_links"},
        )["summary_web_links"]
        empty = _field_comparison_scores_v3(
            definitions,
            {"summary_web_links": []},
            {"summary_web_links": True},
            {"summary_web_links": ["https://baseline.test/profile"]},
            {"summary_web_links"},
        )["summary_web_links"]

        self.assertEqual(1.0, present["completeness_score"])
        self.assertIsNone(present["accuracy_score"])
        self.assertEqual(
            "BASELINE_COMPARISON_DISABLED", present["reason_code"]
        )
        self.assertEqual(0.0, empty["completeness_score"])


    def test_stage4_confirmed_no_hit_has_zero_retrieval_and_nonmatched_quality(self):
        """确认无 HIT 后检索成功率为0，非命中完整度仍正常计算。"""

        scenario = self._build_data_processing_phase0_scenario()
        process_id = scenario["process_id"]
        with self.store.transaction() as connection:
            for index in range(1, 11):
                connection.execute(
                    """
                    UPDATE run_queries SET person_id = ?,
                        person_id_source = 'MANUAL'
                    WHERE run_id = 'run-data-processing-phase0'
                      AND query_id = ?
                    """,
                    (
                        f"person-phase0-{index:02d}",
                        f"phase0-query-{index:02d}",
                    ),
                )
        for index in range(1, 11):
            query_id = f"phase0-query-{index:02d}"
            context = self.service.get_query_classification_context(
                process_id,
                query_id,
            )
            classifications = [
                {
                    "candidate_pk": candidate["candidate_pk"],
                    "judgement": "NOT_HIT",
                    "reason": "MANUAL",
                    "evidence": "阶段4确认非目标人物",
                }
                for candidate in context["candidates"]
                if candidate["detail_status"] == "SUCCESS"
            ]
            self.service.save_query_classification(
                process_id,
                query_id,
                classifications,
                primary_hit_candidate_pk=None,
                confirm_no_hit=True,
                reviewer="stage4",
                review_note="批量确认无命中",
                expected_versions={
                    candidate["candidate_pk"]: (
                        candidate["reviewed_at"] or ""
                    )
                    for candidate in context["candidates"]
                },
            )
        metrics = self.service.calculate_process_metrics(process_id)
        self.assertEqual("READY", metrics["retrieval_success"]["status"])
        self.assertEqual(0.0, metrics["retrieval_success"]["value"])
        self.assertEqual(
            "NOT_APPLICABLE",
            metrics["matched_completeness"]["status"],
        )
        self.assertIn(
            "NO_HIT_CONFIRMED",
            metrics["matched_completeness"]["reason_codes"],
        )
        self.assertEqual(
            "READY",
            metrics["nonmatched_completeness"]["status"],
        )
        self.assertEqual(
            45,
            metrics["nonmatched_completeness"]["denominator"],
        )

    def _build_person_link_phase1_scenario(self) -> dict:
        """创建阶段1人物关联、同名建议、审计和报告过期测试场景。"""

        dataset_path = self.root / "phase1-tasks.jsonl"
        write_jsonl(
            dataset_path,
            [
                {
                    "input_id": "query-alice",
                    "query_stage": "FULL_NAME",
                    "clues": [
                        {
                            "type": "FULL_NAME",
                            "full_name_query": {
                                "full_name": " Alice  Example "
                            },
                        }
                    ],
                    "additional_details": [],
                },
                {
                    "input_id": "query-bob",
                    "query_stage": "FULL_NAME_SOCIAL",
                    "clues": [
                        {"type": "FULL_NAME", "value": "Bob Example"},
                        {
                            "type": "SOCIAL_LINK",
                            "value": "https://social.example.test/bob",
                        },
                    ],
                    "additional_details": [],
                },
                {
                    "input_id": "query-charlie",
                    "query_stage": "FULL_NAME",
                    "clues": [
                        {"type": "FULL_NAME", "value": "Charlie Missing"}
                    ],
                    "additional_details": [],
                },
            ],
        )
        dataset = self.service.import_dataset_jsonl(
            dataset_path,
            name="阶段1人物关联 Dataset",
            dataset_id="dataset-person-links",
        )
        baseline_path = self.root / "phase1-baseline.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": "person-alice",
                    "display_name": "Alice Example",
                    "fields": {"summary_display_name": "Alice Example"},
                    "baseline_available_fields": ["summary_display_name"],
                },
                {
                    "person_id": "person-bob-1",
                    "display_name": "Bob Example",
                    "fields": {"summary_display_name": "Bob Example"},
                    "baseline_available_fields": ["summary_display_name"],
                },
                {
                    "person_id": "person-bob-2",
                    "display_name": "Bob Example",
                    "fields": {"summary_display_name": "Bob Example"},
                    "baseline_available_fields": ["summary_display_name"],
                },
            ],
        )
        baseline = self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段1人物关联 Baseline",
            baseline_version="baseline-person-links",
        )
        run_id = self.service.create_execution_run(
            evaluation_id="eval-import",
            dataset_id=dataset.object_id,
            run_label="阶段1人物关联 Run",
            system_version="phase1",
            evaluation_phase="PHASE_1_BASELINE",
            run_id="run-person-links",
        )
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE runs SET status = 'COMPLETED'
                WHERE run_id = ?
                """,
                (run_id,),
            )
            connection.execute(
                """
                INSERT INTO raw_records(
                    raw_id, run_id, query_id, candidate_pk, stage,
                    sequence_no, payload_json, collected_at
                ) VALUES (
                    'raw-person-link', ?, 'query-alice', NULL,
                    'Input', 1, '{"frozen":true}', 'now'
                )
                """,
                (run_id,),
            )
        schema_version = self.service.ensure_default_field_schema()
        process = self.service.process_run(
            run_id=run_id,
            schema_version=schema_version,
            baseline_version=baseline.object_id,
            process_id="process-person-links",
        )
        with self.store.transaction() as connection:
            connection.execute(
                """
                INSERT INTO reports(
                    report_id, evaluation_id, baseline_process_id,
                    candidate_process_id, report_type, status, metrics_json,
                    html_file, excel_file, created_at
                ) VALUES (
                    'report-person-links', 'eval-import', NULL, ?,
                    'SINGLE', 'READY', '{}', 'phase1/report.html', NULL, 'now'
                )
                """,
                (process.process_id,),
            )
        return {
            "dataset_id": dataset.object_id,
            "baseline_version": baseline.object_id,
            "run_id": run_id,
            "process_id": process.process_id,
        }

    def test_stage1_person_link_context_only_suggests_unique_exact_names(self):
        """唯一规范化姓名可建议；同名和无匹配不自动选中。"""

        scenario = self._build_person_link_phase1_scenario()
        context = self.service.get_run_person_link_context(
            scenario["run_id"],
            scenario["baseline_version"],
        )
        queries = {item["query_id"]: item for item in context["queries"]}

        self.assertEqual(
            {
                "query_count": 3,
                "linked_count": 0,
                "unlinked_count": 3,
                "invalid_count": 0,
                "unique_suggestion_count": 1,
            },
            context["summary"],
        )
        self.assertEqual("Alice Example", queries["query-alice"]["query_name"])
        self.assertEqual(
            ["person-alice"],
            [
                item["person_id"]
                for item in queries["query-alice"]["suggestions"]
            ],
        )
        self.assertTrue(queries["query-alice"]["has_unique_suggestion"])
        self.assertEqual(
            ["person-bob-1", "person-bob-2"],
            [
                item["person_id"]
                for item in queries["query-bob"]["suggestions"]
            ],
        )
        self.assertFalse(queries["query-bob"]["has_unique_suggestion"])
        self.assertEqual([], queries["query-charlie"]["suggestions"])
        self.assertIsNone(queries["query-alice"]["current_person_id"])

    def test_stage1_person_link_update_is_atomic_audited_and_stales_report(self):
        """批量修改与 Dataset 同步同事务完成，Raw 不变且报告过期。"""

        scenario = self._build_person_link_phase1_scenario()
        result = self.service.update_run_query_person_links(
            scenario["run_id"],
            scenario["baseline_version"],
            [
                {
                    "query_id": "query-alice",
                    "expected_person_id": None,
                    "person_id": "person-alice",
                },
                {
                    "query_id": "query-bob",
                    "expected_person_id": None,
                    "person_id": "person-bob-1",
                },
            ],
            sync_dataset=True,
            note="阶段1批量关联",
        )

        self.assertEqual(2, result["updated_count"])
        self.assertEqual(2, result["dataset_synced_count"])
        run_links = {
            row["query_id"]: (row["person_id"], row["person_id_source"])
            for row in self.store.fetch_all(
                """
                SELECT query_id, person_id, person_id_source
                FROM run_queries WHERE run_id = ?
                ORDER BY query_id
                """,
                (scenario["run_id"],),
            )
        }
        self.assertEqual(
            ("person-alice", "MANUAL_RUN"),
            run_links["query-alice"],
        )
        self.assertEqual(
            ("person-bob-1", "MANUAL_RUN"),
            run_links["query-bob"],
        )
        dataset_link = self.store.fetch_one(
            """
            SELECT person_id, person_id_source FROM dataset_queries
            WHERE dataset_id = ? AND query_id = 'query-alice'
            """,
            (scenario["dataset_id"],),
        )
        self.assertEqual("person-alice", dataset_link["person_id"])
        self.assertEqual("MANUAL_DATASET", dataset_link["person_id_source"])
        history = self.store.fetch_all(
            """
            SELECT * FROM run_query_person_history
            WHERE run_id = ? ORDER BY query_id
            """,
            (scenario["run_id"],),
        )
        self.assertEqual(2, len(history))
        self.assertTrue(
            all(row["change_source"] == "MANUAL_BULK" for row in history)
        )
        self.assertTrue(all(row["sync_dataset"] == 1 for row in history))
        self.assertTrue(all(row["note"] == "阶段1批量关联" for row in history))
        self.assertEqual(
            "STALE",
            self.store.fetch_one(
                """
                SELECT status FROM reports
                WHERE report_id = 'report-person-links'
                """
            )["status"],
        )
        self.assertEqual(
            '{"frozen":true}',
            self.store.fetch_one(
                """
                SELECT payload_json FROM raw_records
                WHERE raw_id = 'raw-person-link'
                """
            )["payload_json"],
        )

    def test_stage1_person_link_clear_conflict_and_batch_failure_are_safe(self):
        """清除写审计；旧页面冲突和非法批次均不会发生部分更新。"""

        scenario = self._build_person_link_phase1_scenario()
        self.service.update_run_query_person_links(
            scenario["run_id"],
            scenario["baseline_version"],
            [
                {
                    "query_id": "query-alice",
                    "expected_person_id": None,
                    "person_id": "person-alice",
                }
            ],
        )
        self.service.update_run_query_person_links(
            scenario["run_id"],
            scenario["baseline_version"],
            [
                {
                    "query_id": "query-alice",
                    "expected_person_id": "person-alice",
                    "person_id": "",
                }
            ],
            note="清除错误关联",
        )
        cleared = self.store.fetch_one(
            """
            SELECT person_id, person_id_source FROM run_queries
            WHERE run_id = ? AND query_id = 'query-alice'
            """,
            (scenario["run_id"],),
        )
        self.assertIsNone(cleared["person_id"])
        self.assertEqual("UNSPECIFIED", cleared["person_id_source"])
        self.assertEqual(
            "CLEAR_LINK",
            self.store.fetch_one(
                """
                SELECT change_source FROM run_query_person_history
                WHERE run_id = ? AND query_id = 'query-alice'
                ORDER BY changed_at DESC, rowid DESC LIMIT 1
                """,
                (scenario["run_id"],),
            )["change_source"],
        )

        with self.assertRaisesRegex(
            PersonLinkValidationError,
            "query-alice.*已被其他页面修改",
        ):
            self.service.update_run_query_person_links(
                scenario["run_id"],
                scenario["baseline_version"],
                [
                    {
                        "query_id": "query-alice",
                        "expected_person_id": "person-alice",
                        "person_id": "person-bob-1",
                    }
                ],
            )
        with self.assertRaises(PersonLinkValidationError):
            self.service.update_run_query_person_links(
                scenario["run_id"],
                scenario["baseline_version"],
                [
                    {
                        "query_id": "query-alice",
                        "expected_person_id": None,
                        "person_id": "person-alice",
                    },
                    {
                        "query_id": "query-bob",
                        "expected_person_id": None,
                        "person_id": "missing-person",
                    },
                ],
            )
        self.assertIsNone(
            self.store.fetch_one(
                """
                SELECT person_id FROM run_queries
                WHERE run_id = ? AND query_id = 'query-alice'
                """,
                (scenario["run_id"],),
            )["person_id"]
        )

    def test_dataset_jsonl_import_preview_duplicate_and_atomic_failure(self):
        """Dataset 导入支持预览、SHA 去重，坏记录不会污染事务。"""

        source = self.root / "tasks.jsonl"
        write_jsonl(
            source,
            [
                {
                    "input_id": "query-name",
                    "person_id": "person-1",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME"}],
                    "additional_details": [],
                    "tags": ["baseline"],
                },
                {
                    "input_id": "query-social",
                    "person_id": "person-1",
                    "query_stage": "FULL_NAME_SOCIAL",
                    "clues": [
                        {"type": "FULL_NAME"},
                        {"type": "SOCIAL_LINK"},
                    ],
                    "additional_details": [],
                },
            ],
        )

        preview = self.service.preview_dataset_jsonl(source)
        result = self.service.import_dataset_jsonl(
            source,
            name="测试 Dataset",
            dataset_id="dataset-jsonl",
        )

        self.assertEqual(2, preview.valid_count)
        self.assertEqual([], preview.errors)
        self.assertEqual(2, result.imported_count)
        self.assertEqual(
            2,
            self.store.fetch_one(
                "SELECT COUNT(*) AS count FROM dataset_queries"
            )["count"],
        )
        self.assertTrue((self.root / "data" / result.archived_files[0]).is_file())
        self.assertIn("tasks.jsonl", Path(result.archived_files[0]).name)
        with self.assertRaises(DuplicateImportError):
            self.service.import_dataset_jsonl(
                source,
                name="重复 Dataset",
                dataset_id="dataset-duplicate",
            )

        invalid = self.root / "invalid-tasks.jsonl"
        write_jsonl(
            invalid,
            [
                {
                    "input_id": "valid-before-error",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME"}],
                },
                {
                    "input_id": "bad-social",
                    "query_stage": "FULL_NAME_SOCIAL",
                    "clues": [{"type": "FULL_NAME"}],
                },
            ],
        )
        invalid_preview = self.service.preview_dataset_jsonl(invalid)
        with self.assertRaises(ImportValidationError):
            self.service.import_dataset_jsonl(
                invalid,
                name="坏 Dataset",
                dataset_id="dataset-invalid",
            )

        self.assertEqual(1, len(invalid_preview.errors))
        self.assertIsNone(
            self.store.fetch_one(
                "SELECT dataset_id FROM datasets WHERE dataset_id = ?",
                ("dataset-invalid",),
            )
        )

    def test_field_path_and_builtin_normalizers_are_restricted_and_stable(self):
        """路径仅支持点号/索引/通配，内置转换器产生稳定结果。"""

        source = {
            "ui_sections": {
                "insights": {
                    "data": {
                        "items": [
                            {"description": " first "},
                            {"description": "second"},
                        ]
                    }
                },
                "social": {
                    "data": {
                        "profiles": [
                            {
                                "url": (
                                    " HTTPS://WWW.Example.COM/user/"
                                    "?utm_source=test&tab=posts#bio "
                                )
                            },
                            {"url": "https://example.com/other/"},
                        ]
                    }
                },
                "profile": {
                    "data": {
                        "sections": [
                            {
                                "title": "Identity",
                                "items": [
                                    {"label": "Full Name", "value": "Example"},
                                    {"label": "Location", "value": "Shanghai"},
                                ],
                            }
                        ]
                    }
                },
            }
        }

        self.assertEqual(
            "first",
            normalize_field_value(
                extract_source_path(
                    source,
                    "ui_sections.insights.data.items[0].description",
                ),
                "trim_text",
            ),
        )
        urls = extract_source_path(
            source,
            "ui_sections.social.data.profiles[*].url",
        )
        self.assertEqual(
            [
                "https://example.com/user?tab=posts",
                "https://example.com/other",
            ],
            normalize_field_value(urls, "social_url"),
        )
        self.assertEqual(75.0, normalize_field_value(0.75, "percentage"))
        # Summary 时间字段可能以 Unix 毫秒数返回；trim_text 应保留该标量，
        # 不能将真实接口响应记录为字段处理错误。
        self.assertEqual(
            "1785210772179",
            normalize_field_value(1785210772179, "trim_text"),
        )
        self.assertEqual(
            {
                "Identity": {
                    "Full Name": "Example",
                    "Location": "Shanghai",
                }
            },
            normalize_field_value(
                extract_source_path(
                    source,
                    "ui_sections.profile.data.sections",
                ),
                "profile_sections",
            ),
        )
        with self.assertRaises(FieldSchemaValidationError):
            extract_source_path(source, "ui_sections.social.data.profiles[?].url")
        self.assertIsNone(
            extract_source_path(
                {
                    "ui_sections": {
                        "insights": {"data": {"items": []}},
                    }
                },
                "ui_sections.insights.data.items[0].description",
                missing_policy="EMPTY",
            )
        )
        self.assertIsNone(
            extract_source_path(
                {
                    "ui_sections": {
                        "summary": {"data": {"primary_image": None}},
                    }
                },
                "ui_sections.summary.data.primary_image.url",
                missing_policy="EMPTY",
            )
        )
        self.assertEqual(
            [],
            extract_source_path(
                {
                    "ui_sections": {
                        "social": {"data": {"profiles": []}},
                    }
                },
                "ui_sections.social.data.profiles[*].url",
                missing_policy="EMPTY",
            ),
        )
        with self.assertRaises(FieldSchemaValidationError):
            extract_source_path(
                {
                    "ui_sections": {
                        "insights": {"data": {"items": "not-an-array"}},
                    }
                },
                "ui_sections.insights.data.items[0].description",
                missing_policy="EMPTY",
            )

    def test_field_schema_v2_infers_old_config_scope_and_missing_policy(self):
        """旧配置自动补齐作用域与缺失策略，非法枚举仍拒绝发布。"""

        base_definition = {
            "field_key": "llm_cost",
            "display_name": "LLM Cost",
            "module": "Task",
            "source_stage": "GetTask",
            "source_path": "task.llm_cost",
            "data_type": "number",
            "array_mode": "preserve",
            "empty_rule": "default",
            "normalizer": "number",
            "scoring_role": ["display"],
            "compare_mode": "exact",
            "enabled": True,
            "sort_order": 10,
        }
        normalized = validate_field_definitions(
            [
                base_definition,
                {
                    **base_definition,
                    "field_key": "summary_name",
                    "display_name": "Summary Name",
                    "module": "Summary",
                    "source_stage": "GetTaskCandidateDetail",
                    "source_path": "ui_sections.summary.data.display_name",
                    "data_type": "string",
                    "normalizer": "trim_text",
                    "sort_order": 20,
                },
            ]
        )

        self.assertEqual("QUERY", normalized[0]["value_scope"])
        self.assertEqual("EMPTY", normalized[0]["missing_policy"])
        self.assertEqual("CANDIDATE", normalized[1]["value_scope"])
        self.assertEqual("EMPTY", normalized[1]["missing_policy"])
        with self.store.transaction() as connection:
            connection.execute(
                """
                INSERT INTO field_schemas(
                    schema_version, name, definitions_json, created_by,
                    created_at, is_active
                ) VALUES (
                    'field-schema-default-v1', '默认字段配置', ?,
                    'system', '2026-01-01T00:00:00+00:00', 1
                )
                """,
                (json.dumps(normalized, ensure_ascii=False),),
            )
        v2_version = self.service.ensure_default_field_schema()
        self.assertEqual("field-schema-default-v2", v2_version)
        self.assertEqual(
            "field-schema-default-v2",
            self.store.fetch_one(
                "SELECT schema_version FROM field_schemas WHERE is_active = 1"
            )["schema_version"],
        )
        with self.assertRaises(FieldSchemaValidationError):
            validate_field_definitions(
                [{**base_definition, "value_scope": "RUN"}]
            )
        with self.assertRaises(FieldSchemaValidationError):
            validate_field_definitions(
                [{**base_definition, "missing_policy": "IGNORE"}]
            )

    def test_process_run_snapshots_query_fields_and_treats_optional_ui_as_empty(self):
        """GetTask 公共字段按 Query 落库，可选 UI 缺失不再形成错误。"""

        schema_version = self.service.ensure_default_field_schema()
        results_path = self.root / "query-field-results.jsonl"
        write_jsonl(
            results_path,
            [
                {
                    "result_schema_version": "1.3",
                    "input_id": "query-fields-v2",
                    "person_id": "person-fields-v2",
                    "task_id": "task-fields-v2",
                    "query_stage": "FULL_NAME",
                    "query_status": "SUCCESS",
                    "candidate_count_total": 1,
                    "candidate_count_listed": 1,
                    "detail_success_count": 1,
                    "detail_failure_count": 0,
                    "raw": {
                        "create_intent_task": {
                            "sequence_no": 1,
                            "response_body": api_body(
                                {"task_id": "task-fields-v2"}
                            ),
                        },
                        "get_task_history": [
                            {
                                "sequence_no": 1,
                                "response_body": api_body(
                                    {
                                        "status": "SUCCEEDED",
                                        "task_id": "task-fields-v2",
                                        "llm_cost": 1.25,
                                        "third_party_cost": 2.5,
                                        "total_cost": 3.75,
                                        "pdl_called": True,
                                        "search_duration_ms": 4321,
                                    }
                                ),
                            }
                        ],
                        "list_task_candidates": {
                            "sequence_no": 1,
                            "response_body": api_body({"items": []}),
                        },
                    },
                    "results": [
                        {
                            "candidate_rank": 1,
                            "candidate_id": "candidate-fields-v2",
                            "rank_score": 0.95,
                            "detail_status": "SUCCESS",
                            "detail_error": "",
                            "list_item_raw": {
                                "candidate_id": "candidate-fields-v2"
                            },
                            "detail_data_raw": {
                                "ui_sections": {
                                    "insights": {
                                        "status": "empty",
                                        "data": {"items": []},
                                    },
                                    "summary": {
                                        "status": "data",
                                        "data": {
                                            "confidence_level": "HIGH",
                                            "primary_image": None,
                                        },
                                    },
                                }
                            },
                            "ui_sections": {
                                "insights": {
                                    "status": "empty",
                                    "data": {"items": []},
                                },
                                "summary": {
                                    "status": "data",
                                    "data": {
                                        "confidence_level": "HIGH",
                                        "primary_image": None,
                                    },
                                },
                            },
                        }
                    ],
                }
            ],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="query fields v2",
            system_version="v2",
            run_id="run-query-fields-v2",
        )

        processed = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
        )
        query_row = self.store.fetch_one(
            "SELECT * FROM processed_queries WHERE process_id = ?",
            (processed.process_id,),
        )
        candidate_row = self.store.fetch_one(
            """
            SELECT pc.* FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ? AND c.candidate_id = ?
            """,
            (processed.process_id, "candidate-fields-v2"),
        )
        query_fields = json.loads(query_row["fields_json"])
        candidate_fields = json.loads(candidate_row["fields_json"])
        candidate_empty = json.loads(candidate_row["empty_fields_json"])

        self.assertEqual("HAS_CANDIDATES", query_row["result_status"])
        self.assertEqual(1.25, query_fields["llm_cost"])
        self.assertEqual(2.5, query_fields["third_party_cost"])
        self.assertEqual(4321, query_fields["search_duration_ms"])
        self.assertEqual("HIGH", candidate_fields["candidate_confidence"])
        self.assertIsNone(candidate_fields["insights_description"])
        self.assertEqual([], candidate_fields["insights_links"])
        self.assertIsNone(candidate_fields["summary_primary_image_url"])
        self.assertTrue(candidate_empty["insights_description"])
        self.assertTrue(candidate_empty["insights_links"])
        self.assertTrue(candidate_empty["summary_primary_image_url"])
        self.assertEqual([], json.loads(candidate_row["processing_errors_json"]))
        self.assertEqual(0, processed.error_count)

    def test_field_schema_versions_process_and_reprocess_without_raw_mutation(self):
        """配置只新增版本；重新处理保留旧结果，字段错误不修改 Raw。"""

        default_version = self.service.ensure_default_field_schema()
        custom_definitions = [
            {
                "field_key": "summary_name",
                "display_name": "Summary Name",
                "module": "Summary",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.summary.data.display_name",
                "data_type": "string",
                "array_mode": "preserve",
                "empty_rule": "default",
                "normalizer": "trim_text",
                "scoring_role": ["display"],
                "compare_mode": "normalized_text",
                "enabled": True,
                "sort_order": 10,
            },
            {
                "field_key": "social_urls",
                "display_name": "Social URLs",
                "module": "Social",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.social.data.profiles[*].url",
                "data_type": "array",
                "array_mode": "collect",
                "empty_rule": "default",
                "normalizer": "social_url",
                "scoring_role": ["identity", "completeness", "accuracy"],
                "compare_mode": "url_set",
                "enabled": True,
                "sort_order": 20,
            },
            {
                "field_key": "future_kept",
                "display_name": "Future Kept",
                "module": "Candidate",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "future_field.kept",
                "data_type": "boolean",
                "array_mode": "preserve",
                "empty_rule": "default",
                "normalizer": "identity",
                "scoring_role": ["display"],
                "compare_mode": "exact",
                "enabled": True,
                "sort_order": 30,
            },
        ]
        custom_version = self.service.publish_field_schema(
            name="测试字段配置",
            definitions=custom_definitions,
            created_by="tester",
            schema_version="field-schema-test-001",
        )
        with self.assertRaises(FieldSchemaValidationError):
            self.service.publish_field_schema(
                name="重复字段配置",
                definitions=[custom_definitions[0], custom_definitions[0]],
                schema_version="field-schema-invalid",
            )
        results_path = self.root / "process-results.jsonl"
        write_jsonl(
            results_path,
            [
                {
                    "result_schema_version": "1.3",
                    "input_id": "query-process",
                    "task_id": "task-process",
                    "query_stage": "FULL_NAME_SOCIAL",
                    "query_status": "PARTIAL_DETAIL_FAILED",
                    "results": [
                        {
                            "candidate_rank": 1,
                            "candidate_id": "candidate-success",
                            "detail_status": "SUCCESS",
                            "detail_error": "",
                            "detail_data_raw": {
                                "ui_sections": {
                                    "summary": {
                                        "data": {"display_name": " Example Person "}
                                    },
                                    "social": {
                                        "data": {
                                            "profiles": [
                                                {
                                                    "url": (
                                                        "https://WWW.Example.com/person/"
                                                        "?utm_source=test"
                                                    )
                                                }
                                            ]
                                        }
                                    },
                                },
                                "future_field": {"kept": True},
                            },
                            "ui_sections": {
                                "summary": {
                                    "data": {"display_name": " Example Person "}
                                },
                                "social": {
                                    "data": {
                                        "profiles": [
                                            {
                                                "url": (
                                                    "https://WWW.Example.com/person/"
                                                    "?utm_source=test"
                                                )
                                            }
                                        ]
                                    }
                                },
                            },
                        },
                        {
                            "candidate_rank": 2,
                            "candidate_id": "candidate-failed",
                            "detail_status": "FAILED",
                            "detail_error": "timeout",
                            "detail_data_raw": None,
                            "ui_sections": None,
                        },
                    ],
                }
            ],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="process",
            system_version="v1",
            run_id="run-process",
        )
        raw_before = [
            row["payload_json"]
            for row in self.store.fetch_all(
                "SELECT payload_json FROM raw_records WHERE run_id = ? ORDER BY raw_id",
                (imported.object_id,),
            )
        ]

        first_process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=custom_version,
        )
        successful = self.store.fetch_one(
            """
            SELECT pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ? AND c.candidate_id = 'candidate-success'
            """,
            (first_process.process_id,),
        )
        failed = self.store.fetch_one(
            """
            SELECT pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ? AND c.candidate_id = 'candidate-failed'
            """,
            (first_process.process_id,),
        )
        fields = json.loads(successful["fields_json"])
        self.assertEqual("Example Person", fields["summary_name"])
        self.assertEqual(
            ["https://example.com/person"],
            fields["social_urls"],
        )
        self.assertIs(True, fields["future_kept"])
        self.assertEqual({}, json.loads(failed["fields_json"]))
        self.assertEqual({}, json.loads(failed["empty_fields_json"]))
        self.assertEqual(
            "DETAIL_FAILED",
            json.loads(failed["processing_errors_json"])[0]["code"],
        )

        broken_definitions = [
            *custom_definitions,
            {
                **custom_definitions[0],
                "field_key": "future_missing",
                "display_name": "Future Missing",
                "source_path": "ui_sections.future.data.value",
                "sort_order": 40,
            },
        ]
        broken_version = self.service.publish_field_schema(
            name="包含未来字段",
            definitions=broken_definitions,
            schema_version="field-schema-test-002",
        )
        second_process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=broken_version,
        )
        raw_after = [
            row["payload_json"]
            for row in self.store.fetch_all(
                "SELECT payload_json FROM raw_records WHERE run_id = ? ORDER BY raw_id",
                (imported.object_id,),
            )
        ]

        self.assertNotEqual(first_process.process_id, second_process.process_id)
        self.assertEqual(1, second_process.error_count)
        second_successful = self.store.fetch_one(
            """
            SELECT pc.empty_fields_json, pc.processing_errors_json
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ? AND c.candidate_id = 'candidate-success'
            """,
            (second_process.process_id,),
        )
        self.assertTrue(
            json.loads(second_successful["empty_fields_json"])["future_missing"]
        )
        self.assertEqual(
            [],
            json.loads(second_successful["processing_errors_json"]),
        )
        self.assertEqual(raw_before, raw_after)
        self.assertEqual(
            2,
            self.store.fetch_one(
                "SELECT COUNT(*) AS count FROM process_runs WHERE run_id = ?",
                (imported.object_id,),
            )["count"],
        )
        self.assertEqual(
            0,
            self.store.fetch_one(
                "SELECT is_active FROM field_schemas WHERE schema_version = ?",
                (default_version,),
            )["is_active"],
        )

    def test_jsonl_import_preserves_public_fields_and_admin_raw(self):
        """历史导入应保留公共字段和三个 Admin Raw 阶段，且不发起网络请求。"""

        results_path = self.root / "public-info-results.jsonl"
        write_jsonl(
            results_path,
            [
                {
                    "result_schema_version": "1.3.1",
                    "input_id": "query-public-info",
                    "task_id": "task-public-info",
                    "query_status": "NO_CANDIDATE",
                    "result_status": "NO_CANDIDATES",
                    "task_fields": {
                        "llm_cost": None,
                        "third_party_cost": None,
                        "total_cost": None,
                        "pdl_called": None,
                        "search_duration_ms": None,
                    },
                    "public_fields": {
                        "public_info_status": "COMPLETE",
                        "cache_hit": True,
                        "field_mapping_status": "NOT_MAPPED",
                    },
                    "raw": {
                        "create_intent_task": {"stage": "CreateIntentTask"},
                        "get_task_history": [{"stage": "GetTask"}],
                        "admin_login": [{"stage": "AdminLogin"}],
                        "get_search_task_debug": {
                            "stage": "GetSearchTaskDebug"
                        },
                        "get_provider_cost_summary": {
                            "stage": "GetProviderCostSummary"
                        },
                    },
                    "results": [],
                }
            ],
        )

        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="公共信息导入",
            system_version="v1.3",
            run_id="run-public-info-import",
        )

        query = self.store.fetch_one(
            "SELECT public_fields_json FROM run_queries WHERE run_id = ?",
            (imported.object_id,),
        )
        raw_stages = {
            row["stage"]
            for row in self.store.fetch_all(
                "SELECT stage FROM raw_records WHERE run_id = ?",
                (imported.object_id,),
            )
        }
        self.assertEqual(
            "COMPLETE",
            json.loads(query["public_fields_json"])["public_info_status"],
        )
        self.assertTrue(json.loads(query["public_fields_json"])["cache_hit"])
        self.assertTrue(
            {
                "AdminLogin",
                "GetSearchTaskDebug",
                "GetProviderCostSummary",
            }.issubset(raw_stages)
        )

    def test_reprocess_backfills_confirmed_admin_fields_without_network(self):
        """无成本重处理仅从已入库 Admin Raw 补齐五个任务公共字段。"""

        results_path = self.root / "admin-mapping-results.jsonl"
        debug_body = api_body(
            {
                "debug": {
                    "task": {
                        "start_time": "2026-08-05T10:14:19.126Z",
                        "finish_time": "2026-08-05T10:14:30.721Z",
                    },
                    "diagnosis": {"pdl_called": True},
                }
            }
        )
        cost_body = api_body(
            {
                "cost_summary": {
                    "by_provider": [
                        {
                            "provider": "llm_search:deepseek",
                            "currency": "USD",
                            "total_cost_microunit": "383",
                        },
                        {
                            "provider": "people_data_labs",
                            "currency": "USD",
                            "total_cost_microunit": "1390000",
                        },
                    ],
                    "by_search": [
                        {
                            "task_id": "task-admin-mapping",
                            "currency": "USD",
                            "total_cost_microunit": "1390383",
                        }
                    ],
                    "totals": [],
                }
            }
        )
        write_jsonl(
            results_path,
            [
                {
                    "result_schema_version": "1.3.1",
                    "input_id": "query-admin-mapping",
                    "task_id": "task-admin-mapping",
                    "query_status": "NO_CANDIDATE",
                    "result_status": "NO_CANDIDATES",
                    "task_fields": {
                        "llm_cost": None,
                        "third_party_cost": None,
                        "total_cost": None,
                        "pdl_called": None,
                        "search_duration_ms": None,
                    },
                    "public_fields": {"field_mapping_status": "NOT_MAPPED"},
                    "raw": {
                        "get_search_task_debug": {
                            "response_body": debug_body,
                        },
                        "get_provider_cost_summary": {
                            "response_body": cost_body,
                        },
                    },
                    "results": [],
                }
            ],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="Admin 字段重处理",
            system_version="v1.3",
            run_id="run-admin-mapping",
        )
        schema_version = self.service.ensure_default_field_schema()
        raw_before = [
            row["payload_json"]
            for row in self.store.fetch_all(
                "SELECT payload_json FROM raw_records WHERE run_id = ? ORDER BY raw_id",
                (imported.object_id,),
            )
        ]

        process = self.service.reprocess_existing_run(
            run_id=imported.object_id,
            schema_version=schema_version,
        )

        query = self.store.fetch_one(
            """
            SELECT llm_cost, third_party_cost, total_cost, pdl_called,
                   search_duration_ms, public_fields_json
            FROM run_queries WHERE run_id = ? AND query_id = ?
            """,
            (imported.object_id, "query-admin-mapping"),
        )
        processed = self.store.fetch_one(
            "SELECT fields_json FROM processed_queries WHERE process_id = ?",
            (process.process_id,),
        )
        raw_after = [
            row["payload_json"]
            for row in self.store.fetch_all(
                "SELECT payload_json FROM raw_records WHERE run_id = ? ORDER BY raw_id",
                (imported.object_id,),
            )
        ]
        self.assertEqual(0.000383, query["llm_cost"])
        self.assertEqual(1.39, query["third_party_cost"])
        self.assertEqual(1.390383, query["total_cost"])
        self.assertEqual(1, query["pdl_called"])
        self.assertEqual(11595, query["search_duration_ms"])
        self.assertEqual(
            "COMPLETE",
            json.loads(query["public_fields_json"])["field_mapping_status"],
        )
        processed_fields = json.loads(processed["fields_json"])
        self.assertEqual(1.390383, processed_fields["total_cost"])
        self.assertIs(processed_fields["pdl_called"], True)
        self.assertEqual(raw_before, raw_after)

    def test_execution_failure_persists_collected_public_info(self):
        """失败终态仍应保存公共字段和 Admin Raw，同时保持 EXECUTION_FAILED。"""

        source = self.root / "failed-public-info-tasks.jsonl"
        write_jsonl(
            source,
            [
                {
                    "input_id": "query-failed-public",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME"}],
                }
            ],
        )
        self.service.import_dataset_jsonl(
            source,
            name="失败终态公共信息",
            dataset_id="dataset-failed-public",
        )
        run_id = self.service.create_execution_run(
            evaluation_id="eval-import",
            dataset_id="dataset-failed-public",
            run_label="candidate",
            system_version="web-v1",
            evaluation_phase="PHASE_1_BASELINE",
            run_id="run-failed-public",
        )
        error = FlowError("GetTask", "任务进入失败终态: FAILED", "task-failed")
        error.public_fields = {
            "public_info_status": "COMPLETE",
            "cache_hit": False,
        }
        error.task_fields = {
            "llm_cost": None,
            "third_party_cost": None,
            "total_cost": None,
            "pdl_called": None,
            "search_duration_ms": None,
        }
        error.raw_records = [
            {
                "stage": "GetSearchTaskDebug",
                "sequence_no": 1,
                "response": {"code": 0},
            },
            {
                "stage": "GetProviderCostSummary",
                "sequence_no": 1,
                "response": {"code": 0},
            },
        ]

        self.service._persist_execution_failure(
            run_id=run_id,
            query_id="query-failed-public",
            error=error,
        )

        query = self.store.fetch_one(
            "SELECT status, result_status, public_fields_json FROM run_queries "
            "WHERE run_id = ? AND query_id = ?",
            (run_id, "query-failed-public"),
        )
        raw_stages = {
            row["stage"]
            for row in self.store.fetch_all(
                "SELECT stage FROM raw_records WHERE run_id = ?",
                (run_id,),
            )
        }
        self.assertEqual(("FAILED", "EXECUTION_FAILED"), (query["status"], query["result_status"]))
        self.assertEqual(
            "COMPLETE",
            json.loads(query["public_fields_json"])["public_info_status"],
        )
        self.assertEqual(
            {"GetSearchTaskDebug", "GetProviderCostSummary"},
            raw_stages,
        )

    def test_execution_run_persists_progress_candidates_raw_and_files(self):
        """Web 执行复用采集核心，并按 Query 事务保存结果和全部 Raw。"""

        source = self.root / "execution-tasks.jsonl"
        write_jsonl(
            source,
            [
                {
                    "input_id": "query-execution",
                    "person_id": "person-execution",
                    "query_stage": "FULL_NAME",
                    "match_strategy": "UNION",
                    "clues": [{"type": "FULL_NAME", "value": "Example Person"}],
                    "additional_details": [],
                }
            ],
        )
        self.service.import_dataset_jsonl(
            source,
            name="执行数据集",
            dataset_id="dataset-execution",
        )
        run_id = self.service.create_execution_run(
            evaluation_id="eval-import",
            dataset_id="dataset-execution",
            run_label="candidate",
            system_version="web-v1",
            evaluation_phase="PHASE_2_POST_OPTIMIZATION",
            run_id="run-execution",
        )
        client = FakeExecutionClient(
            [
                api_body({"task_id": "task-execution"}),
                api_body({"status": "SEARCHING"}),
                api_body(
                    {
                        "status": "SUCCEEDED",
                        "candidate_count": 1,
                        "llm_cost": 1.25,
                        "third_party_cost": 2.5,
                        "total_cost": 3.75,
                        "pdl_called": True,
                        "search_duration_ms": 4321,
                    }
                ),
                api_body(
                    {
                        "items": [
                            {
                                "candidate_id": "candidate-execution",
                                "rank_score": 0.88,
                            }
                        ]
                    }
                ),
                api_body(
                    {
                        "ui_sections": {
                            "summary": {
                                "status": "data",
                                "data": {"display_name": "Example Person"},
                            }
                        },
                        "future_field": {"kept": True},
                    }
                ),
            ]
        )

        self.service.execute_run(run_id, client, sleep_fn=lambda _: None)

        run = self.store.fetch_one("SELECT * FROM runs WHERE run_id = ?", (run_id,))
        query = self.store.fetch_one(
            "SELECT * FROM run_queries WHERE run_id = ? AND query_id = ?",
            (run_id, "query-execution"),
        )
        candidate = self.store.fetch_one(
            "SELECT * FROM candidates WHERE run_id = ?",
            (run_id,),
        )
        raw = self.store.fetch_all(
            """
            SELECT stage, sequence_no, payload_json
            FROM raw_records WHERE run_id = ?
            ORDER BY collected_at, sequence_no
            """,
            (run_id,),
        )

        self.assertEqual("COMPLETED", run["status"])
        self.assertEqual("PHASE_2_POST_OPTIMIZATION", run["evaluation_phase"])
        self.assertEqual(1, run["success_queries"])
        self.assertEqual("SUCCESS", query["status"])
        self.assertEqual("HAS_CANDIDATES", query["result_status"])
        self.assertIsNone(query["llm_cost"])
        self.assertIsNone(query["third_party_cost"])
        self.assertIsNone(query["total_cost"])
        self.assertIsNone(query["pdl_called"])
        self.assertIsNone(query["search_duration_ms"])
        self.assertEqual(
            "NOT_CONFIGURED",
            json.loads(query["public_fields_json"])["public_info_status"],
        )
        self.assertEqual("candidate-execution", candidate["candidate_id"])
        self.assertEqual(0.88, candidate["rank_score"])
        self.assertEqual(
            [
                "CreateIntentTask",
                "GetTask",
                "GetTask",
                "ListTaskCandidates",
                "GetTaskCandidateDetail",
            ],
            [row["stage"] for row in raw],
        )
        self.assertNotIn(
            "auth_token",
            json.dumps([json.loads(row["payload_json"]) for row in raw]),
        )
        self.assertTrue((self.root / "data" / run["results_file"]).is_file())
        self.assertTrue((self.root / "data" / run["failures_file"]).is_file())
        saved_result = json.loads(
            (self.root / "data" / run["results_file"]).read_text(
                encoding="utf-8"
            )
        )
        self.assertEqual("1.3.1", saved_result["result_schema_version"])
        self.assertEqual("HAS_CANDIDATES", saved_result["result_status"])

    def test_execution_run_failure_is_recorded_and_next_query_continues(self):
        """一个 Query 失败后继续执行后续 Query，并形成 PARTIAL_FAILED Run。"""

        source = self.root / "partial-execution-tasks.jsonl"
        write_jsonl(
            source,
            [
                {
                    "input_id": "query-failed",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME", "value": "Failed"}],
                },
                {
                    "input_id": "query-success",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME", "value": "Success"}],
                },
            ],
        )
        self.service.import_dataset_jsonl(
            source,
            name="部分失败数据集",
            dataset_id="dataset-partial-execution",
        )
        run_id = self.service.create_execution_run(
            evaluation_id="eval-import",
            dataset_id="dataset-partial-execution",
            run_label="candidate",
            system_version="web-v1",
            evaluation_phase="PHASE_1_BASELINE",
            run_id="run-partial-execution",
        )
        client = FakeExecutionClient(
            [
                FlowError("CreateIntentTask", "first query failed"),
                api_body({"task_id": "task-success"}),
                api_body({"status": "NO_RESULT"}),
            ]
        )

        self.service.execute_run(run_id, client, sleep_fn=lambda _: None)

        run = self.store.fetch_one("SELECT * FROM runs WHERE run_id = ?", (run_id,))
        queries = self.store.fetch_all(
            """
            SELECT query_id, status, result_status FROM run_queries
            WHERE run_id = ? ORDER BY query_id
            """,
            (run_id,),
        )
        self.assertEqual("PARTIAL_FAILED", run["status"])
        self.assertEqual(1, run["success_queries"])
        self.assertEqual(1, run["failed_queries"])
        self.assertEqual(
            [("query-failed", "FAILED"), ("query-success", "NO_CANDIDATE")],
            [(row["query_id"], row["status"]) for row in queries],
        )
        self.assertEqual(
            ["EXECUTION_FAILED", "NO_CANDIDATES"],
            [row["result_status"] for row in queries],
        )
        self.assertEqual(
            1,
            self.store.fetch_one(
                "SELECT COUNT(*) AS count FROM failures WHERE run_id = ?",
                (run_id,),
            )["count"],
        )

    def test_failed_query_can_retry_in_same_run(self):
        """失败 Query 重跑后仍归属原 Run，且不影响已完成的其他 Query。"""

        source = self.root / "retry-execution-tasks.jsonl"
        write_jsonl(
            source,
            [
                {
                    "input_id": "query-retry",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME", "value": "Retry"}],
                },
                {
                    "input_id": "query-kept",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME", "value": "Kept"}],
                },
            ],
        )
        self.service.import_dataset_jsonl(
            source, name="重跑数据集", dataset_id="dataset-retry-execution"
        )
        run_id = self.service.create_execution_run(
            evaluation_id="eval-import",
            dataset_id="dataset-retry-execution",
            run_label="candidate",
            system_version="web-v1",
            evaluation_phase="PHASE_1_BASELINE",
            run_id="run-retry-execution",
        )
        self.service.execute_run(
            run_id,
            FakeExecutionClient(
                [
                    FlowError("CreateIntentTask", "first query failed"),
                    api_body({"task_id": "task-kept"}),
                    api_body({"status": "NO_RESULT"}),
                ]
            ),
            sleep_fn=lambda _: None,
        )

        self.service.execute_query_retry(
            run_id,
            "query-retry",
            FakeExecutionClient(
                [
                    api_body({"task_id": "task-retry"}),
                    api_body({"status": "NO_RESULT"}),
                ]
            ),
            sleep_fn=lambda _: None,
        )

        run = self.store.fetch_one("SELECT * FROM runs WHERE run_id = ?", (run_id,))
        queries = self.store.fetch_all(
            "SELECT query_id, status, result_status FROM run_queries WHERE run_id = ? ORDER BY query_id",
            (run_id,),
        )
        self.assertEqual("COMPLETED", run["status"])
        self.assertEqual((2, 0), (run["success_queries"], run["failed_queries"]))
        self.assertEqual(
            [
                ("query-kept", "NO_CANDIDATE", "NO_CANDIDATES"),
                ("query-retry", "NO_CANDIDATE", "NO_CANDIDATES"),
            ],
            [(row["query_id"], row["status"], row["result_status"]) for row in queries],
        )

    def test_only_one_active_execution_and_startup_recovery(self):
        """同一时间只创建一个执行 Run，重启时标记遗留运行已中断。"""

        source = self.root / "active-tasks.jsonl"
        write_jsonl(
            source,
            [
                {
                    "input_id": "query-active",
                    "query_stage": "FULL_NAME",
                    "clues": [{"type": "FULL_NAME"}],
                }
            ],
        )
        self.service.import_dataset_jsonl(
            source,
            name="互斥数据集",
            dataset_id="dataset-active",
        )
        with self.assertRaises(ImportValidationError):
            self.service.create_execution_run(
                evaluation_id="eval-import",
                dataset_id="dataset-active",
                run_label="invalid-phase",
                system_version="web-v0",
                evaluation_phase="UNSPECIFIED",
                run_id="run-invalid-phase",
            )
        first_run = self.service.create_execution_run(
            evaluation_id="eval-import",
            dataset_id="dataset-active",
            run_label="candidate",
            system_version="web-v1",
            evaluation_phase="PHASE_1_BASELINE",
            run_id="run-active",
        )

        with self.assertRaises(ActiveRunError):
            self.service.create_execution_run(
                evaluation_id="eval-import",
                dataset_id="dataset-active",
                run_label="candidate-2",
                system_version="web-v2",
                evaluation_phase="PHASE_2_POST_OPTIMIZATION",
            )

        with self.store.transaction() as connection:
            connection.execute(
                "UPDATE runs SET status = 'RUNNING' WHERE run_id = ?",
                (first_run,),
            )
        recovered = self.service.recover_interrupted_runs()

        self.assertEqual(1, recovered)
        self.assertEqual(
            "INTERRUPTED",
            self.store.fetch_one(
                "SELECT status FROM runs WHERE run_id = ?",
                (first_run,),
            )["status"],
        )

    def test_results_jsonl_import_unifies_legacy_v13_and_failures(self):
        """v1.2 与 v1.3 结果进入相同 Run/Query/Candidate，并保留 Raw 状态。"""

        results_path = self.root / "results.jsonl"
        failures_path = self.root / "failures.jsonl"
        metadata_path = self.root / "metadata.jsonl"
        write_jsonl(
            results_path,
            [
                {
                    "input_id": "legacy-query",
                    "task_id": "task-legacy",
                    "candidate_count_total": 1,
                    "results": [
                        {
                            "candidate_id": "legacy-candidate",
                            "ui_sections": {"summary": {"status": "data"}},
                        }
                    ],
                },
                {
                    "result_schema_version": "1.3",
                    "run_id": "source-run",
                    "input_id": "partial-query",
                    "task_id": "task-v13",
                    "query_stage": "FULL_NAME_SOCIAL",
                    "query_status": "PARTIAL_DETAIL_FAILED",
                    "result_status": "HAS_CANDIDATES",
                    "candidate_count_total": 2,
                    "candidate_count_listed": 2,
                    "detail_success_count": 1,
                    "detail_failure_count": 1,
                    "task_fields": {
                        "llm_cost": 1.0,
                        "third_party_cost": 2.0,
                        "total_cost": 3.0,
                        "pdl_called": False,
                        "search_duration_ms": 1200,
                        "future_public_field": "kept",
                    },
                    "raw": {
                        "create_intent_task": {
                            "sequence_no": 1,
                            "request_params": {"clues": []},
                            "response_body": {
                                "future_field": "kept",
                                "auth_token": "must-not-be-stored",
                            },
                        },
                        "get_task_history": [],
                        "list_task_candidates": {},
                    },
                    "results": [
                        {
                            "candidate_rank": 1,
                            "candidate_id": "failed-candidate",
                            "rank_score": 0.8,
                            "detail_status": "FAILED",
                            "detail_error": "timeout",
                            "list_item_raw": {"candidate_id": "failed-candidate"},
                            "detail_data_raw": None,
                            "ui_sections": None,
                        },
                        {
                            "candidate_rank": 2,
                            "candidate_id": "success-candidate",
                            "rank_score": 0.7,
                            "detail_status": "SUCCESS",
                            "detail_error": "",
                            "list_item_raw": {"candidate_id": "success-candidate"},
                            "detail_data_raw": {"ui_sections": {}},
                            "ui_sections": {},
                        },
                    ],
                },
            ],
        )
        write_jsonl(
            failures_path,
            [
                {
                    "failure_schema_version": "1.3",
                    "input_id": "partial-query",
                    "task_id": "task-v13",
                    "candidate_id": "failed-candidate",
                    "scope": "CANDIDATE",
                    "stage": "GetTaskCandidateDetail",
                    "error": "timeout",
                }
            ],
        )
        write_jsonl(
            metadata_path,
            [
                {
                    "query_id": "legacy-query",
                    "person_id": "person-legacy",
                    "query_stage": "FULL_NAME",
                },
                {
                    "query_id": "partial-query",
                    "person_id": "person-v13",
                    "query_stage": "FULL_NAME_SOCIAL",
                },
            ],
        )

        result = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="candidate",
            system_version="v1.3",
            evaluation_phase="PHASE_1_BASELINE",
            failures_path=failures_path,
            metadata_path=metadata_path,
            run_id="run-jsonl",
        )

        queries = self.store.fetch_all(
            """
            SELECT query_id, status, result_status, llm_cost,
                   third_party_cost, total_cost, pdl_called,
                   search_duration_ms, public_fields_json
            FROM run_queries WHERE run_id = ? ORDER BY query_id
            """,
            ("run-jsonl",),
        )
        candidates = self.store.fetch_all(
            """
            SELECT query_id, candidate_rank, detail_status
            FROM candidates WHERE run_id = ?
            ORDER BY query_id, candidate_rank
            """,
            ("run-jsonl",),
        )
        raw_rows = self.store.fetch_all(
            "SELECT query_id, stage, payload_json FROM raw_records WHERE run_id = ?",
            ("run-jsonl",),
        )

        self.assertEqual(2, result.imported_count)
        self.assertEqual(
            [
                ("legacy-query", "SUCCESS"),
                ("partial-query", "PARTIAL_DETAIL_FAILED"),
            ],
            [(row["query_id"], row["status"]) for row in queries],
        )
        self.assertEqual(
            ["HAS_CANDIDATES", "HAS_CANDIDATES"],
            [row["result_status"] for row in queries],
        )
        self.assertEqual(2.0, queries[1]["third_party_cost"])
        self.assertEqual(1200, queries[1]["search_duration_ms"])
        self.assertEqual(
            {"future_public_field": "kept"},
            json.loads(queries[1]["public_fields_json"]),
        )
        self.assertEqual(
            "PHASE_1_BASELINE",
            self.store.fetch_one(
                "SELECT evaluation_phase FROM runs WHERE run_id = ?",
                ("run-jsonl",),
            )["evaluation_phase"],
        )
        self.assertEqual(
            [
                ("legacy-query", 1, "SUCCESS"),
                ("partial-query", 1, "FAILED"),
                ("partial-query", 2, "SUCCESS"),
            ],
            [
                (row["query_id"], row["candidate_rank"], row["detail_status"])
                for row in candidates
            ],
        )
        self.assertTrue(
            any(
                row["query_id"] == "legacy-query"
                and "LEGACY_PARTIAL_RAW" in row["payload_json"]
                for row in raw_rows
            )
        )
        self.assertTrue(
            any(
                row["query_id"] == "partial-query"
                and row["stage"] == "CreateIntentTask"
                for row in raw_rows
            )
        )
        self.assertTrue(
            (self.root / "data" / "raw" / "eval-import" / "run-jsonl" / "results.jsonl").is_file()
        )
        normalized_text = (
            self.root
            / "data"
            / "raw"
            / "eval-import"
            / "run-jsonl"
            / "results.jsonl"
        ).read_text(encoding="utf-8")
        self.assertNotIn("must-not-be-stored", normalized_text)
        self.assertIn("future_field", normalized_text)
        with self.assertRaises(DuplicateImportError):
            self.service.import_results_jsonl(
                results_path,
                evaluation_id="eval-import",
                run_label="candidate",
                system_version="v1.3",
                failures_path=failures_path,
                metadata_path=metadata_path,
            )

    def test_normalized_excel_import_creates_run_query_candidate_and_raw_marker(self):
        """规范 Excel 导入统一模型，并重组 Raw Sheet 的分块内容。"""

        result = self.service.import_results_excel(
            IMPORT_WORKBOOK,
            evaluation_id="eval-import",
            run_label="candidate",
            system_version="excel-v1",
            run_id="run-excel",
        )

        candidate = self.store.fetch_one(
            """
            SELECT candidate_id, detail_data_json
            FROM candidates WHERE run_id = ? AND query_id = ?
            """,
            ("run-excel", "excel-query"),
        )
        run = self.store.fetch_one(
            "SELECT source_type, message FROM runs WHERE run_id = ?",
            ("run-excel",),
        )

        self.assertEqual(1, result.imported_count)
        self.assertEqual("excel-candidate", candidate["candidate_id"])
        self.assertIn("Example Person", candidate["detail_data_json"])
        self.assertIn('"profile_data"', candidate["detail_data_json"])
        self.assertIn('"sections"', candidate["detail_data_json"])
        self.assertEqual("EXCEL_IMPORT", run["source_type"])
        self.assertEqual("LEGACY_PARTIAL_RAW", run["message"])

    def test_excel_import_preserves_no_candidate_status_and_task_fields(self):
        """无候选人的规范化 Excel 仍可导入 Query 公共字段和评估阶段。"""

        workbook_path = self.root / "no-candidate-results.xlsx"
        workbook = Workbook()
        candidates = workbook.active
        candidates.title = "候选结果"
        candidates.append(
            ["query_id", "task_id", "candidate_id", "candidate_rank"]
        )
        queries = workbook.create_sheet("Query对比")
        queries.append(
            [
                "run_label",
                "query_id",
                "person_id",
                "query_type",
                "current_status",
                "result_status",
                "llm_cost",
                "third_party_cost",
                "total_cost",
                "pdl_called",
                "search_duration_ms",
            ]
        )
        queries.append(
            [
                "candidate",
                "excel-no-candidate",
                "person-no-candidate",
                "FULL_NAME",
                "NO_CANDIDATE",
                "NO_CANDIDATES",
                1.0,
                2.0,
                3.0,
                False,
                900,
            ]
        )
        workbook.save(workbook_path)
        workbook.close()

        imported = self.service.import_results_excel(
            workbook_path,
            evaluation_id="eval-import",
            run_label="candidate",
            system_version="excel-v2",
            evaluation_phase="PHASE_2_POST_OPTIMIZATION",
            run_id="run-excel-no-candidate",
        )
        query = self.store.fetch_one(
            """
            SELECT * FROM run_queries
            WHERE run_id = ? AND query_id = ?
            """,
            ("run-excel-no-candidate", "excel-no-candidate"),
        )
        run = self.store.fetch_one(
            "SELECT * FROM runs WHERE run_id = ?",
            ("run-excel-no-candidate",),
        )

        self.assertEqual(1, imported.imported_count)
        self.assertEqual("NO_CANDIDATES", query["result_status"])
        self.assertEqual(2.0, query["third_party_cost"])
        self.assertEqual(900, query["search_duration_ms"])
        self.assertEqual(
            "PHASE_2_POST_OPTIMIZATION",
            run["evaluation_phase"],
        )

    def test_invalid_result_record_does_not_create_partial_run_or_archive(self):
        """结果文件含坏记录时整包拒绝，不留下 Run、Query 或归档目录。"""

        results_path = self.root / "invalid-results.jsonl"
        write_jsonl(
            results_path,
            [
                {
                    "input_id": "valid-query",
                    "task_id": "task-valid",
                    "results": [],
                },
                {
                    "input_id": "invalid-query",
                    "task_id": "task-invalid",
                    "results": "not-an-array",
                },
            ],
        )

        with self.assertRaises(ImportValidationError):
            self.service.import_results_jsonl(
                results_path,
                evaluation_id="eval-import",
                run_label="candidate",
                system_version="invalid",
                run_id="run-invalid-results",
            )

        self.assertIsNone(
            self.store.fetch_one(
                "SELECT run_id FROM runs WHERE run_id = ?",
                ("run-invalid-results",),
            )
        )
        self.assertFalse(
            (self.root / "data" / "imports" / "run-invalid-results").exists()
        )

    def test_invalid_explicit_result_status_is_rejected(self):
        """历史结果显式状态必须使用规范化枚举，不能静默回退。"""

        results_path = self.root / "invalid-result-status.jsonl"
        write_jsonl(
            results_path,
            [
                {
                    "input_id": "invalid-result-status",
                    "task_id": "task-invalid-result-status",
                    "query_status": "NO_CANDIDATE",
                    "result_status": "SUCCESS",
                    "results": [],
                }
            ],
        )
        with self.assertRaises(ImportValidationError):
            self.service.import_results_jsonl(
                results_path,
                evaluation_id="eval-import",
                run_label="invalid-result-status",
                system_version="v1",
                run_id="run-invalid-result-status",
            )

    def test_failure_only_history_import_sets_execution_failed_status(self):
        """只有 failures 的旧结果包仍能生成失败 Query 和规范化状态。"""

        results_path = self.root / "failure-only-results.jsonl"
        failures_path = self.root / "failure-only-failures.jsonl"
        results_path.write_text("", encoding="utf-8")
        write_jsonl(
            failures_path,
            [
                {
                    "input_id": "failure-only-query",
                    "task_id": "task-failure-only",
                    "scope": "QUERY",
                    "stage": "GetTask",
                    "error": "polling timeout",
                }
            ],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="failure-only",
            system_version="legacy",
            failures_path=failures_path,
            run_id="run-failure-only",
        )
        query = self.store.fetch_one(
            """
            SELECT status, result_status FROM run_queries
            WHERE run_id = ? AND query_id = ?
            """,
            (imported.object_id, "failure-only-query"),
        )

        self.assertEqual("FAILED", query["status"])
        self.assertEqual("EXECUTION_FAILED", query["result_status"])

    def test_dataset_and_baseline_excel_and_jsonl_import(self):
        """Excel Query、Excel 基准及 JSONL 基准均可进入对应统一表。"""

        dataset_result = self.service.import_dataset_excel(
            IMPORT_WORKBOOK,
            name="Excel Dataset",
            dataset_id="dataset-excel",
        )
        baseline_excel = self.service.import_baseline_excel(
            IMPORT_WORKBOOK,
            name="Excel Baseline",
            baseline_version="baseline-excel",
        )
        baseline_jsonl_path = self.root / "baseline.jsonl"
        write_jsonl(
            baseline_jsonl_path,
            [
                {
                    "person_id": "person-jsonl",
                    "display_name": "JSONL Person",
                    "fields": {"social_urls": ["https://social.example.test/user"]},
                    "baseline_available_fields": [
                        "social_urls",
                        "social_urls",
                        "future_confirmed_field",
                    ],
                    "evidence": {},
                }
            ],
        )
        baseline_jsonl = self.service.import_baseline_jsonl(
            baseline_jsonl_path,
            name="JSONL Baseline",
            baseline_version="baseline-jsonl",
        )

        self.assertEqual(1, dataset_result.imported_count)
        self.assertEqual(1, baseline_excel.imported_count)
        self.assertEqual(1, baseline_jsonl.imported_count)
        excel_person = self.store.fetch_one(
            """
            SELECT available_fields_json, available_fields_source
            FROM baseline_people WHERE baseline_version = 'baseline-excel'
            """
        )
        jsonl_person = self.store.fetch_one(
            """
            SELECT available_fields_json, available_fields_source
            FROM baseline_people WHERE baseline_version = 'baseline-jsonl'
            """
        )
        self.assertEqual(
            "DERIVED_LEGACY",
            excel_person["available_fields_source"],
        )
        self.assertGreater(
            len(json.loads(excel_person["available_fields_json"])),
            0,
        )
        self.assertEqual(
            ["social_urls", "future_confirmed_field"],
            json.loads(jsonl_person["available_fields_json"]),
        )
        self.assertEqual("IMPORT", jsonl_person["available_fields_source"])
        with self.assertRaises(DuplicateImportError):
            self.service.import_baseline_excel(
                IMPORT_WORKBOOK,
                name="重复 Excel Baseline",
                baseline_version="baseline-excel-duplicate",
            )
        self.assertEqual(
            2,
            self.store.fetch_one(
                "SELECT COUNT(*) AS count FROM baseline_people"
            )["count"],
        )

    def test_baseline_available_fields_empty_and_manual_update_are_traceable(self):
        """显式空数组保持未就绪，人工更新标记来源并令旧报告过期。"""

        baseline_path = self.root / "baseline-available.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": "person-available",
                    "display_name": "Available Person",
                    "fields": {
                        "summary_display_name": "Available Person",
                        "unknown_baseline_field": "known by tester",
                    },
                    "baseline_available_fields": [],
                    "evidence": {},
                }
            ],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="Available Baseline",
            baseline_version="baseline-available",
        )
        imported = self.store.fetch_one(
            """
            SELECT available_fields_json, available_fields_source
            FROM baseline_people
            WHERE baseline_version = ? AND person_id = ?
            """,
            ("baseline-available", "person-available"),
        )
        self.assertEqual([], json.loads(imported["available_fields_json"]))
        self.assertEqual("IMPORT", imported["available_fields_source"])

        self.service.update_baseline_available_fields(
            "baseline-available",
            "person-available",
            ["summary_display_name", "unknown_baseline_field"],
        )
        updated = self.store.fetch_one(
            """
            SELECT available_fields_json, available_fields_source
            FROM baseline_people
            WHERE baseline_version = ? AND person_id = ?
            """,
            ("baseline-available", "person-available"),
        )
        self.assertEqual(
            ["summary_display_name", "unknown_baseline_field"],
            json.loads(updated["available_fields_json"]),
        )
        self.assertEqual("MANUAL", updated["available_fields_source"])

        invalid_path = self.root / "baseline-invalid-available.jsonl"
        write_jsonl(
            invalid_path,
            [
                {
                    "person_id": "person-invalid",
                    "fields": {},
                    "baseline_available_fields": ["valid", ""],
                }
            ],
        )
        with self.assertRaises(ImportValidationError):
            self.service.import_baseline_jsonl(
                invalid_path,
                name="Invalid Available Baseline",
                baseline_version="baseline-invalid-available",
            )

    def test_baseline_excel_available_fields_supports_json_and_comma_values(self):
        """Excel 可用字段支持 JSON 数组和英文逗号分隔，坏 JSON 整批拒绝。"""

        workbook_path = self.root / "baseline-available.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "基准数据"
        sheet.append(
            [
                "person_id",
                "display_name",
                "baseline_available_fields",
                "summary_display_name",
                "profile_location",
            ]
        )
        sheet.append(
            [
                "person-json-array",
                "JSON Array",
                '["summary_display_name","profile_location"]',
                "JSON Array",
                "Shanghai",
            ]
        )
        sheet.append(
            [
                "person-comma",
                "Comma",
                "summary_display_name, profile_location",
                "Comma",
                "Beijing",
            ]
        )
        workbook.save(workbook_path)
        workbook.close()

        imported = self.service.import_baseline_excel(
            workbook_path,
            name="Excel Available Baseline",
            baseline_version="baseline-excel-available",
        )
        rows = self.store.fetch_all(
            """
            SELECT person_id, available_fields_json, available_fields_source
            FROM baseline_people
            WHERE baseline_version = ?
            ORDER BY person_id
            """,
            ("baseline-excel-available",),
        )
        self.assertEqual(2, imported.imported_count)
        self.assertEqual(
            ["summary_display_name", "profile_location"],
            json.loads(rows[0]["available_fields_json"]),
        )
        self.assertEqual(
            ["summary_display_name", "profile_location"],
            json.loads(rows[1]["available_fields_json"]),
        )
        self.assertTrue(
            all(row["available_fields_source"] == "IMPORT" for row in rows)
        )

        invalid_path = self.root / "baseline-invalid-json.xlsx"
        invalid_workbook = Workbook()
        invalid_sheet = invalid_workbook.active
        invalid_sheet.title = "基准数据"
        invalid_sheet.append(
            ["person_id", "baseline_available_fields", "display_name"]
        )
        invalid_sheet.append(["person-invalid", '["broken"', "Invalid"])
        invalid_workbook.save(invalid_path)
        invalid_workbook.close()
        with self.assertRaises(ImportValidationError):
            self.service.import_baseline_excel(
                invalid_path,
                name="Invalid Excel Available Baseline",
                baseline_version="baseline-invalid-excel-available",
            )

    def test_stage4_metrics_v2_uses_person_denominator_and_partial_task_values(self):
        """指标 v2 可手算覆盖状态、人物分母、部分成本、PDL、置信度和分组。"""

        schema_version = self.service.ensure_default_field_schema()
        baseline_path = self.root / "baseline-metrics-v2.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": f"person-metrics-{index}",
                    "fields": {
                        "social_urls": [
                            f"https://social.example.test/person-{index}"
                        ]
                    },
                    "baseline_available_fields": ["social_urls"],
                }
                for index in range(1, 4)
            ],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="指标 v2 基准",
            baseline_version="baseline-metrics-v2",
        )
        results_path = self.root / "results-metrics-v2.jsonl"
        matched_sections = {
            "social": {
                "data": {
                    "profiles": [
                        {"url": "https://social.example.test/person-1"}
                    ]
                }
            },
            "summary": {"data": {"confidence_level": "HIGH"}},
        }
        write_jsonl(
            results_path,
            [
                {
                    "result_schema_version": "1.3.1",
                    "input_id": "query-metrics-has",
                    "person_id": "person-metrics-1",
                    "query_stage": "FULL_NAME",
                    "query_status": "SUCCESS",
                    "result_status": "HAS_CANDIDATES",
                    "task_fields": {
                        "llm_cost": 1,
                        "third_party_cost": 2,
                        "total_cost": 3,
                        "pdl_called": True,
                        "search_duration_ms": 100,
                    },
                    "results": [
                        {
                            "candidate_rank": 1,
                            "candidate_id": "candidate-metrics-hit",
                            "detail_status": "SUCCESS",
                            "detail_error": "",
                            "list_item_raw": {
                                "candidate_id": "candidate-metrics-hit"
                            },
                            "detail_data_raw": {
                                "ui_sections": matched_sections
                            },
                            "ui_sections": matched_sections,
                        }
                    ],
                },
                {
                    "result_schema_version": "1.3.1",
                    "input_id": "query-metrics-none",
                    "person_id": "person-metrics-2",
                    "query_stage": "FULL_NAME",
                    "query_status": "NO_CANDIDATE",
                    "result_status": "NO_CANDIDATES",
                    "task_fields": {
                        "llm_cost": 0,
                        "third_party_cost": None,
                        "total_cost": 0,
                        "pdl_called": False,
                        "search_duration_ms": 200,
                    },
                    "results": [],
                },
                {
                    "result_schema_version": "1.3.1",
                    "input_id": "query-metrics-failed",
                    "person_id": "person-metrics-3",
                    "query_stage": "FULL_NAME",
                    "query_status": "FAILED",
                    "result_status": "EXECUTION_FAILED",
                    "task_fields": {
                        "llm_cost": -1,
                        "search_duration_ms": "invalid-duration",
                    },
                    "results": [],
                },
            ],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="指标 v2",
            system_version="metrics-v2-system",
            evaluation_phase="PHASE_2_POST_OPTIMIZATION",
            run_id="run-metrics-v2",
        )
        process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
            baseline_version="baseline-metrics-v2",
            process_id="process-metrics-v2",
        )
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE process_runs SET rule_version = 'field-processing-v2'
                WHERE process_id = ?
                """,
                (process.process_id,),
            )
        candidate = self.store.fetch_one(
            "SELECT candidate_pk FROM candidates WHERE run_id = ?",
            (imported.object_id,),
        )
        context = self.service.get_review_context(
            process.process_id,
            candidate["candidate_pk"],
        )
        self.service.save_review(
            process_id=process.process_id,
            candidate_pk=candidate["candidate_pk"],
            judgement="HIT",
            reason="SOCIAL_MATCH",
            evidence="Social URL 与基准一致",
            reviewer="metrics-tester",
            review_note="阶段4可手算测试",
            field_scores=context["field_scores"],
            expected_reviewed_at=context["reviewed_at"],
        )

        metrics = self.service.calculate_process_metrics(process.process_id)

        self.assertEqual("metrics-v2", metrics["metrics_rule_version"])
        result_metrics = metrics["result_status_metrics"]
        self.assertEqual(3, result_metrics["total_formal_queries"])
        self.assertEqual(1, result_metrics["has_candidates_count"])
        self.assertEqual(1, result_metrics["no_candidates_count"])
        self.assertEqual(1, result_metrics["execution_failed_count"])
        self.assertEqual(0.5, result_metrics["has_result_rate"])
        self.assertAlmostEqual(1 / 3, result_metrics["execution_failed_rate"])
        self.assertEqual(1, metrics["matched_completeness"]["denominator"])
        self.assertEqual(1.0, metrics["matched_completeness"]["value"])
        self.assertEqual(3, metrics["retrieval_success"]["denominator"])
        self.assertAlmostEqual(1 / 3, metrics["retrieval_success"]["value"])
        llm_cost = metrics["cost_metrics"]["llm_cost"]
        self.assertEqual("PARTIAL", llm_cost["status"])
        self.assertEqual(2, llm_cost["value_count"])
        self.assertEqual(0, llm_cost["missing_count"])
        self.assertEqual(1, llm_cost["invalid_count"])
        self.assertEqual(1.0, llm_cost["total"])
        self.assertEqual(0.5, llm_cost["average"])
        self.assertEqual(
            1,
            metrics["cost_metrics"]["search_duration_ms"]["invalid_count"],
        )
        self.assertEqual(2, metrics["pdl_metrics"]["known_count"])
        self.assertEqual(1, metrics["pdl_metrics"]["unknown_count"])
        self.assertEqual(0.5, metrics["pdl_metrics"]["call_rate"])
        self.assertEqual(1, metrics["confidence_metrics"]["overall"]["HIGH"])
        self.assertEqual(1, metrics["confidence_metrics"]["matched"]["HIGH"])
        self.assertEqual(1, len(metrics["grouped_metrics"]))
        self.assertEqual(
            "PHASE_2_POST_OPTIMIZATION",
            metrics["grouped_metrics"][0]["evaluation_phase"],
        )

        self.service.update_baseline_available_fields(
            "baseline-metrics-v2",
            "person-metrics-1",
            [],
        )
        empty_available_metrics = self.service.calculate_process_metrics(
            process.process_id
        )
        self.assertEqual(
            "NOT_READY",
            empty_available_metrics["matched_completeness"]["status"],
        )
        self.assertIsNone(
            empty_available_metrics["matched_completeness"]["value"]
        )

        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE process_runs
                SET rule_version = 'field-processing-v1'
                WHERE process_id = ?
                """,
                (process.process_id,),
            )
        legacy_metrics = self.service.calculate_process_metrics(
            process.process_id
        )
        self.assertEqual("metrics-v1", legacy_metrics["metrics_rule_version"])
        self.assertAlmostEqual(
            1 / 22,
            legacy_metrics["matched_completeness"]["value"],
        )

        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE process_runs
                SET rule_version = 'unknown-processing-rule'
                WHERE process_id = ?
                """,
                (process.process_id,),
            )
        with self.assertRaises(ReviewValidationError):
            self.service.calculate_process_metrics(process.process_id)

    def test_stage5_social_review_metrics_and_report_stale(self):
        """Social 建议、人工复核、四项指标和报告过期保持同一数据来源。"""

        definitions = [
            {
                "field_key": "social_urls",
                "display_name": "Social URLs",
                "module": "Social",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.social.data.profiles[*].url",
                "data_type": "array",
                "array_mode": "collect",
                "empty_rule": "default",
                "normalizer": "social_url",
                "scoring_role": ["identity", "completeness", "accuracy"],
                "compare_mode": "url_set",
                "enabled": True,
                "sort_order": 10,
            },
            {
                "field_key": "display_name",
                "display_name": "Display Name",
                "module": "Summary",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.summary.data.display_name",
                "data_type": "string",
                "array_mode": "preserve",
                "empty_rule": "default",
                "normalizer": "trim_text",
                "scoring_role": ["completeness", "accuracy"],
                "compare_mode": "normalized_text",
                "enabled": True,
                "sort_order": 20,
            },
            {
                "field_key": "biography",
                "display_name": "Biography",
                "module": "Profile",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.profile.data.biography",
                "data_type": "string",
                "array_mode": "preserve",
                "empty_rule": "default",
                "normalizer": "trim_text",
                "scoring_role": ["completeness", "accuracy"],
                "compare_mode": "manual",
                "enabled": True,
                "sort_order": 30,
            },
        ]
        schema_version = self.service.publish_field_schema(
            name="阶段5字段配置",
            definitions=definitions,
            schema_version="field-schema-stage5",
        )
        baseline_path = self.root / "baseline-stage5.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": "person-stage5",
                    "display_name": "Alice",
                    "fields": {
                        "social_urls": [
                            "https://www.example.test/alice/?utm_source=base"
                        ],
                        "display_name": "Alice",
                        "biography": "Known biography",
                    },
                    "evidence": {},
                }
            ],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段5基准",
            baseline_version="baseline-stage5",
        )
        results_path = self.root / "results-stage5.jsonl"

        def candidate(
            rank: int,
            candidate_id: str,
            urls: list[str],
            display_name: str,
            biography: str = "",
        ) -> dict:
            """构造包含可评分字段的阶段5候选人。"""

            ui_sections = {
                "social": {
                    "data": {
                        "profiles": [{"url": url} for url in urls],
                    }
                },
                "summary": {"data": {"display_name": display_name}},
                "profile": {"data": {"biography": biography}},
            }
            return {
                "candidate_rank": rank,
                "candidate_id": candidate_id,
                "detail_status": "SUCCESS",
                "detail_error": "",
                "list_item_raw": {"candidate_id": candidate_id},
                "detail_data_raw": {"ui_sections": ui_sections},
                "ui_sections": ui_sections,
            }

        write_jsonl(
            results_path,
            [
                {
                    "result_schema_version": "1.3",
                    "input_id": "query-stage5",
                    "person_id": "person-stage5",
                    "query_stage": "FULL_NAME_SOCIAL",
                    "task_id": "task-stage5",
                    "query_status": "SUCCESS",
                    "results": [
                        candidate(
                            1,
                            "candidate-match",
                            ["https://example.test/alice/"],
                            " Alice ",
                            "Returned biography",
                        ),
                        candidate(
                            2,
                            "candidate-conflict",
                            ["https://example.test/other"],
                            "Other",
                        ),
                        candidate(
                            3,
                            "candidate-match-conflict",
                            [
                                "https://example.test/alice",
                                "https://example.test/other",
                            ],
                            "Alice",
                        ),
                        candidate(4, "candidate-no-social", [], "Alice"),
                    ],
                }
            ],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="candidate",
            system_version="stage5-v1",
            run_id="run-stage5",
        )
        process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
            baseline_version="baseline-stage5",
            process_id="process-stage5",
        )
        review_rows = self.store.fetch_all(
            """
            SELECT c.candidate_id, r.judgement, r.reviewed_at,
                   r.classification_source, r.is_primary_hit
            FROM reviews AS r
            JOIN candidates AS c ON c.candidate_pk = r.candidate_pk
            WHERE r.process_id = ?
            ORDER BY c.candidate_rank
            """,
            (process.process_id,),
        )
        self.assertEqual(
            ["HIT", "NOT_HIT", "NOT_HIT", "SUSPECTED"],
            [row["judgement"] for row in review_rows],
        )
        self.assertTrue(all(row["reviewed_at"] is not None for row in review_rows))
        self.assertTrue(
            all(row["classification_source"] == "RULE" for row in review_rows)
        )
        self.assertEqual(1, review_rows[0]["is_primary_hit"])
        self.assertTrue(
            all(not row["is_primary_hit"] for row in review_rows[1:])
        )

        contexts = {}
        for row in self.store.fetch_all(
            """
            SELECT candidate_pk, candidate_id
            FROM candidates WHERE run_id = ? ORDER BY candidate_rank
            """,
            (imported.object_id,),
        ):
            contexts[row["candidate_id"]] = self.service.get_review_context(
                process.process_id,
                row["candidate_pk"],
            )
        match_context = contexts["candidate-match"]
        match_scores = match_context["field_scores"]
        self.assertEqual(1.0, match_scores["social_urls"]["completeness_score"])
        self.assertEqual(1.0, match_scores["social_urls"]["accuracy_score"])
        self.assertIsNone(match_scores["biography"]["completeness_score"])

        match_scores["biography"]["completeness_score"] = 0.5
        match_scores["biography"]["accuracy_score"] = 1.0
        match_review = self.service.save_review(
            process_id=process.process_id,
            candidate_pk=match_context["candidate_pk"],
            judgement="HIT",
            reason="SOCIAL_MATCH",
            evidence="Social URL 与基准一致",
            reviewer="tester",
            review_note="人工确认 biography 得分",
            field_scores=match_scores,
            expected_reviewed_at=match_context["reviewed_at"],
        )
        for candidate_id, judgement in (
            ("candidate-conflict", "NOT_HIT"),
            ("candidate-match-conflict", "SUSPECTED"),
            ("candidate-no-social", "NOT_HIT"),
        ):
            context = contexts[candidate_id]
            self.service.save_review(
                process_id=process.process_id,
                candidate_pk=context["candidate_pk"],
                judgement=judgement,
                reason="MANUAL",
                evidence="人工确认",
                reviewer="tester",
                review_note="阶段5测试复核",
                field_scores=context["field_scores"],
                expected_reviewed_at=context["reviewed_at"],
            )

        metrics = self.service.calculate_process_metrics(process.process_id)
        self.assertTrue(metrics["formal_ready"])
        self.assertEqual(1, metrics["retrieval_success"]["numerator"])
        self.assertEqual(1, metrics["retrieval_success"]["denominator"])
        # 命中资料完整度仅判断资料是否真实返回，不再受人工填写的
        # 基准覆盖分影响；三个启用字段均有返回，故为 1.0。
        self.assertAlmostEqual(1.0, metrics["matched_completeness"]["value"])
        self.assertAlmostEqual(1.0, metrics["matched_accuracy"]["value"])
        self.assertAlmostEqual(
            5 / 9,
            metrics["nonmatched_data_completeness"]["value"],
        )
        self.assertAlmostEqual(
            1 / 2,
            metrics["nonmatched_baseline_overlap"]["value"],
        )
        self.assertEqual("NOT_CONNECTED", metrics["cost_status"]["status"])
        self.assertIsNone(metrics["cost_status"]["llm_cost_total"])
        self.assertIsNone(metrics["cost_status"]["total_cost_total"])
        self.assertIsNone(metrics["cost_status"]["pdl_called_total"])

        with self.store.transaction() as connection:
            connection.execute(
                """
                INSERT INTO reports(
                    report_id, evaluation_id, baseline_process_id,
                    candidate_process_id, report_type, status, metrics_json,
                    html_file, excel_file, created_at
                ) VALUES (
                    'report-stage5', 'eval-import', NULL, ?,
                    'SINGLE', 'READY', '{}', 'report.html', NULL, ?
                )
                """,
                (process.process_id, match_review["reviewed_at"]),
            )
        self.service.save_review(
            process_id=process.process_id,
            candidate_pk=match_context["candidate_pk"],
            judgement="HIT",
            reason="MANUAL",
            evidence="二次人工确认",
            reviewer="tester-2",
            review_note="修改后报告应过期",
            field_scores=match_scores,
            expected_reviewed_at=match_review["reviewed_at"],
        )
        self.assertEqual(
            "STALE",
            self.store.fetch_one(
                "SELECT status FROM reports WHERE report_id = 'report-stage5'"
            )["status"],
        )
        with self.assertRaises(ReviewValidationError):
            self.service.save_review(
                process_id=process.process_id,
                candidate_pk=match_context["candidate_pk"],
                judgement="HIT",
                reason="MANUAL",
                evidence="旧页面提交",
                reviewer="stale-tester",
                review_note="不应覆盖新复核",
                field_scores=match_scores,
                expected_reviewed_at=match_review["reviewed_at"],
            )

    def test_stage5_evaluation_thresholds_validate_and_persist(self):
        """Evaluation 参考线允许部分配置，并拒绝越界比例和负成本。"""

        thresholds = self.service.update_evaluation_thresholds(
            "eval-import",
            {
                "FULL_NAME": {
                    "min_retrieval_success": 0.7,
                    "min_matched_completeness": None,
                    "min_matched_accuracy": 0.9,
                    "max_average_total_cost": 5,
                    "max_average_search_duration_ms": 3000,
                },
                "FULL_NAME_SOCIAL": {},
            },
        )
        stored = json.loads(
            self.store.fetch_one(
                """
                SELECT thresholds_json FROM evaluations
                WHERE evaluation_id = 'eval-import'
                """
            )["thresholds_json"]
        )

        self.assertEqual(thresholds, stored)
        self.assertEqual(
            0.7,
            stored["FULL_NAME"]["min_retrieval_success"],
        )
        self.assertIsNone(
            stored["FULL_NAME"]["min_matched_completeness"]
        )
        with self.assertRaises(ReviewValidationError):
            self.service.update_evaluation_thresholds(
                "eval-import",
                {
                    "FULL_NAME": {
                        "min_retrieval_success": 1.1,
                    }
                },
            )
        with self.assertRaises(ReviewValidationError):
            self.service.update_evaluation_thresholds(
                "eval-import",
                {
                    "FULL_NAME": {
                        "max_average_total_cost": -0.01,
                    }
                },
            )
        synthetic_metrics = {
            "metrics_rule_version": "metrics-v2",
            "grouped_metrics": [
                {
                    "query_stage": "FULL_NAME",
                    "query_count": 10,
                    "quality_metrics": {
                        "retrieval_success": {"value": 0.8},
                        "matched_completeness": {"value": 0.6},
                        "matched_accuracy": {"value": 0.95},
                    },
                    "cost_metrics": {
                        "total_cost": {"average": 4.0},
                        "search_duration_ms": {"average": 2500},
                    },
                }
            ],
        }
        failed = self.service.assess_evaluation_thresholds(
            synthetic_metrics,
            {
                "FULL_NAME": {
                    "min_retrieval_success": 0.9,
                }
            },
        )
        passed = self.service.assess_evaluation_thresholds(
            synthetic_metrics,
            {
                "FULL_NAME": {
                    "min_retrieval_success": 0.7,
                    "max_average_total_cost": 5,
                }
            },
        )
        self.assertEqual(
            "FAIL",
            failed["stages"]["FULL_NAME"]["items"][
                "min_retrieval_success"
            ]["status"],
        )
        self.assertEqual("继续优化", failed["recommendation"])
        self.assertEqual("建议上线", passed["recommendation"])

    def test_stage3_threshold_profiles_are_versioned_reusable_and_archivable(self):
        """参考线方案不可覆盖、可复用，归档后不可再关联新 Evaluation。"""

        thresholds_v1 = {
            "FULL_NAME": {
                "min_retrieval_success": 0.7,
                "max_average_total_cost": 5,
            },
            "FULL_NAME_SOCIAL": {},
        }
        profile_v1 = self.service.create_threshold_profile(
            profile_id="release-standard-v1",
            name="发布验收参考线",
            description="首版",
            version=1,
            thresholds=thresholds_v1,
        )
        profile_v2 = self.service.create_threshold_profile(
            profile_id="release-standard-v2",
            name="发布验收参考线",
            description="第二版",
            version=2,
            thresholds={
                "FULL_NAME": {
                    "min_retrieval_success": 0.8,
                },
                "FULL_NAME_SOCIAL": {
                    "min_matched_accuracy": 0.9,
                },
            },
            based_on_profile_id=profile_v1["profile_id"],
        )
        with self.assertRaises(ReviewValidationError):
            self.service.create_threshold_profile(
                profile_id="release-standard-v2-copy",
                name="发布验收参考线",
                description="重复版本",
                version=2,
                thresholds=thresholds_v1,
            )

        self.service.assign_evaluation_threshold_profile(
            "eval-import",
            profile_v1["profile_id"],
        )
        self.service.create_evaluation(
            evaluation_id="eval-reuse",
            name="复用评测",
            notes="",
            threshold_profile_id=profile_v1["profile_id"],
        )
        first = self.store.fetch_one(
            """
            SELECT threshold_profile_id, thresholds_json
            FROM evaluations WHERE evaluation_id = 'eval-import'
            """
        )
        reused = self.store.fetch_one(
            """
            SELECT threshold_profile_id, thresholds_json
            FROM evaluations WHERE evaluation_id = 'eval-reuse'
            """
        )
        self.assertEqual(profile_v1["profile_id"], first["threshold_profile_id"])
        self.assertEqual(first["thresholds_json"], reused["thresholds_json"])

        self.service.archive_threshold_profile(profile_v2["profile_id"])
        with self.assertRaises(ReviewValidationError):
            self.service.assign_evaluation_threshold_profile(
                "eval-import",
                profile_v2["profile_id"],
            )
        archived = self.store.fetch_one(
            """
            SELECT status FROM threshold_profiles WHERE profile_id = ?
            """,
            (profile_v2["profile_id"],),
        )
        self.assertEqual("ARCHIVED", archived["status"])

        self.service.update_evaluation_thresholds(
            "eval-import",
            {"FULL_NAME": {"min_retrieval_success": 0.6}},
        )
        custom = self.store.fetch_one(
            """
            SELECT threshold_profile_id, thresholds_json
            FROM evaluations WHERE evaluation_id = 'eval-import'
            """
        )
        self.assertIsNone(custom["threshold_profile_id"])
        self.assertEqual(
            0.6,
            json.loads(custom["thresholds_json"])["FULL_NAME"][
                "min_retrieval_success"
            ],
        )

    def test_stage5_process_pairing_uses_person_and_query_stage(self):
        """版本配对要求处理配置一致，并按 person_id + query_stage 分类。"""

        definitions = [
            {
                "field_key": "social_urls",
                "display_name": "Social URLs",
                "module": "Social",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.social.data.profiles[*].url",
                "data_type": "array",
                "array_mode": "collect",
                "empty_rule": "default",
                "normalizer": "social_url",
                "scoring_role": ["identity", "completeness", "accuracy"],
                "compare_mode": "url_set",
                "enabled": True,
                "sort_order": 10,
            }
        ]
        schema_version = self.service.publish_field_schema(
            name="配对字段配置",
            definitions=definitions,
            schema_version="field-schema-pair",
        )
        baseline_path = self.root / "baseline-pair.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": "person-pair",
                    "fields": {
                        "social_urls": ["https://social.example.test/person"]
                    },
                    "evidence": {},
                }
            ],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="配对基准",
            baseline_version="baseline-pair",
        )

        process_ids = []
        for label, candidate_id in (
            ("baseline", "candidate-pair-a"),
            ("candidate", "candidate-pair-b"),
        ):
            results_path = self.root / f"{label}-pair.jsonl"
            ui_sections = {
                "social": {
                    "data": {
                        "profiles": [
                            {"url": "https://social.example.test/person/"}
                        ]
                    }
                }
            }
            write_jsonl(
                results_path,
                [
                    {
                        "result_schema_version": "1.3",
                        "input_id": f"query-{label}",
                        "person_id": "person-pair",
                        "query_stage": "FULL_NAME_SOCIAL",
                        "task_id": f"task-{label}",
                        "query_status": "SUCCESS",
                        "results": [
                            {
                                "candidate_rank": 1,
                                "candidate_id": candidate_id,
                                "detail_status": "SUCCESS",
                                "detail_error": "",
                                "list_item_raw": {"candidate_id": candidate_id},
                                "detail_data_raw": {
                                    "ui_sections": ui_sections,
                                },
                                "ui_sections": ui_sections,
                            }
                        ],
                    }
                ],
            )
            imported = self.service.import_results_jsonl(
                results_path,
                evaluation_id="eval-import",
                run_label=label,
                system_version=f"{label}-v1",
                run_id=f"run-pair-{label}",
            )
            process = self.service.process_run(
                run_id=imported.object_id,
                schema_version=schema_version,
                baseline_version="baseline-pair",
                process_id=f"process-pair-{label}",
            )
            candidate_row = self.store.fetch_one(
                "SELECT candidate_pk FROM candidates WHERE run_id = ?",
                (imported.object_id,),
            )
            context = self.service.get_review_context(
                process.process_id,
                candidate_row["candidate_pk"],
            )
            self.service.save_review(
                process_id=process.process_id,
                candidate_pk=candidate_row["candidate_pk"],
                judgement="HIT",
                reason="SOCIAL_MATCH",
                evidence="链接一致",
                reviewer="tester",
                review_note="配对测试",
                field_scores=context["field_scores"],
                expected_reviewed_at=context["reviewed_at"],
            )
            process_ids.append(process.process_id)

        paired = self.service.compare_processes(
            process_ids[0],
            process_ids[1],
        )
        self.assertFalse(paired["formal_ready"])
        self.assertEqual(0, paired["same_condition"]["coverage"]["paired_count"])
        self.assertEqual(
            "INPUT_SIGNATURE_UNAVAILABLE",
            paired["not_comparable"]["queries"][0]["reason"],
        )

        dataset_ids = []
        for label, full_name in (
            ("baseline", "Person Pair"),
            ("candidate", "Person Pair"),
        ):
            dataset_path = self.root / f"dataset-pair-{label}.jsonl"
            write_jsonl(
                dataset_path,
                [
                    {
                        "input_id": f"query-{label}",
                        "person_id": "person-pair",
                        "query_stage": "FULL_NAME_SOCIAL",
                        "match_strategy": "UNION",
                        "clues": [
                            {"type": "FULL_NAME", "value": full_name},
                            {
                                "type": "SOCIAL_LINK",
                                "value": "https://social.example.test/person",
                            },
                        ],
                        "additional_details": [],
                    }
                ],
            )
            dataset = self.service.import_dataset_jsonl(
                dataset_path,
                name=f"{label} 配对输入",
                dataset_id=f"dataset-pair-{label}",
            )
            dataset_ids.append(dataset.object_id)
        with self.store.transaction() as connection:
            connection.execute(
                "UPDATE runs SET dataset_id = ? WHERE run_id = 'run-pair-baseline'",
                (dataset_ids[0],),
            )
            connection.execute(
                "UPDATE runs SET dataset_id = ? WHERE run_id = 'run-pair-candidate'",
                (dataset_ids[1],),
            )
        paired = self.service.compare_processes(process_ids[0], process_ids[1])
        self.assertTrue(paired["formal_ready"])
        self.assertEqual(1, paired["same_condition"]["category_counts"]["持续命中"])
        self.assertEqual(0, paired["same_condition"]["category_counts"]["新增命中"])
        self.assertEqual(
            {
                "person_id": "person-pair",
                "query_stage": "FULL_NAME_SOCIAL",
            },
            {
                "person_id": paired["same_condition"]["pairs"][0]["person_id"],
                "query_stage": paired["same_condition"]["pairs"][0]["query_stage"],
            },
        )
        self.assertIn(
            "matched_completeness_delta",
            paired["same_condition"]["pairs"][0],
        )

        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE dataset_queries
                SET clues_json = ?
                WHERE dataset_id = ? AND query_id = 'query-candidate'
                """,
                (
                    json.dumps(
                        [{"type": "FULL_NAME", "value": "Different Person"}]
                    ),
                    dataset_ids[1],
                ),
            )
        mismatched = self.service.compare_processes(
            process_ids[0],
            process_ids[1],
        )
        self.assertEqual(0, mismatched["same_condition"]["coverage"]["paired_count"])
        self.assertEqual(
            "INPUT_SIGNATURE_MISMATCH",
            mismatched["not_comparable"]["queries"][0]["reason"],
        )

    def test_stage6_report_model_snapshot_static_file_and_export_records(self):
        """阶段6报告共用模型可快照、配对、下钻并生成静态/Excel输入。"""

        definitions = [
            {
                "field_key": "social_urls",
                "display_name": "Social URLs",
                "module": "Social",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.social.data.profiles[*].url",
                "data_type": "array",
                "array_mode": "collect",
                "empty_rule": "default",
                "normalizer": "social_url",
                "scoring_role": ["identity", "completeness", "accuracy"],
                "compare_mode": "url_set",
                "enabled": True,
                "sort_order": 10,
            },
            {
                "field_key": "display_name",
                "display_name": "Display Name",
                "module": "Summary",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": "ui_sections.summary.data.display_name",
                "data_type": "string",
                "array_mode": "preserve",
                "empty_rule": "default",
                "normalizer": "trim_text",
                "scoring_role": ["completeness", "accuracy"],
                "compare_mode": "normalized_text",
                "enabled": True,
                "sort_order": 20,
            },
        ]
        schema_version = self.service.publish_field_schema(
            name="阶段6字段",
            definitions=definitions,
            schema_version="field-schema-stage6",
        )
        baseline_path = self.root / "baseline-stage6.jsonl"
        write_jsonl(
            baseline_path,
            [
                {
                    "person_id": "person-stage6",
                    "display_name": "Report Person",
                    "fields": {
                        "social_urls": [
                            "https://social.example.test/report-person"
                        ],
                        "display_name": "Report Person",
                    },
                    "evidence": {},
                }
            ],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段6基准",
            baseline_version="baseline-stage6",
        )

        process_ids = []
        for run_label in ("baseline", "candidate"):
            results_path = self.root / f"results-stage6-{run_label}.jsonl"
            records = []
            query_stages = (
                ("FULL_NAME",)
                if run_label == "baseline"
                else ("FULL_NAME", "FULL_NAME_SOCIAL")
            )
            for index, query_stage in enumerate(query_stages, start=1):
                returned_url = (
                    "https://social.example.test/report-person/"
                    if run_label == "candidate" or query_stage == "FULL_NAME"
                    else "https://social.example.test/other"
                )
                candidate_id = f"candidate-{run_label}-{index}"
                ui_sections = {
                    "social": {
                        "data": {
                            "profiles": [{"url": returned_url}],
                        }
                    },
                    "summary": {
                        "data": {"display_name": "Report Person"},
                    },
                }
                records.append(
                    {
                        "result_schema_version": "1.3",
                        "input_id": f"query-{run_label}-{index}",
                        "person_id": "person-stage6",
                        "query_stage": query_stage,
                        "task_id": f"task-{run_label}-{index}",
                        "query_status": "SUCCESS",
                        "results": [
                            {
                                "candidate_rank": 1,
                                "candidate_id": candidate_id,
                                "rank_score": 0.9,
                                "detail_status": "SUCCESS",
                                "detail_error": "",
                                "list_item_raw": {
                                    "candidate_id": candidate_id,
                                },
                                "detail_data_raw": {
                                    "ui_sections": ui_sections,
                                },
                                "ui_sections": ui_sections,
                            }
                        ],
                    }
                )
            write_jsonl(results_path, records)
            imported = self.service.import_results_jsonl(
                results_path,
                evaluation_id="eval-import",
                run_label=run_label,
                system_version=f"{run_label}-stage6",
                run_id=f"run-stage6-{run_label}",
            )
            dataset_path = self.root / f"dataset-stage6-{run_label}.jsonl"
            write_jsonl(
                dataset_path,
                [
                    {
                        "input_id": f"query-{run_label}-{index}",
                        "person_id": "person-stage6",
                        "query_stage": query_stage,
                        "match_strategy": "UNION",
                        "clues": [
                            {
                                "type": "FULL_NAME",
                                "value": "Report Person",
                            },
                            *(
                                [
                                    {
                                        "type": "SOCIAL_LINK",
                                        "value": (
                                            "https://social.example.test/"
                                            "report-person"
                                        ),
                                    }
                                ]
                                if query_stage == "FULL_NAME_SOCIAL"
                                else []
                            ),
                        ],
                        "additional_details": [],
                    }
                    for index, query_stage in enumerate(
                        query_stages,
                        start=1,
                    )
                ],
            )
            dataset = self.service.import_dataset_jsonl(
                dataset_path,
                name=f"阶段6 {run_label} Dataset",
                dataset_id=f"dataset-stage6-{run_label}",
            )
            with self.store.transaction() as connection:
                connection.execute(
                    "UPDATE runs SET dataset_id = ? WHERE run_id = ?",
                    (dataset.object_id, imported.object_id),
                )
            process = self.service.process_run(
                run_id=imported.object_id,
                schema_version=schema_version,
                baseline_version="baseline-stage6",
                process_id=f"process-stage6-{run_label}",
            )
            candidate_rows = self.store.fetch_all(
                """
                SELECT candidate_pk, query_id FROM candidates
                WHERE run_id = ? ORDER BY query_id
                """,
                (imported.object_id,),
            )
            for candidate_row in candidate_rows:
                context = self.service.get_review_context(
                    process.process_id,
                    candidate_row["candidate_pk"],
                )
                judgement = (
                    "NOT_HIT"
                    if run_label == "baseline"
                    and context["query_stage"] == "FULL_NAME_SOCIAL"
                    else "HIT"
                )
                self.service.save_review(
                    process_id=process.process_id,
                    candidate_pk=candidate_row["candidate_pk"],
                    judgement=judgement,
                    reason="MANUAL",
                    evidence="阶段6固定夹具",
                    reviewer="report-tester",
                    review_note="报告测试",
                    field_scores=context["field_scores"],
                    expected_reviewed_at=context["reviewed_at"],
                )
            process_ids.append(process.process_id)

        self.service.create_threshold_profile(
            profile_id="report-threshold-v1",
            name="报告参考线",
            description="阶段6报告测试",
            version=1,
            thresholds={
                "FULL_NAME": {
                    "min_retrieval_success": 0.5,
                    "max_average_total_cost": 1,
                },
                "FULL_NAME_SOCIAL": {
                    "min_retrieval_success": 0.5,
                },
            },
        )
        self.service.assign_evaluation_threshold_profile(
            "eval-import",
            "report-threshold-v1",
        )
        report = self.service.create_report(
            candidate_process_id=process_ids[1],
            baseline_process_id=process_ids[0],
            data_marker="MOCK",
            report_id="report-stage6",
        )
        self.assertEqual("COMPARE", report.model["metadata"]["report_type"])
        self.assertEqual(
            "report-model-v3",
            report.model["metadata"]["report_model_version"],
        )
        self.assertEqual(
            "report-threshold-v1",
            report.model["metadata"]["threshold_profile_id"],
        )
        self.assertEqual(
            "报告参考线",
            report.model["metadata"]["threshold_profile_name"],
        )
        self.assertEqual(
            1,
            report.model["metadata"]["threshold_profile_version"],
        )
        self.assertTrue(report.model["summary"]["formal_ready"])
        self.assertEqual(
            {"FULL_NAME", "FULL_NAME_SOCIAL"},
            set(report.model["query_stage_metrics"]),
        )
        self.assertEqual(
            1,
            report.model["comparison"]["new_clue"]["query_stage_metrics"][
                "FULL_NAME_SOCIAL"
            ]["query_count"],
        )
        self.assertIn(
            "reason_codes",
            report.model["comparison"]["new_clue"]["query_stage_metrics"][
                "FULL_NAME_SOCIAL"
            ]["quality_metrics"]["retrieval_success"],
        )
        self.assertEqual(
            1,
            report.model["comparison"]["same_condition"]["coverage"][
                "paired_count"
            ],
        )
        self.assertEqual(
            "PASS",
            report.model["threshold_assessment"]["stages"]["FULL_NAME"][
                "items"
            ]["min_retrieval_success"]["status"],
        )
        self.assertEqual(
            "NOT_READY",
            report.model["threshold_assessment"]["stages"]["FULL_NAME"][
                "items"
            ]["max_average_total_cost"]["status"],
        )
        self.assertEqual(
            "暂不能判断",
            report.model["threshold_assessment"]["recommendation"],
        )
        self.assertIn("Summary", report.model["module_metrics"])
        self.assertIn("display_name", report.model["field_metrics"])
        self.assertEqual(
            "candidate-candidate-2",
            report.model["case_groups"]["新增线索"][0][
                "candidate_candidate_id"
            ],
        )
        stored = self.store.fetch_one(
            "SELECT * FROM reports WHERE report_id = 'report-stage6'"
        )
        self.assertEqual("READY", stored["status"])
        self.assertEqual(
            report.model,
            json.loads(stored["metrics_json"]),
        )
        self.service.create_threshold_profile(
            profile_id="report-threshold-v2",
            name="报告参考线",
            description="后续新报告使用",
            version=2,
            thresholds={
                "FULL_NAME": {
                    "min_retrieval_success": 0.95,
                }
            },
            based_on_profile_id="report-threshold-v1",
        )
        self.service.assign_evaluation_threshold_profile(
            "eval-import",
            "report-threshold-v2",
        )
        unchanged_report = self.store.fetch_one(
            """
            SELECT status, metrics_json FROM reports
            WHERE report_id = 'report-stage6'
            """
        )
        self.assertEqual("READY", unchanged_report["status"])
        self.assertEqual(
            "report-threshold-v1",
            json.loads(unchanged_report["metrics_json"])["metadata"][
                "threshold_profile_id"
            ],
        )
        self.service.update_evaluation_thresholds(
            "eval-import",
            {
                "FULL_NAME": {
                    "min_retrieval_success": 0.95,
                }
            },
        )
        self.assertEqual(
            0.5,
            json.loads(
                self.store.fetch_one(
                    """
                    SELECT metrics_json FROM reports
                    WHERE report_id = 'report-stage6'
                    """
                )["metrics_json"]
            )["threshold_assessment"]["threshold_snapshot"]["FULL_NAME"][
                "min_retrieval_success"
            ],
        )
        report_dir = (
            self.root
            / "data"
            / "reports"
            / "eval-import"
            / "report-stage6"
        )
        self.assertTrue((report_dir / "report_model.json").is_file())
        export_records = [
            json.loads(line)
            for line in (report_dir / "processed_export.jsonl")
            .read_text(encoding="utf-8")
            .splitlines()
            if line
        ]
        self.assertEqual(
            2,
            sum(
                record["record_type"] == "candidate"
                for record in export_records
            ),
        )
        self.assertTrue(
            any(record["record_type"] == "query" for record in export_records)
        )
        html_path = self.service.save_report_html(
            report.report_id,
            "<!doctype html><title>阶段6报告</title>",
        )
        self.assertEqual(report.html_file, html_path)
        self.assertTrue(
            (report_dir / Path(report.html_file).name).is_file()
        )
        excel_path = self.service.export_report_excel(report.report_id)
        self.assertTrue(excel_path.endswith("_report.xlsx"))
        workbook = load_workbook(
            report_dir / Path(excel_path).name,
            read_only=True,
            data_only=True,
        )
        try:
            self.assertTrue(
                {
                    "说明",
                    "核心指标",
                    "Query明细",
                    "候选结果",
                    "同条件对比",
                    "新增线索",
                    "模块字段统计",
                    "失败记录",
                    "人工复核",
                }.issubset(workbook.sheetnames)
            )
            self.assertEqual(
                2,
                sum(
                    1
                    for _ in workbook["候选结果"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                ),
            )
            query_rows = list(
                workbook["Query明细"].iter_rows(values_only=True)
            )
            query_record = dict(zip(query_rows[0], query_rows[1]))
            self.assertEqual(
                "HAS_CANDIDATES",
                query_record["result_status"],
            )
            self.assertIsNone(query_record["llm_cost"])
            self.assertIsNone(query_record["third_party_cost"])
            self.assertIsNone(query_record["total_cost"])
            self.assertIsNone(query_record["pdl_called"])
            self.assertIsNone(query_record["search_duration_ms"])

            core_rows = list(
                workbook["核心指标"].iter_rows(values_only=True)
            )
            core_headers = core_rows[0]
            self.assertTrue(
                any(
                    row[core_headers.index("metric_key")]
                    == "retrieval_success"
                    for row in core_rows[1:]
                )
            )
        finally:
            workbook.close()

    def test_stage7_fixed_fixture_completes_import_to_report_pipeline(self):
        """固定脱敏夹具贯通导入、处理、复核、指标和报告导出。"""

        schema_version = self.service.publish_field_schema(
            name="阶段7端到端字段",
            schema_version="field-schema-stage7-e2e",
            definitions=[
                {
                    "field_key": "social_urls",
                    "display_name": "Social URLs",
                    "module": "Social",
                    "source_stage": "GetTaskCandidateDetail",
                    "source_path": "ui_sections.social.data.profiles[*].url",
                    "data_type": "array",
                    "array_mode": "collect",
                    "empty_rule": "default",
                    "normalizer": "social_url",
                    "scoring_role": [
                        "identity",
                        "completeness",
                        "accuracy",
                    ],
                    "compare_mode": "url_set",
                    "enabled": True,
                    "sort_order": 10,
                },
                {
                    "field_key": "future_label",
                    "display_name": "Future Label",
                    "module": "Profile",
                    "source_stage": "GetTaskCandidateDetail",
                    "source_path": "ui_sections.future_module.data.label",
                    "data_type": "string",
                    "array_mode": "preserve",
                    "empty_rule": "default",
                    "normalizer": "trim_text",
                    "scoring_role": ["completeness", "accuracy"],
                    "compare_mode": "normalized_text",
                    "enabled": True,
                    "sort_order": 20,
                },
            ],
        )
        baseline = self.service.import_baseline_jsonl(
            E2E_FIXTURE / "baseline.jsonl",
            name="阶段7固定基准",
            baseline_version="baseline-stage7-e2e",
        )
        self.assertEqual(2, baseline.imported_count)

        imported_runs = {}
        for label in ("baseline", "candidate"):
            imported_runs[label] = self.service.import_results_jsonl(
                E2E_FIXTURE / f"{label}_results.jsonl",
                failures_path=E2E_FIXTURE / f"{label}_failures.jsonl",
                metadata_path=(
                    E2E_FIXTURE / "candidate_metadata.jsonl"
                    if label == "candidate"
                    else None
                ),
                evaluation_id="eval-import",
                run_label=label,
                system_version=f"{label}-stage7",
                run_id=f"run-stage7-{label}",
            )
        self.assertEqual(4, imported_runs["baseline"].imported_count)
        self.assertEqual(4, imported_runs["candidate"].imported_count)
        for label, imported in imported_runs.items():
            dataset_path = self.root / f"dataset-stage7-{label}.jsonl"
            write_jsonl(
                dataset_path,
                [
                    {
                        "input_id": f"{label}-person-{person_index:03d}-{suffix}",
                        "person_id": f"person-e2e-{person_index:03d}",
                        "query_stage": query_stage,
                        "match_strategy": "UNION",
                        "clues": [
                            {
                                "type": "FULL_NAME",
                                "value": f"Example Person {person_index}",
                            },
                            *(
                                [
                                    {
                                        "type": "SOCIAL_LINK",
                                        "value": (
                                            "https://social.example.test/"
                                            f"person-{person_index}"
                                        ),
                                    }
                                ]
                                if query_stage == "FULL_NAME_SOCIAL"
                                else []
                            ),
                        ],
                        "additional_details": [],
                    }
                    for person_index in (1, 2)
                    for suffix, query_stage in (
                        ("name", "FULL_NAME"),
                        ("social", "FULL_NAME_SOCIAL"),
                    )
                ],
            )
            dataset = self.service.import_dataset_jsonl(
                dataset_path,
                name=f"阶段7 {label} Dataset",
                dataset_id=f"dataset-stage7-{label}",
            )
            with self.store.transaction() as connection:
                connection.execute(
                    "UPDATE runs SET dataset_id = ? WHERE run_id = ?",
                    (dataset.object_id, imported.object_id),
                )

        process_ids = {}
        for label, imported in imported_runs.items():
            process_ids[label] = self.service.process_run(
                run_id=imported.object_id,
                schema_version=schema_version,
                baseline_version="baseline-stage7-e2e",
                process_id=f"process-stage7-{label}",
            ).process_id

        future_fields = self.store.fetch_one(
            """
            SELECT pc.fields_json
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ?
              AND c.candidate_id = 'candidate-candidate-001-name'
            """,
            (process_ids["candidate"],),
        )
        self.assertEqual(
            "alpha",
            json.loads(future_fields["fields_json"])["future_label"],
        )
        detail_failure = self.store.fetch_one(
            """
            SELECT pc.processing_errors_json
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ?
              AND c.candidate_id = 'baseline-candidate-detail-failed'
            """,
            (process_ids["baseline"],),
        )
        self.assertEqual(
            "DETAIL_FAILED",
            json.loads(detail_failure["processing_errors_json"])[0]["code"],
        )

        judgements = {
            "baseline-candidate-001-name": "HIT",
            "baseline-candidate-001-social": "NOT_HIT",
            "baseline-candidate-002-name": "HIT",
            "baseline-candidate-002-social": "HIT",
            "candidate-candidate-001-name": "HIT",
            "candidate-candidate-001-social": "HIT",
            "candidate-candidate-002-name": "SUSPECTED",
        }
        for label, process_id in process_ids.items():
            candidates = self.store.fetch_all(
                """
                SELECT candidate_pk, candidate_id, detail_status
                FROM candidates WHERE run_id = ? ORDER BY candidate_rank
                """,
                (imported_runs[label].object_id,),
            )
            for candidate in candidates:
                if candidate["detail_status"] != "SUCCESS":
                    continue
                context = self.service.get_review_context(
                    process_id,
                    candidate["candidate_pk"],
                )
                self.service.save_review(
                    process_id=process_id,
                    candidate_pk=candidate["candidate_pk"],
                    judgement=judgements[candidate["candidate_id"]],
                    reason="MANUAL",
                    evidence="stage7 fixed fixture",
                    reviewer="stage7-tester",
                    review_note="端到端验收",
                    field_scores=context["field_scores"],
                    expected_reviewed_at=context["reviewed_at"],
                )

        comparison = self.service.compare_processes(
            process_ids["baseline"],
            process_ids["candidate"],
        )
        self.assertEqual(1, comparison["category_counts"]["新增命中"])
        self.assertEqual(2, comparison["category_counts"]["退化未命中"])
        self.assertTrue(comparison["formal_ready"])
        self.assertEqual(
            "NOT_CONNECTED",
            comparison["candidate_metrics"]["cost_status"]["status"],
        )
        failed_query = self.store.fetch_one(
            """
            SELECT person_id, query_stage, status
            FROM run_queries
            WHERE run_id = 'run-stage7-candidate'
              AND query_id = 'candidate-person-002-social'
            """
        )
        self.assertEqual(
            ("person-e2e-002", "FULL_NAME_SOCIAL", "FAILED"),
            (
                failed_query["person_id"],
                failed_query["query_stage"],
                failed_query["status"],
            ),
        )

        report = self.service.create_report(
            candidate_process_id=process_ids["candidate"],
            baseline_process_id=process_ids["baseline"],
            data_marker="MOCK",
            report_id="report-stage7-e2e",
        )
        self.assertTrue(
            report.model["field_alignment_matrix"]["fields"]
        )
        export_records = [
            json.loads(line)
            for line in (
                self.root
                / "data"
                / "reports"
                / "eval-import"
                / report.report_id
                / "processed_export.jsonl"
            ).read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        query_export = next(
            item for item in export_records
            if item["record_type"] == "query"
        )
        candidate_export = next(
            item for item in export_records
            if item["record_type"] == "candidate"
            and item["detail_status"] == "SUCCESS"
        )
        self.assertEqual("MATCHED", query_export["baseline_match_status"])
        self.assertIn(
            candidate_export["classification_source"],
            {"MANUAL", "RULE"},
        )
        html_path = self.service.save_report_html(
            report.report_id,
            "<!doctype html><meta charset=\"utf-8\">"
            "<title>searchTool v1.3 阶段7验收</title>",
        )
        excel_path = self.service.export_report_excel(report.report_id)
        report_directory = (
            self.root
            / "data"
            / "reports"
            / "eval-import"
            / report.report_id
        )
        self.assertTrue(
            (report_directory / Path(html_path).name).is_file()
        )
        workbook = load_workbook(
            report_directory / Path(excel_path).name,
            read_only=True,
            data_only=True,
        )
        try:
            self.assertIn("Query明细", workbook.sheetnames)
            self.assertIn("同条件对比", workbook.sheetnames)
            self.assertIn("新增线索", workbook.sheetnames)
            self.assertIn("模块字段统计", workbook.sheetnames)
            self.assertTrue(
                {
                    "Report_Summary",
                    "Query_Person_Links",
                    "Identity_Classification",
                    "Field_Matrix",
                    "Field_Metrics",
                    "Module_Metrics",
                    "Not_Ready_Reasons",
                }.issubset(workbook.sheetnames)
            )
            self.assertIn("future_label", next(
                workbook["候选结果"].iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                )
            ))
            self.assertEqual(
                3,
                sum(
                    1
                    for _ in workbook["候选结果"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                ),
            )
        finally:
            workbook.close()

    def test_stage7_synthetic_100_people_two_versions_capacity(self):
        """100人×两种条件×两个版本可导入且磁盘占用保持可控。"""

        started_at = time.monotonic()
        for version in ("baseline", "candidate"):
            records = []
            for person_index in range(1, 101):
                for query_stage in ("FULL_NAME", "FULL_NAME_SOCIAL"):
                    suffix = "name" if query_stage == "FULL_NAME" else "social"
                    query_id = f"{version}-person-{person_index:03d}-{suffix}"
                    candidate_id = f"{version}-candidate-{person_index:03d}-{suffix}"
                    ui_sections = {
                        "future_module": {
                            "data": {"label": f"person-{person_index:03d}"}
                        }
                    }
                    records.append(
                        {
                            "result_schema_version": "1.3",
                            "input_id": query_id,
                            "person_id": f"person-{person_index:03d}",
                            "query_stage": query_stage,
                            "task_id": f"task-{query_id}",
                            "query_status": "SUCCESS",
                            "candidate_count_total": 1,
                            "task_fields": {
                                "llm_cost": None,
                                "total_cost": None,
                                "pdl_called": None,
                            },
                            "raw": {
                                "create_intent_task": {
                                    "task_id": f"task-{query_id}"
                                }
                            },
                            "results": [
                                {
                                    "candidate_rank": 1,
                                    "candidate_id": candidate_id,
                                    "rank_score": 0.8,
                                    "detail_status": "SUCCESS",
                                    "detail_error": "",
                                    "list_item_raw": {
                                        "candidate_id": candidate_id
                                    },
                                    "detail_data_raw": {
                                        "ui_sections": ui_sections
                                    },
                                    "ui_sections": ui_sections,
                                }
                            ],
                        }
                    )
            source = self.root / f"stage7-scale-{version}.jsonl"
            write_jsonl(source, records)
            imported = self.service.import_results_jsonl(
                source,
                evaluation_id="eval-import",
                run_label=version,
                system_version=f"{version}-scale",
                run_id=f"run-stage7-scale-{version}",
            )
            self.assertEqual(200, imported.imported_count)

        elapsed_seconds = time.monotonic() - started_at
        self.assertEqual(
            400,
            self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM run_queries
                WHERE run_id LIKE 'run-stage7-scale-%'
                """
            )["count"],
        )
        self.assertEqual(
            400,
            self.store.fetch_one(
                """
                SELECT COUNT(*) AS count FROM candidates
                WHERE run_id LIKE 'run-stage7-scale-%'
                """
            )["count"],
        )
        self.assertLess(elapsed_seconds, 30)
        self.assertLess(
            self.store.db_path.stat().st_size,
            25 * 1024 * 1024,
        )


    def test_stage2_reprocess_and_query_classification_without_http(self):
        """重处理不调用接口，并支持 Query 级主命中、待判定和并发保护。"""

        schema_version = self.service.ensure_default_field_schema()
        baseline_path = self.root / "stage2-baseline.jsonl"
        write_jsonl(
            baseline_path,
            [{
                "person_id": "person-stage2",
                "display_name": "Stage2 Person",
                "fields": {"social_urls": ["https://example.test/stage2"]},
                "baseline_available_fields": ["social_urls"],
            }],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段2基准",
            baseline_version="baseline-stage2",
        )
        results_path = self.root / "stage2-results.jsonl"
        records = []
        for rank, candidate_id, rank_score, detail_status in (
            (1, "candidate-primary", 0.2, "SUCCESS"),
            (2, "candidate-secondary", 0.9, "SUCCESS"),
            (3, "candidate-failed", 0.1, "FAILED"),
        ):
            ui_sections = {
                "social": {
                    "data": {
                        "profiles": [
                            {"url": f"https://example.test/{candidate_id}"}
                        ]
                    }
                }
            }
            records.append({
                "candidate_rank": rank,
                "candidate_id": candidate_id,
                "detail_status": detail_status,
                "detail_error": "" if detail_status == "SUCCESS" else "failed",
                "rank_score": rank_score,
                "list_item_raw": {
                    "candidate_id": candidate_id,
                    "rank_score": rank_score,
                },
                "detail_data_raw": (
                    {"ui_sections": ui_sections}
                    if detail_status == "SUCCESS" else {}
                ),
                "ui_sections": ui_sections if detail_status == "SUCCESS" else {},
            })
        write_jsonl(
            results_path,
            [{
                "result_schema_version": "1.3",
                "input_id": "query-stage2",
                "person_id": "person-stage2",
                "query_stage": "FULL_NAME_SOCIAL",
                "task_id": "task-stage2",
                "query_status": "SUCCESS",
                "results": records,
            }],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="stage2",
            system_version="v1",
            run_id="run-stage2",
        )
        original = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
            baseline_version="baseline-stage2",
            process_id="process-stage2-original",
        )

        original_execute_run = self.service.execute_run
        self.service.execute_run = lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("重处理不允许调用 execute_run")
        )
        try:
            reprocessed = self.service.reprocess_existing_run(
                run_id=imported.object_id,
                schema_version=schema_version,
                baseline_version="baseline-stage2",
                process_id="process-stage2-new",
            )
        finally:
            self.service.execute_run = original_execute_run
        self.assertNotEqual(original.process_id, reprocessed.process_id)
        self.assertEqual(3, reprocessed.candidate_count)
        automatic_rows = self.store.fetch_all(
            """
            SELECT c.candidate_id, rv.classification_source, rv.reviewed_at,
                   rv.judgement, rv.is_primary_hit
            FROM reviews AS rv
            JOIN candidates AS c ON c.candidate_pk = rv.candidate_pk
            WHERE rv.process_id = ? ORDER BY c.candidate_rank
            """,
            (reprocessed.process_id,),
        )
        by_candidate_id = {
            row["candidate_id"]: row for row in automatic_rows
        }
        self.assertEqual("RULE", by_candidate_id["candidate-primary"]["classification_source"])
        self.assertEqual("NOT_HIT", by_candidate_id["candidate-primary"]["judgement"])
        self.assertIsNotNone(by_candidate_id["candidate-primary"]["reviewed_at"])
        self.assertEqual("RULE", by_candidate_id["candidate-secondary"]["classification_source"])
        self.assertIsNone(by_candidate_id["candidate-failed"]["reviewed_at"])
        self.assertEqual("SUGGESTED", by_candidate_id["candidate-failed"]["classification_source"])

        context = self.service.get_query_classification_context(
            reprocessed.process_id,
            "query-stage2",
        )
        by_id = {item["candidate_id"]: item for item in context["candidates"]}
        failed = by_id["candidate-failed"]
        with self.assertRaises(ReviewValidationError):
            self.service.save_query_classification(
                reprocessed.process_id,
                "query-stage2",
                classifications=[{
                    "candidate_pk": failed["candidate_pk"],
                    "judgement": "HIT",
                    "reason": "MANUAL",
                }],
                primary_hit_candidate_pk=failed["candidate_pk"],
                confirm_no_hit=False,
                expected_versions={
                    item["candidate_pk"]: item["reviewed_at"] or ""
                    for item in context["candidates"]
                },
            )

        primary = by_id["candidate-primary"]
        secondary = by_id["candidate-secondary"]
        with self.store.transaction() as connection:
            connection.execute(
                """
                INSERT INTO reports(
                    report_id, evaluation_id, baseline_process_id,
                    candidate_process_id, report_type, status, metrics_json,
                    html_file, excel_file, created_at
                ) VALUES (
                    'report-stage2', 'eval-import', NULL, ?,
                    'SINGLE', 'READY', '{}', 'report.html', NULL, ?
                )
                """,
                (reprocessed.process_id, "2026-07-28T00:00:00+00:00"),
            )
        saved = self.service.save_query_classification(
            reprocessed.process_id,
            "query-stage2",
            classifications=[
                {
                    "candidate_pk": primary["candidate_pk"],
                    "judgement": "HIT",
                    "reason": "MANUAL",
                    "evidence": "人工确认",
                },
                {
                    "candidate_pk": secondary["candidate_pk"],
                    "judgement": "HIT",
                    "reason": "MANUAL",
                    "evidence": "重复命中",
                },
            ],
            primary_hit_candidate_pk=primary["candidate_pk"],
            confirm_no_hit=False,
            reviewer="stage2",
            review_note="仅确认主命中",
            expected_versions={
                item["candidate_pk"]: item["reviewed_at"] or ""
                for item in context["candidates"]
            },
        )
        self.assertEqual("HIT_CONFIRMED", saved["identity_state"])
        self.assertEqual(0, saved["pending_count"])
        primary_flags = self.store.fetch_all(
            """
            SELECT c.candidate_id, rv.is_primary_hit
            FROM reviews AS rv
            JOIN candidates AS c ON c.candidate_pk = rv.candidate_pk
            WHERE rv.process_id = ?
              AND c.candidate_id IN ('candidate-primary', 'candidate-secondary')
            ORDER BY c.candidate_id
            """,
            (reprocessed.process_id,),
        )
        self.assertEqual(
            {"candidate-primary": 0, "candidate-secondary": 1},
            {row["candidate_id"]: row["is_primary_hit"] for row in primary_flags},
        )
        self.assertEqual(
            "STALE",
            self.store.fetch_one(
                "SELECT status FROM reports WHERE report_id = 'report-stage2'"
            )["status"],
        )
        with self.assertRaises(ReviewValidationError):
            self.service.save_query_classification(
                reprocessed.process_id,
                "query-stage2",
                classifications=[{
                    "candidate_pk": secondary["candidate_pk"],
                    "judgement": "NOT_HIT",
                    "reason": "MANUAL",
                }],
                primary_hit_candidate_pk=primary["candidate_pk"],
                confirm_no_hit=False,
                expected_versions={primary["candidate_pk"]: ""},
            )

        refreshed = self.service.get_query_classification_context(
            reprocessed.process_id,
            "query-stage2",
        )
        versions = {
            item["candidate_pk"]: item["reviewed_at"] or ""
            for item in refreshed["candidates"]
        }
        no_hit = self.service.save_query_classification(
            reprocessed.process_id,
            "query-stage2",
            classifications=[
                {
                    "candidate_pk": primary["candidate_pk"],
                    "judgement": "NOT_HIT",
                    "reason": "MANUAL",
                },
                {
                    "candidate_pk": secondary["candidate_pk"],
                    "judgement": "NOT_HIT",
                    "reason": "MANUAL",
                },
            ],
            primary_hit_candidate_pk=None,
            confirm_no_hit=True,
            expected_versions=versions,
        )
        self.assertEqual("NO_HIT_CONFIRMED", no_hit["identity_state"])
        self.assertEqual(0, no_hit["pending_count"])

    def test_stage3_field_matrix_and_processing_v3_module_semantics(self):
        """字段矩阵发现冲突，processing-v3 统一 URL 并按模块语义判空。"""

        definitions = [
            dict(item)
            for item in DEFAULT_FIELD_DEFINITIONS
            if item["field_key"] in {
                "task_id",
                "photos_status",
                "photos_data",
                "social_status",
                "social_urls",
                "summary_social_links",
            }
        ]
        for definition in definitions:
            if definition["field_key"] == "summary_social_links":
                definition["source_path"] = (
                    "ui_sections.summary.data.social_links[*].url"
                )
                definition["normalizer"] = "social_url"
                definition["compare_mode"] = "url_set"
                definition["scoring_role"] = ["completeness", "accuracy"]
            if definition["field_key"] == "photos_data":
                definition["scoring_role"] = ["completeness"]
                definition["compare_mode"] = "manual"
        schema_version = self.service.publish_field_schema(
            name="阶段3字段配置",
            definitions=definitions,
            schema_version="field-schema-stage3",
        )
        baseline_path = self.root / "stage3-baseline.jsonl"
        write_jsonl(
            baseline_path,
            [{
                "person_id": "person-stage3",
                "display_name": "Stage3 Person",
                "fields": {
                    "social_urls": ["https://x.com/stage3"],
                    "summary_social_links": ["https://x.com/stage3"],
                    "baseline_only_field": "prepared",
                },
                "baseline_available_fields": [
                    "social_urls",
                    "summary_social_links",
                    "baseline_only_field",
                ],
            }],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段3基准",
            baseline_version="baseline-stage3",
        )
        before = self.service.build_field_comparison_matrix(
            schema_version,
            "baseline-stage3",
        )
        before_by_key = {
            item["field_key"]: item for item in before["fields"]
        }
        self.assertEqual(
            "BASELINE_ENABLED_NOT_EXTRACTED",
            before_by_key["baseline_only_field"]["status"],
        )
        self.assertTrue(
            any(
                issue["field_key"] == "baseline_only_field"
                and issue["severity"] == "ERROR"
                for issue in self.service.validate_process_field_alignment(
                    schema_version,
                    "baseline-stage3",
                )
            )
        )

        results_path = self.root / "stage3-results.jsonl"
        ui_sections = {
            "photos": {
                "status": "empty",
                "data": {"count": 0, "image_urls": []},
            },
            "social": {
                "status": "data",
                "data": {
                    "profiles": [{
                        "url": "https://x.com/stage3/",
                        "platform": "x",
                    }]
                },
            },
            "summary": {
                "data": {
                    "social_links": [{
                        "platform": "x",
                        "url": "https://x.com/stage3/",
                    }]
                }
            },
        }
        write_jsonl(
            results_path,
            [{
                "result_schema_version": "1.3",
                "input_id": "query-stage3",
                "person_id": "person-stage3",
                "query_stage": "FULL_NAME_SOCIAL",
                "task_id": "task-stage3",
                "query_status": "SUCCESS",
                "results": [{
                    "candidate_rank": 1,
                    "candidate_id": "candidate-stage3",
                    "detail_status": "SUCCESS",
                    "detail_error": "",
                    "list_item_raw": {"candidate_id": "candidate-stage3"},
                    "detail_data_raw": {"ui_sections": ui_sections},
                    "ui_sections": ui_sections,
                }],
            }],
        )
        imported = self.service.import_results_jsonl(
            results_path,
            evaluation_id="eval-import",
            run_label="stage3",
            system_version="v3",
            run_id="run-stage3",
        )
        process = self.service.process_run(
            run_id=imported.object_id,
            schema_version=schema_version,
            baseline_version="baseline-stage3",
            process_id="process-stage3",
        )
        self.assertEqual(
            "field-processing-v4",
            self.store.fetch_one(
                "SELECT rule_version FROM process_runs WHERE process_id = ?",
                (process.process_id,),
            )["rule_version"],
        )
        candidate = self.store.fetch_one(
            """
            SELECT pc.* FROM processed_candidates AS pc
            WHERE pc.process_id = ?
            """,
            (process.process_id,),
        )
        fields = json.loads(candidate["fields_json"])
        empty_fields = json.loads(candidate["empty_fields_json"])
        self.assertEqual(
            ["https://twitter.com/stage3"],
            fields["summary_social_links"],
        )
        self.assertTrue(empty_fields["photos_data"])
        after = self.service.build_field_comparison_matrix(
            schema_version,
            "baseline-stage3",
            process_id=process.process_id,
        )
        after_by_key = {item["field_key"]: item for item in after["fields"]}
        self.assertEqual(
            1.0,
            after_by_key["summary_social_links"]["candidate_return_rate"],
        )
        self.assertEqual(
            "COMPARABLE",
            after_by_key["summary_social_links"]["status"],
        )

    def test_stage3_field_matrix_handles_500_fields_without_n_plus_one(self):
        """500字段矩阵在固定查询次数下完成，验证首版规模基础。"""

        definitions = [
            {
                "field_key": f"field_{index:03d}",
                "display_name": f"Field {index:03d}",
                "module": "Summary",
                "source_stage": "GetTaskCandidateDetail",
                "source_path": f"ui_sections.summary.data.field_{index:03d}",
                "data_type": "string",
                "array_mode": "preserve",
                "empty_rule": "default",
                "normalizer": "trim_text",
                "scoring_role": ["completeness"],
                "compare_mode": "normalized_text",
                "enabled": True,
                "sort_order": index,
                "value_scope": "CANDIDATE",
                "missing_policy": "EMPTY",
            }
            for index in range(500)
        ]
        self.service.publish_field_schema(
            name="阶段3 500字段配置",
            definitions=definitions,
            schema_version="field-schema-stage3-500",
        )
        baseline_path = self.root / "stage3-500-baseline.jsonl"
        write_jsonl(
            baseline_path,
            [{
                "person_id": "person-stage3-500",
                "display_name": "Scale Person",
                "fields": {
                    f"field_{index:03d}": f"value-{index}"
                    for index in range(500)
                },
                "baseline_available_fields": [
                    f"field_{index:03d}" for index in range(500)
                ],
            }],
        )
        self.service.import_baseline_jsonl(
            baseline_path,
            name="阶段3 500字段基准",
            baseline_version="baseline-stage3-500",
        )
        started_at = time.monotonic()
        matrix = self.service.build_field_comparison_matrix(
            "field-schema-stage3-500",
            "baseline-stage3-500",
        )
        self.assertEqual(500, len(matrix["fields"]))
        self.assertLess(time.monotonic() - started_at, 2)


if __name__ == "__main__":
    unittest.main()
