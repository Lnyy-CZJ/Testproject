"""Tests for the offline JSONL-to-Excel exporter.

The tests use synthetic data and never read credentials, call HTTP APIs, or mutate
the real output directory.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from result_to_excel import prepare_arguments

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORTER = PROJECT_ROOT / "result_to_excel.py"
BASELINE_FIXTURE = PROJECT_ROOT / "tests" / "fixtures" / "v1_3_baseline"


def write_jsonl(path: Path, records: list[dict]) -> None:
    """Write synthetic JSONL records used by a test.

    Args:
        path: Destination JSONL file.
        records: JSON objects to write, one per line.

    Returns:
        None.
    """

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(record, ensure_ascii=False) + "\n" for record in records),
        encoding="utf-8",
    )


def candidate(
    candidate_id: str,
    profile_label: str = "Full Name",
    photo_payload: str = "",
    rank_score: float | None = None,
) -> dict:
    """Build a representative candidate fixture.

    Args:
        candidate_id: Candidate identifier used for traceability checks.
        profile_label: Dynamic Profile label included in this candidate.
        photo_payload: Optional large value used to exercise Raw sheet splitting.
        rank_score: Optional ranking score returned by ListTaskCandidates.

    Returns:
        A result item with all five ui_sections modules.
    """

    photos_data = {
        "baseline_photo_url": "https://example.test/base.jpg",
        "identity_match_rate": 0.75,
        "authenticity_photos": [{"url": "https://example.test/a.jpg"}],
        "match_photos": [],
    }
    if photo_payload:
        photos_data["authenticity_photos"] = [{"payload": photo_payload}]

    result = {
        "candidate_id": candidate_id,
        "ui_sections": {
            "insights": {
                "status": "data",
                "data": {
                    "count": 1,
                    "items": [
                        {
                            "description": "Known public profile",
                            "links": [
                                {
                                    "platform": "web",
                                    "title": "Evidence",
                                    "type": "source",
                                    "url": "https://example.test/evidence",
                                }
                            ],
                        }
                    ],
                },
            },
            "photos": {"status": "data", "data": photos_data},
            "profile": {
                "status": "data",
                "data": {
                    "sections": [
                        {
                            "title": "Identity",
                            "items": [{"label": profile_label, "value": "Example Person"}],
                        }
                    ]
                },
            },
            "social": {
                "status": "data",
                "data": {
                    "profiles": [
                        {
                            "display_handle": "first",
                            "platform": "linkedin",
                            "url": "https://linkedin.test/first",
                        },
                        {
                            "display_handle": "second",
                            "platform": "x",
                            "url": "https://x.test/second",
                        },
                    ]
                },
            },
            "summary": {
                "status": "data",
                "data": {
                    "avatar_url": "https://example.test/avatar.jpg",
                    "confidence_level": "HIGH",
                    "primary_image": {"url": "https://example.test/primary.jpg"},
                    "social_links": [
                        {
                            "platform": "linkedin",
                            "title": "LinkedIn",
                            "url": "https://linkedin.test/first",
                        }
                    ],
                    "web_links": [
                        {
                            "platform": "web",
                            "title": "Official site",
                            "url": "https://example.test/profile",
                        }
                    ],
                    "display_name": "Example Person",
                    "location": "Shanghai",
                    "match_score": 95,
                    "is_top_result": True,
                    "is_best_match": True,
                },
            },
        },
    }
    if rank_score is not None:
        result["rank_score"] = rank_score
    return result


class ResultToExcelTests(unittest.TestCase):
    """End-to-end tests for model construction and workbook export."""

    def test_explicit_run_directory_prevents_env_result_file_injection(self):
        """验证显式 Run 目录优先于 .env 中配置的结果文件。"""

        with tempfile.TemporaryDirectory() as temp_dir:
            env_file = Path(temp_dir) / ".env"
            env_file.write_text(
                "EXCEL_RESULTS_FILE=output/unrelated_results.jsonl\n",
                encoding="utf-8",
            )
            with patch.dict(os.environ, {}, clear=True):
                prepared = prepare_arguments(
                    [
                        "single",
                        "--run-dir",
                        "output/explicit-run",
                        "--env-file",
                        str(env_file),
                    ]
                )

        self.assertIn("--run-dir", prepared)
        self.assertNotIn("--results-file", prepared)

    def test_phase0_baseline_fixture_remains_exportable(self):
        """验证阶段0冻结的当前 JSONL 结构可以稳定生成导出模型。"""

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            model_path = root / "model.json"
            self.run_exporter(
                [
                    "single",
                    "--results-file",
                    str(BASELINE_FIXTURE / "results.jsonl"),
                    "--failures-file",
                    str(BASELINE_FIXTURE / "failures.jsonl"),
                    "--run-label",
                    "baseline",
                    "--system-version",
                    "phase0",
                    "--evaluation-id",
                    "eval-phase0",
                    "--metadata",
                    str(BASELINE_FIXTURE / "query_metadata.jsonl"),
                    "--output",
                    str(root / "baseline.xlsx"),
                ],
                model_path,
            )

            model = json.loads(model_path.read_text(encoding="utf-8"))

        self.assertEqual(2, len(model["candidateRows"]))
        self.assertEqual(3, len(model["queryRows"]))
        self.assertEqual(1, len(model["failureRows"]))
        self.assertEqual(
            [1, 2],
            [row["candidate_rank"] for row in model["candidateRows"]],
        )

    def run_exporter(
        self,
        arguments: list[str],
        model_path: Path,
        *,
        create_workbook: bool = False,
    ) -> subprocess.CompletedProcess[str]:
        """Run the public Python entry point with a test-only model sidecar.

        Args:
            arguments: CLI arguments after the script name.
            model_path: Path receiving the internal model for deterministic assertions.
            create_workbook: Whether to run the slower artifact-tool export path.

        Returns:
            Completed subprocess result.

        Raises:
            AssertionError: When the exporter returns a non-zero status.
        """

        env = os.environ.copy()
        env["SEARCHTOOL_MODEL_OUTPUT"] = str(model_path)
        if not create_workbook:
            env["SEARCHTOOL_SKIP_WORKBOOK"] = "1"
        result = subprocess.run(
            [sys.executable, str(EXPORTER), *arguments],
            cwd=PROJECT_ROOT,
            env=env,
            text=True,
            capture_output=True,
            timeout=120,
            check=False,
        )
        self.assertEqual(0, result.returncode, msg=f"stdout={result.stdout}\nstderr={result.stderr}")
        return result

    def test_single_run_extracts_requested_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            run_dir = root / "current"
            write_jsonl(
                run_dir / "results.jsonl",
                [
                    {
                        "input_id": "query-1",
                        "task_id": "task-1",
                        "candidate_count_total": 8,
                        "results": [candidate("c1", photo_payload="R" * 33000, rank_score=0.91)],
                    }
                ],
            )
            write_jsonl(run_dir / "failures.jsonl", [])
            metadata_path = root / "metadata.jsonl"
            write_jsonl(
                metadata_path,
                [
                    {
                        "query_id": "query-1",
                        "person_id": "person-1",
                        "query_type": "Q1",
                        "person_group": "C",
                        "difficulty": "medium",
                        "tags": ["common_name", "few_clues"],
                    }
                ],
            )
            output_path = root / "single.xlsx"
            model_path = root / "model.json"

            self.run_exporter(
                [
                    "single",
                    "--results-file",
                    str(run_dir / "results.jsonl"),
                    "--failures-file",
                    str(run_dir / "failures.jsonl"),
                    "--run-label",
                    "current",
                    "--system-version",
                    "v1",
                    "--evaluation-id",
                    "eval-1",
                    "--metadata",
                    str(metadata_path),
                    "--output",
                    str(output_path),
                ],
                model_path,
                create_workbook=True,
            )

            model = json.loads(model_path.read_text(encoding="utf-8"))
            row = model["candidateRows"][0]
            self.assertEqual("query-1", row["query_id"])
            self.assertEqual("person-1", row["person_id"])
            self.assertEqual(8, row["candidate_count_total"])
            self.assertEqual(1, row["candidate_rank"])
            self.assertEqual(0.91, row["rank_score"])
            self.assertEqual("Known public profile", row["insights_description"])
            self.assertEqual(
                "Evidence：https://example.test/evidence",
                row["insights_links"],
            )
            self.assertEqual("data", row["photos_status"])
            self.assertEqual("Example Person", row["profile.Identity.Full Name"])
            self.assertEqual("first\nsecond", row["social_display_handles"])
            self.assertEqual("linkedin\nx", row["social_platforms"])
            self.assertEqual("HIGH", row["summary_confidence_level"])
            self.assertEqual("https://example.test/primary.jpg", row["summary_primary_image_url"])
            self.assertEqual(
                "LinkedIn：https://linkedin.test/first",
                row["summary_social_links"],
            )
            self.assertEqual(
                "Official site：https://example.test/profile",
                row["summary_web_links"],
            )
            self.assertEqual(
                model["candidateHeaders"].index("task_id") + 1,
                model["candidateHeaders"].index("candidate_id"),
            )
            for removed_header in [
                "system_version",
                "person_group",
                "difficulty",
                "tags",
                "insights_data",
                "photos_data",
                "profile_data",
                "social_profiles",
                "summary_display_name",
                "identity_judgement",
                "review_comment",
            ]:
                self.assertNotIn(removed_header, model["candidateHeaders"])
            field_catalog = {item["当前表头"]: item for item in model["fieldCatalogRows"]}
            self.assertEqual("整数或空值", field_catalog["candidate_count_total"]["内容格式"])
            self.assertIn("ListTaskCandidates", field_catalog["rank_score"]["来源路径/生成规则"])
            self.assertEqual("是", field_catalog["rank_score"]["是否保留"])
            self.assertEqual("否", field_catalog["system_version"]["是否保留"])
            self.assertIn("title：url", field_catalog["insights_links"]["处理规则/备注"])
            self.assertTrue(output_path.is_file())
            with zipfile.ZipFile(output_path) as archive:
                self.assertIn("xl/workbook.xml", archive.namelist())
                workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
                self.assertIn("Raw数据", workbook_xml)

    def test_single_mode_reads_file_settings_from_env_file(self):
        """Explicit .env export settings can replace all single-mode file arguments."""

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            run_dir = root / "env-run"
            write_jsonl(
                run_dir / "20260722_tasks_v01_results.jsonl",
                [{"input_id": "query-env", "task_id": "task-env", "results": []}],
            )
            write_jsonl(run_dir / "20260722_tasks_v01_failures.jsonl", [])
            output_path = root / "from-env.xlsx"
            model_path = root / "model.json"
            env_path = root / ".env"
            env_path.write_text(
                "\n".join(
                    [
                        f"EXCEL_RESULTS_FILE={run_dir / '20260722_tasks_v01_results.jsonl'}",
                        f"EXCEL_FAILURES_FILE={run_dir / '20260722_tasks_v01_failures.jsonl'}",
                        f"EXCEL_OUTPUT_FILE={output_path}",
                        "EXCEL_RUN_LABEL=current",
                        "EXCEL_SYSTEM_VERSION=v-env",
                        "EXCEL_EVALUATION_ID=eval-env",
                    ]
                ),
                encoding="utf-8",
            )

            env = os.environ.copy()
            env["SEARCHTOOL_MODEL_OUTPUT"] = str(model_path)
            env["SEARCHTOOL_SKIP_WORKBOOK"] = "1"
            result = subprocess.run(
                [sys.executable, str(EXPORTER), "single", "--env-file", str(env_path)],
                cwd=PROJECT_ROOT,
                env=env,
                text=True,
                capture_output=True,
                timeout=30,
                check=False,
            )

            self.assertEqual(0, result.returncode, msg=result.stderr)
            model = json.loads(model_path.read_text(encoding="utf-8"))
            self.assertEqual("query-env", model["queryRows"][0]["query_id"])

    def test_compare_uses_union_of_profile_columns_and_aligns_queries(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            baseline = root / "baseline"
            candidate_run = root / "candidate"
            write_jsonl(
                baseline / "results.jsonl",
                [{"input_id": "query-1", "task_id": "task-b", "results": [candidate("b1", "Full Name")]}],
            )
            write_jsonl(baseline / "failures.jsonl", [])
            write_jsonl(
                candidate_run / "results.jsonl",
                [{"input_id": "query-1", "task_id": "task-c", "results": [candidate("c1", "Location")]}],
            )
            write_jsonl(candidate_run / "failures.jsonl", [])
            output_path = root / "compare.xlsx"
            model_path = root / "model.json"

            self.run_exporter(
                [
                    "compare",
                    "--baseline-dir",
                    str(baseline),
                    "--baseline-version",
                    "base",
                    "--candidate-dir",
                    str(candidate_run),
                    "--candidate-version",
                    "next",
                    "--evaluation-id",
                    "eval-2",
                    "--output",
                    str(output_path),
                ],
                model_path,
            )

            model = json.loads(model_path.read_text(encoding="utf-8"))
            self.assertEqual(
                ["profile.Identity.Full Name", "profile.Identity.Location"],
                model["profileColumns"],
            )
            self.assertEqual(["baseline", "candidate"], [row["run_label"] for row in model["candidateRows"]])
            query_row = model["queryRows"][0]
            self.assertEqual("SUCCESS", query_row["baseline_status"])
            self.assertEqual("SUCCESS", query_row["candidate_status"])
            self.assertEqual(1, query_row["baseline_candidate_count"])
            self.assertEqual(1, query_row["candidate_candidate_count"])

    def test_processed_mode_uses_dynamic_fields_reviews_and_raw_sheet(self):
        """阶段6 processed 模式展开 ReportModel v2 并保留复核和超长值。"""

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            processed_path = root / "processed_export.jsonl"
            report_model_path = root / "report_model.json"
            output_path = root / "processed.xlsx"
            model_path = root / "model.json"
            long_value = "长" * 33000
            write_jsonl(
                processed_path,
                [
                    {
                        "record_type": "candidate",
                        "evaluation_id": "eval-processed",
                        "process_id": "process-processed",
                        "run_id": "run-processed",
                        "run_label": "candidate",
                        "system_version": "v-next",
                        "query_id": "query-processed",
                        "person_id": "person-processed",
                        "query_stage": "FULL_NAME_SOCIAL",
                        "task_id": "task-processed",
                        "query_status": "SUCCESS",
                        "candidate_pk": "candidate-pk-1",
                        "candidate_id": "candidate-1",
                        "candidate_rank": 1,
                        "rank_score": 0.95,
                        "detail_status": "SUCCESS",
                        "judgement": "HIT",
                        "reason": "SOCIAL_MATCH",
                        "reviewer": "tester",
                        "review_note": "已确认",
                        "reviewed_at": "2026-07-24T00:00:00+00:00",
                        "fields": {
                            "display_name": "=unsafe",
                            "profile_data": long_value,
                        },
                        "empty_fields": {
                            "display_name": False,
                            "profile_data": False,
                        },
                        "field_scores": {
                            "display_name": {
                                "completeness_score": 1.0,
                                "accuracy_score": 1.0,
                            }
                        },
                        "processing_errors": [],
                    },
                    {
                        "record_type": "candidate",
                        "evaluation_id": "eval-processed",
                        "process_id": "process-processed",
                        "run_id": "run-processed",
                        "run_label": "candidate",
                        "system_version": "v-next",
                        "query_id": "query-processed",
                        "person_id": "person-processed",
                        "query_stage": "FULL_NAME_SOCIAL",
                        "task_id": "task-processed",
                        "query_status": "SUCCESS",
                        "candidate_pk": "candidate-pk-2",
                        "candidate_id": "candidate-2",
                        "candidate_rank": 2,
                        "rank_score": 0.5,
                        "detail_status": "FAILED",
                        "judgement": "PENDING_REVIEW",
                        "reason": "NO_STRONG_FIELD",
                        "reviewer": "",
                        "review_note": "",
                        "reviewed_at": None,
                        "fields": {},
                        "empty_fields": {},
                        "field_scores": {},
                        "processing_errors": [
                            {
                                "code": "DETAIL_FAILED",
                                "error": "timeout",
                            }
                        ],
                    },
                    {
                        "record_type": "query",
                        "run_id": "run-processed",
                        "run_label": "candidate",
                        "system_version": "v-next",
                        "evaluation_phase": "PHASE_2_POST_OPTIMIZATION",
                        "query_id": "query-processed",
                        "person_id": "person-processed",
                        "query_stage": "FULL_NAME_SOCIAL",
                        "task_id": "task-processed",
                        "query_status": "SUCCESS",
                        "result_status": "HAS_CANDIDATES",
                        "candidate_count_total": 2,
                        "candidate_count_listed": 2,
                        "detail_success_count": 1,
                        "detail_failure_count": 1,
                        "llm_cost": 1.25,
                        "third_party_cost": None,
                        "total_cost": 3.75,
                        "pdl_called": False,
                        "search_duration_ms": 1250,
                        "retrieval_success": True,
                        "matched_completeness": 0.1,
                        "matched_accuracy": 1.0,
                        "formal_ready": True,
                    },
                    {
                        "record_type": "failure",
                        "query_id": "query-processed",
                        "candidate_id": "candidate-2",
                        "scope": "CANDIDATE",
                        "stage": "GetTaskCandidateDetail",
                        "error": "timeout",
                        "created_at": "2026-07-24T00:00:00Z",
                    },
                ],
            )
            report_model_path.write_text(
                json.dumps(
                    {
                        "metadata": {
                            "report_id": "report-processed",
                            "evaluation_id": "eval-processed",
                            "report_type": "SINGLE",
                            "report_model_version": "report-model-v2",
                            "metrics_rule_version": "metrics-v2",
                            "evaluation_phase": "PHASE_2_POST_OPTIMIZATION",
                            "candidate_system_version": "v-next",
                            "generated_at": "2026-07-24T00:00:00+00:00",
                        },
                        "summary": {"formal_ready": True},
                        "result_status_metrics": {
                            "total_formal_queries": 1,
                            "has_candidates_count": 1,
                            "no_candidates_count": 0,
                            "execution_failed_count": 0,
                            "has_result_rate": 1.0,
                            "no_result_rate": 0.0,
                            "execution_failed_rate": 0.0,
                        },
                        "quality_metrics": {
                            "retrieval_success": {
                                "status": "READY",
                                "numerator": 1,
                                "denominator": 1,
                                "value": 1.0,
                                "preview_value": 1.0,
                            },
                            "matched_completeness": {
                                "status": "READY",
                                "numerator": 0.1,
                                "denominator": 1,
                                "value": 0.1,
                                "preview_value": 0.1,
                            },
                            "matched_accuracy": {
                                "status": "READY",
                                "numerator": 1,
                                "denominator": 1,
                                "value": 1.0,
                                "preview_value": 1.0,
                            },
                        },
                        "cost_metrics": {
                            "llm_cost": {
                                "status": "COMPLETE",
                                "task_count": 1,
                                "value_count": 1,
                                "missing_count": 0,
                                "invalid_count": 0,
                                "total": 1.25,
                                "average": 1.25,
                                "minimum": 1.25,
                                "maximum": 1.25,
                            },
                            "third_party_cost": {
                                "status": "NOT_CONNECTED",
                                "task_count": 1,
                                "value_count": 0,
                                "missing_count": 1,
                                "invalid_count": 0,
                                "total": None,
                                "average": None,
                                "minimum": None,
                                "maximum": None,
                            },
                            "total_cost": {
                                "status": "COMPLETE",
                                "task_count": 1,
                                "value_count": 1,
                                "missing_count": 0,
                                "invalid_count": 0,
                                "total": 3.75,
                                "average": 3.75,
                                "minimum": 3.75,
                                "maximum": 3.75,
                            },
                            "search_duration_ms": {
                                "status": "COMPLETE",
                                "task_count": 1,
                                "value_count": 1,
                                "missing_count": 0,
                                "invalid_count": 0,
                                "total": 1250,
                                "average": 1250,
                                "minimum": 1250,
                                "maximum": 1250,
                            },
                        },
                        "pdl_metrics": {
                            "true_count": 0,
                            "false_count": 1,
                            "unknown_count": 0,
                            "known_count": 1,
                            "call_rate": 0.0,
                        },
                        "confidence_metrics": {
                            "overall": {"HIGH": 1, "UNKNOWN": 1},
                            "matched": {"HIGH": 1},
                            "nonmatched": {"UNKNOWN": 1},
                        },
                        "threshold_assessment": {
                            "recommendation": "建议上线",
                            "recommendation_code": "RECOMMEND_RELEASE",
                            "stages": {
                                "FULL_NAME_SOCIAL": {
                                    "query_count": 1,
                                    "items": {
                                        "min_retrieval_success": {
                                            "threshold": 0.8,
                                            "actual": 1.0,
                                            "direction": "MINIMUM",
                                            "status": "PASS",
                                            "reason": "达到参考线",
                                        }
                                    },
                                }
                            },
                        },
                        "comparison": {
                            "same_condition": {
                                "pairs": [
                                    {
                                        "person_id": "person-processed",
                                        "query_stage": "FULL_NAME_SOCIAL",
                                        "baseline_query_id": "query-old",
                                        "candidate_query_id": "query-processed",
                                        "baseline_hit": False,
                                        "candidate_hit": True,
                                        "category": "新增命中",
                                        "baseline_matched_completeness": 0.0,
                                        "candidate_matched_completeness": 0.1,
                                        "matched_completeness_delta": 0.1,
                                        "baseline_matched_accuracy": None,
                                        "candidate_matched_accuracy": 1.0,
                                        "matched_accuracy_delta": None,
                                        "baseline_total_cost": None,
                                        "candidate_total_cost": 3.75,
                                        "total_cost_delta": None,
                                    }
                                ]
                            },
                            "new_clue": {
                                "queries": [
                                    {
                                        "person_id": "person-new",
                                        "query_stage": "FULL_NAME_SOCIAL",
                                        "candidate_query_id": "query-new",
                                        "result_status": "NO_CANDIDATES",
                                        "retrieval_success": False,
                                        "matched_completeness": None,
                                        "matched_accuracy": None,
                                        "candidate_confidence": None,
                                        "candidate_count": 0,
                                        "task_fields": {
                                            "llm_cost": None,
                                            "third_party_cost": None,
                                            "total_cost": None,
                                            "pdl_called": None,
                                            "search_duration_ms": None,
                                        },
                                    }
                                ]
                            },
                            "not_comparable": {"queries": []},
                        },
                        "module_metrics": {
                            "Summary": {
                                "module": "Summary",
                                "returned_candidate_count": 1,
                                "candidate_count": 2,
                                "return_rate": 0.5,
                                "hit_return_rate": 1.0,
                                "nonmatched_return_rate": 0.0,
                            }
                        },
                        "field_metrics": {
                            "display_name": {
                                "field_key": "display_name",
                                "display_name": "Display Name",
                                "module": "Summary",
                                "returned_count": 1,
                                "empty_count": 1,
                                "candidate_count": 2,
                                "return_rate": 0.5,
                                "hit_completeness": 1.0,
                                "hit_accuracy": 1.0,
                            },
                            "profile_data": {
                                "field_key": "profile_data",
                                "display_name": "Profile Data",
                                "module": "Profile",
                            },
                        },
                        "warnings": ["第三方成本未接入。"],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            self.run_exporter(
                [
                    "processed",
                    "--input-file",
                    str(processed_path),
                    "--report-model",
                    str(report_model_path),
                    "--output",
                    str(output_path),
                ],
                model_path,
                create_workbook=True,
            )

            model = json.loads(model_path.read_text(encoding="utf-8"))
            self.assertEqual(2, len(model["candidateRows"]))
            self.assertEqual(1, len(model["queryRows"]))
            self.assertEqual(2, len(model["reviewRows"]))
            self.assertIn("display_name", model["candidateHeaders"])
            self.assertIn("profile_data", model["candidateHeaders"])
            self.assertTrue(model["rawRows"])
            self.assertEqual(1, len(model["sameConditionRows"]))
            self.assertEqual(1, len(model["newClueRows"]))
            self.assertTrue(
                any(
                    row["metric_key"] == "total_cost"
                    and row["total"] == 3.75
                    for row in model["coreMetricRows"]
                )
            )
            self.assertTrue(
                any(
                    row["row_type"] == "FIELD"
                    and row["field_key"] == "display_name"
                    for row in model["moduleFieldRows"]
                )
            )
            self.assertTrue(output_path.is_file())
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                self.assertTrue(
                    {
                        "说明",
                        "核心指标",
                        "Query明细",
                        "候选结果",
                        "同条件对比",
                        "新增线索",
                        "模块字段统计",
                        "失败记录",
                        "人工复核",
                        "Raw数据",
                    }.issubset(workbook.sheetnames)
                )
                query_rows = list(
                    workbook["Query明细"].iter_rows(values_only=True)
                )
                query_record = dict(zip(query_rows[0], query_rows[1]))
                self.assertEqual(
                    "PHASE_2_POST_OPTIMIZATION",
                    query_record["evaluation_phase"],
                )
                self.assertEqual("HAS_CANDIDATES", query_record["result_status"])
                self.assertEqual(1.25, query_record["llm_cost"])
                self.assertIsNone(query_record["third_party_cost"])
                self.assertEqual(3.75, query_record["total_cost"])
                self.assertIs(query_record["pdl_called"], False)
                self.assertEqual(1250, query_record["search_duration_ms"])

                candidate_rows = list(
                    workbook["候选结果"].iter_rows(values_only=False)
                )
                candidate_headers = [cell.value for cell in candidate_rows[0]]
                reviewed_at_index = candidate_headers.index("reviewed_at")
                reviewed_at_cell = candidate_rows[1][reviewed_at_index]
                self.assertEqual(
                    datetime(2026, 7, 24, 8, 0, 0),
                    reviewed_at_cell.value,
                )
                self.assertEqual(
                    "yyyy-mm-dd hh:mm:ss",
                    reviewed_at_cell.number_format,
                )

                failure_rows = list(
                    workbook["失败记录"].iter_rows(values_only=False)
                )
                failure_headers = [cell.value for cell in failure_rows[0]]
                failure_created_at = failure_rows[1][
                    failure_headers.index("created_at")
                ]
                self.assertEqual(
                    datetime(2026, 7, 24, 8, 0, 0),
                    failure_created_at.value,
                )
                self.assertEqual(
                    "yyyy-mm-dd hh:mm:ss",
                    failure_created_at.number_format,
                )

                note_rows = list(workbook["说明"].iter_rows(values_only=False))
                generated_at_cell = next(
                    row[1]
                    for row in note_rows[1:]
                    if row[0].value == "生成时间"
                )
                self.assertEqual(
                    datetime(2026, 7, 24, 8, 0, 0),
                    generated_at_cell.value,
                )
                self.assertEqual(
                    "yyyy-mm-dd hh:mm:ss",
                    generated_at_cell.number_format,
                )

                core_rows = list(
                    workbook["核心指标"].iter_rows(values_only=True)
                )
                core_headers = core_rows[0]
                total_cost = next(
                    dict(zip(core_headers, row))
                    for row in core_rows[1:]
                    if row[core_headers.index("metric_key")] == "total_cost"
                )
                self.assertEqual(3.75, total_cost["total"])
                self.assertIsNone(
                    next(
                        dict(zip(core_headers, row))
                        for row in core_rows[1:]
                        if row[core_headers.index("metric_key")]
                        == "third_party_cost"
                    )["total"]
                )
            finally:
                workbook.close()

    def test_processed_mode_keeps_report_model_v1_metrics_exportable(self):
        """旧 ReportModel v1 仍能在核心指标中看到原有质量指标。"""

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            processed_path = root / "processed_export.jsonl"
            report_model_path = root / "report_model.json"
            model_path = root / "model.json"
            write_jsonl(
                processed_path,
                [
                    {
                        "record_type": "query",
                        "query_id": "legacy-query",
                        "person_id": "legacy-person",
                        "query_stage": "FULL_NAME",
                        "retrieval_success": False,
                        "formal_ready": False,
                    }
                ],
            )
            report_model_path.write_text(
                json.dumps(
                    {
                        "metadata": {
                            "report_id": "legacy-report",
                            "report_type": "SINGLE",
                        },
                        "summary": {
                            "formal_ready": False,
                            "candidate": {
                                "retrieval_success": {
                                    "numerator": 0,
                                    "denominator": 1,
                                    "value": None,
                                    "preview_value": 0.0,
                                },
                                "matched_completeness": {
                                    "numerator": 0,
                                    "denominator": 0,
                                    "value": None,
                                    "preview_value": None,
                                },
                            },
                        },
                        "paired_metrics": None,
                        "module_metrics": {},
                        "field_metrics": {},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            self.run_exporter(
                [
                    "processed",
                    "--input-file",
                    str(processed_path),
                    "--report-model",
                    str(report_model_path),
                    "--output",
                    str(root / "legacy.xlsx"),
                ],
                model_path,
            )
            model = json.loads(model_path.read_text(encoding="utf-8"))

        retrieval = next(
            row
            for row in model["coreMetricRows"]
            if row["metric_key"] == "retrieval_success"
        )
        self.assertEqual(0, retrieval["numerator"])
        self.assertEqual(1, retrieval["denominator"])
        self.assertIsNone(retrieval["value"])
        self.assertEqual(0.0, retrieval["preview_value"])

    def test_processed_mode_exports_report_model_v3_audit_sheets(self):
        """ReportModel v3 导出人物关联、身份归类、字段矩阵和原因 Sheet。"""

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            processed_path = root / "processed_export.jsonl"
            report_model_path = root / "report_model.json"
            output_path = root / "report-v3.xlsx"
            model_path = root / "model.json"
            write_jsonl(
                processed_path,
                [
                    {
                        "record_type": "query",
                        "run_id": "run-v3",
                        "query_id": "query-v3",
                        "person_id": "person-v3",
                        "person_id_source": "MANUAL",
                        "baseline_display_name": "Example Person",
                        "baseline_match_status": "MATCHED",
                        "query_stage": "FULL_NAME",
                        "result_status": "HAS_CANDIDATES",
                        "identity_state": "PENDING",
                        "formal_ready": False,
                    },
                    {
                        "record_type": "candidate",
                        "process_id": "process-v3",
                        "run_id": "run-v3",
                        "query_id": "query-v3",
                        "person_id": "person-v3",
                        "candidate_pk": "candidate-pk-v3",
                        "candidate_id": "candidate-v3",
                        "candidate_rank": 1,
                        "detail_status": "SUCCESS",
                        "judgement": "PENDING_REVIEW",
                        "classification_source": "SUGGESTED",
                        "is_primary_hit": False,
                        "fields": {"summary_name": "Example Person"},
                        "empty_fields": {"summary_name": False},
                        "field_scores": {},
                        "processing_errors": [],
                    },
                ],
            )
            report_model_path.write_text(
                json.dumps(
                    {
                        "metadata": {
                            "report_id": "report-v3",
                            "evaluation_id": "eval-v3",
                            "report_type": "SINGLE",
                            "report_model_version": "report-model-v3",
                            "metrics_rule_version": "metrics-v3",
                            "generated_at": "2026-07-28T00:00:00+00:00",
                        },
                        "summary": {"formal_ready": False},
                        "execution_summary": {
                            "query_count": 1,
                            "has_candidates_count": 1,
                            "no_candidates_count": 0,
                            "execution_failed_count": 0,
                            "detail_success_count": 1,
                            "detail_failure_count": 0,
                        },
                        "identity_summary": {
                            "query_count": 1,
                            "classified_query_count": 0,
                            "pending_query_count": 1,
                            "primary_hit_query_count": 0,
                            "no_hit_query_count": 0,
                        },
                        "quality_metrics": {
                            "retrieval_success": {
                                "status": "NOT_READY",
                                "value": None,
                                "preview_value": None,
                                "numerator": 0,
                                "denominator": 0,
                                "reason_codes": ["IDENTITY_PENDING"],
                                "reasons": ["query-v3 有1个候选人待身份归类"],
                            }
                        },
                        "field_alignment_matrix": {
                            "schema_version": "schema-v3",
                            "baseline_version": "baseline-v3",
                            "fields": [
                                {
                                    "field_key": "summary_name",
                                    "display_name": "Summary Name",
                                    "module": "Summary",
                                    "value_scope": "CANDIDATE",
                                    "enabled": True,
                                    "baseline_available_count": 1,
                                    "baseline_person_count": 1,
                                    "candidate_nonempty_count": 1,
                                    "candidate_count": 1,
                                    "candidate_return_rate": 1.0,
                                    "completeness_enabled": True,
                                    "accuracy_enabled": True,
                                    "identity_enabled": True,
                                    "compare_mode": "exact",
                                    "normalizer": "trim_text",
                                    "status": "COMPARABLE",
                                    "issues": [],
                                }
                            ],
                        },
                        "field_metrics": {
                            "task_id": {
                                "field_key": "task_id",
                                "display_name": "Task ID",
                                "module": "Task",
                                "value_scope": "QUERY",
                                "returned_count": 1,
                                "empty_count": 0,
                                "entity_count": 1,
                                "return_rate": 1.0,
                                "status": "READY",
                                "reason_codes": [],
                                "reasons": [],
                            },
                            "summary_name": {
                                "field_key": "summary_name",
                                "display_name": "Summary Name",
                                "module": "Summary",
                                "value_scope": "CANDIDATE",
                                "returned_count": 1,
                                "empty_count": 0,
                                "entity_count": 1,
                                "return_rate": 1.0,
                                "status": "READY",
                                "reason_codes": [],
                                "reasons": [],
                            },
                        },
                        "module_metrics": {
                            "Summary": {
                                "module": "Summary",
                                "data_count": 1,
                                "empty_count": 0,
                                "unknown_count": 0,
                                "candidate_count": 1,
                                "data_rate": 1.0,
                                "status": "READY",
                                "reason_codes": [],
                                "reasons": [],
                            }
                        },
                        "not_ready_reasons": [
                            {
                                "metric": "retrieval_success",
                                "reason_code": "IDENTITY_PENDING",
                                "reason": "候选人身份归类尚未完成",
                                "details": ["query-v3 有1个候选人待身份归类"],
                            }
                        ],
                        "comparison": {},
                        "warnings": [],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            self.run_exporter(
                [
                    "processed",
                    "--input-file",
                    str(processed_path),
                    "--report-model",
                    str(report_model_path),
                    "--output",
                    str(output_path),
                ],
                model_path,
                create_workbook=True,
            )

            model = json.loads(model_path.read_text(encoding="utf-8"))
            self.assertEqual("QUERY", model["fieldMetricRows"][0]["value_scope"])
            self.assertEqual(
                "IDENTITY_PENDING",
                model["notReadyReasonRows"][0]["reason_code"],
            )
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                self.assertTrue(
                    {
                        "Report_Summary",
                        "Query_Person_Links",
                        "Identity_Classification",
                        "Field_Matrix",
                        "Field_Metrics",
                        "Module_Metrics",
                        "Not_Ready_Reasons",
                        "候选结果",
                    }.issubset(workbook.sheetnames)
                )
                field_rows = list(
                    workbook["Field_Metrics"].iter_rows(values_only=True)
                )
                field_headers = field_rows[0]
                task_row = dict(zip(field_headers, field_rows[1]))
                self.assertEqual("QUERY", task_row["value_scope"])
                self.assertEqual(1.0, task_row["return_rate"])
                reason_rows = list(
                    workbook["Not_Ready_Reasons"].iter_rows(
                        values_only=True
                    )
                )
                self.assertEqual(
                    "IDENTITY_PENDING",
                    reason_rows[1][reason_rows[0].index("reason_code")],
                )
            finally:
                workbook.close()

    def test_long_json_is_split_and_reconstructable(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            run_dir = root / "current"
            large_value = "中" * 70000
            write_jsonl(
                run_dir / "results.jsonl",
                [
                    {
                        "input_id": "query-long",
                        "task_id": "task-long",
                        "results": [candidate("long-candidate", photo_payload=large_value)],
                    }
                ],
            )
            output_path = root / "long.xlsx"
            model_path = root / "model.json"

            self.run_exporter(
                [
                    "single",
                    "--run-dir",
                    str(run_dir),
                    "--run-label",
                    "current",
                    "--system-version",
                    "v1",
                    "--evaluation-id",
                    "eval-long",
                    "--output",
                    str(output_path),
                ],
                model_path,
            )

            model = json.loads(model_path.read_text(encoding="utf-8"))
            raw_rows = [
                row
                for row in model["rawRows"]
                if row["field_name"] == "photos_authenticity_photos"
            ]
            self.assertEqual(3, len(raw_rows))
            self.assertTrue(all(len(row["content"]) <= 30000 for row in raw_rows))
            reconstructed = "".join(row["content"] for row in sorted(raw_rows, key=lambda row: row["chunk_index"]))
            expected = json.dumps(
                candidate("long-candidate", photo_payload=large_value)["ui_sections"]["photos"]["data"][
                    "authenticity_photos"
                ],
                ensure_ascii=False,
                separators=(",", ":"),
            )
            self.assertEqual(expected, reconstructed)
            self.assertTrue(
                model["candidateRows"][0]["photos_authenticity_photos"].startswith(
                    "[超长内容见 Raw数据] RAW:"
                )
            )

    def test_report_model_v5_exports_quality_comparison_and_rule_sheets(self):
        """v5 Excel 兼容导出质量、字段对比和规则快照审计 Sheet。"""

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            processed_path = root / "processed.jsonl"
            report_model_path = root / "report-model.json"
            output_path = root / "report.xlsx"
            model_path = root / "model.json"
            write_jsonl(processed_path, [{
                "record_type": "candidate",
                "query_id": "query-v4",
                "person_id": "person-v4",
                "candidate_pk": "candidate-pk-v4",
                "candidate_id": "candidate-v4",
                "is_primary_hit": True,
                "fields": {"summary_display_name": "Example Person"},
                "field_scores": {
                    "summary_display_name": {
                        "baseline_field_key": "summary_display_name",
                        "baseline_value": "Example Person",
                        "returned_value": "Example Person",
                        "baseline_available": True,
                        "returned_nonempty": True,
                        "completeness_score": 1.0,
                        "accuracy_score": 1.0,
                        "comparison_status": "READY",
                        "reason_code": "EXACT_MATCH",
                    }
                },
                "processing_errors": [],
            }])
            report_model_path.write_text(json.dumps({
                "metadata": {
                    "report_model_version": "report-model-v5",
                    "metrics_rule_version": "metrics-v4",
                    "rule_version": "field-processing-v5",
                    "schema_version": "field-schema-v3",
                    "baseline_version": "baseline-v4",
                    "candidate_process_id": "process-v4",
                },
                "baseline_quality_metrics": {
                    "modules": {
                        module: {
                            "completeness": {"status": "READY", "value": 1.0, "numerator": 1.0, "denominator": 1},
                            "accuracy": {"status": "NOT_APPLICABLE", "numerator": 0, "denominator": 0},
                        }
                        for module in ["Insights", "Photos", "Profile", "Social", "Summary"]
                    },
                },
                "non_hit_data_return": {"modules": {}},
                "regression_metrics": {"status": "NOT_APPLICABLE"},
                "field_metrics": {
                    "summary_display_name": {
                        "display_name": "Summary Display Name",
                        "module": "Summary",
                        "value_scope": "CANDIDATE",
                        "returned_count": 1,
                        "empty_count": 0,
                        "entity_count": 1,
                        "return_rate": 1.0,
                        "status": "READY",
                    }
                },
                "module_metrics": {}, "quality_metrics": {},
                "result_status_metrics": {}, "grouped_metrics": [],
                "warnings": [],
            }, ensure_ascii=False), encoding="utf-8")
            self.run_exporter([
                "processed", "--input-file", str(processed_path),
                "--report-model", str(report_model_path),
                "--output", str(output_path),
            ], model_path, create_workbook=True)
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                self.assertTrue({
                    "Core Metrics", "Module Quality", "Field Comparison",
                    "Field Returns", "Rule Snapshot",
                }.issubset(workbook.sheetnames))
                comparison_rows = list(workbook["Field Comparison"].iter_rows(values_only=True))
                row = dict(zip(comparison_rows[0], comparison_rows[1]))
                self.assertEqual("summary_display_name", row["field_key"])
                self.assertEqual(1, row["accuracy_score"])
                rule_rows = list(workbook["Rule Snapshot"].iter_rows(values_only=True))
                self.assertIn(("report", "metrics_rule_version", "metrics-v4"), rule_rows)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
