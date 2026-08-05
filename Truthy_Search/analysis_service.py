"""searchTool v1.3 历史 JSONL/Excel 校验、归档和统一入库。"""

from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import unicodedata
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook

from analysis_store import AnalysisStore, utc_now_text
from search_tool import (
    RESULT_SCHEMA_VERSION,
    FlowError,
    normalize_result_status,
    process_one,
    sanitize_raw,
)


SUPPORTED_QUERY_STAGES = {"FULL_NAME", "FULL_NAME_SOCIAL"}
EVALUATION_THRESHOLD_FIELDS = {
    "min_retrieval_success": "minimum_ratio",
    "min_matched_completeness": "minimum_ratio",
    "min_matched_accuracy": "minimum_ratio",
    "max_average_total_cost": "maximum_number",
    "max_average_search_duration_ms": "maximum_number",
}
EVALUATION_PHASES = {
    "PHASE_1_BASELINE",
    "PHASE_2_POST_OPTIMIZATION",
    "PHASE_3_TARGETED_ITERATION",
    "UNSPECIFIED",
}
RESULT_STATUSES = {
    "HAS_CANDIDATES",
    "NO_CANDIDATES",
    "EXECUTION_FAILED",
}
TASK_FIELD_KEYS = {
    "llm_cost",
    "third_party_cost",
    "total_cost",
    "pdl_called",
    "search_duration_ms",
}
SAFE_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.-]{0,127}$")
FIELD_KEY_PATTERN = re.compile(r"^[A-Za-z][A-Za-z0-9_.-]{0,127}$")
FIELD_PATH_PART_PATTERN = re.compile(
    r"^(?P<key>[A-Za-z0-9_-]+)(?:\[(?P<index>\d+|\*)\])?$"
)
FIELD_MODULES = {
    "Task",
    "Candidate",
    "Insights",
    "Photos",
    "Profile",
    "Social",
    "Summary",
}
FIELD_SOURCE_STAGES = {
    "GetTask",
    "ListTaskCandidates",
    "GetTaskCandidateDetail",
}
FIELD_DATA_TYPES = {
    "string",
    "number",
    "boolean",
    "object",
    "array",
    "url",
    "image",
}
FIELD_ARRAY_MODES = {"preserve", "first", "collect", "join"}
FIELD_EMPTY_RULES = {"default", "zero_is_empty", "false_is_empty"}
FIELD_NORMALIZERS = {
    "identity",
    "trim_text",
    "number",
    "percentage",
    "social_url",
    "string_list",
    "profile_sections",
}
FIELD_SCORING_ROLES = {
    "identity",
    "completeness",
    "accuracy",
    "display",
    "none",
}
FIELD_COMPARE_MODES = {
    "presence",
    "exact",
    "normalized_text",
    "set",
    "url_set",
    "semantic_text_lite",
    "manual",
}
FIELD_SOURCE_TYPES = {"PATH", "PROFILE_ITEM"}
FIELD_VALUE_SCOPES = {"QUERY", "CANDIDATE"}
FIELD_MISSING_POLICIES = {"EMPTY", "ERROR"}
FINAL_JUDGEMENTS = {"HIT", "NOT_HIT", "SUSPECTED"}
REVIEW_REASONS = {
    "SOCIAL_MATCH",
    "SOCIAL_CONFLICT",
    "PHOTO_MATCH",
    "PHOTO_BELOW_THRESHOLD",
    "NO_STRONG_FIELD",
    "MANUAL",
}
CONTENT_FIELD_COUNT = 22
DEFAULT_FIELD_SCHEMA_VERSION = "field-schema-default-v2"
DEFAULT_FIELD_SCHEMA_V3_VERSION = "field-schema-default-v3"
LEGACY_FIELD_PROCESSING_RULE_VERSION = "field-processing-v1"
V2_FIELD_PROCESSING_RULE_VERSION = "field-processing-v2"
V3_FIELD_PROCESSING_RULE_VERSION = "field-processing-v3"
FIELD_PROCESSING_RULE_VERSION = "field-processing-v4"
V5_FIELD_PROCESSING_RULE_VERSION = "field-processing-v5"
V2_METRICS_RULE_VERSION = "metrics-v2"
V3_METRICS_RULE_VERSION = "metrics-v3"
METRICS_RULE_VERSION = "metrics-v4"
V2_REPORT_MODEL_VERSION = "report-model-v2"
V3_REPORT_MODEL_VERSION = "report-model-v3"
REPORT_MODEL_VERSION = "report-model-v4"
# 报告优化 v5 的完整快照版本，仅供新的 metrics-v4 Process 报告使用，
# 避免把历史 v2/v3 报告错误标记为新模型。
V5_REPORT_MODEL_VERSION = "report-model-v5"

# 核心指标的产品说明属于报告契约，而不是页面文案。将其随报告快照保存，
# 可以保证静态 HTML、Web 页面和后续导出使用同一套名称与计算口径。
REPORT_V5_CORE_METRIC_DEFINITIONS = {
    "candidate_return_rate": {
        "label": "候选人返回率",
        "purpose": "衡量检索请求能否返回可供判断的候选人",
        "formula": "有返回候选人的 Query 数 ÷ 成功执行的 Query 数",
        "breakdown_scope": "QUERY",
    },
    "retrieval_success": {
        "label": "目标人物命中率",
        "purpose": "衡量系统能否在全部有效 Query 中找到目标人物",
        "formula": "命中目标人物的 Query 数 ÷ 成功执行的 Query 数",
        "breakdown_scope": "QUERY",
    },
    "conditional_hit_rate": {
        "label": "有候选人条件下命中率",
        "purpose": "排除无结果 Query，观察候选人召回后的身份识别能力",
        "formula": "命中目标人物的 Query 数 ÷ 有返回候选人的 Query 数",
        "breakdown_scope": "QUERY",
    },
    "matched_accuracy": {
        "label": "命中信息准确度",
        "purpose": "衡量主命中候选人的字段信息与基准人物是否一致",
        "formula": "主 HIT 候选人准确度的宏平均；字段级微平均仅作辅助观察",
        "breakdown_scope": "PRIMARY_HIT",
    },
    "matched_completeness": {
        "label": "命中资料完整度",
        "purpose": "衡量主命中候选人的启用字段是否实际返回",
        "formula": "主 HIT 实际返回字段数 ÷ 主 HIT 启用字段槽位；全部 HIT 与非主 HIT 另作辅助分析",
        "breakdown_scope": "PRIMARY_HIT",
    },
    "nonmatched_data_completeness": {
        "label": "非命中资料完整度",
        "purpose": "衡量非目标候选人自身返回了多少有效资料",
        "formula": "全部非命中/疑似候选人的实际返回字段数 ÷ 启用字段槽位",
        "breakdown_scope": "NONMATCHED",
    },
    "nonmatched_baseline_overlap": {
        "label": "非命中候选人资料相似度",
        "purpose": "衡量非命中候选人的资料与基准人物有多相似，辅助分析同名干扰与误召回",
        "formula": "全部非命中/疑似候选人的资料相似得分 ÷ 可比较资料字段槽位；排除 profile_full_name，不代表身份一致",
        "breakdown_scope": "NONMATCHED",
    },
}

# 姓名通常来自检索条件本身，纳入非命中相似度会把 Query 回显误当成资料相似。
NONMATCHED_SIMILARITY_EXCLUDED_FIELDS = {"profile_full_name"}

# 报告的资料模块概览只统计真实业务原子字段。接口状态、对象容器和
# Insights Items 聚合容器不代表一条独立资料，不能重复进入字段槽位。
REPORT_V5_PROFILE_MODULES = (
    "Insights", "Photos", "Profile", "Social", "Summary",
)
REPORT_V5_MODULE_NON_BUSINESS_FIELDS = {
    "insights_status",
    "insights_items",
    "photos_status",
    "profile_status",
    "social_status",
    "summary_status",
}


def select_primary_hit_candidate(
    candidates: Iterable[dict[str, Any] | sqlite3.Row],
) -> str | None:
    """从已命中候选人中选择唯一主命中。

    功能说明：同一 Query 可能返回多条真实命中。候选人级资料指标应保留
    给所有 ``HIT``，但 Query/整体指标只能选择一条，避免同一人物的重复
    返回放大汇总权重。选择顺序固定为 ``rank_score`` 降序、候选排名升序、
    Candidate 主键升序，确保重处理和人工覆写后的结果稳定可复现。

    参数说明：
        candidates: 仅包含最终 ``HIT`` 候选人的字典或 SQLite 行；每项需要
            ``candidate_pk``，可选 ``rank_score`` 和 ``candidate_rank``。

    返回值：
        str | None: 分数最高的 Candidate 主键；没有有效候选人时返回 None。

    异常说明：
        非数值、布尔值、NaN 或无穷大的 rank_score 统一按缺失分数处理，
        不抛出异常，以保证历史数据可以无成本重处理。
    """

    normalized: list[tuple[float, int, str]] = []
    for candidate in candidates:
        candidate_pk = str(candidate["candidate_pk"] or "").strip()
        if not candidate_pk:
            continue
        raw_score = candidate["rank_score"]
        score = (
            float(raw_score)
            if isinstance(raw_score, (int, float))
            and not isinstance(raw_score, bool)
            and math.isfinite(float(raw_score))
            else float("-inf")
        )
        raw_rank = candidate["candidate_rank"]
        rank = (
            raw_rank
            if isinstance(raw_rank, int) and not isinstance(raw_rank, bool)
            else 2**31 - 1
        )
        normalized.append((score, rank, candidate_pk))
    if not normalized:
        return None
    return min(normalized, key=lambda item: (-item[0], item[1], item[2]))[2]


def _default_field(
    field_key: str,
    display_name: str,
    module: str,
    source_path: str,
    data_type: str,
    sort_order: int,
    *,
    source_stage: str = "GetTaskCandidateDetail",
    array_mode: str = "preserve",
    normalizer: str = "identity",
    scoring_role: list[str] | None = None,
    compare_mode: str = "exact",
    value_scope: str | None = None,
    missing_policy: str | None = None,
    enabled: bool = True,
    display_enabled: bool | None = None,
    baseline_field_key: str | None = None,
    baseline_compare_enabled: bool = False,
    run_compare_enabled: bool | None = None,
    identity_enabled: bool | None = None,
    completeness_enabled: bool | None = None,
    accuracy_enabled: bool | None = None,
    source_type: str = "PATH",
    source_options: dict[str, Any] | None = None,
    similarity_threshold: float = 0.6,
    include_v3: bool = False,
) -> dict[str, Any]:
    """构造默认字段配置，并显式保存 v3 字段职责与兼容属性。

    功能说明：v2 默认配置继续使用原有字段集合；v3 目录通过同一构造器
    生成完整属性。评分角色与独立开关同时保存，确保旧处理逻辑可读取、
    后续字段比较可直接使用开关。
    """

    roles = scoring_role or ["display"]
    final_enabled = bool(enabled)
    final_identity = (
        "identity" in roles if identity_enabled is None else identity_enabled
    )
    final_completeness = (
        "completeness" in roles
        if completeness_enabled is None
        else completeness_enabled
    )
    final_accuracy = (
        "accuracy" in roles if accuracy_enabled is None else accuracy_enabled
    )

    result = {
        "field_key": field_key,
        "display_name": display_name,
        "module": module,
        "source_stage": source_stage,
        "source_path": source_path,
        "data_type": data_type,
        "array_mode": array_mode,
        "empty_rule": "default",
        "normalizer": normalizer,
        "scoring_role": roles,
        "compare_mode": compare_mode,
        "enabled": final_enabled,
        "sort_order": sort_order,
        "value_scope": value_scope or (
            "QUERY" if module == "Task" else "CANDIDATE"
        ),
        "missing_policy": missing_policy or (
            "ERROR" if field_key in {"task_id", "candidate_id"} else "EMPTY"
        ),
    }
    if include_v3:
        result.update(
            {
                "source_type": source_type,
                "source_options": source_options or {},
                "display_enabled": (
                    final_enabled
                    if display_enabled is None
                    else bool(display_enabled)
                ),
                "baseline_field_key": baseline_field_key or "",
                "baseline_compare_enabled": bool(baseline_compare_enabled),
                "run_compare_enabled": (
                    final_enabled
                    if run_compare_enabled is None
                    else bool(run_compare_enabled)
                ),
                "identity_enabled": bool(final_identity),
                "completeness_enabled": bool(final_completeness),
                "accuracy_enabled": bool(final_accuracy),
                "similarity_threshold": similarity_threshold,
                "field_schema_version": "v3",
            }
        )
    return result


DEFAULT_FIELD_DEFINITIONS = [
    _default_field(
        "task_id",
        "Task ID",
        "Task",
        "task.task_id",
        "string",
        10,
        source_stage="GetTask",
    ),
    _default_field(
        "llm_cost",
        "LLM Cost",
        "Task",
        "task.llm_cost",
        "number",
        20,
        source_stage="GetTask",
    ),
    _default_field(
        "third_party_cost",
        "Third-party Cost",
        "Task",
        "task.third_party_cost",
        "number",
        25,
        source_stage="GetTask",
    ),
    _default_field(
        "total_cost",
        "Total Cost",
        "Task",
        "task.total_cost",
        "number",
        30,
        source_stage="GetTask",
    ),
    _default_field(
        "search_duration_ms",
        "Search Duration (ms)",
        "Task",
        "task.search_duration_ms",
        "number",
        45,
        source_stage="GetTask",
    ),
    _default_field(
        "pdl_called",
        "PDL Called",
        "Task",
        "task.pdl_called",
        "boolean",
        40,
        source_stage="GetTask",
    ),
    _default_field(
        "candidate_id",
        "Candidate ID",
        "Candidate",
        "candidate.candidate_id",
        "string",
        50,
        source_stage="ListTaskCandidates",
    ),
    _default_field(
        "candidate_rank",
        "Candidate Rank",
        "Candidate",
        "candidate.candidate_rank",
        "number",
        60,
        source_stage="ListTaskCandidates",
    ),
    _default_field(
        "rank_score",
        "Rank Score",
        "Candidate",
        "candidate.rank_score",
        "number",
        70,
        source_stage="ListTaskCandidates",
    ),
    _default_field(
        "insights_status",
        "Insights Status",
        "Insights",
        "ui_sections.insights.status",
        "string",
        100,
        normalizer="trim_text",
    ),
    _default_field(
        "insights_description",
        "Insights Description",
        "Insights",
        "ui_sections.insights.data.items[0].description",
        "string",
        110,
        normalizer="trim_text",
        compare_mode="normalized_text",
    ),
    _default_field(
        "insights_links",
        "Insights Links",
        "Insights",
        "ui_sections.insights.data.items[0].links",
        "array",
        120,
        array_mode="collect",
        compare_mode="set",
    ),
    _default_field(
        "insights_data",
        "Insights Data",
        "Insights",
        "ui_sections.insights.data",
        "object",
        130,
    ),
    _default_field(
        "photos_status",
        "Photos Status",
        "Photos",
        "ui_sections.photos.status",
        "string",
        200,
        normalizer="trim_text",
    ),
    _default_field(
        "photos_data",
        "Photos Data",
        "Photos",
        "ui_sections.photos.data",
        "object",
        210,
    ),
    _default_field(
        "photos_identity_match_rate",
        "Photo Identity Match Rate",
        "Photos",
        "ui_sections.photos.data.identity_match_rate",
        "number",
        220,
        normalizer="percentage",
    ),
    _default_field(
        "profile_status",
        "Profile Status",
        "Profile",
        "ui_sections.profile.status",
        "string",
        300,
        normalizer="trim_text",
    ),
    _default_field(
        "profile_data",
        "Profile Data",
        "Profile",
        "ui_sections.profile.data",
        "object",
        310,
    ),
    _default_field(
        "profile_sections",
        "Profile Sections",
        "Profile",
        "ui_sections.profile.data.sections",
        "object",
        320,
        normalizer="profile_sections",
        compare_mode="manual",
    ),
    _default_field(
        "social_status",
        "Social Status",
        "Social",
        "ui_sections.social.status",
        "string",
        400,
        normalizer="trim_text",
    ),
    _default_field(
        "social_display_handles",
        "Social Display Handles",
        "Social",
        "ui_sections.social.data.profiles[*].display_handle",
        "array",
        410,
        array_mode="collect",
        normalizer="string_list",
        compare_mode="set",
    ),
    _default_field(
        "social_platforms",
        "Social Platforms",
        "Social",
        "ui_sections.social.data.profiles[*].platform",
        "array",
        420,
        array_mode="collect",
        normalizer="string_list",
        compare_mode="set",
    ),
    _default_field(
        "social_urls",
        "Social URLs",
        "Social",
        "ui_sections.social.data.profiles[*].url",
        "array",
        430,
        array_mode="collect",
        normalizer="social_url",
        scoring_role=["identity", "completeness", "accuracy"],
        compare_mode="url_set",
    ),
    _default_field(
        "summary_avatar_url",
        "Summary Avatar URL",
        "Summary",
        "ui_sections.summary.data.avatar_url",
        "image",
        500,
    ),
    _default_field(
        "candidate_confidence",
        "Summary Confidence Level",
        "Summary",
        "ui_sections.summary.data.confidence_level",
        "string",
        510,
        normalizer="trim_text",
    ),
    _default_field(
        "summary_primary_image_url",
        "Summary Primary Image URL",
        "Summary",
        "ui_sections.summary.data.primary_image.url",
        "image",
        520,
    ),
    _default_field(
        "summary_social_links",
        "Summary Social Links",
        "Summary",
        "ui_sections.summary.data.social_links",
        "array",
        530,
        array_mode="collect",
        compare_mode="set",
    ),
    _default_field(
        "summary_web_links",
        "Summary Web Links",
        "Summary",
        "ui_sections.summary.data.web_links",
        "array",
        540,
        array_mode="collect",
        compare_mode="set",
    ),
    _default_field(
        "summary_display_name",
        "Summary Display Name",
        "Summary",
        "ui_sections.summary.data.display_name",
        "string",
        550,
        normalizer="trim_text",
        compare_mode="normalized_text",
    ),
    _default_field(
        "summary_location",
        "Summary Location",
        "Summary",
        "ui_sections.summary.data.location",
        "string",
        560,
        normalizer="trim_text",
        compare_mode="normalized_text",
    ),
    _default_field(
        "summary_match_score",
        "Summary Match Score",
        "Summary",
        "ui_sections.summary.data.match_score",
        "number",
        570,
        normalizer="percentage",
    ),
    _default_field(
        "summary_is_top_result",
        "Summary Is Top Result",
        "Summary",
        "ui_sections.summary.data.is_top_result",
        "boolean",
        580,
    ),
    _default_field(
        "summary_is_best_match",
        "Summary Is Best Match",
        "Summary",
        "ui_sections.summary.data.is_best_match",
        "boolean",
        590,
    ),
]


def _v3_catalog_field(
    field_key: str,
    display_name: str,
    module: str,
    source_path: str,
    data_type: str,
    sort_order: int,
    **kwargs: Any,
) -> dict[str, Any]:
    """构造字段目录中的 v3 字段，并应用首版安全默认开关。

    所有 ``ui_sections`` 字段默认提取和展示、可参与版本比较；除已确认
    的 Social URL 和照片身份相似度外，不自动进入身份和质量指标，避免
    新接口字段在口径未确认时影响既有报告。
    """

    return _default_field(
        field_key,
        display_name,
        module,
        source_path,
        data_type,
        sort_order,
        missing_policy=kwargs.pop("missing_policy", "EMPTY"),
        compare_mode=kwargs.pop("compare_mode", "manual"),
        normalizer=kwargs.pop("normalizer", "identity"),
        array_mode=kwargs.pop("array_mode", "preserve"),
        scoring_role=kwargs.pop("scoring_role", ["display"]),
        include_v3=True,
        **kwargs,
    )


# FieldSchema v3 是一个新快照，不能修改 DEFAULT_FIELD_DEFINITIONS（v2）。
# 目录记录已确认的 ui_sections 原子字段；Profile 中用于正式评估的稳定
# 标签通过受控 PROFILE_ITEM 定义，其余动态标签仍由页面按建议新增，避免
# 为接口未知标签硬编码字段。
DEFAULT_FIELD_DEFINITIONS_V3 = [
    _v3_catalog_field(
        "task_id", "Task ID", "Task", "task.task_id", "string", 10,
        source_stage="GetTask", missing_policy="ERROR",
    ),
    _v3_catalog_field(
        "llm_cost", "LLM Cost", "Task", "task.llm_cost", "number", 20,
        source_stage="GetTask", enabled=False, display_enabled=True,
        run_compare_enabled=False,
    ),
    _v3_catalog_field(
        "third_party_cost", "Third-party Cost", "Task", "task.third_party_cost", "number", 30,
        source_stage="GetTask", enabled=False, display_enabled=True,
        run_compare_enabled=False,
    ),
    _v3_catalog_field(
        "total_cost", "Total Cost", "Task", "task.total_cost", "number", 40,
        source_stage="GetTask", enabled=False, display_enabled=True,
        run_compare_enabled=False,
    ),
    _v3_catalog_field(
        "pdl_called", "PDL Called", "Task", "task.pdl_called", "boolean", 50,
        source_stage="GetTask", enabled=False, display_enabled=True,
        run_compare_enabled=False,
    ),
    _v3_catalog_field(
        "search_duration_ms", "Search Duration (ms)", "Task", "task.search_duration_ms", "number", 60,
        source_stage="GetTask",
    ),
    _v3_catalog_field(
        "candidate_id", "Candidate ID", "Candidate", "candidate.candidate_id", "string", 100,
        source_stage="ListTaskCandidates", missing_policy="ERROR",
    ),
    _v3_catalog_field(
        "candidate_rank", "Candidate Rank", "Candidate", "candidate.candidate_rank", "number", 110,
        source_stage="ListTaskCandidates",
    ),
    _v3_catalog_field(
        "rank_score", "Rank Score", "Candidate", "candidate.rank_score", "number", 120,
        source_stage="ListTaskCandidates",
    ),
    _v3_catalog_field("insights_status", "Insights Status", "Insights", "ui_sections.insights.status", "string", 200, normalizer="trim_text"),
    _v3_catalog_field("insights_count", "Insights Count", "Insights", "ui_sections.insights.data.count", "number", 210),
    _v3_catalog_field("insights_items", "Insights Items", "Insights", "ui_sections.insights.data.items", "array", 220, array_mode="collect"),
    _v3_catalog_field("insights_description", "Insights Description", "Insights", "ui_sections.insights.data.items[*].description", "array", 230, array_mode="collect", normalizer="string_list", compare_mode="semantic_text_lite"),
    _v3_catalog_field("insights_links", "Insights Links", "Insights", "ui_sections.insights.data.items[*].links", "array", 240, array_mode="collect"),
    _v3_catalog_field("photos_status", "Photos Status", "Photos", "ui_sections.photos.status", "string", 300, normalizer="trim_text"),
    _v3_catalog_field("photos_authenticity_photos", "Authenticity Photos", "Photos", "ui_sections.photos.data.authenticity_photos", "array", 310, array_mode="collect"),
    _v3_catalog_field("photos_baseline_photo_url", "Baseline Photo URL", "Photos", "ui_sections.photos.data.baseline_photo_url", "image", 320),
    _v3_catalog_field(
        "photos_identity_match_rate", "Photo Identity Match Rate", "Photos", "ui_sections.photos.data.identity_match_rate", "number", 330,
        normalizer="percentage", compare_mode="manual", baseline_field_key="photos_identity_match_rate",
        baseline_compare_enabled=True, identity_enabled=True,
    ),
    _v3_catalog_field("photos_match_photos", "Match Photos", "Photos", "ui_sections.photos.data.match_photos", "array", 340, array_mode="collect"),
    _v3_catalog_field("profile_status", "Profile Status", "Profile", "ui_sections.profile.status", "string", 400, normalizer="trim_text"),
    _v3_catalog_field("profile_data", "Profile Data", "Profile", "ui_sections.profile.data", "object", 410),
    _v3_catalog_field("profile_sections", "Profile Sections", "Profile", "ui_sections.profile.data.sections", "object", 420, normalizer="profile_sections"),
    _v3_catalog_field(
        "profile_full_name", "Profile Full Name", "Profile",
        "ui_sections.profile.data.sections", "string", 430,
        source_type="PROFILE_ITEM",
        source_options={"section": "Identity", "label": "Full Name"},
        normalizer="trim_text", compare_mode="normalized_text",
        scoring_role=["completeness", "accuracy"],
        baseline_field_key="profile_full_name", baseline_compare_enabled=True,
        completeness_enabled=True, accuracy_enabled=True,
    ),
    _v3_catalog_field(
        "profile_age", "Profile Age", "Profile",
        "ui_sections.profile.data.sections", "number", 440,
        source_type="PROFILE_ITEM",
        source_options={"section": "Identity", "label": "Age"},
        normalizer="number", compare_mode="exact",
        scoring_role=["completeness", "accuracy"],
        baseline_field_key="profile_age", baseline_compare_enabled=True,
        completeness_enabled=True, accuracy_enabled=True,
    ),
    _v3_catalog_field(
        "profile_location", "Profile Location", "Profile",
        "ui_sections.profile.data.sections", "string", 450,
        source_type="PROFILE_ITEM",
        source_options={"section": "Identity", "label": "Location"},
        normalizer="trim_text", compare_mode="semantic_text_lite",
        scoring_role=["completeness", "accuracy"],
        baseline_field_key="profile_location", baseline_compare_enabled=True,
        completeness_enabled=True, accuracy_enabled=True,
    ),
    _v3_catalog_field(
        "profile_jobs", "Profile Current Role", "Profile",
        "ui_sections.profile.data.sections", "array", 460,
        source_type="PROFILE_ITEM",
        source_options={
            "section": "Career",
            "labels": ["Current Role", "Other Roles"],
        },
        array_mode="collect", compare_mode="semantic_text_lite",
        scoring_role=["completeness", "accuracy"],
        baseline_field_key="profile_jobs", baseline_compare_enabled=True,
        completeness_enabled=True, accuracy_enabled=True,
    ),
    _v3_catalog_field(
        "profile_education", "Profile Education", "Profile",
        "ui_sections.profile.data.sections", "string", 470,
        source_type="PROFILE_ITEM",
        source_options={"section": "Background", "label": "Education"},
        normalizer="trim_text", compare_mode="semantic_text_lite",
        scoring_role=["completeness", "accuracy"],
        baseline_field_key="profile_education", baseline_compare_enabled=True,
        completeness_enabled=True, accuracy_enabled=True,
    ),
    _v3_catalog_field("social_status", "Social Status", "Social", "ui_sections.social.status", "string", 500, normalizer="trim_text"),
    _v3_catalog_field("social_private_accounts", "Private Accounts", "Social", "ui_sections.social.data.private_accounts", "array", 510, array_mode="collect"),
    _v3_catalog_field("social_display_handles", "Social Display Handles", "Social", "ui_sections.social.data.profiles[*].display_handle", "array", 520, array_mode="collect", normalizer="string_list", compare_mode="set"),
    _v3_catalog_field("social_platforms", "Social Platforms", "Social", "ui_sections.social.data.profiles[*].platform", "array", 530, array_mode="collect", normalizer="string_list", compare_mode="set"),
    _v3_catalog_field(
        "social_urls", "Social URLs", "Social", "ui_sections.social.data.profiles[*].url", "array", 540,
        array_mode="collect", normalizer="social_url", compare_mode="url_set",
        scoring_role=["identity", "completeness", "accuracy"], baseline_field_key="social_urls",
        baseline_compare_enabled=True, identity_enabled=True, completeness_enabled=True, accuracy_enabled=True,
    ),
    _v3_catalog_field("social_usernames", "Social Usernames", "Social", "ui_sections.social.data.profiles[*].username", "array", 550, array_mode="collect", normalizer="string_list", compare_mode="set"),
    _v3_catalog_field("summary_status", "Summary Status", "Summary", "ui_sections.summary.status", "string", 600, normalizer="trim_text"),
    _v3_catalog_field("summary_age", "Summary Age", "Summary", "ui_sections.summary.data.age", "number", 610),
    _v3_catalog_field("summary_avatar_url", "Summary Avatar URL", "Summary", "ui_sections.summary.data.avatar_url", "image", 620),
    _v3_catalog_field("summary_candidate_id", "Summary Candidate ID", "Summary", "ui_sections.summary.data.candidate_id", "string", 630, normalizer="trim_text", compare_mode="normalized_text"),
    _v3_catalog_field("candidate_confidence", "Summary Confidence Level", "Summary", "ui_sections.summary.data.confidence_level", "string", 640, normalizer="trim_text", compare_mode="normalized_text"),
    _v3_catalog_field("summary_disclaimers", "Summary Disclaimers", "Summary", "ui_sections.summary.data.disclaimers", "array", 650, array_mode="collect"),
    _v3_catalog_field("summary_display_name", "Summary Display Name", "Summary", "ui_sections.summary.data.display_name", "string", 660, normalizer="trim_text", compare_mode="normalized_text"),
    _v3_catalog_field("summary_education", "Summary Education", "Summary", "ui_sections.summary.data.education", "string", 670, normalizer="trim_text", compare_mode="semantic_text_lite"),
    _v3_catalog_field("summary_generate_time", "Summary Generate Time", "Summary", "ui_sections.summary.data.generate_time", "string", 680, normalizer="trim_text"),
    _v3_catalog_field("summary_headline", "Summary Headline", "Summary", "ui_sections.summary.data.headline", "string", 690, normalizer="trim_text", compare_mode="semantic_text_lite"),
    _v3_catalog_field("summary_is_best_match", "Summary Is Best Match", "Summary", "ui_sections.summary.data.is_best_match", "boolean", 700),
    _v3_catalog_field("summary_is_top_result", "Summary Is Top Result", "Summary", "ui_sections.summary.data.is_top_result", "boolean", 710),
    _v3_catalog_field("summary_jobs", "Summary Jobs", "Summary", "ui_sections.summary.data.jobs", "array", 720, array_mode="collect", compare_mode="semantic_text_lite"),
    _v3_catalog_field("summary_location", "Summary Location", "Summary", "ui_sections.summary.data.location", "string", 730, normalizer="trim_text", compare_mode="semantic_text_lite"),
    _v3_catalog_field("summary_match_reasons", "Summary Match Reasons", "Summary", "ui_sections.summary.data.match_reasons", "array", 740, array_mode="collect"),
    _v3_catalog_field("summary_match_score", "Summary Match Score", "Summary", "ui_sections.summary.data.match_score", "number", 750, normalizer="percentage"),
    _v3_catalog_field("summary_more_social_count", "More Social Count", "Summary", "ui_sections.summary.data.more_social_count", "number", 760),
    _v3_catalog_field("summary_person_id", "Summary Person ID", "Summary", "ui_sections.summary.data.person_id", "string", 770, normalizer="trim_text", compare_mode="normalized_text"),
    _v3_catalog_field("summary_primary_image", "Summary Primary Image", "Summary", "ui_sections.summary.data.primary_image", "object", 780),
    _v3_catalog_field("summary_profile_url", "Summary Profile URL", "Summary", "ui_sections.summary.data.profile_url", "url", 790),
    _v3_catalog_field("summary_report_expires_at", "Summary Report Expires At", "Summary", "ui_sections.summary.data.report_expires_at", "string", 800, normalizer="trim_text"),
    _v3_catalog_field("summary_social_platforms", "Summary Social Platforms", "Summary", "ui_sections.summary.data.social_links[*].platform", "array", 810, array_mode="collect", normalizer="string_list", compare_mode="set"),
    _v3_catalog_field("summary_social_titles", "Summary Social Titles", "Summary", "ui_sections.summary.data.social_links[*].title", "array", 820, array_mode="collect", normalizer="string_list"),
    _v3_catalog_field("summary_social_links", "Summary Social Links", "Summary", "ui_sections.summary.data.social_links[*].url", "array", 830, array_mode="collect", normalizer="social_url", compare_mode="url_set"),
    _v3_catalog_field("summary_social_platform_list", "Summary Social Platform List", "Summary", "ui_sections.summary.data.social_platforms", "array", 840, array_mode="collect", normalizer="string_list", compare_mode="set"),
    _v3_catalog_field("summary_web_platforms", "Web Link Platforms", "Summary", "ui_sections.summary.data.web_links[*].platform", "array", 850, array_mode="collect", normalizer="string_list", compare_mode="set"),
    _v3_catalog_field("summary_web_titles", "Web Link Titles", "Summary", "ui_sections.summary.data.web_links[*].title", "array", 860, array_mode="collect", normalizer="string_list"),
    _v3_catalog_field("summary_web_links", "Summary Web Links", "Summary", "ui_sections.summary.data.web_links[*].url", "array", 870, array_mode="collect", normalizer="social_url", compare_mode="url_set"),
]


class ImportValidationError(ValueError):
    """导入文件存在格式或业务校验错误。"""

    def __init__(self, errors: Iterable[str]) -> None:
        """保存可供页面预览的全部错误。"""

        self.errors = list(errors)
        super().__init__("；".join(self.errors))


class DuplicateImportError(ValueError):
    """相同 SHA-256 内容已经导入，禁止静默创建副本。"""


class ActiveRunError(RuntimeError):
    """已有待执行或执行中的采集 Run，禁止并行启动第二个 Run。"""


class FieldSchemaValidationError(ValueError):
    """字段配置、路径或转换器输入不符合受限规则。"""

    def __init__(self, errors: str | Iterable[str]) -> None:
        """保存适合 Web 展示的一个或多个字段错误。"""

        self.errors = [errors] if isinstance(errors, str) else list(errors)
        super().__init__("；".join(self.errors))


class ReviewValidationError(ValueError):
    """候选人复核内容非法或页面版本已经过期。"""

    def __init__(self, errors: str | Iterable[str]) -> None:
        """保存可直接展示给测试人员的复核错误。"""

        self.errors = [errors] if isinstance(errors, str) else list(errors)
        super().__init__("；".join(self.errors))


class PersonLinkValidationError(ValueError):
    """历史 Run Query 人物关联输入非法或页面版本已经过期。"""

    def __init__(self, errors: str | Iterable[str]) -> None:
        """保存可直接展示给测试人员的一个或多个关联错误。"""

        self.errors = [errors] if isinstance(errors, str) else list(errors)
        super().__init__("；".join(self.errors))


@dataclass
class ImportPreview:
    """导入前校验摘要。"""

    checksum: str
    valid_count: int
    errors: list[str]


@dataclass
class ImportResult:
    """一次成功导入的最小返回信息。"""

    object_id: str
    imported_count: int
    checksum: str
    archived_files: list[str]


@dataclass
class ProcessResult:
    """一次字段处理完成后的摘要。"""

    process_id: str
    candidate_count: int
    error_count: int
    status: str
    warnings: list[str] = field(default_factory=list)


@dataclass
class ReportResult:
    """一次报告快照创建后的可展示和可导出结果。"""

    report_id: str
    model: dict[str, Any]
    html_file: str
    excel_file: str | None


def _source_path_tokens(path: str) -> list[tuple[str, str | int | None]]:
    """解析受限字段路径，不接受表达式、过滤器或函数调用。"""

    if not isinstance(path, str) or not path.strip():
        raise FieldSchemaValidationError("source_path 不能为空")
    tokens: list[tuple[str, str | int | None]] = []
    for part in path.split("."):
        match = FIELD_PATH_PART_PATTERN.fullmatch(part)
        if match is None:
            raise FieldSchemaValidationError(
                f"source_path 语法无效: {path}；只支持点号、[数字] 和 [*]"
            )
        tokens.append(("key", match.group("key")))
        index = match.group("index")
        if index == "*":
            tokens.append(("wildcard", None))
        elif index is not None:
            tokens.append(("index", int(index)))
    return tokens


def extract_source_path(
    source: Any,
    path: str,
    *,
    missing_policy: str = "ERROR",
) -> Any:
    """按受限路径从 JSON 兼容对象提取值。

    参数说明:
        source: Candidate Detail、Task 和 Candidate 组成的处理源对象。
        path: 点号、固定索引和数组通配组成的路径。
        missing_policy: ``EMPTY`` 将缺失键、空父节点和越界索引返回为空；
            ``ERROR`` 保持旧版严格行为。

    返回值:
        无通配时返回单值，缺失可返回 ``None``；存在通配时返回保持原
        顺序的数组，缺失或空数组返回 ``[]``。

    异常说明:
        FieldSchemaValidationError: 策略/路径语法非法，严格模式下路径缺失，
        或已有数据的结构类型不匹配。
    """

    if missing_policy not in FIELD_MISSING_POLICIES:
        raise FieldSchemaValidationError(
            f"不支持的 missing_policy: {missing_policy}"
        )
    values = [source]
    wildcard_used = False
    for token_type, token_value in _source_path_tokens(path):
        next_values: list[Any] = []
        if token_type == "key":
            key = str(token_value)
            for value in values:
                if value is None and missing_policy == "EMPTY":
                    continue
                if not isinstance(value, dict):
                    raise FieldSchemaValidationError(
                        f"source_path 字段父节点不是对象: {path}"
                    )
                if key not in value:
                    if missing_policy == "EMPTY":
                        continue
                    raise FieldSchemaValidationError(
                        f"source_path 找不到字段: {path}"
                    )
                next_values.append(value[key])
        elif token_type == "index":
            index = int(token_value)
            for value in values:
                if value is None and missing_policy == "EMPTY":
                    continue
                if not isinstance(value, list):
                    raise FieldSchemaValidationError(
                        f"source_path 索引目标不是数组: {path}"
                    )
                if index >= len(value):
                    if missing_policy == "EMPTY":
                        continue
                    raise FieldSchemaValidationError(
                        f"source_path 数组索引不存在: {path}"
                    )
                next_values.append(value[index])
        else:
            wildcard_used = True
            for value in values:
                if value is None and missing_policy == "EMPTY":
                    continue
                if not isinstance(value, list):
                    raise FieldSchemaValidationError(
                        f"source_path 通配目标不是数组: {path}"
                    )
                next_values.extend(value)
        values = next_values
    if wildcard_used:
        return values
    if not values and missing_policy == "EMPTY":
        return None
    if len(values) != 1:
        raise FieldSchemaValidationError(f"source_path 无法得到单一值: {path}")
    return values[0]


def _profile_selector_text(value: Any) -> str:
    """规范化 Profile 分组标题和标签，供受控选择器进行稳定匹配。"""

    return " ".join(
        unicodedata.normalize("NFKC", str(value)).strip().casefold().split()
    )


def extract_profile_item(
    source: Any,
    source_path: str,
    source_options: dict[str, Any],
    *,
    missing_policy: str = "EMPTY",
) -> tuple[Any, bool]:
    """按 ``section + label/labels`` 从 Profile sections 提取受控字段。

    参数说明：``source_path`` 指向 sections 数组；``source_options`` 必须
    包含 section，以及单个 label 或多个 labels。返回值的第二项仅表示
    同一个已选标签重复出现；多标签聚合本身不视为重复。找不到项目遵循
    EMPTY/ERROR 策略，不接受任意 JSONPath 条件或脚本。
    """

    sections = extract_source_path(
        source, source_path, missing_policy=missing_policy
    )
    if sections is None:
        return None, False
    if not isinstance(sections, list):
        raise FieldSchemaValidationError("PROFILE_ITEM 的 sections 必须是数组")
    wanted_section = _profile_selector_text(source_options.get("section", ""))
    single_label = source_options.get("label")
    multiple_labels = source_options.get("labels")
    raw_labels = (
        [single_label]
        if isinstance(single_label, str)
        else multiple_labels if isinstance(multiple_labels, list) else []
    )
    wanted_labels = {
        _profile_selector_text(label)
        for label in raw_labels
        if isinstance(label, str) and label.strip()
    }
    values: list[Any] = []
    seen_labels: set[str] = set()
    duplicated = False
    for section in sections:
        if not isinstance(section, dict):
            continue
        if _profile_selector_text(section.get("title", "")) != wanted_section:
            continue
        items = section.get("items", [])
        if not isinstance(items, list):
            continue
        for item in items:
            item_label = (
                _profile_selector_text(item.get("label", ""))
                if isinstance(item, dict) else ""
            )
            if isinstance(item, dict) and item_label in wanted_labels:
                if item_label in seen_labels:
                    duplicated = True
                seen_labels.add(item_label)
                values.append(sanitize_raw(item.get("value")))
    if not values:
        if missing_policy == "ERROR":
            raise FieldSchemaValidationError(
                "PROFILE_ITEM 找不到指定的 section / label"
            )
        return None, False
    return (values[0] if len(values) == 1 else values), duplicated


def _normalize_social_url(value: str) -> str:
    """规范化单个 Social URL，移除格式差异但保留账号语义。

    X 与 Twitter 是同一平台的历史域名；Twitter、Facebook、Instagram、LinkedIn
    和 TikTok 的账号路径不区分大小写。因此这些不影响账号身份的差异统一处理，
    避免同一主页被误判为不同链接。
    """

    text = value.strip()
    parsed = urlsplit(text)
    scheme = parsed.scheme.lower()
    if scheme not in {"http", "https"} or not parsed.hostname:
        raise FieldSchemaValidationError(f"不是有效的 http/https URL: {value!r}")
    hostname = parsed.hostname.lower()
    if hostname.startswith("www."):
        hostname = hostname[4:]
    hostname = {
        "x.com": "twitter.com",
        "fb.com": "facebook.com",
    }.get(hostname, hostname)
    netloc = hostname
    try:
        port = parsed.port
    except ValueError as exc:
        raise FieldSchemaValidationError(
            f"URL 端口无效: {value!r}"
        ) from exc
    if port is not None:
        netloc = f"{hostname}:{port}"
    tracking_names = {"fbclid", "gclid", "mc_cid", "mc_eid"}
    query = [
        (key, item)
        for key, item in parse_qsl(parsed.query, keep_blank_values=True)
        if not key.lower().startswith("utm_")
        and key.lower() not in tracking_names
    ]
    path = parsed.path.rstrip("/")
    if hostname in {
        "twitter.com",
        "facebook.com",
        "instagram.com",
        "linkedin.com",
        "tiktok.com",
    }:
        path = path.casefold()
    return urlunsplit(
        (scheme, netloc, path, urlencode(query, doseq=True), "")
    )


def _normalize_profile_sections(value: Any) -> dict[str, Any]:
    """把 Profile sections/items 转为 section → label → value 对象。"""

    if isinstance(value, dict):
        # 部分接口版本已经直接返回按 section 分组的对象；保持结构并交给
        # processing-v3 的模块语义判空，避免把合法扩展格式记为处理错误。
        return sanitize_raw(value)
    if not isinstance(value, list):
        raise FieldSchemaValidationError("profile_sections 输入必须是数组或对象")
    result: dict[str, dict[str, Any]] = {}
    for section_index, section in enumerate(value):
        if not isinstance(section, dict):
            raise FieldSchemaValidationError(
                f"profile_sections[{section_index}] 必须是对象"
            )
        title = section.get("title")
        items = section.get("items")
        if not isinstance(title, str) or not title.strip():
            raise FieldSchemaValidationError(
                f"profile_sections[{section_index}] 缺少 title"
            )
        if not isinstance(items, list):
            raise FieldSchemaValidationError(
                f"profile_sections[{section_index}].items 必须是数组"
            )
        section_values: dict[str, Any] = {}
        for item_index, item in enumerate(items):
            if not isinstance(item, dict):
                raise FieldSchemaValidationError(
                    f"profile_sections[{section_index}].items[{item_index}] "
                    "必须是对象"
                )
            label = item.get("label")
            if not isinstance(label, str) or not label.strip():
                raise FieldSchemaValidationError(
                    f"profile_sections[{section_index}].items[{item_index}] "
                    "缺少 label"
                )
            section_values[label.strip()] = sanitize_raw(item.get("value"))
        result[title.strip()] = section_values
    return result


def normalize_field_value(value: Any, normalizer: str) -> Any:
    """执行白名单内置转换器，不运行用户代码或任意表达式。"""

    if normalizer == "identity":
        return sanitize_raw(value)
    if normalizer == "trim_text":
        # 接口中的日期时间字段有时以 Unix 毫秒时间戳返回。它们语义上仍是
        # 可展示的文本字段；将数值标量安全转为文本，避免把正常响应误记为
        # FIELD_PROCESSING_ERROR。对象、数组和布尔值仍拒绝，防止结构错误被
        # 静默掩盖。
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if not math.isfinite(float(value)):
                raise FieldSchemaValidationError("trim_text 输入必须是有限文本或数值")
            return str(value).strip()
        if not isinstance(value, str):
            raise FieldSchemaValidationError("trim_text 输入必须是字符串")
        return value.strip()
    if normalizer in {"number", "percentage"}:
        if isinstance(value, bool):
            raise FieldSchemaValidationError(f"{normalizer} 不接受布尔值")
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise FieldSchemaValidationError(
                f"{normalizer} 输入必须是有限数值"
            ) from exc
        if not math.isfinite(number):
            raise FieldSchemaValidationError(
                f"{normalizer} 输入必须是有限数值"
            )
        if normalizer == "percentage":
            if 0 <= number <= 1:
                number *= 100
            if not 0 <= number <= 100:
                raise FieldSchemaValidationError("percentage 必须在 0–100 范围")
        return number
    if normalizer == "social_url":
        if isinstance(value, list):
            return [_normalize_social_url(str(item)) for item in value]
        if not isinstance(value, str):
            raise FieldSchemaValidationError(
                "social_url 输入必须是字符串或字符串数组"
            )
        return _normalize_social_url(value)
    if normalizer == "string_list":
        items = value if isinstance(value, list) else [value]
        normalized: list[str] = []
        for item in items:
            if not isinstance(item, (str, int, float, bool)):
                raise FieldSchemaValidationError(
                    "string_list 只接受标量或标量数组"
                )
            text = str(item).strip()
            if text and text not in normalized:
                normalized.append(text)
        return normalized
    if normalizer == "profile_sections":
        return _normalize_profile_sections(value)
    raise FieldSchemaValidationError(f"不支持的 normalizer: {normalizer}")


def _is_empty_value(value: Any, empty_rule: str) -> bool:
    """按字段空值规则判断空值，默认不把0和False当空。"""

    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    if isinstance(value, (list, dict)) and not value:
        return True
    if empty_rule == "zero_is_empty" and value == 0 and not isinstance(value, bool):
        return True
    if empty_rule == "false_is_empty" and value is False:
        return True
    return False


def _has_module_business_value(value: Any) -> bool:
    """递归判断模块容器是否包含业务值，忽略纯计数和空容器。

    数值0和布尔False在普通字段中仍是有效值；这里只忽略模块容器中的
    ``count=0``，避免 ``{"count": 0, "items": []}`` 被当成有数据。
    """

    if value is None or value == "" or value == [] or value == {}:
        return False
    if isinstance(value, dict):
        return any(
            _has_module_business_value(item)
            for key, item in value.items()
            if key not in {"count", "total_count"}
        )
    if isinstance(value, list):
        return any(_has_module_business_value(item) for item in value)
    return True


def _apply_module_empty_rules(
    fields: dict[str, Any],
    empty_fields: dict[str, bool],
    definitions: list[dict[str, Any]],
) -> dict[str, bool]:
    """按 processing-v3 的模块业务语义修正 Candidate 字段空值。

    状态字段只用于解释模块，不会因 ``status=empty`` 自身有文本而把业务
    模块判为有数据。函数仅覆盖模块内容字段，不修改已提取的原始值。
    """

    result = dict(empty_fields)
    definitions_by_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for definition in definitions:
        if (
            definition.get("enabled")
            and definition.get("value_scope") == "CANDIDATE"
        ):
            definitions_by_module[definition["module"]].append(definition)
    for module in ("Insights", "Photos", "Profile", "Social"):
        module_key = module.lower()
        status = str(fields.get(f"{module_key}_status") or "").strip().lower()
        content_definitions = [
            definition
            for definition in definitions_by_module.get(module, [])
            if definition["field_key"] != f"{module_key}_status"
        ]
        if not content_definitions:
            continue
        if status == "empty":
            for definition in content_definitions:
                result[definition["field_key"]] = True
            continue
        if status == "data":
            content_has_value = any(
                _has_module_business_value(fields.get(definition["field_key"]))
                for definition in content_definitions
            )
            if not content_has_value:
                for definition in content_definitions:
                    result[definition["field_key"]] = True
    return result


def _field_value_shape(value: Any) -> str:
    """返回矩阵结构对齐使用的稳定类型标签。"""

    if value is None:
        return "empty"
    if isinstance(value, list):
        nonempty = [item for item in value if not _is_empty_value(item, "default")]
        if not nonempty:
            return "array_empty"
        item_shapes = {_field_value_shape(item) for item in nonempty}
        return "array_" + (
            next(iter(item_shapes)) if len(item_shapes) == 1 else "mixed"
        )
    if isinstance(value, dict):
        return "object"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    return "string"


def _matrix_sample(value: Any, limit: int = 180) -> Any:
    """生成字段矩阵的安全截断样例，不改变原始快照。"""

    if value is None:
        return None
    sanitized = sanitize_raw(value)
    text = json.dumps(sanitized, ensure_ascii=False, separators=(",", ":"))
    if len(text) <= limit:
        return sanitized
    return text[: limit - 1] + "…"


def _apply_array_mode(value: Any, array_mode: str) -> Any:
    """应用保留、首项、收集或合并数组策略。"""

    if array_mode == "preserve":
        return value
    if array_mode == "first":
        if not isinstance(value, list):
            raise FieldSchemaValidationError("array_mode=first 的输入必须是数组")
        return value[0] if value else None
    if array_mode == "collect":
        return value if isinstance(value, list) else [value]
    if array_mode == "join":
        if not isinstance(value, list):
            raise FieldSchemaValidationError("array_mode=join 的输入必须是数组")
        if not all(isinstance(item, (str, int, float, bool)) for item in value):
            raise FieldSchemaValidationError("array_mode=join 只支持标量数组")
        return "\n".join(str(item) for item in value)
    raise FieldSchemaValidationError(f"不支持的 array_mode: {array_mode}")


def _validate_field_data_type(value: Any, data_type: str) -> Any:
    """校验规范化后的值类型，并返回可稳定 JSON 序列化的值。"""

    if value is None:
        return None
    if data_type in {"string", "url", "image"}:
        if not isinstance(value, str):
            raise FieldSchemaValidationError(f"{data_type} 字段必须是字符串")
        if value and data_type in {"url", "image"}:
            parsed = urlsplit(value)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                raise FieldSchemaValidationError(
                    f"{data_type} 字段必须是 http/https URL"
                )
        return value
    if data_type == "number":
        if (
            not isinstance(value, (int, float))
            or isinstance(value, bool)
            or not math.isfinite(float(value))
        ):
            raise FieldSchemaValidationError("number 字段必须是有限数值")
        return value
    if data_type == "boolean":
        if isinstance(value, bool):
            return value
        if value in {0, 1}:
            return bool(value)
        raise FieldSchemaValidationError("boolean 字段必须是布尔值或0/1")
    if data_type == "object":
        if not isinstance(value, dict):
            raise FieldSchemaValidationError("object 字段必须是对象")
        return value
    if data_type == "array":
        if not isinstance(value, list):
            raise FieldSchemaValidationError("array 字段必须是数组")
        return value
    raise FieldSchemaValidationError(f"不支持的 data_type: {data_type}")


def validate_field_definitions(
    definitions: Any,
) -> list[dict[str, Any]]:
    """校验 FieldSchema v3，并在内存中兼容读取旧定义。

    功能说明：字段配置仍保存在既有 ``definitions_json`` 中。旧 v1/v2
    定义缺少的 v3 开关和来源属性在此推导，绝不回写历史 Schema；新发布
    的定义则校验完整职责，避免复杂 JSON 或未接通成本字段误入正式评分。
    """

    if not isinstance(definitions, list) or not definitions:
        raise FieldSchemaValidationError("字段配置必须是非空数组")
    required = {
        "field_key",
        "display_name",
        "module",
        "source_stage",
        "source_path",
        "data_type",
        "array_mode",
        "empty_rule",
        "normalizer",
        "scoring_role",
        "compare_mode",
        "enabled",
        "sort_order",
    }
    normalized: list[dict[str, Any]] = []
    errors: list[str] = []
    seen_keys: set[str] = set()
    for index, raw_definition in enumerate(definitions, start=1):
        label = f"第 {index} 个字段"
        if not isinstance(raw_definition, dict):
            errors.append(f"{label}必须是对象")
            continue
        missing = sorted(required - set(raw_definition))
        if missing:
            errors.append(f"{label}缺少配置项: {', '.join(missing)}")
            continue
        definition = {key: sanitize_raw(raw_definition[key]) for key in required}
        field_key = definition["field_key"]
        is_v3 = any(
            key in raw_definition
            for key in {
                "source_type", "source_options", "display_enabled",
                "baseline_field_key", "baseline_compare_enabled",
                "run_compare_enabled", "identity_enabled",
                "completeness_enabled", "accuracy_enabled",
                "similarity_threshold",
            }
        )
        definition["value_scope"] = sanitize_raw(
            raw_definition.get(
                "value_scope",
                "QUERY" if definition["module"] == "Task" else "CANDIDATE",
            )
        )
        definition["missing_policy"] = sanitize_raw(
            raw_definition.get(
                "missing_policy",
                "ERROR"
                if field_key in {"task_id", "candidate_id"}
                else "EMPTY",
            )
        )
        legacy_roles = definition["scoring_role"]
        definition["source_type"] = sanitize_raw(
            raw_definition.get("source_type", "PATH")
        )
        definition["source_options"] = sanitize_raw(
            raw_definition.get("source_options", {})
        )
        definition["display_enabled"] = sanitize_raw(
            raw_definition.get("display_enabled", definition["enabled"])
        )
        default_baseline_compare = any(
            role in {"identity", "completeness", "accuracy"}
            for role in legacy_roles
        ) if isinstance(legacy_roles, list) else False
        definition["baseline_compare_enabled"] = sanitize_raw(
            raw_definition.get(
                "baseline_compare_enabled", default_baseline_compare
            )
        )
        definition["run_compare_enabled"] = sanitize_raw(
            raw_definition.get("run_compare_enabled", definition["enabled"])
        )
        definition["identity_enabled"] = sanitize_raw(
            raw_definition.get(
                "identity_enabled",
                "identity" in legacy_roles
                if isinstance(legacy_roles, list) else False,
            )
        )
        definition["completeness_enabled"] = sanitize_raw(
            raw_definition.get(
                "completeness_enabled",
                "completeness" in legacy_roles
                if isinstance(legacy_roles, list) else False,
            )
        )
        definition["accuracy_enabled"] = sanitize_raw(
            raw_definition.get(
                "accuracy_enabled",
                "accuracy" in legacy_roles
                if isinstance(legacy_roles, list) else False,
            )
        )
        definition["baseline_field_key"] = sanitize_raw(
            raw_definition.get(
                "baseline_field_key",
                field_key if default_baseline_compare else "",
            )
        )
        definition["similarity_threshold"] = sanitize_raw(
            raw_definition.get("similarity_threshold", 0.6)
        )
        definition["field_schema_version"] = sanitize_raw(
            raw_definition.get("field_schema_version", "v2")
        )
        if not isinstance(field_key, str) or not FIELD_KEY_PATTERN.fullmatch(
            field_key
        ):
            errors.append(f"{label} field_key 格式无效")
        elif field_key in seen_keys:
            errors.append(f"{label} field_key 重复: {field_key}")
        else:
            seen_keys.add(field_key)
        if (
            not isinstance(definition["display_name"], str)
            or not definition["display_name"].strip()
        ):
            errors.append(f"{label} display_name 不能为空")
        if (
            not isinstance(definition["module"], str)
            or definition["module"] not in FIELD_MODULES
        ):
            errors.append(f"{label} module 不受支持")
        if (
            not isinstance(definition["source_stage"], str)
            or definition["source_stage"] not in FIELD_SOURCE_STAGES
        ):
            errors.append(f"{label} source_stage 不受支持")
        if (
            not isinstance(definition["source_type"], str)
            or definition["source_type"] not in FIELD_SOURCE_TYPES
        ):
            errors.append(f"{label} source_type 不受支持")
        if not isinstance(definition["source_options"], dict):
            errors.append(f"{label} source_options 必须是对象")
        if definition["field_schema_version"] not in {"v2", "v3"}:
            errors.append(f"{label} field_schema_version 不受支持")
        if (
            not isinstance(definition["value_scope"], str)
            or definition["value_scope"] not in FIELD_VALUE_SCOPES
        ):
            errors.append(f"{label} value_scope 不受支持")
        if (
            not isinstance(definition["missing_policy"], str)
            or definition["missing_policy"] not in FIELD_MISSING_POLICIES
        ):
            errors.append(f"{label} missing_policy 不受支持")
        if (
            definition["value_scope"] == "QUERY"
            and definition["source_stage"] != "GetTask"
        ):
            errors.append(f"{label} QUERY 字段仅支持 GetTask")
        try:
            _source_path_tokens(definition["source_path"])
        except FieldSchemaValidationError as exc:
            errors.extend(f"{label} {item}" for item in exc.errors)
        if definition["source_type"] == "PROFILE_ITEM":
            section = definition["source_options"].get("section")
            item_label = definition["source_options"].get("label")
            item_labels = definition["source_options"].get("labels")
            if not isinstance(section, str) or not section.strip():
                errors.append(f"{label} PROFILE_ITEM 必须配置 source_options.section")
            has_single_label = isinstance(item_label, str) and bool(item_label.strip())
            has_multiple_labels = (
                isinstance(item_labels, list)
                and bool(item_labels)
                and all(
                    isinstance(item, str) and item.strip()
                    for item in item_labels
                )
            )
            if has_single_label == has_multiple_labels:
                errors.append(
                    f"{label} PROFILE_ITEM 必须且只能配置 source_options.label 或 labels"
                )
        if (
            not isinstance(definition["data_type"], str)
            or definition["data_type"] not in FIELD_DATA_TYPES
        ):
            errors.append(f"{label} data_type 不受支持")
        if (
            field_key == "candidate_confidence"
            and definition["data_type"] != "string"
        ):
            errors.append(f"{label} candidate_confidence 必须是 string")
        if (
            field_key
            in {
                "llm_cost",
                "third_party_cost",
                "total_cost",
                "search_duration_ms",
            }
            and definition["data_type"] != "number"
        ):
            errors.append(f"{label} 成本/耗时字段必须是 number")
        if field_key == "pdl_called" and definition["data_type"] != "boolean":
            errors.append(f"{label} pdl_called 必须是 boolean")
        if (
            not isinstance(definition["array_mode"], str)
            or definition["array_mode"] not in FIELD_ARRAY_MODES
        ):
            errors.append(f"{label} array_mode 不受支持")
        if (
            not isinstance(definition["empty_rule"], str)
            or definition["empty_rule"] not in FIELD_EMPTY_RULES
        ):
            errors.append(f"{label} empty_rule 不受支持")
        if (
            not isinstance(definition["normalizer"], str)
            or definition["normalizer"] not in FIELD_NORMALIZERS
        ):
            errors.append(f"{label} normalizer 不受支持")
        roles = definition["scoring_role"]
        if (
            not isinstance(roles, list)
            or not roles
            or any(
                not isinstance(role, str)
                or role not in FIELD_SCORING_ROLES
                for role in roles
            )
        ):
            errors.append(f"{label} scoring_role 必须是受支持的非空数组")
        if (
            not isinstance(definition["compare_mode"], str)
            or definition["compare_mode"] not in FIELD_COMPARE_MODES
        ):
            errors.append(f"{label} compare_mode 不受支持")
        if not isinstance(definition["enabled"], bool):
            errors.append(f"{label} enabled 必须是布尔值")
        for switch_name in (
            "display_enabled", "baseline_compare_enabled",
            "run_compare_enabled", "identity_enabled",
            "completeness_enabled", "accuracy_enabled",
        ):
            if not isinstance(definition[switch_name], bool):
                errors.append(f"{label} {switch_name} 必须是布尔值")
        if (
            not isinstance(definition["baseline_field_key"], str)
            or (
                definition["baseline_field_key"]
                and not FIELD_KEY_PATTERN.fullmatch(
                    definition["baseline_field_key"]
                )
            )
        ):
            errors.append(f"{label} baseline_field_key 格式无效")
        if definition["baseline_compare_enabled"] and not definition[
            "baseline_field_key"
        ]:
            errors.append(f"{label} 开启基准对比必须配置 baseline_field_key")
        if (
            not isinstance(definition["similarity_threshold"], (int, float))
            or isinstance(definition["similarity_threshold"], bool)
            or not 0 <= float(definition["similarity_threshold"]) <= 1
        ):
            errors.append(f"{label} similarity_threshold 必须在 0 到 1 之间")
        if definition["accuracy_enabled"] and definition["compare_mode"] == "presence":
            errors.append(f"{label} 准确度字段不能使用 presence")
        if (
            definition["accuracy_enabled"]
            and definition["data_type"] in {"object", "array"}
            and definition["compare_mode"] == "exact"
        ):
            errors.append(f"{label} 复杂对象不能使用 exact 自动准确度")
        if definition["identity_enabled"] and field_key not in {
            "social_urls", "summary_social_links", "photos_identity_match_rate",
        }:
            errors.append(f"{label} identity_enabled 仅支持 Social URL 或照片相似度")
        if (
            not isinstance(definition["sort_order"], int)
            or isinstance(definition["sort_order"], bool)
        ):
            errors.append(f"{label} sort_order 必须是整数")
        if is_v3 and all(
            isinstance(definition[item], bool)
            for item in (
                "identity_enabled", "completeness_enabled", "accuracy_enabled",
            )
        ):
            definition["scoring_role"] = [
                role
                for role, enabled in (
                    ("identity", definition["identity_enabled"]),
                    ("completeness", definition["completeness_enabled"]),
                    ("accuracy", definition["accuracy_enabled"]),
                )
                if enabled
            ] or ["display"]
        definition["display_name"] = str(definition["display_name"]).strip()
        normalized.append(definition)
    if errors:
        raise FieldSchemaValidationError(errors)
    return sorted(normalized, key=lambda item: (item["sort_order"], item["field_key"]))


def json_text(value: Any) -> str:
    """将 JSON 兼容值序列化为稳定、紧凑的 UTF-8 文本。"""

    return json.dumps(
        sanitize_raw(value),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )


def read_jsonl(path: Path) -> tuple[list[Any], list[str]]:
    """读取 JSONL 并返回合法对象与可预览的行级错误。"""

    records: list[Any] = []
    errors: list[str] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = raw_line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                errors.append(f"第 {line_number} 行 JSON 格式错误: {exc.msg}")
    return records, errors


def file_checksum(files: Iterable[tuple[str, Path]]) -> str:
    """按文件角色和内容计算稳定 SHA-256，支持多文件导入包去重。"""

    digest = hashlib.sha256()
    for role, path in files:
        digest.update(role.encode("utf-8"))
        digest.update(b"\0")
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        digest.update(b"\0")
    return digest.hexdigest()


def validate_storage_id(value: str, label: str) -> str:
    """校验会进入数据库主键和归档目录的外部 ID。"""

    if not SAFE_ID_PATTERN.fullmatch(value) or ".." in value:
        raise ValueError(
            f"{label} 只能包含字母、数字、点、下划线和连字符，且不能包含 '..'"
        )
    return value


def cell_value(value: Any) -> Any:
    """把 Excel 文本中的 JSON 或多行值恢复为可用 Python 值。"""

    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return ""
    if text[:1] in {"[", "{"}:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    if "\n" in text:
        return [item.strip() for item in text.splitlines() if item.strip()]
    return text


def normalize_available_fields(value: Any) -> list[str]:
    """校验并去重 Baseline 可用字段键。

    参数说明:
        value: 必须为字符串数组；允许显式空数组表示完整度未就绪。

    返回值:
        保持首次出现顺序的去重字段键数组。

    异常说明:
        ValueError: 输入不是数组、包含空字符串或非字符串。
    """

    if not isinstance(value, list):
        raise ValueError("baseline_available_fields 必须是字符串数组")
    normalized: list[str] = []
    for item in value:
        if not isinstance(item, str) or not item.strip():
            raise ValueError(
                "baseline_available_fields 不能包含空值或非字符串"
            )
        field_key = item.strip()
        if field_key not in normalized:
            normalized.append(field_key)
    return normalized


def normalize_evaluation_thresholds(value: Any) -> dict[str, dict[str, float | None]]:
    """校验并补齐 Evaluation 的分检索条件参考线。

    参数说明:
        value: 以 ``FULL_NAME``、``FULL_NAME_SOCIAL`` 为键的对象；每个
            阶段可以只提交部分阈值，空字符串和 ``None`` 均表示未配置。

    返回值:
        包含两个 Query Stage 和全部支持字段的稳定对象，数值统一为 float。

    异常说明:
        ReviewValidationError: 结构、字段名、数值类型或取值范围非法。
    """

    if value is None:
        value = {}
    if not isinstance(value, dict):
        raise ReviewValidationError("Evaluation 参考线必须是对象")
    unknown_stages = set(value) - SUPPORTED_QUERY_STAGES
    if unknown_stages:
        raise ReviewValidationError(
            "参考线包含不支持的 query_stage: "
            + ", ".join(sorted(unknown_stages))
        )
    normalized: dict[str, dict[str, float | None]] = {}
    errors: list[str] = []
    for query_stage in sorted(SUPPORTED_QUERY_STAGES):
        raw_stage = value.get(query_stage, {})
        if raw_stage is None:
            raw_stage = {}
        if not isinstance(raw_stage, dict):
            errors.append(f"{query_stage} 参考线必须是对象")
            raw_stage = {}
        unknown_fields = set(raw_stage) - set(EVALUATION_THRESHOLD_FIELDS)
        if unknown_fields:
            errors.append(
                f"{query_stage} 包含未知参考线: "
                + ", ".join(sorted(unknown_fields))
            )
        stage_values: dict[str, float | None] = {}
        for field_key, rule in EVALUATION_THRESHOLD_FIELDS.items():
            raw_value = raw_stage.get(field_key)
            if raw_value in (None, ""):
                stage_values[field_key] = None
                continue
            if isinstance(raw_value, bool):
                errors.append(f"{query_stage}.{field_key} 必须是数值")
                stage_values[field_key] = None
                continue
            try:
                number = float(raw_value)
            except (TypeError, ValueError):
                errors.append(f"{query_stage}.{field_key} 必须是数值")
                stage_values[field_key] = None
                continue
            if not math.isfinite(number):
                errors.append(f"{query_stage}.{field_key} 必须是有限数值")
            elif rule == "minimum_ratio" and not 0 <= number <= 1:
                errors.append(f"{query_stage}.{field_key} 必须在0到1之间")
            elif rule == "maximum_number" and number < 0:
                errors.append(f"{query_stage}.{field_key} 不能为负数")
            stage_values[field_key] = number
        normalized[query_stage] = stage_values
    if errors:
        raise ReviewValidationError(errors)
    return normalized


def _derived_available_fields(fields: dict[str, Any]) -> list[str]:
    """从旧 Baseline 非空值推导兼容字段键，不改变显式空数组语义。"""

    return [
        str(field_key)
        for field_key, value in fields.items()
        if not _is_empty_value(value, "default")
    ]


def _value_items(value: Any) -> list[Any]:
    """把单值或数组统一为可比较数组，空值返回空数组。"""

    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def _social_url_items(value: Any) -> list[str]:
    """从 Social 字段提取 URL，兼容字符串列表与含 ``url`` 的对象列表。

    功能说明:
        ``social_urls`` 通常为 URL 字符串数组，而 ``summary_social_links`` 通常为
        ``{platform, title, url}`` 对象数组。身份规则只比较账号 URL，不能把对象的
        Python 文本表示当作 URL，否则会使整组 Social 证据失效。

    参数说明:
        value: 单个 URL、URL 数组、Social Link 对象或对象数组。

    返回值:
        list[str]: 可参与 URL 规范化的非空 URL 字符串；缺失 ``url`` 的对象忽略。
    """

    urls: list[str] = []
    for item in _value_items(value):
        candidate = item.get("url") if isinstance(item, dict) else item
        if isinstance(candidate, str) and candidate.strip():
            urls.append(candidate)
    return urls


def _stable_item(value: Any) -> str:
    """将集合比较项转换为稳定 JSON 文本，支持对象和标量。"""

    return json.dumps(
        sanitize_raw(value),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )


def _normalized_text(value: Any) -> str:
    """用于准确率比较的轻量文本规范化，不改变原始处理结果。"""

    return " ".join(str(value).strip().casefold().split())


def _semantic_text_lite_score(baseline_value: Any, returned_value: Any) -> float:
    """计算无需外部模型的可解释轻量文本相似度。

    规则依次采用 Unicode 规范化、完全匹配、包含关系和词元/双字符集合
    Jaccard。该分数只服务资料字段准确度，绝不替代 Social 或照片身份规则。
    """

    def normalized(value: Any) -> str:
        text = unicodedata.normalize("NFKC", str(value)).casefold()
        text = re.sub(r"[^\w\u4e00-\u9fff]+", " ", text)
        return " ".join(text.split())

    baseline_text = normalized(baseline_value)
    returned_text = normalized(returned_value)
    if not baseline_text or not returned_text:
        return 0.0
    if baseline_text == returned_text:
        return 1.0
    containment = 0.8 if (
        min(len(baseline_text), len(returned_text)) >= 4
        and (
            baseline_text in returned_text
            or returned_text in baseline_text
        )
    ) else 0.0
    if " " in baseline_text or " " in returned_text:
        baseline_units = set(baseline_text.split())
        returned_units = set(returned_text.split())
    else:
        baseline_units = {
            baseline_text[index:index + 2]
            for index in range(max(1, len(baseline_text) - 1))
        }
        returned_units = {
            returned_text[index:index + 2]
            for index in range(max(1, len(returned_text) - 1))
        }
    union = baseline_units | returned_units
    overlap = len(baseline_units & returned_units) / len(union) if union else 0.0
    return max(containment, overlap)


def _compare_field_values(
    baseline_value: Any,
    returned_value: Any,
    compare_mode: str,
) -> tuple[float | None, float | None, str]:
    """计算单字段建议完整度和准确率。

    返回值:
        ``(完整度, 准确率, 错误)``。基准缺失或 manual 模式返回 ``None``；
        返回为空时准确率分母为0，因此准确率保持 ``None``。
    """

    if _is_empty_value(baseline_value, "default"):
        return None, None, ""
    if compare_mode == "manual":
        return None, None, ""
    returned_empty = _is_empty_value(returned_value, "default")
    if returned_empty:
        return 0.0, None, ""
    try:
        if compare_mode == "presence":
            return 1.0, None, ""
        if compare_mode == "normalized_text":
            matched = _normalized_text(baseline_value) == _normalized_text(
                returned_value
            )
            score = 1.0 if matched else 0.0
            return score, score, ""
        if compare_mode == "exact":
            matched = sanitize_raw(baseline_value) == sanitize_raw(returned_value)
            score = 1.0 if matched else 0.0
            return score, score, ""
        if compare_mode in {"set", "url_set"}:
            baseline_items = _value_items(baseline_value)
            returned_items = _value_items(returned_value)
            if compare_mode == "url_set":
                baseline_items = [
                    _normalize_social_url(str(item)) for item in baseline_items
                ]
                returned_items = [
                    _normalize_social_url(str(item)) for item in returned_items
                ]
            baseline_set = {_stable_item(item) for item in baseline_items}
            returned_set = {_stable_item(item) for item in returned_items}
            intersection_count = len(baseline_set & returned_set)
            completeness = (
                intersection_count / len(baseline_set)
                if baseline_set
                else None
            )
            accuracy = (
                intersection_count / len(returned_set)
                if returned_set
                else None
            )
            return completeness, accuracy, ""
        if compare_mode == "semantic_text_lite":
            score = _semantic_text_lite_score(baseline_value, returned_value)
            return 1.0, score, ""
    except (FieldSchemaValidationError, TypeError, ValueError) as exc:
        return None, None, str(exc)
    return None, None, f"不支持的 compare_mode: {compare_mode}"


def _social_platform(url: str) -> str:
    """从规范化 Social URL 提取平台，合并少量已知等价域名。"""

    hostname = (urlsplit(url).hostname or "").lower()
    if hostname.startswith("www."):
        hostname = hostname[4:]
    aliases = {
        "x.com": "twitter",
        "twitter.com": "twitter",
        "facebook.com": "facebook",
        "fb.com": "facebook",
        "instagram.com": "instagram",
        "linkedin.com": "linkedin",
        "tiktok.com": "tiktok",
        "youtube.com": "youtube",
        "youtu.be": "youtube",
    }
    return aliases.get(hostname, hostname)


def _candidate_identity_rule(
    baseline_urls: Any,
    returned_urls: Any,
    photo_identity_match_rate: Any,
) -> tuple[str, str, str]:
    """按既定优先级自动判定候选人是否命中 Baseline Person。

    功能说明:
        先执行 Social Link 强绑定规则，再使用照片相似度作为兜底。Social 同平台
        链接冲突优先级最高，不能被名称、照片或其他普通字段覆盖。

    参数说明:
        baseline_urls: 基准候选人的 social_urls 与 summary_social_links 合集。
        returned_urls: 当前检索候选人的 social_urls 与 summary_social_links 合集。
        photo_identity_match_rate: 已规范化为 0–100 的照片身份相似度，可为空。

    返回值:
        tuple[str, str, str]: judgement、reason 与可回查的 JSON 证据。
        judgement 仅为 HIT、NOT_HIT 或 SUSPECTED，调用方可将其保存为 RULE 终判。

    异常说明:
        Social URL 格式非法时不抛出异常；该规则视为不可比较并继续执行照片规则，
        以免单个扩展字段格式影响整个 Process。
    """

    try:
        baseline = {
            _normalize_social_url(item) for item in _social_url_items(baseline_urls)
        }
        returned = {
            _normalize_social_url(item) for item in _social_url_items(returned_urls)
        }
    except FieldSchemaValidationError as exc:
        baseline = set()
        returned = set()
        social_error = str(exc)
    else:
        social_error = ""
    matched = sorted(baseline & returned)
    conflicts = sorted(
        returned_url
        for returned_url in returned
        if returned_url not in baseline
        and any(
            _social_platform(returned_url) == _social_platform(baseline_url)
            for baseline_url in baseline
        )
    )
    try:
        photo_rate = float(photo_identity_match_rate)
        if not math.isfinite(photo_rate):
            photo_rate = None
        elif 0 <= photo_rate <= 1:
            photo_rate *= 100
        elif not 0 <= photo_rate <= 100:
            photo_rate = None
    except (TypeError, ValueError):
        photo_rate = None
    evidence = json_text(
        {
            "matched_social_urls": matched,
            "conflicting_social_urls": conflicts,
            "photo_identity_match_rate": photo_rate,
            "social_rule_error": social_error or None,
        }
    )
    # Social 冲突优先于所有后续证据，包括同一候选人上的其他匹配链接和照片相似度。
    if conflicts:
        return "NOT_HIT", "SOCIAL_CONFLICT", evidence
    if matched:
        return "HIT", "SOCIAL_MATCH", evidence
    if photo_rate is not None and photo_rate >= 80:
        return "HIT", "PHOTO_MATCH", evidence
    if not returned and photo_rate is None:
        return "SUSPECTED", "NO_STRONG_FIELD", evidence
    if photo_rate is not None and photo_rate < 80:
        return "NOT_HIT", "PHOTO_BELOW_THRESHOLD", evidence
    return "NOT_HIT", "NO_STRONG_FIELD", evidence


def _suggested_field_scores(
    definitions: list[dict[str, Any]],
    fields: dict[str, Any],
    empty_fields: dict[str, bool],
    baseline_fields: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    """生成可被人工覆盖的字段完整度和准确率建议。"""

    scores: dict[str, dict[str, Any]] = {}
    for definition in definitions:
        roles = definition["scoring_role"]
        if not definition["enabled"] or not (
            {"completeness", "accuracy"} & set(roles)
        ):
            continue
        field_key = definition["field_key"]
        baseline_available = (
            field_key in baseline_fields
            and not _is_empty_value(baseline_fields.get(field_key), "default")
        )
        returned_value = fields.get(field_key)
        returned_nonempty = not empty_fields.get(
            field_key,
            _is_empty_value(returned_value, "default"),
        )
        completeness, accuracy, comparison_error = _compare_field_values(
            baseline_fields.get(field_key),
            returned_value,
            definition["compare_mode"],
        )
        if "completeness" not in roles:
            completeness = None
        if "accuracy" not in roles:
            accuracy = None
        scores[field_key] = {
            "compare_mode": definition["compare_mode"],
            "baseline_available": baseline_available,
            "returned_nonempty": returned_nonempty,
            "suggested_completeness_score": completeness,
            "suggested_accuracy_score": accuracy,
            "completeness_score": completeness,
            "accuracy_score": accuracy,
            "manual_override": False,
            "review_note": "",
            "comparison_error": comparison_error,
        }
    return scores


def _field_comparison_scores_v3(
    definitions: list[dict[str, Any]],
    fields: dict[str, Any],
    empty_fields: dict[str, bool],
    baseline_fields: dict[str, Any],
    baseline_available_fields: set[str],
) -> dict[str, dict[str, Any]]:
    """生成 field-processing-v5 的基准人物字段对比快照。

    功能说明：完整度与基准对比是独立开关。开启完整度的 Candidate 字段即使
    未开启 ``baseline_compare_enabled``，也要按自身是否返回真实值生成 1/0；
    只有开启基准对比的字段才使用 ``baseline_field_key`` 计算准确度和两侧
    比较结果。这样“已提取且启用完整度”的字段不会因未参与基准对比而没有
    评分。本函数纯计算，不访问网络、不修改 Raw。
    """

    scores: dict[str, dict[str, Any]] = {}
    for definition in definitions:
        if (
            not definition["enabled"]
            or definition["value_scope"] != "CANDIDATE"
            or not (
                definition["completeness_enabled"]
                or definition["baseline_compare_enabled"]
            )
        ):
            continue
        field_key = definition["field_key"]
        baseline_field_key = definition["baseline_field_key"]
        baseline_value = baseline_fields.get(baseline_field_key)
        baseline_available = baseline_field_key in baseline_available_fields
        returned_value = fields.get(field_key)
        returned_nonempty = not empty_fields.get(
            field_key, _is_empty_value(returned_value, "default")
        )
        base_item = {
            "baseline_field_key": baseline_field_key,
            "baseline_available": baseline_available,
            "baseline_value": sanitize_raw(baseline_value),
            "returned_value": sanitize_raw(returned_value),
            "returned_nonempty": returned_nonempty,
            "compare_mode": definition["compare_mode"],
            "similarity_threshold": definition["similarity_threshold"],
            "completeness_enabled": definition["completeness_enabled"],
            "accuracy_enabled": definition["accuracy_enabled"],
            "suggested_completeness_score": None,
            "suggested_accuracy_score": None,
            "completeness_score": None,
            "accuracy_score": None,
            # 资料完整度只回答“有没有返回”；基准覆盖度保留原完整度算法，
            # 用于 URL/集合交集和基准字段覆盖分析，二者不能混用。
            "data_completeness_score": None,
            "baseline_coverage_score": None,
            "manual_override": False,
            "review_note": "",
            "comparison_status": "NOT_APPLICABLE",
            "reason_code": "",
            "comparison_error": "",
        }
        if not definition["baseline_compare_enabled"]:
            # 完整度只回答“候选人是否返回该字段”，不依赖 Baseline 是否配置。
            # 准确度需要两侧值，因此在未启用基准对比时明确标记为不适用。
            completeness = 1.0 if returned_nonempty else 0.0
            if not definition["completeness_enabled"]:
                completeness = None
            base_item.update(
                suggested_completeness_score=completeness,
                completeness_score=completeness,
                data_completeness_score=completeness,
                comparison_status="NOT_APPLICABLE",
                reason_code="BASELINE_COMPARISON_DISABLED",
            )
            scores[field_key] = base_item
            continue
        if not baseline_available:
            base_item.update(
                data_completeness_score=(
                    1.0 if returned_nonempty and definition["completeness_enabled"]
                    else 0.0 if definition["completeness_enabled"] else None
                ),
                comparison_status="NOT_APPLICABLE",
                reason_code="BASELINE_FIELD_UNAVAILABLE",
            )
            scores[field_key] = base_item
            continue
        if _is_empty_value(baseline_value, "default"):
            base_item.update(
                data_completeness_score=(
                    1.0 if returned_nonempty and definition["completeness_enabled"]
                    else 0.0 if definition["completeness_enabled"] else None
                ),
                comparison_status="NOT_APPLICABLE",
                reason_code="BASELINE_VALUE_EMPTY",
            )
            scores[field_key] = base_item
            continue
        completeness, accuracy, comparison_error = _compare_field_values(
            baseline_value, returned_value, definition["compare_mode"]
        )
        if definition["compare_mode"] == "manual":
            completeness = 1.0 if returned_nonempty else 0.0
            accuracy = None
        if not definition["completeness_enabled"]:
            completeness = None
        if not definition["accuracy_enabled"]:
            accuracy = None
        reason_code = "READY"
        status = "READY"
        if comparison_error:
            reason_code = "FIELD_COMPARISON_ERROR"
            status = "ERROR"
        elif not returned_nonempty:
            reason_code = "CANDIDATE_VALUE_EMPTY"
            status = "READY"
        elif definition["compare_mode"] == "manual":
            reason_code = "MANUAL_ACCURACY"
        elif definition["compare_mode"] == "semantic_text_lite":
            if accuracy is not None and accuracy < definition["similarity_threshold"]:
                reason_code = "BELOW_SIMILARITY_THRESHOLD"
                status = "WARNING"
            else:
                reason_code = "SEMANTIC_SIMILARITY"
        base_item.update(
            suggested_completeness_score=completeness,
            suggested_accuracy_score=accuracy,
            completeness_score=completeness,
            accuracy_score=accuracy,
            data_completeness_score=(
                1.0 if returned_nonempty and definition["completeness_enabled"]
                else 0.0 if definition["completeness_enabled"] else None
            ),
            baseline_coverage_score=completeness,
            comparison_status=status,
            reason_code=reason_code,
            comparison_error=comparison_error,
        )
        scores[field_key] = base_item
    return scores


class AnalysisService:
    """编排导入、Web 执行、字段处理、复核和指标计算。"""

    def __init__(
        self,
        store: AnalysisStore,
        data_dir: Path | str,
        *,
        import_dir: Path | str | None = None,
        raw_dir: Path | str | None = None,
        report_dir: Path | str | None = None,
    ) -> None:
        """绑定已初始化 Store 与受控数据目录。

        参数说明:
            store: 已初始化的 SQLite Store。
            data_dir: 所有数据库文件引用的共同根目录。
            import_dir: 可选上传归档目录，必须位于 data_dir 内。
            raw_dir: 可选标准化结果目录，必须位于 data_dir 内。
            report_dir: 可选报告目录；Web 可将其配置到 output/reports。

        异常说明:
            ValueError: import_dir 或 raw_dir 越出 data_dir。
        """

        self.store = store
        self.data_dir = Path(data_dir).resolve()
        self.import_dir = (
            Path(import_dir).resolve()
            if import_dir is not None
            else self.data_dir / "imports"
        )
        self.raw_dir = (
            Path(raw_dir).resolve()
            if raw_dir is not None
            else self.data_dir / "raw"
        )
        self.report_dir = (
            Path(report_dir).resolve()
            if report_dir is not None
            else self.data_dir / "reports"
        )
        for label, directory in (
            ("import_dir", self.import_dir),
            ("raw_dir", self.raw_dir),
        ):
            try:
                directory.relative_to(self.data_dir)
            except ValueError as exc:
                raise ValueError(f"{label} 必须位于 SEARCH_DATA_DIR 内") from exc
        # 单进程 Web 中串行保护“检查活动 Run → 创建 Run”，避免并发 POST 穿透。
        self._execution_lock = threading.Lock()

    def update_evaluation_thresholds(
        self,
        evaluation_id: str,
        thresholds: Any,
    ) -> dict[str, dict[str, float | None]]:
        """校验并保存 Evaluation 参考线。

        参数说明:
            evaluation_id: 已存在的 Evaluation 标识。
            thresholds: 分 Query Stage 的参考线对象，允许部分字段为空。

        返回值:
            已补齐并持久化的稳定阈值对象。

        异常说明:
            ReviewValidationError: Evaluation 不存在或阈值结构/范围非法。
        """

        normalized = normalize_evaluation_thresholds(thresholds)
        now = utc_now_text()
        with self.store.transaction() as connection:
            cursor = connection.execute(
                """
                UPDATE evaluations
                SET thresholds_json = ?, threshold_profile_id = NULL,
                    updated_at = ?
                WHERE evaluation_id = ?
                """,
                (json_text(normalized), now, evaluation_id),
            )
            if cursor.rowcount != 1:
                raise ReviewValidationError(
                    f"Evaluation 不存在: {evaluation_id}"
                )
        return normalized

    def create_threshold_profile(
        self,
        *,
        profile_id: str,
        name: str,
        description: str,
        version: int,
        thresholds: Any,
        based_on_profile_id: str | None = None,
    ) -> dict[str, Any]:
        """创建不可变参考线方案版本，拒绝覆盖已有标识或名称版本。"""

        try:
            validate_storage_id(profile_id, "profile_id")
            if based_on_profile_id:
                validate_storage_id(
                    based_on_profile_id,
                    "based_on_profile_id",
                )
        except ValueError as exc:
            raise ReviewValidationError(str(exc)) from exc
        normalized_name = name.strip()
        if not normalized_name:
            raise ReviewValidationError("参考线方案名称不能为空")
        if isinstance(version, bool) or not isinstance(version, int):
            raise ReviewValidationError("参考线方案版本必须是正整数")
        if version < 1:
            raise ReviewValidationError("参考线方案版本必须是正整数")
        normalized = normalize_evaluation_thresholds(thresholds)
        now = utc_now_text()
        try:
            with self.store.transaction() as connection:
                if based_on_profile_id:
                    source = connection.execute(
                        """
                        SELECT profile_id FROM threshold_profiles
                        WHERE profile_id = ?
                        """,
                        (based_on_profile_id,),
                    ).fetchone()
                    if source is None:
                        raise ReviewValidationError(
                            f"来源参考线方案不存在: {based_on_profile_id}"
                        )
                connection.execute(
                    """
                    INSERT INTO threshold_profiles(
                        profile_id, name, description, version,
                        thresholds_json, based_on_profile_id, status,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, 'ACTIVE', ?)
                    """,
                    (
                        profile_id,
                        normalized_name,
                        description.strip(),
                        version,
                        json_text(normalized),
                        based_on_profile_id,
                        now,
                    ),
                )
        except sqlite3.IntegrityError as exc:
            raise ReviewValidationError(
                "profile_id 或方案名称 + 版本已存在，不能覆盖旧版本"
            ) from exc
        return {
            "profile_id": profile_id,
            "name": normalized_name,
            "description": description.strip(),
            "version": version,
            "thresholds": normalized,
            "based_on_profile_id": based_on_profile_id,
            "status": "ACTIVE",
            "created_at": now,
        }

    def archive_threshold_profile(self, profile_id: str) -> None:
        """归档一个方案版本；保留记录和历史 Evaluation 关联。"""

        with self.store.transaction() as connection:
            cursor = connection.execute(
                """
                UPDATE threshold_profiles SET status = 'ARCHIVED'
                WHERE profile_id = ?
                """,
                (profile_id,),
            )
            if cursor.rowcount != 1:
                raise ReviewValidationError(
                    f"参考线方案不存在: {profile_id}"
                )

    def assign_evaluation_threshold_profile(
        self,
        evaluation_id: str,
        profile_id: str | None,
    ) -> dict[str, dict[str, float | None]]:
        """原子复制 ACTIVE 方案快照到 Evaluation；空值表示取消方案。"""

        now = utc_now_text()
        with self.store.transaction() as connection:
            normalized = normalize_evaluation_thresholds({})
            normalized_profile_id = profile_id or None
            if normalized_profile_id:
                profile = connection.execute(
                    """
                    SELECT thresholds_json FROM threshold_profiles
                    WHERE profile_id = ? AND status = 'ACTIVE'
                    """,
                    (normalized_profile_id,),
                ).fetchone()
                if profile is None:
                    raise ReviewValidationError(
                        "参考线方案不存在、已归档或不可选择"
                    )
                try:
                    normalized = normalize_evaluation_thresholds(
                        json.loads(profile["thresholds_json"])
                    )
                except (TypeError, json.JSONDecodeError) as exc:
                    raise ReviewValidationError(
                        "参考线方案快照 JSON 已损坏"
                    ) from exc
            cursor = connection.execute(
                """
                UPDATE evaluations
                SET threshold_profile_id = ?, thresholds_json = ?,
                    updated_at = ?
                WHERE evaluation_id = ?
                """,
                (
                    normalized_profile_id,
                    json_text(normalized),
                    now,
                    evaluation_id,
                ),
            )
            if cursor.rowcount != 1:
                raise ReviewValidationError(
                    f"Evaluation 不存在: {evaluation_id}"
                )
        return normalized

    def create_evaluation(
        self,
        *,
        evaluation_id: str,
        name: str,
        notes: str,
        threshold_profile_id: str | None = None,
    ) -> None:
        """创建 Evaluation，并在同一事务复制所选 ACTIVE 方案快照。"""

        try:
            validate_storage_id(evaluation_id, "evaluation_id")
        except ValueError as exc:
            raise ReviewValidationError(str(exc)) from exc
        if not name.strip():
            raise ReviewValidationError("评测名称不能为空")
        normalized = normalize_evaluation_thresholds({})
        normalized_profile_id = threshold_profile_id or None
        now = utc_now_text()
        with self.store.transaction() as connection:
            if normalized_profile_id:
                profile = connection.execute(
                    """
                    SELECT thresholds_json FROM threshold_profiles
                    WHERE profile_id = ? AND status = 'ACTIVE'
                    """,
                    (normalized_profile_id,),
                ).fetchone()
                if profile is None:
                    raise ReviewValidationError(
                        "参考线方案不存在、已归档或不可选择"
                    )
                try:
                    normalized = normalize_evaluation_thresholds(
                        json.loads(profile["thresholds_json"])
                    )
                except (TypeError, json.JSONDecodeError) as exc:
                    raise ReviewValidationError(
                        "参考线方案快照 JSON 已损坏"
                    ) from exc
            try:
                connection.execute(
                    """
                    INSERT INTO evaluations(
                        evaluation_id, name, notes, thresholds_json,
                        threshold_profile_id, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        evaluation_id,
                        name.strip(),
                        notes.strip(),
                        json_text(normalized),
                        normalized_profile_id,
                        now,
                        now,
                    ),
                )
            except sqlite3.IntegrityError as exc:
                raise ReviewValidationError(
                    f"Evaluation 已存在: {evaluation_id}"
                ) from exc

    def _source_path(self, source: Path | str, suffixes: set[str]) -> Path:
        """解析并校验只读导入文件。"""

        path = Path(source).resolve()
        if not path.is_file():
            raise ImportValidationError([f"导入文件不存在: {path}"])
        if path.suffix.lower() not in suffixes:
            raise ImportValidationError(
                [f"不支持的文件类型: {path.suffix or '无后缀'}"]
            )
        return path

    def _archive_sources(
        self,
        object_id: str,
        sources: list[tuple[str, Path]],
    ) -> tuple[Path, list[str]]:
        """复制原文件到唯一归档目录，不覆盖任何既有归档。"""

        archive_dir = self.import_dir / object_id
        if archive_dir.exists():
            raise DuplicateImportError(f"归档目录已存在: {archive_dir}")
        archive_dir.mkdir(parents=True)
        archived_files: list[str] = []
        for role, source in sources:
            # 角色前缀避免同名文件冲突，同时保留原始文件名便于审计。
            target = archive_dir / f"{role}__{source.name}"
            shutil.copy2(source, target)
            archived_files.append(target.relative_to(self.data_dir).as_posix())
        return archive_dir, archived_files

    def _cleanup_created(self, *directories: Path) -> None:
        """仅清理由当前失败导入创建、且位于 data_dir 内的目录。"""

        for directory in directories:
            try:
                directory.resolve().relative_to(self.data_dir)
            except ValueError:
                continue
            if directory.exists():
                shutil.rmtree(directory)

    def _write_jsonl(self, path: Path, records: Iterable[dict[str, Any]]) -> None:
        """写入规范化脱敏 JSONL，供恢复和后续导入使用。"""

        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as handle:
            for record in records:
                handle.write(json_text(record))
                handle.write("\n")

    def _append_jsonl(self, path: Path, record: dict[str, Any]) -> None:
        """追加一条脱敏 JSONL，供后台执行逐 Query 持久化。"""

        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json_text(record))
            handle.write("\n")

    def ensure_default_field_schema(self) -> str:
        """幂等创建 v2 默认配置，仅替换系统自带的 v1 活跃配置。

        用户发布的活跃配置始终保留；只有当前仍使用系统默认 v1 时，新安装
        才把默认选择升级到 v2，旧配置记录不会删除。
        """

        existing = self.store.fetch_one(
            """
            SELECT schema_version FROM field_schemas
            WHERE schema_version = ?
            """,
            (DEFAULT_FIELD_SCHEMA_VERSION,),
        )
        if existing is not None:
            return existing["schema_version"]
        active = self.store.fetch_one(
            """
            SELECT schema_version, created_by FROM field_schemas
            WHERE is_active = 1 LIMIT 1
            """
        )
        activate_default = active is None or (
            active["schema_version"] == "field-schema-default-v1"
            and active["created_by"] == "system"
        )
        definitions = validate_field_definitions(DEFAULT_FIELD_DEFINITIONS)
        now = utc_now_text()
        try:
            with self.store.transaction() as connection:
                if activate_default and active is not None:
                    connection.execute(
                        "UPDATE field_schemas SET is_active = 0 WHERE is_active = 1"
                    )
                connection.execute(
                    """
                    INSERT INTO field_schemas(
                        schema_version, name, definitions_json, created_by,
                        created_at, is_active
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        DEFAULT_FIELD_SCHEMA_VERSION,
                        "默认字段配置",
                        json_text(definitions),
                        "system",
                        now,
                        int(activate_default),
                    ),
                )
        except sqlite3.IntegrityError:
            # 多线程同时初始化时，唯一主键保证只保留一份默认配置。
            existing = self.store.fetch_one(
                """
                SELECT schema_version FROM field_schemas
                WHERE schema_version = ?
                """,
                (DEFAULT_FIELD_SCHEMA_VERSION,),
            )
            if existing is None:
                raise
        return DEFAULT_FIELD_SCHEMA_VERSION

    def ensure_default_field_schema_v3(self) -> str:
        """幂等创建 FieldSchema v3 默认字段目录。

        功能说明：v3 作为新的不可变版本写入，不修改 v2 定义。仅当当前
        活跃配置仍是系统自带 v1/v2 时才切换到 v3；用户已发布的配置始终
        保持当前状态，避免升级阶段意外改变其后续处理选择。

        返回值：已存在或新建的 ``field-schema-default-v3`` 版本标识。
        异常说明：数据库写入失败时向上抛出，调用方不会得到半成品版本。
        """

        existing = self.store.fetch_one(
            "SELECT schema_version FROM field_schemas WHERE schema_version = ?",
            (DEFAULT_FIELD_SCHEMA_V3_VERSION,),
        )
        if existing is not None:
            return existing["schema_version"]
        active = self.store.fetch_one(
            """
            SELECT schema_version, created_by FROM field_schemas
            WHERE is_active = 1 LIMIT 1
            """
        )
        activate_default = active is None or (
            active["schema_version"] in {
                "field-schema-default-v1", DEFAULT_FIELD_SCHEMA_VERSION,
            }
            and active["created_by"] == "system"
        )
        definitions = validate_field_definitions(DEFAULT_FIELD_DEFINITIONS_V3)
        now = utc_now_text()
        try:
            with self.store.transaction() as connection:
                if activate_default and active is not None:
                    connection.execute(
                        "UPDATE field_schemas SET is_active = 0 WHERE is_active = 1"
                    )
                connection.execute(
                    """
                    INSERT INTO field_schemas(
                        schema_version, name, definitions_json, created_by,
                        created_at, is_active
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        DEFAULT_FIELD_SCHEMA_V3_VERSION,
                        "默认字段配置 v3",
                        json_text(definitions),
                        "system",
                        now,
                        int(activate_default),
                    ),
                )
        except sqlite3.IntegrityError:
            # 并发初始化时由主键保证仅保留一份 v3 快照。
            existing = self.store.fetch_one(
                "SELECT schema_version FROM field_schemas WHERE schema_version = ?",
                (DEFAULT_FIELD_SCHEMA_V3_VERSION,),
            )
            if existing is None:
                raise
        return DEFAULT_FIELD_SCHEMA_V3_VERSION

    def publish_field_schema(
        self,
        *,
        name: str,
        definitions: Any,
        created_by: str = "",
        schema_version: str | None = None,
        activate: bool = True,
    ) -> str:
        """校验并发布不可变字段配置新版本。

        参数说明:
            name: 配置版本的可读名称。
            definitions: 完整字段定义数组，不接受增量补丁。
            created_by: 单用户模式下可选的维护人文本。
            schema_version: 测试或迁移可指定；默认按时间和随机后缀生成。
            activate: 是否切换为新处理任务的默认版本。

        返回值:
            新生成的 schema_version。

        异常说明:
            FieldSchemaValidationError: 名称或任一字段配置不合法。
            sqlite3.IntegrityError: 指定版本已经存在；旧版本绝不更新。
        """

        if not isinstance(name, str) or not name.strip():
            raise FieldSchemaValidationError("字段配置名称不能为空")
        normalized = validate_field_definitions(definitions)
        generated_version = (
            schema_version
            or "field-schema-"
            + datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S-")
            + uuid.uuid4().hex[:6]
        )
        try:
            validate_storage_id(generated_version, "schema_version")
        except ValueError as exc:
            raise FieldSchemaValidationError(str(exc)) from exc
        now = utc_now_text()
        with self.store.transaction() as connection:
            if activate:
                connection.execute("UPDATE field_schemas SET is_active = 0")
            connection.execute(
                """
                INSERT INTO field_schemas(
                    schema_version, name, definitions_json, created_by,
                    created_at, is_active
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    generated_version,
                    name.strip(),
                    json_text(normalized),
                    created_by.strip() if isinstance(created_by, str) else "",
                    now,
                    int(activate),
                ),
            )
        return generated_version

    def discover_field_candidates(
        self,
        *,
        schema_version: str,
        process_id: str | None = None,
        baseline_version: str | None = None,
    ) -> dict[str, Any]:
        """从已入库数据发现尚未登记的 ui_sections 子字段建议。

        功能说明：仅读取指定 Process 的 Candidate Detail 与可选 Baseline，
        生成可由测试人员复制到新 Schema 的建议；不写数据库、不修改 Raw、
        不自动发布字段。Profile 建议使用 ``PROFILE_ITEM`` 受控选择器。

        返回值：包含候选建议、来源计数和当前 Schema 已登记字段集合的摘要。
        异常说明：Schema 或 Process 不存在时抛出 FieldSchemaValidationError。
        """

        schema = self.store.fetch_one(
            "SELECT definitions_json FROM field_schemas WHERE schema_version = ?",
            (schema_version,),
        )
        if schema is None:
            raise FieldSchemaValidationError(f"字段配置不存在: {schema_version}")
        try:
            definitions = validate_field_definitions(
                json.loads(schema["definitions_json"])
            )
        except (TypeError, json.JSONDecodeError) as exc:
            raise FieldSchemaValidationError("字段配置快照 JSON 已损坏") from exc
        known_paths = {
            item["source_path"] for item in definitions
            if item["source_type"] == "PATH"
        }
        known_profile = {
            (
                _profile_selector_text(item["source_options"].get("section", "")),
                _profile_selector_text(item["source_options"].get("label", "")),
            )
            for item in definitions
            if item["source_type"] == "PROFILE_ITEM"
        }
        known_keys = {item["field_key"] for item in definitions}
        suggestions: dict[tuple[str, str], dict[str, Any]] = {}

        def decode_json(value: Any) -> Any:
            """读取历史 JSON 文本；损坏数据仅跳过发现，不影响页面。"""

            try:
                return json.loads(value) if isinstance(value, str) else value
            except (TypeError, json.JSONDecodeError):
                return None

        def field_key_for(module: str, name: str) -> str:
            base = re.sub(
                r"[^a-z0-9]+", "_", f"{module}_{name}".casefold()
            ).strip("_") or "discovered_field"
            candidate = base[:110]
            suffix = 2
            while candidate in known_keys or any(
                item["field_key"] == candidate
                for item in suggestions.values()
            ):
                candidate = f"{base[:100]}_{suffix}"
                suffix += 1
            return candidate

        def inferred_type(value: Any) -> str:
            """把发现样例映射为安全的首版字段类型。"""

            if isinstance(value, bool):
                return "boolean"
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return "number"
            if isinstance(value, list):
                return "array"
            if isinstance(value, dict):
                return "object"
            return "string"

        def add_path(
            module: str,
            path: str,
            display_name: str,
            sample: Any = None,
        ) -> None:
            if path in known_paths:
                return
            key = ("PATH", path)
            if key not in suggestions:
                suggestions[key] = {
                    "field_key": field_key_for(module, display_name),
                    "display_name": display_name,
                    "module": module,
                    "source_type": "PATH",
                    "source_path": path,
                    "source_options": {},
                    "publishable": True,
                    "source_count": 0,
                    "sample": sanitize_raw(sample),
                    "data_type": inferred_type(sample),
                }
            suggestions[key]["source_count"] += 1

        def add_profile(section: Any, label: Any, sample: Any = None) -> None:
            if not isinstance(section, str) or not section.strip():
                return
            if not isinstance(label, str) or not label.strip():
                return
            normalized = (
                _profile_selector_text(section), _profile_selector_text(label)
            )
            if normalized in known_profile:
                return
            key = ("PROFILE_ITEM", f"{normalized[0]}::{normalized[1]}")
            if key not in suggestions:
                suggestions[key] = {
                    "field_key": field_key_for("profile", f"{section}_{label}"),
                    "display_name": f"{section.strip()} / {label.strip()}",
                    "module": "Profile",
                    "source_type": "PROFILE_ITEM",
                    "source_path": "ui_sections.profile.data.sections",
                    "source_options": {
                        "section": section.strip(), "label": label.strip(),
                    },
                    "publishable": True,
                    "source_count": 0,
                    "sample": sanitize_raw(sample),
                    "data_type": inferred_type(sample),
                }
            suggestions[key]["source_count"] += 1

        candidate_count = 0
        if process_id:
            process = self.store.fetch_one(
                "SELECT process_id FROM process_runs WHERE process_id = ?",
                (process_id,),
            )
            if process is None:
                raise FieldSchemaValidationError(f"处理结果不存在: {process_id}")
            rows = self.store.fetch_all(
                """
                SELECT c.ui_sections_json
                FROM processed_candidates AS pc
                JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
                WHERE pc.process_id = ? AND c.detail_status = 'SUCCESS'
                """,
                (process_id,),
            )
            for row in rows:
                sections = decode_json(row["ui_sections_json"])
                if not isinstance(sections, dict):
                    continue
                candidate_count += 1
                for module_key, module in sections.items():
                    module_name = module_key.capitalize()
                    if module_name not in {
                        "Insights", "Photos", "Profile", "Social", "Summary"
                    } or not isinstance(module, dict):
                        continue
                    data = module.get("data")
                    if isinstance(data, dict):
                        for data_key, data_value in data.items():
                            add_path(
                                module_name,
                                f"ui_sections.{module_key}.data.{data_key}",
                                f"{module_name} {data_key}",
                                data_value,
                            )
                    if module_key == "insights" and isinstance(data, dict):
                        for item in data.get("items", []):
                            if isinstance(item, dict):
                                for item_key, item_value in item.items():
                                    add_path(
                                        "Insights",
                                        f"ui_sections.insights.data.items[*].{item_key}",
                                        f"Insights Item {item_key}",
                                        item_value,
                                    )
                    if module_key == "profile" and isinstance(data, dict):
                        for section in data.get("sections", []):
                            if isinstance(section, dict):
                                for item in section.get("items", []):
                                    if isinstance(item, dict):
                                        add_profile(
                                            section.get("title"),
                                            item.get("label"),
                                            item.get("value"),
                                        )

        baseline_unknown: list[str] = []
        if baseline_version:
            if self.store.fetch_one(
                "SELECT baseline_version FROM baseline_sets WHERE baseline_version = ?",
                (baseline_version,),
            ) is None:
                raise FieldSchemaValidationError(f"基准版本不存在: {baseline_version}")
            for row in self.store.fetch_all(
                "SELECT fields_json FROM baseline_people WHERE baseline_version = ?",
                (baseline_version,),
            ):
                values = decode_json(row["fields_json"])
                if isinstance(values, dict):
                    baseline_unknown.extend(
                        key for key in values if key not in known_keys
                    )
        discovered = sorted(
            suggestions.values(),
            key=lambda item: (
                item["module"], item["display_name"], item["field_key"]
            ),
        )
        next_sort_order = 9000
        for item in discovered:
            item["definition"] = _v3_catalog_field(
                item["field_key"],
                item["display_name"],
                item["module"],
                item["source_path"],
                item["data_type"],
                next_sort_order,
                source_type=item["source_type"],
                source_options=item["source_options"],
                array_mode=(
                    "collect" if item["data_type"] == "array" else "preserve"
                ),
            )
            next_sort_order += 10
        return {
            "schema_version": schema_version,
            "process_id": process_id or "",
            "candidate_count": candidate_count,
            "suggestions": discovered,
            "baseline_only_field_keys": sorted(set(baseline_unknown)),
        }

    def build_field_comparison_matrix(
        self,
        schema_version: str,
        baseline_version: str,
        *,
        process_id: str | None = None,
        person_id: str | None = None,
    ) -> dict[str, Any]:
        """批量聚合 Baseline、FieldSchema 与 Process 的字段对比矩阵。

        功能说明:
            一次读取字段配置、Baseline 人物和可选 Process 快照，在内存中
            按 field_key 计算覆盖、开关、样例、结构和冲突。未知 Baseline
            字段也保留为矩阵行，以暴露“已准备但未提取”的问题。

        参数说明:
            schema_version: 不可变 FieldSchema 版本。
            baseline_version: Baseline Version。
            process_id: 可选 Process；提供后统计实际 Candidate/Query 返回率。
            person_id: 可选单人物过滤，用于人物级字段工作台。

        返回值:
            字段行、状态/严重级别计数和所用版本信息。

        异常说明:
            FieldSchemaValidationError: 版本不存在、不匹配或快照损坏。
        """

        schema = self.store.fetch_one(
            "SELECT * FROM field_schemas WHERE schema_version = ?",
            (schema_version,),
        )
        if schema is None:
            raise FieldSchemaValidationError(
                f"字段配置不存在: {schema_version}"
            )
        baseline = self.store.fetch_one(
            "SELECT * FROM baseline_sets WHERE baseline_version = ?",
            (baseline_version,),
        )
        if baseline is None:
            raise FieldSchemaValidationError(
                f"基准版本不存在: {baseline_version}"
            )
        try:
            definitions = validate_field_definitions(
                json.loads(schema["definitions_json"])
            )
        except (TypeError, json.JSONDecodeError) as exc:
            raise FieldSchemaValidationError(
                f"字段配置快照 JSON 已损坏: {schema_version}"
            ) from exc
        baseline_sql = """
            SELECT person_id, fields_json, available_fields_json
            FROM baseline_people WHERE baseline_version = ?
        """
        baseline_parameters: list[Any] = [baseline_version]
        if person_id:
            baseline_sql += " AND person_id = ?"
            baseline_parameters.append(person_id)
        baseline_rows = self.store.fetch_all(
            baseline_sql + " ORDER BY person_id",
            baseline_parameters,
        )
        if person_id and not baseline_rows:
            raise FieldSchemaValidationError(
                f"人物 {person_id} 不属于基准版本 {baseline_version}"
            )
        baseline_values: dict[str, list[Any]] = defaultdict(list)
        baseline_available: dict[str, int] = defaultdict(int)
        all_field_keys = [item["field_key"] for item in definitions]
        try:
            for row in baseline_rows:
                fields = json.loads(row["fields_json"] or "{}")
                available = json.loads(row["available_fields_json"] or "[]")
                if not isinstance(fields, dict) or not isinstance(available, list):
                    raise TypeError("Baseline 字段或可用字段结构无效")
                for field_key, value in fields.items():
                    baseline_values[field_key].append(value)
                    if field_key not in all_field_keys:
                        all_field_keys.append(field_key)
                for field_key in available:
                    baseline_available[str(field_key)] += 1
                    if field_key not in all_field_keys:
                        all_field_keys.append(str(field_key))
        except (TypeError, json.JSONDecodeError) as exc:
            raise FieldSchemaValidationError(
                f"Baseline 字段快照已损坏: {baseline_version}"
            ) from exc

        process = None
        candidate_rows: list[sqlite3.Row] = []
        query_rows: list[sqlite3.Row] = []
        if process_id:
            process = self.store.fetch_one(
                """
                SELECT process_id, schema_version, baseline_version, rule_version
                FROM process_runs WHERE process_id = ?
                """,
                (process_id,),
            )
            if process is None:
                raise FieldSchemaValidationError(
                    f"处理结果不存在: {process_id}"
                )
            if process["schema_version"] != schema_version:
                raise FieldSchemaValidationError(
                    f"Process {process_id} 使用 {process['schema_version']}，"
                    f"不能按 {schema_version} 聚合"
                )
            candidate_rows = self.store.fetch_all(
                """
                SELECT pc.fields_json, pc.empty_fields_json, c.detail_status
                FROM processed_candidates AS pc
                JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
                WHERE pc.process_id = ?
                """,
                (process_id,),
            )
            query_rows = self.store.fetch_all(
                """
                SELECT fields_json, empty_fields_json
                FROM processed_queries WHERE process_id = ?
                """,
                (process_id,),
            )
        process_values: dict[str, list[Any]] = defaultdict(list)
        process_nonempty: dict[str, int] = defaultdict(int)
        candidate_count = sum(
            row["detail_status"] == "SUCCESS" for row in candidate_rows
        )
        query_count = len(query_rows)
        try:
            for row in candidate_rows:
                if row["detail_status"] != "SUCCESS":
                    continue
                fields = json.loads(row["fields_json"] or "{}")
                empty = json.loads(row["empty_fields_json"] or "{}")
                for field_key, value in fields.items():
                    process_values[field_key].append(value)
                    if not empty.get(field_key, _is_empty_value(value, "default")):
                        process_nonempty[field_key] += 1
            for row in query_rows:
                fields = json.loads(row["fields_json"] or "{}")
                empty = json.loads(row["empty_fields_json"] or "{}")
                for field_key, value in fields.items():
                    process_values[field_key].append(value)
                    if not empty.get(field_key, _is_empty_value(value, "default")):
                        process_nonempty[field_key] += 1
        except (TypeError, json.JSONDecodeError) as exc:
            raise FieldSchemaValidationError(
                f"Process 字段快照已损坏: {process_id}"
            ) from exc

        definitions_by_key = {
            item["field_key"]: item for item in definitions
        }
        fields_result: list[dict[str, Any]] = []
        severity_counts = {"ERROR": 0, "WARNING": 0, "INFO": 0}
        status_counts: dict[str, int] = defaultdict(int)
        for field_key in all_field_keys:
            definition = definitions_by_key.get(field_key)
            roles = set(definition["scoring_role"]) if definition else set()
            enabled = bool(definition and definition["enabled"])
            baseline_field_key = (
                definition["baseline_field_key"]
                if definition and definition["baseline_field_key"]
                else field_key
            )
            baseline_nonempty_values = [
                value
                for value in baseline_values.get(baseline_field_key, [])
                if not _is_empty_value(value, "default")
            ]
            candidate_nonempty_values = [
                value
                for value in process_values.get(field_key, [])
                if not _is_empty_value(value, "default")
            ]
            baseline_shapes = {
                _field_value_shape(value)
                for value in baseline_nonempty_values
            }
            candidate_shapes = {
                _field_value_shape(value)
                for value in candidate_nonempty_values
            }
            issues: list[dict[str, str]] = []
            available_count = baseline_available.get(baseline_field_key, 0)
            url_object_path_risk = bool(
                definition
                and baseline_shapes == {"array_string"}
                and field_key
                in {"summary_social_links", "summary_web_links"}
                and not definition["source_path"].endswith("[*].url")
            )
            if available_count and not enabled:
                status = "BASELINE_ENABLED_NOT_EXTRACTED"
                issues.append({
                    "severity": "ERROR",
                    "code": "BASELINE_ENABLED_NOT_EXTRACTED",
                    "message": "Baseline 已启用，但 Candidate 未配置提取",
                })
            elif url_object_path_risk:
                status = "STRUCTURE_MISMATCH"
                issues.append({
                    "severity": "ERROR",
                    "code": "URL_OBJECT_ARRAY_NOT_EXTRACTED",
                    "message": (
                        "Baseline 是 URL 字符串数组，Candidate 路径必须先"
                        "使用 [*].url 提取对象中的 URL"
                    ),
                })
            elif (
                baseline_shapes
                and candidate_shapes
                and baseline_shapes.isdisjoint(candidate_shapes)
            ):
                status = "STRUCTURE_MISMATCH"
                issues.append({
                    "severity": "ERROR",
                    "code": "STRUCTURE_MISMATCH",
                    "message": (
                        "Baseline 与 Candidate 结构不一致："
                        f"{sorted(baseline_shapes)} / {sorted(candidate_shapes)}"
                    ),
                })
            elif available_count and "completeness" not in roles:
                status = "BASELINE_ENABLED_NOT_SCORED"
                issues.append({
                    "severity": "WARNING",
                    "code": "BASELINE_ENABLED_NOT_SCORED",
                    "message": "Baseline 已启用，但字段未参与完整度",
                })
            elif available_count and not baseline_nonempty_values:
                status = "NO_BASELINE_DATA"
                issues.append({
                    "severity": "WARNING",
                    "code": "NO_BASELINE_DATA",
                    "message": "字段已设为 Baseline 可用，但所有值均为空",
                })
            elif (
                process_id
                and enabled
                and not candidate_nonempty_values
            ):
                status = "NO_CANDIDATE_DATA"
                issues.append({
                    "severity": "INFO",
                    "code": "NO_CANDIDATE_DATA",
                    "message": "当前 Process 没有返回该字段",
                })
            elif "accuracy" in roles and definition["compare_mode"] == "manual":
                status = "MANUAL_ACCURACY"
                issues.append({
                    "severity": "WARNING",
                    "code": "MANUAL_ACCURACY",
                    "message": "准确率需要在候选人详细复核中人工评分",
                })
            elif {"completeness", "accuracy"}.issubset(roles):
                status = "COMPARABLE"
            elif "completeness" in roles:
                status = "COMPLETENESS_ONLY"
            else:
                status = "DISPLAY_ONLY"
            issue_codes = {issue["code"] for issue in issues}
            if (
                available_count
                and not baseline_nonempty_values
                and "NO_BASELINE_DATA" not in issue_codes
            ):
                issues.append({
                    "severity": "WARNING",
                    "code": "NO_BASELINE_DATA",
                    "message": "字段已设为 Baseline 可用，但所有值均为空",
                })
            if (
                process_id
                and enabled
                and not candidate_nonempty_values
                and "NO_CANDIDATE_DATA" not in issue_codes
            ):
                issues.append({
                    "severity": "INFO",
                    "code": "NO_CANDIDATE_DATA",
                    "message": "当前 Process 没有返回该字段",
                })
            if enabled and not (
                {"completeness", "accuracy", "identity"} & roles
            ):
                issues.append({
                    "severity": "INFO",
                    "code": "CANDIDATE_EXTRACTED_NOT_SCORED",
                    "message": "Candidate 已提取，但字段仅用于展示",
                })
            if definition and (
                (definition["module"] == "Task")
                != (definition["value_scope"] == "QUERY")
            ):
                issues.append({
                    "severity": "ERROR",
                    "code": "VALUE_SCOPE_MISMATCH",
                    "message": "字段 Module 与 value_scope 不一致",
                })
                status = "STRUCTURE_MISMATCH"
            entity_count = (
                query_count
                if definition and definition["value_scope"] == "QUERY"
                else candidate_count
            )
            for issue in issues:
                severity_counts[issue["severity"]] += 1
            status_counts[status] += 1
            fields_result.append({
                "field_key": field_key,
                "display_name": (
                    definition["display_name"] if definition else field_key
                ),
                "module": (
                    definition["module"] if definition else "未配置字段"
                ),
                "value_scope": (
                    definition["value_scope"] if definition else "CANDIDATE"
                ),
                "enabled": enabled,
                "display_enabled": bool(
                    definition and definition["display_enabled"]
                ),
                "baseline_compare_enabled": bool(
                    definition and definition["baseline_compare_enabled"]
                ),
                "run_compare_enabled": bool(
                    definition and definition["run_compare_enabled"]
                ),
                "baseline_field_key": baseline_field_key if definition else "",
                "baseline_nonempty_count": len(baseline_nonempty_values),
                "baseline_available_count": available_count,
                "baseline_person_count": len(baseline_rows),
                "candidate_nonempty_count": process_nonempty.get(field_key, 0),
                "candidate_count": entity_count,
                "candidate_return_rate": (
                    process_nonempty.get(field_key, 0) / entity_count
                    if entity_count
                    else None
                ),
                "completeness_enabled": "completeness" in roles,
                "accuracy_enabled": "accuracy" in roles,
                "identity_enabled": "identity" in roles,
                "similarity_threshold": (
                    definition["similarity_threshold"] if definition else 0.6
                ),
                "compare_mode": (
                    definition["compare_mode"] if definition else ""
                ),
                "normalizer": (
                    definition["normalizer"] if definition else ""
                ),
                "baseline_sample": _matrix_sample(
                    baseline_nonempty_values[0]
                    if baseline_nonempty_values else None
                ),
                "candidate_sample": _matrix_sample(
                    candidate_nonempty_values[0]
                    if candidate_nonempty_values else None
                ),
                "baseline_shapes": sorted(baseline_shapes),
                "candidate_shapes": sorted(candidate_shapes),
                "status": status,
                "issues": issues,
            })
        fields_result.sort(
            key=lambda item: (
                definitions_by_key.get(item["field_key"], {}).get(
                    "sort_order",
                    999999,
                ),
                item["field_key"],
            )
        )
        return {
            "schema_version": schema_version,
            "schema_name": schema["name"],
            "baseline_version": baseline_version,
            "baseline_name": baseline["name"],
            "process_id": process_id,
            "person_id": person_id,
            "baseline_person_count": len(baseline_rows),
            "candidate_count": candidate_count,
            "query_count": query_count,
            "fields": fields_result,
            "status_counts": dict(status_counts),
            "severity_counts": severity_counts,
        }

    def validate_process_field_alignment(
        self,
        schema_version: str,
        baseline_version: str,
        *,
        process_id: str | None = None,
    ) -> list[dict[str, Any]]:
        """返回处理前字段对齐问题，不创建 Process 或修改配置。"""

        matrix = self.build_field_comparison_matrix(
            schema_version,
            baseline_version,
            process_id=process_id,
        )
        return [
            {
                "field_key": field["field_key"],
                **issue,
                "status": field["status"],
            }
            for field in matrix["fields"]
            for issue in field["issues"]
        ]

    def get_field_comparison_matrix(
        self,
        schema_version: str,
        baseline_version: str,
        *,
        process_id: str | None = None,
        person_id: str | None = None,
    ) -> dict[str, Any]:
        """返回字段矩阵及 v1.3 MVP 首页使用的配置冲突摘要。"""

        matrix = self.build_field_comparison_matrix(
            schema_version,
            baseline_version,
            process_id=process_id,
            person_id=person_id,
        )
        baseline_available_keys = {
            field["field_key"]
            for field in matrix["fields"]
            if field["baseline_available_count"] > 0
        }
        completeness_keys = {
            field["field_key"]
            for field in matrix["fields"]
            if field["enabled"] and field["completeness_enabled"]
        }
        return {
            **matrix,
            "baseline_available_field_count": len(
                baseline_available_keys
            ),
            "completeness_field_count": len(completeness_keys),
            "baseline_only_conflict_count": len(
                baseline_available_keys - completeness_keys
            ),
        }

    def _processing_source(self, candidate: sqlite3.Row) -> dict[str, Any]:
        """组合 Candidate Detail 与 List 字段为 Candidate 处理源。"""

        detail_data: dict[str, Any] = {}
        if candidate["detail_data_json"]:
            parsed_detail = json.loads(candidate["detail_data_json"])
            if not isinstance(parsed_detail, dict):
                raise FieldSchemaValidationError(
                    "Candidate detail_data_json 必须是对象"
                )
            detail_data = parsed_detail
        ui_sections: dict[str, Any] = {}
        if candidate["ui_sections_json"]:
            parsed_sections = json.loads(candidate["ui_sections_json"])
            if not isinstance(parsed_sections, dict):
                raise FieldSchemaValidationError(
                    "Candidate ui_sections_json 必须是对象"
                )
            ui_sections = parsed_sections
        source = dict(detail_data)
        if ui_sections or not isinstance(source.get("ui_sections"), dict):
            source["ui_sections"] = ui_sections
        source["candidate"] = {
            "candidate_id": candidate["candidate_id"],
            "candidate_rank": candidate["candidate_rank"],
            "rank_score": candidate["rank_score"],
            "detail_status": candidate["detail_status"],
            "list_item": (
                json.loads(candidate["list_item_json"])
                if candidate["list_item_json"]
                else {}
            ),
        }
        return sanitize_raw(source)

    def _process_field_values(
        self,
        source: dict[str, Any],
        definitions: list[dict[str, Any]],
        value_scope: str,
    ) -> tuple[dict[str, Any], dict[str, bool], list[dict[str, Any]]]:
        """处理指定作用域的字段，空路径与真实结构错误分别记录。"""

        fields: dict[str, Any] = {}
        empty_fields: dict[str, bool] = {}
        errors: list[dict[str, Any]] = []
        for definition in definitions:
            if (
                not definition["enabled"]
                or definition["value_scope"] != value_scope
            ):
                continue
            field_key = definition["field_key"]
            source_path = definition["source_path"]
            try:
                duplicated_profile_item = False
                if definition["source_type"] == "PROFILE_ITEM":
                    raw_value, duplicated_profile_item = extract_profile_item(
                        source,
                        source_path,
                        definition["source_options"],
                        missing_policy=definition["missing_policy"],
                    )
                else:
                    raw_value = extract_source_path(
                        source,
                        source_path,
                        missing_policy=definition["missing_policy"],
                    )
                if raw_value is None:
                    # 数组型可选字段使用空数组，避免 collect 把空值变成 [null]。
                    array_value = (
                        []
                        if definition["data_type"] == "array"
                        else None
                    )
                else:
                    array_value = _apply_array_mode(
                        raw_value,
                        definition["array_mode"],
                    )
                normalized_value = (
                    None
                    if array_value is None
                    else normalize_field_value(
                        array_value,
                        definition["normalizer"],
                    )
                )
                normalized_value = _validate_field_data_type(
                    normalized_value,
                    definition["data_type"],
                )
                fields[field_key] = sanitize_raw(normalized_value)
                empty_fields[field_key] = _is_empty_value(
                    normalized_value,
                    definition["empty_rule"],
                )
                if duplicated_profile_item:
                    errors.append(
                        {
                            "code": "DUPLICATE_PROFILE_ITEM",
                            "field_key": field_key,
                            "source_path": source_path,
                            "error": "Profile 中存在多个相同 section / label，已保留全部值",
                        }
                    )
            except FieldSchemaValidationError as exc:
                errors.append(
                    {
                        "code": "FIELD_PROCESSING_ERROR",
                        "field_key": field_key,
                        "source_path": source_path,
                        "error": str(exc),
                    }
                )
        return fields, empty_fields, errors

    @staticmethod
    def _get_task_response(payload: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]] | None:
        """从 Raw 包装或直接响应中提取 GetTask 的标准响应体与 data。"""

        response_body = payload.get("response_body", payload)
        if not isinstance(response_body, dict):
            return None
        responses = response_body.get("responses")
        if not isinstance(responses, list) or not responses:
            return None
        response = responses[0]
        if not isinstance(response, dict) or response.get("success") is False:
            return None
        data = response.get("data")
        if not isinstance(data, dict):
            return None
        return response_body, data

    def _query_processing_source(
        self,
        query: sqlite3.Row,
    ) -> dict[str, Any]:
        """使用最后一条成功 GetTask Raw 构造 Query 字段处理源。

        缺少 Raw 时保留数据库中的 Task 基础值；Raw JSON 本身损坏时抛出
        结构错误，由调用方记录为 Query 处理错误而不阻断 Candidate。
        """

        selected_body: dict[str, Any] = {}
        selected_data: dict[str, Any] = {}
        fallback: tuple[dict[str, Any], dict[str, Any]] | None = None
        rows = self.store.fetch_all(
            """
            SELECT payload_json FROM raw_records
            WHERE run_id = ? AND query_id = ? AND stage = 'GetTask'
            ORDER BY sequence_no, collected_at, raw_id
            """,
            (query["run_id"], query["query_id"]),
        )
        for row in rows:
            payload = json.loads(row["payload_json"])
            if not isinstance(payload, dict):
                raise FieldSchemaValidationError(
                    "GetTask Raw payload_json 必须是对象"
                )
            parsed = self._get_task_response(payload)
            if parsed is None:
                continue
            fallback = parsed
            if parsed[1].get("status") == "SUCCEEDED":
                selected_body, selected_data = parsed
        if not selected_data and fallback is not None:
            selected_body, selected_data = fallback

        public_fields: dict[str, Any] = {}
        if query["public_fields_json"]:
            parsed_public = json.loads(query["public_fields_json"])
            if not isinstance(parsed_public, dict):
                raise FieldSchemaValidationError(
                    "run_queries.public_fields_json 必须是对象"
                )
            public_fields = parsed_public
        task = {
            "task_id": query["task_id"],
            "llm_cost": query["llm_cost"],
            "third_party_cost": query["third_party_cost"],
            "total_cost": query["total_cost"],
            "pdl_called": query["pdl_called"],
            "search_duration_ms": query["search_duration_ms"],
        }
        task.update(public_fields)
        task.update(selected_data)
        if not task.get("task_id"):
            task["task_id"] = query["task_id"]
        result_status = (
            query["result_status"]
            or normalize_result_status(
                query["status"],
                query["candidate_count_listed"],
            )
        )
        return sanitize_raw(
            {
                "query": {
                    "query_id": query["query_id"],
                    "person_id": query["person_id"],
                    "query_stage": query["query_stage"],
                    "task_id": query["task_id"],
                    "result_status": result_status,
                },
                "task": task,
                "raw": selected_body,
            }
        )

    def _process_query_fields(
        self,
        query: sqlite3.Row,
        definitions: list[dict[str, Any]],
    ) -> tuple[dict[str, Any], dict[str, bool], list[dict[str, Any]]]:
        """处理单条 Query 公共字段，Raw 结构错误不影响其他 Query。"""

        try:
            source = self._query_processing_source(query)
        except (
            FieldSchemaValidationError,
            json.JSONDecodeError,
            TypeError,
        ) as exc:
            return (
                {},
                {},
                [
                    {
                        "code": "INVALID_TASK_DATA",
                        "field_key": "",
                        "source_path": "",
                        "error": str(exc),
                    }
                ],
            )
        return self._process_field_values(source, definitions, "QUERY")

    def _process_candidate_fields(
        self,
        candidate: sqlite3.Row,
        definitions: list[dict[str, Any]],
    ) -> tuple[dict[str, Any], dict[str, bool], list[dict[str, Any]]]:
        """处理单名候选人的字段，任一字段错误不影响其他字段。"""

        if candidate["detail_status"] != "SUCCESS":
            return (
                {},
                {},
                [
                    {
                        "code": "DETAIL_FAILED",
                        "field_key": "",
                        "source_path": "",
                        "error": candidate["detail_error"]
                        or "Candidate Detail 未成功",
                    }
                ],
            )
        try:
            source = self._processing_source(candidate)
        except (
            FieldSchemaValidationError,
            json.JSONDecodeError,
            TypeError,
        ) as exc:
            return (
                {},
                {},
                [
                    {
                        "code": "INVALID_DETAIL_DATA",
                        "field_key": "",
                        "source_path": "",
                        "error": str(exc),
                    }
                ],
            )
        fields, empty_fields, errors = self._process_field_values(
            source,
            definitions,
            "CANDIDATE",
        )
        return (
            fields,
            _apply_module_empty_rules(fields, empty_fields, definitions),
            errors,
        )

    def process_run(
        self,
        *,
        run_id: str,
        schema_version: str,
        baseline_version: str | None = None,
        process_id: str | None = None,
    ) -> ProcessResult:
        """使用字段配置快照处理一个 Run，并保留所有旧处理结果。

        功能说明:
            读取 Candidate Detail 和关联 Task/List 字段，逐候选人生成结构化
            值、空值标记和字段级错误；关联基准时同时生成候选人判定及字段
            得分建议。处理只新增版本数据，不更新 Candidate 或 Raw。

        参数说明:
            run_id: 已完成或已导入的 Run。
            schema_version: 已发布的不可变字段配置版本。
            baseline_version: 可选基准版本；为空时仍可处理但不能形成正式指标。
            process_id: 测试或迁移可指定的唯一处理标识。

        返回值:
            包含 process_id、候选人数、错误数和状态的摘要。

        异常说明:
            FieldSchemaValidationError: Run、配置或状态不允许处理。
            非预期处理异常会把 Process 标记 FAILED 并记录 PROCESS Failure。
        """

        run = self.store.fetch_one("SELECT * FROM runs WHERE run_id = ?", (run_id,))
        if run is None:
            raise FieldSchemaValidationError(f"Run 不存在: {run_id}")
        if run["status"] in {"PENDING", "RUNNING"}:
            raise FieldSchemaValidationError(
                f"Run {run_id} 尚未结束，不能启动字段处理"
            )
        schema = self.store.fetch_one(
            "SELECT * FROM field_schemas WHERE schema_version = ?",
            (schema_version,),
        )
        if schema is None:
            raise FieldSchemaValidationError(
                f"字段配置不存在: {schema_version}"
            )
        baseline_version = (baseline_version or "").strip() or None
        if baseline_version is not None and self.store.fetch_one(
            """
            SELECT baseline_version FROM baseline_sets
            WHERE baseline_version = ?
            """,
            (baseline_version,),
        ) is None:
            raise FieldSchemaValidationError(
                f"基准版本不存在: {baseline_version}"
            )
        try:
            definitions = validate_field_definitions(
                json.loads(schema["definitions_json"])
            )
        except json.JSONDecodeError as exc:
            raise FieldSchemaValidationError(
                f"字段配置快照 JSON 已损坏: {schema_version}"
            ) from exc
        candidate_definitions = [
            definition
            for definition in definitions
            if definition["value_scope"] == "CANDIDATE"
        ]
        # 兼容阶段 1 初期已经落库的 default-v3：当时快照可能缺少
        # field_schema_version。不能仅按职责开关识别，因为旧 v2 快照在
        # 校验时也会补齐同名默认属性，从而误路由到新指标。
        is_v3_schema = (
            schema_version == DEFAULT_FIELD_SCHEMA_V3_VERSION
            or any(
                definition.get("field_schema_version") == "v3"
                for definition in definitions
            )
        )
        processing_rule_version = (
            V5_FIELD_PROCESSING_RULE_VERSION
            if is_v3_schema
            else FIELD_PROCESSING_RULE_VERSION
        )
        object_id = process_id or f"process_{uuid.uuid4().hex}"
        try:
            validate_storage_id(object_id, "process_id")
        except ValueError as exc:
            raise FieldSchemaValidationError(str(exc)) from exc
        now = utc_now_text()
        with self.store.transaction() as connection:
            connection.execute(
                """
                INSERT INTO process_runs(
                    process_id, run_id, schema_version, baseline_version,
                    rule_version, status, error_count, created_at
                ) VALUES (?, ?, ?, ?, ?, 'PROCESSING', 0, ?)
                """,
                (
                    object_id,
                    run_id,
                    schema_version,
                    baseline_version,
                    processing_rule_version,
                    now,
                ),
            )

        queries = self.store.fetch_all(
            """
            SELECT * FROM run_queries
            WHERE run_id = ?
            ORDER BY query_id
            """,
            (run_id,),
        )
        candidates = self.store.fetch_all(
            """
            SELECT c.*, rq.person_id, rq.query_stage, rq.task_id
            FROM candidates AS c
            JOIN run_queries AS rq
              ON rq.run_id = c.run_id AND rq.query_id = c.query_id
            WHERE c.run_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (run_id,),
        )
        baseline_people: dict[str, dict[str, Any]] = {}
        baseline_available_by_person: dict[str, set[str]] = {}
        if baseline_version is not None:
            for row in self.store.fetch_all(
                """
                SELECT person_id, fields_json, available_fields_json
                FROM baseline_people WHERE baseline_version = ?
                """,
                (baseline_version,),
            ):
                try:
                    value = json.loads(row["fields_json"])
                except (TypeError, json.JSONDecodeError):
                    value = {}
                baseline_people[row["person_id"]] = (
                    value if isinstance(value, dict) else {}
                )
                try:
                    available = json.loads(row["available_fields_json"] or "[]")
                except (TypeError, json.JSONDecodeError):
                    available = []
                baseline_available_by_person[row["person_id"]] = {
                    item for item in available if isinstance(item, str)
                }
        outputs: list[
            tuple[
                sqlite3.Row,
                dict[str, Any],
                dict[str, bool],
                list[dict[str, Any]],
            ]
        ] = []
        query_outputs: list[
            tuple[
                sqlite3.Row,
                dict[str, Any],
                dict[str, bool],
                list[dict[str, Any]],
            ]
        ] = []
        try:
            for query in queries:
                fields, empty_fields, errors = self._process_query_fields(
                    query,
                    definitions,
                )
                query_outputs.append(
                    (query, fields, empty_fields, errors)
                )
            for candidate in candidates:
                fields, empty_fields, errors = self._process_candidate_fields(
                    candidate,
                    definitions,
                )
                outputs.append(
                    (
                        candidate,
                        fields,
                        empty_fields,
                        errors,
                    )
                )
            error_count = sum(
                len(item[3]) for item in [*query_outputs, *outputs]
            )
            with self.store.transaction() as connection:
                for query, fields, empty_fields, errors in query_outputs:
                    result_status = (
                        query["result_status"]
                        or normalize_result_status(
                            query["status"],
                            query["candidate_count_listed"],
                        )
                    )
                    connection.execute(
                        """
                        INSERT INTO processed_queries(
                            process_id, run_id, query_id, result_status,
                            fields_json, empty_fields_json,
                            processing_errors_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            object_id,
                            run_id,
                            query["query_id"],
                            result_status,
                            json_text(fields),
                            json_text(empty_fields),
                            json_text(errors),
                        ),
                    )
                # 功能说明：同一 Query 可能出现多个强证据命中；按最高 rank_score
                # 固定一个主命中，其他命中仍保留 RULE=HIT，确保指标分母和候选人明细可追溯。
                # 返回值说明：该映射仅用于本次事务内写入 is_primary_hit，不修改 Raw。
                # 异常说明：候选人详情失败和缺少 Baseline 的记录不进入该映射，保持待处理。
                rule_hits_by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
                rule_reviewed_at = utc_now_text()
                for candidate, fields, empty_fields, errors in outputs:
                    candidate_pk = candidate["candidate_pk"]
                    connection.execute(
                        """
                        INSERT INTO processed_candidates(
                            process_id, candidate_pk, fields_json,
                            empty_fields_json, processing_errors_json
                        ) VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            object_id,
                            candidate_pk,
                            json_text(fields),
                            json_text(empty_fields),
                            json_text(errors),
                        ),
                    )
                    baseline_fields = baseline_people.get(
                        candidate["person_id"],
                        {},
                    )
                    if processing_rule_version == V5_FIELD_PROCESSING_RULE_VERSION:
                        field_scores = _field_comparison_scores_v3(
                            candidate_definitions,
                            fields,
                            empty_fields,
                            baseline_fields,
                            baseline_available_by_person.get(
                                candidate["person_id"], set()
                            ),
                        )
                    else:
                        field_scores = _suggested_field_scores(
                            candidate_definitions,
                            fields,
                            empty_fields,
                            baseline_fields,
                        )
                    baseline_identity_urls: list[Any] = []
                    returned_identity_urls: list[Any] = []
                    for definition in candidate_definitions:
                        if (
                            definition["enabled"]
                            and definition["identity_enabled"]
                            and definition["compare_mode"] == "url_set"
                        ):
                            baseline_identity_urls.extend(
                                _value_items(
                                    baseline_fields.get(definition["field_key"])
                                )
                            )
                            returned_identity_urls.extend(
                                _value_items(fields.get(definition["field_key"]))
                            )
                    if candidate["detail_status"] != "SUCCESS":
                        judgement, reason, evidence = (
                            "PENDING_REVIEW",
                            "NO_STRONG_FIELD",
                            candidate["detail_error"] or "Candidate Detail 未成功",
                        )
                    elif not baseline_fields:
                        judgement, reason, evidence = (
                            "PENDING_REVIEW",
                            "NO_STRONG_FIELD",
                            "缺少该人物的基准数据",
                        )
                    else:
                        judgement, reason, evidence = (
                            _candidate_identity_rule(
                                baseline_identity_urls,
                                returned_identity_urls,
                                (
                                    fields.get("photos_identity_match_rate")
                                    if any(
                                        item["field_key"]
                                        == "photos_identity_match_rate"
                                        and item["enabled"]
                                        and item["identity_enabled"]
                                        for item in candidate_definitions
                                    )
                                    else None
                                ),
                            )
                        )
                    is_rule_final = judgement in FINAL_JUDGEMENTS
                    classification_source = "RULE" if is_rule_final else "SUGGESTED"
                    reviewed_at = rule_reviewed_at if is_rule_final else None
                    if judgement == "HIT" and is_rule_final:
                        rule_hits_by_query[candidate["query_id"]].append(
                            {
                                "candidate_pk": candidate_pk,
                                "rank_score": candidate["rank_score"],
                                "candidate_rank": candidate["candidate_rank"],
                            }
                        )
                    connection.execute(
                        """
                        INSERT INTO reviews(
                            process_id, candidate_pk, judgement, reason,
                            evidence, field_scores_json, reviewer,
                            review_note, reviewed_at, classification_source
                        ) VALUES (?, ?, ?, ?, ?, ?, NULL, '', ?, ?)
                        """,
                        (
                            object_id,
                            candidate_pk,
                            judgement,
                            reason,
                            evidence,
                            json_text(field_scores),
                            reviewed_at,
                            classification_source,
                        ),
                    )
                for hit_candidates in rule_hits_by_query.values():
                    primary_candidate_pk = select_primary_hit_candidate(
                        hit_candidates
                    )
                    if primary_candidate_pk is None:
                        continue
                    connection.execute(
                        """
                        UPDATE reviews
                        SET is_primary_hit = 1
                        WHERE process_id = ? AND candidate_pk = ?
                        """,
                        (object_id, primary_candidate_pk),
                    )
                connection.execute(
                    """
                    UPDATE process_runs
                    SET status = 'COMPLETED', error_count = ?, finished_at = ?
                    WHERE process_id = ?
                    """,
                    (error_count, utc_now_text(), object_id),
                )
                # 同一 Run 重新处理会改变最新分析结果，旧 READY 报告需过期。
                connection.execute(
                    """
                    UPDATE reports
                    SET status = 'STALE'
                    WHERE status = 'READY'
                      AND (
                        candidate_process_id IN (
                            SELECT process_id FROM process_runs
                            WHERE run_id = ? AND process_id <> ?
                        )
                        OR baseline_process_id IN (
                            SELECT process_id FROM process_runs
                            WHERE run_id = ? AND process_id <> ?
                        )
                      )
                    """,
                    (run_id, object_id, run_id, object_id),
                )
        except Exception as exc:
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    UPDATE process_runs
                    SET status = 'FAILED', error_count = 1, finished_at = ?
                    WHERE process_id = ?
                    """,
                    (utc_now_text(), object_id),
                )
                connection.execute(
                    """
                    INSERT INTO failures(
                        failure_id, run_id, query_id, candidate_id,
                        scope, stage, error, created_at
                    ) VALUES (?, ?, NULL, '', 'PROCESS', 'FieldProcessing', ?, ?)
                    """,
                    (
                        f"failure_{uuid.uuid4().hex}",
                        run_id,
                        str(exc)[:2000],
                        utc_now_text(),
                    ),
                )
            raise
        return ProcessResult(
            process_id=object_id,
            candidate_count=len(candidates),
            error_count=error_count,
            status="COMPLETED",
        )

    def reprocess_existing_run(
        self,
        *,
        run_id: str,
        schema_version: str,
        baseline_version: str | None = None,
        process_id: str | None = None,
    ) -> ProcessResult:
        """仅使用已入库数据为历史 Run 新建独立 Process。

        功能说明:
            在调用现有 ``process_run`` 前校验 Run、字段配置、基准人物关联
            和结构化数据来源。该方法不接收 HTTP Client，也不会进入
            ``execute_run`` 或采集 Worker，因此不会产生新的接口费用。

        参数说明:
            run_id: 已结束且已包含 Query/Candidate 结构化记录的 Run。
            schema_version: 本次重处理采用的已发布字段配置版本。
            baseline_version: 可选基准版本；有关联人物时必须包含该人物。
            process_id: 测试或迁移场景可指定的新 Process ID。

        返回值:
            新 Process 摘要；``warnings`` 提示未关联人物或缺少 Raw 的情况。

        异常说明:
            FieldSchemaValidationError: Run 尚未结束、无可处理数据、配置或
            基准不存在，或人物关联不属于所选基准。
        """

        run = self.store.fetch_one(
            "SELECT run_id, status FROM runs WHERE run_id = ?",
            (run_id,),
        )
        if run is None:
            raise FieldSchemaValidationError(f"Run 不存在: {run_id}")
        if run["status"] in {"PENDING", "RUNNING"}:
            raise FieldSchemaValidationError(
                f"Run {run_id} 尚未结束，不能进行无成本重处理"
            )
        query_count = self.store.fetch_one(
            "SELECT COUNT(*) AS count FROM run_queries WHERE run_id = ?",
            (run_id,),
        )["count"]
        if query_count == 0:
            raise FieldSchemaValidationError(
                f"Run {run_id} 没有已入库 Query，无法重处理"
            )
        baseline_version = (baseline_version or "").strip() or None
        warnings: list[str] = []
        if baseline_version is not None:
            if self.store.fetch_one(
                "SELECT 1 FROM baseline_sets WHERE baseline_version = ?",
                (baseline_version,),
            ) is None:
                raise FieldSchemaValidationError(
                    f"基准版本不存在: {baseline_version}"
                )
            invalid_people = self.store.fetch_all(
                """
                SELECT DISTINCT rq.person_id
                FROM run_queries AS rq
                LEFT JOIN baseline_people AS bp
                  ON bp.baseline_version = ?
                 AND bp.person_id = rq.person_id
                WHERE rq.run_id = ?
                  AND rq.person_id IS NOT NULL
                  AND TRIM(rq.person_id) <> ''
                  AND bp.person_id IS NULL
                ORDER BY rq.person_id
                """,
                (baseline_version, run_id),
            )
            if invalid_people:
                raise FieldSchemaValidationError(
                    "以下已关联 person_id 不属于所选基准版本: "
                    + ", ".join(row["person_id"] for row in invalid_people[:10])
                )
        unlinked_count = self.store.fetch_one(
            """
            SELECT COUNT(*) AS count FROM run_queries
            WHERE run_id = ?
              AND (person_id IS NULL OR TRIM(person_id) = '')
            """,
            (run_id,),
        )["count"]
        if unlinked_count:
            warnings.append(
                f"{unlinked_count} 条 Query 尚未关联 person_id，"
                "可以处理字段，但无法形成对应人物的正式对比"
            )
        raw_count = self.store.fetch_one(
            "SELECT COUNT(*) AS count FROM raw_records WHERE run_id = ?",
            (run_id,),
        )["count"]
        if raw_count == 0:
            warnings.append(
                "该 Run 没有 Raw 记录，将使用已入库的 Query/Candidate "
                "结构化字段进行重处理"
            )
        result = self.process_run(
            run_id=run_id,
            schema_version=schema_version,
            baseline_version=baseline_version,
            process_id=process_id,
        )
        result.warnings.extend(warnings)
        return result

    def get_query_classification_context(
        self,
        process_id: str,
        query_id: str,
    ) -> dict[str, Any]:
        """读取 Query、基准人物、全部候选人及当前身份归类。

        返回值:
            页面和事务保存共用的上下文；字段空值和完整度分数均来自当前
            Process 快照，不允许身份归类页面手工改写。

        异常说明:
            ReviewValidationError: Process/Query 不匹配或 JSON 快照损坏。
        """

        query = self.store.fetch_one(
            """
            SELECT pq.*, pr.baseline_version, pr.status AS process_status,
                   rq.person_id, rq.person_id_source, rq.query_stage,
                   rq.public_fields_json, rq.status AS query_status,
                   bp.display_name AS baseline_display_name,
                   bp.fields_json AS baseline_fields_json,
                   bp.available_fields_json
            FROM processed_queries AS pq
            JOIN process_runs AS pr ON pr.process_id = pq.process_id
            JOIN run_queries AS rq
              ON rq.run_id = pq.run_id AND rq.query_id = pq.query_id
            LEFT JOIN baseline_people AS bp
              ON bp.baseline_version = pr.baseline_version
             AND bp.person_id = rq.person_id
            WHERE pq.process_id = ? AND pq.query_id = ?
            """,
            (process_id, query_id),
        )
        if query is None:
            raise ReviewValidationError(
                f"处理结果 {process_id} 中不存在 Query {query_id}"
            )
        rows = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.candidate_id, c.candidate_rank,
                   c.rank_score, c.detail_status, c.detail_error,
                   pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json,
                   rv.judgement, rv.reason, rv.evidence,
                   rv.field_scores_json, rv.reviewer, rv.review_note,
                   rv.reviewed_at, rv.classification_source,
                   rv.is_primary_hit
            FROM candidates AS c
            JOIN processed_candidates AS pc
              ON pc.candidate_pk = c.candidate_pk
             AND pc.process_id = ?
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE c.run_id = ? AND c.query_id = ?
            ORDER BY c.candidate_rank, c.candidate_pk
            """,
            (process_id, query["run_id"], query_id),
        )
        candidates: list[dict[str, Any]] = []
        try:
            for row in rows:
                candidates.append({
                    **dict(row),
                    "fields": json.loads(row["fields_json"] or "{}"),
                    "empty_fields": json.loads(
                        row["empty_fields_json"] or "{}"
                    ),
                    "processing_errors": json.loads(
                        row["processing_errors_json"] or "[]"
                    ),
                    "field_scores": json.loads(
                        row["field_scores_json"] or "{}"
                    ),
                    "is_primary_hit": bool(row["is_primary_hit"]),
                })
            baseline_fields = json.loads(
                query["baseline_fields_json"] or "{}"
            )
            available_fields = json.loads(
                query["available_fields_json"] or "[]"
            )
            input_data = json.loads(query["public_fields_json"] or "{}")
        except (TypeError, json.JSONDecodeError) as exc:
            raise ReviewValidationError(
                f"Query 身份归类依赖的数据快照已损坏: {exc}"
            ) from exc
        final_candidates = [
            item
            for item in candidates
            if item["reviewed_at"] is not None
            and item["classification_source"] in {"MANUAL", "RULE"}
        ]
        primary_hits = [
            item
            for item in final_candidates
            if item["judgement"] == "HIT" and item["is_primary_hit"]
        ]
        successful = [
            item for item in candidates if item["detail_status"] == "SUCCESS"
        ]
        pending_count = sum(
            item["reviewed_at"] is None for item in successful
        )
        if primary_hits:
            identity_state = "HIT_CONFIRMED"
        elif successful and pending_count == 0:
            identity_state = "NO_HIT_CONFIRMED"
        elif not candidates:
            identity_state = "NO_CANDIDATES"
        else:
            identity_state = "PENDING"
        query_ids = [
            row["query_id"]
            for row in self.store.fetch_all(
                """
                SELECT query_id FROM processed_queries
                WHERE process_id = ? ORDER BY query_id
                """,
                (process_id,),
            )
        ]
        index = query_ids.index(query_id)
        return {
            "process_id": process_id,
            "run_id": query["run_id"],
            "query_id": query_id,
            "query_stage": query["query_stage"],
            "query_status": query["query_status"],
            "result_status": query["result_status"],
            "person_id": query["person_id"],
            "person_id_source": query["person_id_source"],
            "input_data": input_data,
            "baseline_version": query["baseline_version"],
            "baseline_display_name": query["baseline_display_name"],
            "baseline_fields": (
                baseline_fields if isinstance(baseline_fields, dict) else {}
            ),
            "baseline_available_fields": (
                available_fields if isinstance(available_fields, list) else []
            ),
            "candidates": candidates,
            "identity_state": identity_state,
            "primary_hit_candidate_pk": (
                primary_hits[0]["candidate_pk"] if primary_hits else None
            ),
            "pending_count": pending_count,
            "previous_query_id": query_ids[index - 1] if index > 0 else None,
            "next_query_id": (
                query_ids[index + 1] if index + 1 < len(query_ids) else None
            ),
        }

    def save_query_classification(
        self,
        process_id: str,
        query_id: str,
        classifications: list[dict[str, Any]],
        *,
        primary_hit_candidate_pk: str | None,
        confirm_no_hit: bool,
        reviewer: str = "",
        review_note: str = "",
        expected_versions: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        """原子保存一个 Query 的候选人身份归类。

        关键逻辑:
            未提交候选人保持原状态；主命中和“确认无命中”互斥；详情失败
            候选人禁止设为 HIT；乐观锁按 ``reviewed_at`` 防止旧页面覆盖。
            本方法只更新身份元数据，既不修改接口原始值，也不接收完整度。
        """

        context = self.get_query_classification_context(process_id, query_id)
        if not isinstance(classifications, list):
            raise ReviewValidationError("候选人身份归类必须是数组")
        candidate_by_pk = {
            item["candidate_pk"]: item for item in context["candidates"]
        }
        primary_hit_candidate_pk = (
            str(primary_hit_candidate_pk).strip()
            if primary_hit_candidate_pk
            else None
        )
        if primary_hit_candidate_pk and confirm_no_hit:
            raise ReviewValidationError("主命中与确认无命中不能同时提交")
        submitted: dict[str, dict[str, Any]] = {}
        errors: list[str] = []
        for raw in classifications:
            if not isinstance(raw, dict):
                errors.append("候选人归类项必须是对象")
                continue
            candidate_pk = str(raw.get("candidate_pk") or "").strip()
            judgement = str(raw.get("judgement") or "").strip()
            reason = str(raw.get("reason") or "MANUAL").strip()
            if candidate_pk not in candidate_by_pk:
                errors.append(f"候选人不属于当前 Query: {candidate_pk}")
                continue
            if candidate_pk in submitted:
                errors.append(f"候选人重复提交: {candidate_pk}")
                continue
            if judgement not in FINAL_JUDGEMENTS:
                errors.append(
                    f"{candidate_pk} 只支持 HIT、NOT_HIT 或 SUSPECTED"
                )
            if reason not in REVIEW_REASONS:
                errors.append(f"{candidate_pk} 的归类原因不受支持")
            if (
                judgement == "HIT"
                and candidate_by_pk[candidate_pk]["detail_status"] != "SUCCESS"
            ):
                errors.append(f"{candidate_pk} 详情失败，不能标记为 HIT")
            submitted[candidate_pk] = {
                "judgement": judgement,
                "reason": reason,
                "evidence": str(raw.get("evidence") or "").strip(),
            }
        hit_pks = [
            candidate_pk
            for candidate_pk, item in submitted.items()
            if item["judgement"] == "HIT"
        ]
        if (
            primary_hit_candidate_pk
            and primary_hit_candidate_pk not in candidate_by_pk
        ):
            errors.append("主命中候选人不属于当前 Query")
        expected_versions = expected_versions or {}
        for candidate_pk in submitted:
            current = candidate_by_pk[candidate_pk]["reviewed_at"] or ""
            if candidate_pk not in expected_versions:
                errors.append(f"{candidate_pk} 缺少页面版本，请刷新后重试")
            elif str(expected_versions[candidate_pk] or "") != current:
                errors.append("该归类已在其他页面更新，请刷新后重新提交")
        if confirm_no_hit:
            for candidate_pk, candidate in candidate_by_pk.items():
                if candidate["detail_status"] != "SUCCESS":
                    continue
                proposed = submitted.get(candidate_pk)
                judgement = (
                    proposed["judgement"]
                    if proposed is not None
                    else (
                        candidate["judgement"]
                        if candidate["reviewed_at"] is not None
                        else "PENDING_REVIEW"
                    )
                )
                if judgement not in {"NOT_HIT", "SUSPECTED"}:
                    errors.append(
                        "确认无命中前，所有详情成功候选人都必须明确归类为"
                        " NOT_HIT 或 SUSPECTED"
                    )
                    break
        if errors:
            raise ReviewValidationError(errors)
        reviewed_at = utc_now_text()
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE reviews SET is_primary_hit = 0
                WHERE process_id = ?
                  AND candidate_pk IN (
                    SELECT candidate_pk FROM candidates
                    WHERE run_id = ? AND query_id = ?
                  )
                """,
                (process_id, context["run_id"], query_id),
            )
            for candidate_pk, item in submitted.items():
                current = candidate_by_pk[candidate_pk]
                cursor = connection.execute(
                    """
                    UPDATE reviews
                    SET judgement = ?, reason = ?, evidence = ?,
                        reviewer = ?, review_note = ?, reviewed_at = ?,
                        classification_source = 'MANUAL',
                        is_primary_hit = ?
                    WHERE process_id = ? AND candidate_pk = ?
                      AND COALESCE(reviewed_at, '') = ?
                    """,
                    (
                        item["judgement"],
                        item["reason"],
                        item["evidence"],
                        reviewer.strip() or None,
                        review_note.strip(),
                        reviewed_at,
                        0,
                        process_id,
                        candidate_pk,
                        current["reviewed_at"] or "",
                    ),
                )
                if cursor.rowcount != 1:
                    raise ReviewValidationError(
                        "该归类已在其他页面更新，请刷新后重新提交"
                    )
            # 功能说明：允许人工确认多条 HIT，但正式汇总只采用 rank_score
            # 最高的一条。页面提交的旧“主命中”单选值仅为兼容旧表单，不能
            # 覆盖统一的评分排序规则。
            hit_rows = connection.execute(
                """
                SELECT c.candidate_pk, c.rank_score, c.candidate_rank
                FROM reviews AS rv
                JOIN candidates AS c ON c.candidate_pk = rv.candidate_pk
                WHERE rv.process_id = ? AND c.run_id = ? AND c.query_id = ?
                  AND rv.judgement = 'HIT' AND rv.reviewed_at IS NOT NULL
                  AND rv.classification_source IN ('MANUAL', 'RULE')
                """,
                (process_id, context["run_id"], query_id),
            ).fetchall()
            selected_primary_pk = select_primary_hit_candidate(hit_rows)
            if selected_primary_pk:
                connection.execute(
                    """
                    UPDATE reviews SET is_primary_hit = 1
                    WHERE process_id = ? AND candidate_pk = ?
                    """,
                    (process_id, selected_primary_pk),
                )
            connection.execute(
                """
                UPDATE reports SET status = 'STALE'
                WHERE status = 'READY'
                  AND (
                    baseline_process_id = ?
                    OR candidate_process_id = ?
                  )
                """,
                (process_id, process_id),
            )
        return self.get_query_classification_context(process_id, query_id)

    def get_process_classification_progress(
        self,
        process_id: str,
    ) -> dict[str, Any]:
        """汇总 Process 的 Query 身份归类进度。

        返回值:
            包含 Query 总数、已确认命中、已确认无命中、待判定数量及逐
            Query 状态。无候选人的 Query 单独计数，不误记为人工无命中。
        """

        query_ids = [
            row["query_id"]
            for row in self.store.fetch_all(
                """
                SELECT query_id FROM processed_queries
                WHERE process_id = ? ORDER BY query_id
                """,
                (process_id,),
            )
        ]
        items = [
            self.get_query_classification_context(process_id, query_id)
            for query_id in query_ids
        ]
        counts = {
            "HIT_CONFIRMED": 0,
            "NO_HIT_CONFIRMED": 0,
            "PENDING": 0,
            "NO_CANDIDATES": 0,
        }
        for item in items:
            counts[item["identity_state"]] += 1
        return {
            "total_count": len(items),
            "hit_confirmed_count": counts["HIT_CONFIRMED"],
            "no_hit_confirmed_count": counts["NO_HIT_CONFIRMED"],
            "pending_count": counts["PENDING"],
            "no_candidates_count": counts["NO_CANDIDATES"],
            "items": items,
        }

    def get_review_context(
        self,
        process_id: str,
        candidate_pk: str,
    ) -> dict[str, Any]:
        """读取候选人的处理结果、基准、建议得分和当前人工复核。

        参数说明:
            process_id: 不可变字段处理结果标识。
            candidate_pk: 候选人内部主键。

        返回值:
            页面和保存校验共用的复核上下文。

        异常说明:
            ReviewValidationError: Process/Candidate 不匹配或配置快照损坏。
        """

        row = self.store.fetch_one(
            """
            SELECT pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json,
                   pr.run_id, pr.schema_version, pr.baseline_version,
                   pr.rule_version, pr.status AS process_status,
                   fs.definitions_json,
                   c.candidate_pk, c.candidate_id, c.candidate_rank,
                   c.detail_status, c.detail_error, c.query_id,
                   rq.person_id, rq.query_stage,
                   bp.display_name AS baseline_display_name,
                   bp.fields_json AS baseline_fields_json,
                   bp.evidence_json AS baseline_evidence_json,
                   rv.judgement AS review_judgement,
                   rv.reason AS review_reason,
                   rv.evidence AS review_evidence,
                   rv.field_scores_json, rv.reviewer, rv.review_note,
                   rv.reviewed_at, rv.is_primary_hit
            FROM processed_candidates AS pc
            JOIN process_runs AS pr ON pr.process_id = pc.process_id
            JOIN field_schemas AS fs
              ON fs.schema_version = pr.schema_version
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            JOIN run_queries AS rq
              ON rq.run_id = c.run_id AND rq.query_id = c.query_id
            LEFT JOIN baseline_people AS bp
              ON bp.baseline_version = pr.baseline_version
             AND bp.person_id = rq.person_id
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ? AND pc.candidate_pk = ?
            """,
            (process_id, candidate_pk),
        )
        if row is None:
            raise ReviewValidationError(
                f"处理结果 {process_id} 中不存在候选人 {candidate_pk}"
            )
        try:
            definitions = validate_field_definitions(
                json.loads(row["definitions_json"])
            )
            fields = json.loads(row["fields_json"])
            empty_fields = json.loads(row["empty_fields_json"])
            baseline_fields = (
                json.loads(row["baseline_fields_json"])
                if row["baseline_fields_json"]
                else {}
            )
        except (
            json.JSONDecodeError,
            FieldSchemaValidationError,
            TypeError,
        ) as exc:
            raise ReviewValidationError(
                f"复核依赖的数据快照已损坏: {exc}"
            ) from exc
        if not isinstance(fields, dict):
            fields = {}
        if not isinstance(empty_fields, dict):
            empty_fields = {}
        if not isinstance(baseline_fields, dict):
            baseline_fields = {}
        if row["field_scores_json"]:
            try:
                field_scores = json.loads(row["field_scores_json"])
            except (TypeError, json.JSONDecodeError) as exc:
                raise ReviewValidationError("字段复核得分快照已损坏") from exc
        else:
            field_scores = _suggested_field_scores(
                definitions,
                fields,
                empty_fields,
                baseline_fields,
            )
        if not isinstance(field_scores, dict):
            raise ReviewValidationError("字段复核得分快照必须是对象")
        return {
            "process_id": process_id,
            "run_id": row["run_id"],
            "candidate_pk": row["candidate_pk"],
            "candidate_id": row["candidate_id"],
            "candidate_rank": row["candidate_rank"],
            "query_id": row["query_id"],
            "person_id": row["person_id"],
            "query_stage": row["query_stage"],
            "detail_status": row["detail_status"],
            "detail_error": row["detail_error"],
            "schema_version": row["schema_version"],
            "baseline_version": row["baseline_version"],
            "baseline_display_name": row["baseline_display_name"],
            "baseline_fields": baseline_fields,
            "baseline_evidence": (
                json.loads(row["baseline_evidence_json"])
                if row["baseline_evidence_json"]
                else {}
            ),
            "fields": fields,
            "empty_fields": empty_fields,
            "definitions": {
                item["field_key"]: item for item in definitions
            },
            "field_scores": field_scores,
            "judgement": row["review_judgement"] or "PENDING_REVIEW",
            "reason": row["review_reason"] or "NO_STRONG_FIELD",
            "evidence": row["review_evidence"] or "",
            "reviewer": row["reviewer"] or "",
            "review_note": row["review_note"] or "",
            "reviewed_at": row["reviewed_at"],
            "is_primary_hit": bool(row["is_primary_hit"]),
            "review_exists": row["review_judgement"] is not None,
            "is_final": row["reviewed_at"] is not None,
        }

    def save_review(
        self,
        *,
        process_id: str,
        candidate_pk: str,
        judgement: str,
        reason: str,
        evidence: str = "",
        reviewer: str = "",
        review_note: str = "",
        field_scores: dict[str, Any] | None = None,
        expected_reviewed_at: str | None = None,
    ) -> dict[str, Any]:
        """事务保存人工最终判定与字段得分，并使关联报告过期。

        ``expected_reviewed_at`` 来自页面隐藏字段；数据库时间变化时拒绝覆盖，
        避免同一用户的多个浏览器标签互相覆盖。
        """

        context = self.get_review_context(process_id, candidate_pk)
        if judgement not in FINAL_JUDGEMENTS:
            raise ReviewValidationError(
                "最终判定只支持 HIT、NOT_HIT 或 SUSPECTED"
            )
        if reason not in REVIEW_REASONS:
            raise ReviewValidationError("复核原因不受支持")
        current = context["reviewed_at"] or ""
        expected = current if expected_reviewed_at is None else expected_reviewed_at
        # 功能说明：Web 页面会提交快照时间以防并发覆盖；服务内部与兼容调用未传
        # expected_reviewed_at 时，允许测试或批处理显式覆写自动 RULE 结果。
        # 参数说明：None 表示未启用并发校验；空字符串仍表示调用方期望旧记录未复核。
        # 异常说明：仅在调用方提供快照且与当前值不一致时拒绝保存。
        if expected != current:
            raise ReviewValidationError(
                "该复核已在其他页面更新，请刷新后重新提交"
            )
        proposed = field_scores if field_scores is not None else context["field_scores"]
        if not isinstance(proposed, dict):
            raise ReviewValidationError("字段得分必须是对象")
        current_scores = context["field_scores"]
        if set(proposed) - set(current_scores):
            raise ReviewValidationError("字段得分包含当前配置之外的字段")
        validated_scores: dict[str, dict[str, Any]] = {}
        overridden = False
        errors: list[str] = []
        for field_key, current_score in current_scores.items():
            raw_score = proposed.get(field_key, current_score)
            if not isinstance(raw_score, dict):
                errors.append(f"{field_key} 字段得分必须是对象")
                continue
            item = dict(current_score)
            for score_name in ("completeness_score", "accuracy_score"):
                value = raw_score.get(score_name)
                if value in (None, ""):
                    item[score_name] = None
                    continue
                if isinstance(value, bool):
                    errors.append(f"{field_key} {score_name} 必须在0到1之间")
                    continue
                try:
                    number = float(value)
                except (TypeError, ValueError):
                    errors.append(f"{field_key} {score_name} 必须在0到1之间")
                    continue
                if not math.isfinite(number) or not 0 <= number <= 1:
                    errors.append(f"{field_key} {score_name} 必须在0到1之间")
                    continue
                item[score_name] = number
            item["review_note"] = str(raw_score.get("review_note", "")).strip()
            item["manual_override"] = (
                item["completeness_score"]
                != item.get("suggested_completeness_score")
                or item["accuracy_score"]
                != item.get("suggested_accuracy_score")
            )
            overridden = overridden or item["manual_override"]
            validated_scores[field_key] = item
        if judgement == "HIT":
            for field_key, definition in context["definitions"].items():
                score = validated_scores.get(field_key)
                if score is None or not score.get("baseline_available"):
                    continue
                roles = definition["scoring_role"]
                if (
                    "completeness" in roles
                    and score.get("completeness_score") is None
                ):
                    errors.append(f"{field_key} 的完整度得分尚未确认")
                if (
                    "accuracy" in roles
                    and score.get("returned_nonempty")
                    and score.get("accuracy_score") is None
                ):
                    errors.append(f"{field_key} 的准确率得分尚未确认")
        if overridden and not review_note.strip():
            errors.append("人工覆盖建议得分时必须填写复核说明")
        if errors:
            raise ReviewValidationError(errors)
        reviewed_at = utc_now_text()
        with self.store.transaction() as connection:
            if context["review_exists"]:
                cursor = connection.execute(
                    """
                    UPDATE reviews
                    SET judgement = ?, reason = ?, evidence = ?,
                        field_scores_json = ?, reviewer = ?,
                        review_note = ?, reviewed_at = ?,
                        classification_source = 'MANUAL'
                    WHERE process_id = ? AND candidate_pk = ?
                      AND COALESCE(reviewed_at, '') = ?
                    """,
                    (
                        judgement,
                        reason,
                        evidence.strip(),
                        json_text(validated_scores),
                        reviewer.strip() or None,
                        review_note.strip(),
                        reviewed_at,
                        process_id,
                        candidate_pk,
                        expected,
                    ),
                )
                if cursor.rowcount != 1:
                    raise ReviewValidationError(
                        "该复核已在其他页面更新，请刷新后重新提交"
                    )
            else:
                try:
                    connection.execute(
                        """
                        INSERT INTO reviews(
                            process_id, candidate_pk, judgement, reason,
                            evidence, field_scores_json, reviewer,
                            review_note, reviewed_at, classification_source
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'MANUAL')
                        """,
                        (
                            process_id,
                            candidate_pk,
                            judgement,
                            reason,
                            evidence.strip(),
                            json_text(validated_scores),
                            reviewer.strip() or None,
                            review_note.strip(),
                            reviewed_at,
                        ),
                    )
                except sqlite3.IntegrityError as exc:
                    raise ReviewValidationError(
                        "该复核已在其他页面更新，请刷新后重新提交"
                    ) from exc
            # 单候选人详细复核也遵循与批量归类相同的主命中选择规则：多个
            # HIT 都保留候选人级评分，仅 rank_score 最高者进入正式汇总。
            connection.execute(
                """
                UPDATE reviews SET is_primary_hit = 0
                WHERE process_id = ?
                  AND candidate_pk IN (
                    SELECT candidate_pk FROM candidates
                    WHERE run_id = ? AND query_id = ?
                  )
                """,
                (process_id, context["run_id"], context["query_id"]),
            )
            hit_rows = connection.execute(
                """
                SELECT c.candidate_pk, c.rank_score, c.candidate_rank
                FROM reviews AS rv
                JOIN candidates AS c ON c.candidate_pk = rv.candidate_pk
                WHERE rv.process_id = ? AND c.run_id = ? AND c.query_id = ?
                  AND rv.judgement = 'HIT' AND rv.reviewed_at IS NOT NULL
                  AND rv.classification_source IN ('MANUAL', 'RULE')
                """,
                (process_id, context["run_id"], context["query_id"]),
            ).fetchall()
            selected_primary_pk = select_primary_hit_candidate(hit_rows)
            if selected_primary_pk:
                connection.execute(
                    """
                    UPDATE reviews SET is_primary_hit = 1
                    WHERE process_id = ? AND candidate_pk = ?
                    """,
                    (process_id, selected_primary_pk),
                )
            connection.execute(
                """
                UPDATE reports SET status = 'STALE'
                WHERE status = 'READY'
                  AND (
                    baseline_process_id = ?
                    OR candidate_process_id = ?
                  )
                """,
                (process_id, process_id),
            )
        return {
            "process_id": process_id,
            "candidate_pk": candidate_pk,
            "judgement": judgement,
            "reviewed_at": reviewed_at,
        }

    def calculate_process_metrics(self, process_id: str) -> dict[str, Any]:
        """按 Process 固化的字段处理规则分派指标口径。

        历史 ``field-processing-v1/v2`` 分别继续使用 metrics-v1/v2；
        新 ``field-processing-v3`` 使用 metrics-v3。未知规则直接拒绝，
        避免历史 Process 被静默套用新口径。
        """

        process = self.store.fetch_one(
            """
            SELECT rule_version FROM process_runs
            WHERE process_id = ?
            """,
            (process_id,),
        )
        if process is None:
            raise ReviewValidationError(f"处理结果不存在: {process_id}")
        if process["rule_version"] == LEGACY_FIELD_PROCESSING_RULE_VERSION:
            metrics = self._calculate_process_metrics_v1(process_id)
            metrics["metrics_rule_version"] = "metrics-v1"
            return metrics
        if process["rule_version"] == V2_FIELD_PROCESSING_RULE_VERSION:
            return self._calculate_process_metrics_v2(process_id)
        if process["rule_version"] in {
            V3_FIELD_PROCESSING_RULE_VERSION,
            FIELD_PROCESSING_RULE_VERSION,
        }:
            return self._calculate_process_metrics_v3(process_id)
        if process["rule_version"] == V5_FIELD_PROCESSING_RULE_VERSION:
            return self._calculate_process_metrics_v4(process_id)
        raise ReviewValidationError(
            f"不支持的指标规则来源: {process['rule_version']}"
        )

    def _calculate_process_metrics_v1(self, process_id: str) -> dict[str, Any]:
        """按最终人工复核计算 Query 与 Run 核心指标。

        未完成复核或缺少基准时仍返回可下钻的预览分子、分母，但正式
        ``value`` 保持 ``None``，避免把系统建议或缺失数据当成正式结论。
        """

        process = self.store.fetch_one(
            """
            SELECT pr.*, r.evaluation_id, r.run_label, r.system_version,
                   fs.definitions_json
            FROM process_runs AS pr
            JOIN runs AS r ON r.run_id = pr.run_id
            JOIN field_schemas AS fs
              ON fs.schema_version = pr.schema_version
            WHERE pr.process_id = ?
            """,
            (process_id,),
        )
        if process is None:
            raise ReviewValidationError(f"处理结果不存在: {process_id}")
        try:
            definitions = validate_field_definitions(
                json.loads(process["definitions_json"])
            )
        except (json.JSONDecodeError, FieldSchemaValidationError) as exc:
            raise ReviewValidationError(f"字段配置快照已损坏: {exc}") from exc
        completeness_keys = {
            item["field_key"]
            for item in definitions
            if item["enabled"] and "completeness" in item["scoring_role"]
        }
        accuracy_keys = {
            item["field_key"]
            for item in definitions
            if item["enabled"] and "accuracy" in item["scoring_role"]
        }
        baseline_people = {
            row["person_id"]
            for row in self.store.fetch_all(
                """
                SELECT person_id FROM baseline_people
                WHERE baseline_version = ?
                """,
                (process["baseline_version"],),
            )
        } if process["baseline_version"] else set()
        queries = self.store.fetch_all(
            """
            SELECT * FROM run_queries
            WHERE run_id = ? ORDER BY query_id
            """,
            (process["run_id"],),
        )
        candidate_rows = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.query_id, c.candidate_id,
                   c.candidate_rank, c.detail_status,
                   pc.fields_json, pc.empty_fields_json,
                   rv.judgement, rv.field_scores_json, rv.reviewed_at
            FROM candidates AS c
            JOIN processed_candidates AS pc
              ON pc.candidate_pk = c.candidate_pk AND pc.process_id = ?
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE c.run_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process_id, process["run_id"]),
        )
        by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in candidate_rows:
            item = dict(row)
            for key, default in (
                ("fields_json", {}),
                ("empty_fields_json", {}),
                ("field_scores_json", {}),
            ):
                try:
                    parsed = json.loads(item[key]) if item[key] else default
                except (TypeError, json.JSONDecodeError):
                    parsed = default
                item[key.removesuffix("_json")] = (
                    parsed if isinstance(parsed, dict) else default
                )
            by_query[row["query_id"]].append(item)

        query_metrics: list[dict[str, Any]] = []
        nonmatched_values: list[float] = []
        for query in queries:
            successful_candidates = [
                item
                for item in by_query.get(query["query_id"], [])
                if item["detail_status"] == "SUCCESS"
            ]
            final_candidates = [
                item for item in successful_candidates if item["reviewed_at"]
            ]
            hit_candidates = [
                item
                for item in final_candidates
                if item["judgement"] == "HIT"
            ]
            hit_completeness: list[float] = []
            hit_accuracy: list[float] = []
            query_nonmatched_values: list[float] = []
            score_complete = True
            for candidate in hit_candidates:
                scores = candidate["field_scores"]
                completeness_scores = [
                    scores.get(field_key, {}).get("completeness_score")
                    for field_key in completeness_keys
                ]
                if any(value is None for value in completeness_scores):
                    score_complete = False
                else:
                    hit_completeness.append(
                        sum(float(value) for value in completeness_scores)
                        / CONTENT_FIELD_COUNT
                    )
                accuracy_scores = [
                    scores.get(field_key, {}).get("accuracy_score")
                    for field_key in accuracy_keys
                    if scores.get(field_key, {}).get("returned_nonempty")
                ]
                if any(value is None for value in accuracy_scores):
                    score_complete = False
                elif accuracy_scores:
                    hit_accuracy.append(
                        sum(float(value) for value in accuracy_scores)
                        / len(accuracy_scores)
                    )
            for candidate in final_candidates:
                if candidate["judgement"] not in {"NOT_HIT", "SUSPECTED"}:
                    continue
                empty_fields = candidate["empty_fields"]
                nonempty_count = sum(
                    1
                    for field_key in completeness_keys
                    if field_key in candidate["fields"]
                    and not empty_fields.get(field_key, True)
                )
                nonmatched_values.append(
                    nonempty_count / CONTENT_FIELD_COUNT
                )
                query_nonmatched_values.append(
                    nonempty_count / CONTENT_FIELD_COUNT
                )
            baseline_available = (
                bool(process["baseline_version"])
                and bool(query["person_id"])
                and query["person_id"] in baseline_people
            )
            review_complete = len(final_candidates) == len(successful_candidates)
            formal_ready = baseline_available and review_complete and score_complete
            query_metrics.append(
                {
                    "query_id": query["query_id"],
                    "person_id": query["person_id"],
                    "query_stage": query["query_stage"],
                    "retrieval_success": bool(hit_candidates),
                    "matched_completeness": (
                        sum(hit_completeness) / len(hit_completeness)
                        if hit_completeness
                        else None
                    ),
                    "matched_accuracy": (
                        sum(hit_accuracy) / len(hit_accuracy)
                        if hit_accuracy
                        else None
                    ),
                    "formal_ready": formal_ready,
                    "pending_review_count": (
                        len(successful_candidates) - len(final_candidates)
                    ),
                    "hit_candidate_ids": [
                        item["candidate_pk"] for item in hit_candidates
                    ],
                    "nonmatched_candidate_ids": [
                        item["candidate_pk"]
                        for item in final_candidates
                        if item["judgement"] in {"NOT_HIT", "SUSPECTED"}
                    ],
                    "nonmatched_completeness_values": query_nonmatched_values,
                    "candidate_ids": [
                        item["candidate_pk"] for item in successful_candidates
                    ],
                }
            )
        formal_ready = bool(query_metrics) and all(
            item["formal_ready"] for item in query_metrics
        )

        def aggregate(
            numerator: float,
            denominator: int,
        ) -> dict[str, Any]:
            """构造含正式值和预览值的统一指标对象。"""

            preview_value = (
                numerator / denominator if denominator else None
            )
            return {
                "numerator": numerator,
                "denominator": denominator,
                "value": preview_value if formal_ready else None,
                "preview_value": preview_value,
            }

        successful_queries = [
            item for item in query_metrics if item["retrieval_success"]
        ]
        matched_completeness_values = [
            item["matched_completeness"]
            for item in successful_queries
            if item["matched_completeness"] is not None
        ]
        matched_accuracy_values = [
            item["matched_accuracy"]
            for item in successful_queries
            if item["matched_accuracy"] is not None
        ]
        task_rows = self.store.fetch_all(
            """
            SELECT llm_cost, total_cost, pdl_called
            FROM run_queries WHERE run_id = ?
            """,
            (process["run_id"],),
        )
        cost_columns = ("llm_cost", "total_cost", "pdl_called")
        all_cost_missing = all(
            row[column] is None
            for row in task_rows
            for column in cost_columns
        )
        all_cost_complete = bool(task_rows) and all(
            row[column] is not None
            for row in task_rows
            for column in cost_columns
        )
        cost_status = {
            "status": (
                "NOT_CONNECTED"
                if all_cost_missing
                else "COMPLETE" if all_cost_complete else "PARTIAL"
            ),
            "task_count": len(task_rows),
            "missing_task_count": sum(
                1
                for row in task_rows
                if any(row[column] is None for column in cost_columns)
            ),
            "llm_cost_total": (
                sum(float(row["llm_cost"]) for row in task_rows)
                if all_cost_complete
                else None
            ),
            "total_cost_total": (
                sum(float(row["total_cost"]) for row in task_rows)
                if all_cost_complete
                else None
            ),
            "pdl_called_total": (
                sum(int(row["pdl_called"]) for row in task_rows)
                if all_cost_complete
                else None
            ),
        }
        return {
            "process_id": process_id,
            "run_id": process["run_id"],
            "evaluation_id": process["evaluation_id"],
            "schema_version": process["schema_version"],
            "baseline_version": process["baseline_version"],
            "rule_version": process["rule_version"],
            "formal_ready": formal_ready,
            "review_status": "REVIEWED" if formal_ready else "PENDING_REVIEW",
            "retrieval_success": aggregate(
                float(len(successful_queries)),
                len(query_metrics),
            ),
            "matched_completeness": aggregate(
                sum(matched_completeness_values),
                len(matched_completeness_values),
            ),
            "matched_accuracy": aggregate(
                sum(matched_accuracy_values),
                len(matched_accuracy_values),
            ),
            "nonmatched_completeness": aggregate(
                sum(nonmatched_values),
                len(nonmatched_values),
            ),
            "cost_status": cost_status,
            "query_metrics": query_metrics,
        }

    @staticmethod
    def _metrics_v2_value(
        numerator: float,
        denominator: int,
        *,
        ready: bool,
        not_ready_reasons: list[str] | None = None,
    ) -> dict[str, Any]:
        """构造可追溯的 v2 指标值。

        参数说明:
            numerator: 已纳入计算的分子。
            denominator: 已纳入计算的分母。
            ready: 依赖的 Baseline、复核和字段得分是否完整。
            not_ready_reasons: 未就绪原因，供页面和 API 下钻。

        返回值:
            包含分子、分母、正式值、预览值和状态的字典。分母为0时正式
            值保持 ``None``，不会把“不适用”展示成0%。
        """

        preview_value = numerator / denominator if denominator else None
        status = (
            "NOT_READY"
            if not ready
            else "READY" if denominator else "NOT_APPLICABLE"
        )
        return {
            "numerator": numerator,
            "denominator": denominator,
            "value": preview_value if ready and denominator else None,
            "preview_value": preview_value,
            "status": status,
            "not_ready_reasons": sorted(set(not_ready_reasons or [])),
        }

    @staticmethod
    def _metrics_v2_result_status(
        query_rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """按规范化 Query 结果状态计算有结果、无结果和执行失败比例。"""

        counts = {
            status: sum(row["result_status"] == status for row in query_rows)
            for status in RESULT_STATUSES
        }
        completed = (
            counts["HAS_CANDIDATES"] + counts["NO_CANDIDATES"]
        )
        total = len(query_rows)
        return {
            "total_formal_queries": total,
            "has_candidates_count": counts["HAS_CANDIDATES"],
            "no_candidates_count": counts["NO_CANDIDATES"],
            "execution_failed_count": counts["EXECUTION_FAILED"],
            "has_result_rate": (
                counts["HAS_CANDIDATES"] / completed if completed else None
            ),
            "no_result_rate": (
                counts["NO_CANDIDATES"] / completed if completed else None
            ),
            "execution_failed_rate": (
                counts["EXECUTION_FAILED"] / total if total else None
            ),
        }

    @staticmethod
    def _metrics_v2_numeric_aggregate(
        query_rows: list[dict[str, Any]],
        field_key: str,
    ) -> dict[str, Any]:
        """独立汇总一个成本或耗时字段，缺失和非法值都不按0补齐。

        负数、布尔值、非数字和非有限数值记为类型错误，不进入汇总；
        合法的0仍是有效值。
        """

        values: list[float] = []
        missing_query_ids: list[str] = []
        invalid_query_ids: list[str] = []
        for row in query_rows:
            value = row["task_fields"].get(field_key)
            if value is None or value == "":
                has_processing_error = any(
                    isinstance(error, dict)
                    and error.get("field_key") == field_key
                    for error in row.get("task_processing_errors", [])
                )
                (
                    invalid_query_ids
                    if has_processing_error
                    else missing_query_ids
                ).append(row["query_id"])
                continue
            if (
                isinstance(value, bool)
                or not isinstance(value, (int, float))
                or not math.isfinite(float(value))
                or float(value) < 0
            ):
                invalid_query_ids.append(row["query_id"])
                continue
            values.append(float(value))
        status = (
            "NOT_CONNECTED"
            if not values
            else "COMPLETE"
            if len(values) == len(query_rows)
            else "PARTIAL"
        )
        return {
            "status": status,
            "task_count": len(query_rows),
            "value_count": len(values),
            "missing_count": len(missing_query_ids),
            "invalid_count": len(invalid_query_ids),
            "total": sum(values) if values else None,
            "average": sum(values) / len(values) if values else None,
            "minimum": min(values) if values else None,
            "maximum": max(values) if values else None,
            "missing_query_ids": missing_query_ids,
            "invalid_query_ids": invalid_query_ids,
        }

    @staticmethod
    def _metrics_v2_pdl(
        query_rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """统计 PDL 的 true、false、未知和只基于已知值的调用率。"""

        true_count = 0
        false_count = 0
        invalid_query_ids: list[str] = []
        unknown_query_ids: list[str] = []
        for row in query_rows:
            value = row["task_fields"].get("pdl_called")
            if value is True:
                true_count += 1
            elif value is False:
                false_count += 1
            else:
                unknown_query_ids.append(row["query_id"])
                has_processing_error = any(
                    isinstance(error, dict)
                    and error.get("field_key") == "pdl_called"
                    for error in row.get("task_processing_errors", [])
                )
                if value not in (None, "") or has_processing_error:
                    invalid_query_ids.append(row["query_id"])
        known_count = true_count + false_count
        return {
            "true_count": true_count,
            "false_count": false_count,
            "unknown_count": len(unknown_query_ids),
            "known_count": known_count,
            "call_rate": true_count / known_count if known_count else None,
            "unknown_query_ids": unknown_query_ids,
            "invalid_query_ids": invalid_query_ids,
        }

    @staticmethod
    def _metrics_v2_confidence(
        candidate_rows: list[dict[str, Any]],
    ) -> dict[str, dict[str, int]]:
        """统计 Candidate Confidence，保留未知新枚举的原始字符串。"""

        def label(row: dict[str, Any]) -> str:
            value = row.get("confidence")
            return value.strip() if isinstance(value, str) and value.strip() else "UNKNOWN"

        def distribution(rows: list[dict[str, Any]]) -> dict[str, int]:
            result: dict[str, int] = {}
            for row in rows:
                key = label(row)
                result[key] = result.get(key, 0) + 1
            return dict(sorted(result.items()))

        return {
            "overall": distribution(candidate_rows),
            "matched": distribution(
                [
                    row
                    for row in candidate_rows
                    if row["reviewed_at"] and row["judgement"] == "HIT"
                ]
            ),
            "nonmatched": distribution(
                [
                    row
                    for row in candidate_rows
                    if row["reviewed_at"]
                    and row["judgement"] in {"NOT_HIT", "SUSPECTED"}
                ]
            ),
        }

    def _metrics_v2_quality(
        self,
        query_rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """从人物级明细聚合正式质量指标，并保留各自独立就绪状态。"""

        retrieval_reasons = [
            reason
            for row in query_rows
            for reason in row["retrieval_not_ready_reasons"]
        ]
        retrieval_ready = bool(query_rows) and not retrieval_reasons
        hit_rows = [row for row in query_rows if row["retrieval_success"]]
        completeness_reasons = [
            reason
            for row in hit_rows
            for reason in row["matched_completeness_not_ready_reasons"]
        ]
        accuracy_reasons = [
            reason
            for row in hit_rows
            for reason in row["matched_accuracy_not_ready_reasons"]
        ]
        completeness_values = [
            float(row["matched_completeness"])
            for row in hit_rows
            if row["matched_completeness"] is not None
        ]
        accuracy_values = [
            float(row["matched_accuracy"])
            for row in hit_rows
            if row["matched_accuracy"] is not None
        ]
        nonmatched_values = [
            float(value)
            for row in query_rows
            for value in row["nonmatched_completeness_values"]
        ]
        return {
            "formal_ready": retrieval_ready and not completeness_reasons and not accuracy_reasons,
            "retrieval_success": self._metrics_v2_value(
                float(sum(row["retrieval_success"] for row in query_rows)),
                len(query_rows),
                ready=retrieval_ready,
                not_ready_reasons=retrieval_reasons,
            ),
            "matched_completeness": self._metrics_v2_value(
                sum(completeness_values),
                len(hit_rows),
                ready=not completeness_reasons,
                not_ready_reasons=completeness_reasons,
            ),
            "matched_accuracy": self._metrics_v2_value(
                sum(accuracy_values),
                len(accuracy_values),
                ready=not accuracy_reasons,
                not_ready_reasons=accuracy_reasons,
            ),
            "nonmatched_completeness": self._metrics_v2_value(
                sum(nonmatched_values),
                len(nonmatched_values),
                ready=True,
            ),
        }

    def _calculate_process_metrics_v2(self, process_id: str) -> dict[str, Any]:
        """按 metrics-v2 计算结果、质量、任务字段、置信度和阶段分组。

        所有 Task 值来自 ``processed_queries`` 不可变快照；完整度分母来自
        每个人物的 ``baseline_available_fields``。依赖缺失时保留预览和
        原因，但正式值保持空。
        """

        process = self.store.fetch_one(
            """
            SELECT pr.*, r.evaluation_id, r.run_label, r.system_version,
                   r.evaluation_phase, fs.definitions_json
            FROM process_runs AS pr
            JOIN runs AS r ON r.run_id = pr.run_id
            JOIN field_schemas AS fs
              ON fs.schema_version = pr.schema_version
            WHERE pr.process_id = ?
            """,
            (process_id,),
        )
        if process is None:
            raise ReviewValidationError(f"处理结果不存在: {process_id}")
        try:
            definitions = validate_field_definitions(
                json.loads(process["definitions_json"])
            )
        except (json.JSONDecodeError, FieldSchemaValidationError) as exc:
            raise ReviewValidationError(f"字段配置快照已损坏: {exc}") from exc
        definitions_by_key = {
            item["field_key"]: item
            for item in definitions
            if item["enabled"] and item["value_scope"] == "CANDIDATE"
        }
        completeness_keys = {
            key
            for key, item in definitions_by_key.items()
            if "completeness" in item["scoring_role"]
        }
        accuracy_keys = {
            key
            for key, item in definitions_by_key.items()
            if "accuracy" in item["scoring_role"]
        }

        baseline_people: dict[str, dict[str, Any]] = {}
        if process["baseline_version"]:
            for row in self.store.fetch_all(
                """
                SELECT person_id, available_fields_json
                FROM baseline_people
                WHERE baseline_version = ?
                """,
                (process["baseline_version"],),
            ):
                try:
                    available_fields = json.loads(row["available_fields_json"])
                except (TypeError, json.JSONDecodeError):
                    available_fields = []
                baseline_people[row["person_id"]] = {
                    "available_fields": (
                        available_fields
                        if isinstance(available_fields, list)
                        else []
                    )
                }

        raw_queries = self.store.fetch_all(
            """
            SELECT rq.query_id, rq.person_id, rq.query_stage,
                   rq.candidate_count_listed, rq.detail_failure_count,
                   pq.result_status, pq.fields_json,
                   pq.empty_fields_json, pq.processing_errors_json
            FROM run_queries AS rq
            JOIN processed_queries AS pq
              ON pq.run_id = rq.run_id AND pq.query_id = rq.query_id
             AND pq.process_id = ?
            WHERE rq.run_id = ?
            ORDER BY rq.query_id
            """,
            (process_id, process["run_id"]),
        )
        query_rows: list[dict[str, Any]] = []
        for row in raw_queries:
            if row["result_status"] not in RESULT_STATUSES:
                raise ReviewValidationError(
                    f"Query {row['query_id']} 结果状态无效: {row['result_status']}"
                )
            item = dict(row)
            try:
                task_fields = json.loads(row["fields_json"])
            except (TypeError, json.JSONDecodeError):
                task_fields = {}
            item["task_fields"] = (
                task_fields if isinstance(task_fields, dict) else {}
            )
            try:
                task_processing_errors = json.loads(
                    row["processing_errors_json"]
                )
            except (TypeError, json.JSONDecodeError):
                task_processing_errors = []
            item["task_processing_errors"] = (
                task_processing_errors
                if isinstance(task_processing_errors, list)
                else []
            )
            query_rows.append(item)

        raw_candidates = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.query_id, c.candidate_id,
                   c.candidate_rank, c.detail_status,
                   pc.fields_json, pc.empty_fields_json,
                   rv.judgement, rv.field_scores_json, rv.reviewed_at
            FROM candidates AS c
            JOIN processed_candidates AS pc
              ON pc.candidate_pk = c.candidate_pk AND pc.process_id = ?
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE c.run_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process_id, process["run_id"]),
        )
        candidate_rows: list[dict[str, Any]] = []
        candidates_by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in raw_candidates:
            item = dict(row)
            for source_name, target_name in (
                ("fields_json", "fields"),
                ("empty_fields_json", "empty_fields"),
                ("field_scores_json", "field_scores"),
            ):
                try:
                    value = (
                        json.loads(row[source_name])
                        if row[source_name]
                        else {}
                    )
                except (TypeError, json.JSONDecodeError):
                    value = {}
                item[target_name] = value if isinstance(value, dict) else {}
            confidence = item["fields"].get("candidate_confidence")
            if not isinstance(confidence, str) or not confidence.strip():
                confidence = item["fields"].get("summary_confidence_level")
            item["confidence"] = confidence
            candidate_rows.append(item)
            candidates_by_query[item["query_id"]].append(item)

        warnings: list[str] = []
        for query in query_rows:
            all_candidates = candidates_by_query.get(query["query_id"], [])
            successful_candidates = [
                item
                for item in all_candidates
                if item["detail_status"] == "SUCCESS"
            ]
            reviewed_candidates = [
                item for item in successful_candidates if item["reviewed_at"]
            ]
            hit_candidates = [
                item
                for item in reviewed_candidates
                if item["judgement"] == "HIT"
            ]
            selected_hit = hit_candidates[0] if hit_candidates else None
            baseline = baseline_people.get(query["person_id"])
            available_fields = (
                list(baseline["available_fields"]) if baseline else []
            )
            retrieval_reasons: list[str] = []
            if baseline is None:
                retrieval_reasons.append(
                    f"{query['query_id']}: 未关联该人物 Baseline"
                )
            if query["result_status"] == "HAS_CANDIDATES":
                pending_count = len(successful_candidates) - len(reviewed_candidates)
                if pending_count:
                    retrieval_reasons.append(
                        f"{query['query_id']}: {pending_count} 个候选人待复核"
                    )
                if not all_candidates:
                    retrieval_reasons.append(
                        f"{query['query_id']}: HAS_CANDIDATES 但没有候选人"
                    )
            else:
                pending_count = 0

            completeness_numerator = 0.0
            completeness_denominator = len(available_fields)
            completeness_missing: list[str] = []
            accuracy_numerator = 0.0
            accuracy_denominator = 0
            accuracy_missing: list[str] = []
            if selected_hit is not None:
                scores = selected_hit["field_scores"]
                for field_key in available_fields:
                    score = scores.get(field_key)
                    if field_key not in completeness_keys or not isinstance(score, dict):
                        completeness_missing.append(field_key)
                        continue
                    value = score.get("completeness_score")
                    if value is None:
                        completeness_missing.append(field_key)
                    else:
                        completeness_numerator += float(value)
                    if field_key not in accuracy_keys or not score.get(
                        "returned_nonempty"
                    ):
                        continue
                    accuracy_denominator += 1
                    accuracy_value = score.get("accuracy_score")
                    if accuracy_value is None:
                        accuracy_missing.append(field_key)
                    else:
                        accuracy_numerator += float(accuracy_value)
                if not available_fields:
                    completeness_missing.append("baseline_available_fields 为空")
                if completeness_missing:
                    warnings.append(
                        f"{query['query_id']}: 完整度字段未就绪 "
                        + ", ".join(completeness_missing)
                    )

            nonmatched_values: list[float] = []
            for candidate in reviewed_candidates:
                if candidate["judgement"] not in {"NOT_HIT", "SUSPECTED"}:
                    continue
                nonempty_count = sum(
                    field_key in candidate["fields"]
                    and not candidate["empty_fields"].get(field_key, True)
                    for field_key in completeness_keys
                )
                if completeness_keys:
                    nonmatched_values.append(
                        nonempty_count / len(completeness_keys)
                    )

            query.update(
                {
                    "candidate_count": len(all_candidates),
                    "successful_candidate_count": len(successful_candidates),
                    "pending_review_count": pending_count,
                    "retrieval_success": bool(selected_hit),
                    "retrieval_not_ready_reasons": retrieval_reasons,
                    "baseline_available_fields": available_fields,
                    "selected_hit_candidate_id": (
                        selected_hit["candidate_pk"] if selected_hit else None
                    ),
                    "matched_completeness_numerator": completeness_numerator,
                    "matched_completeness_denominator": completeness_denominator,
                    "matched_completeness": (
                        completeness_numerator / completeness_denominator
                        if selected_hit
                        and completeness_denominator
                        and not completeness_missing
                        else None
                    ),
                    "matched_completeness_not_ready_reasons": [
                        f"{query['query_id']}: {field_key}"
                        for field_key in completeness_missing
                    ],
                    "matched_accuracy_numerator": accuracy_numerator,
                    "matched_accuracy_denominator": accuracy_denominator,
                    "matched_accuracy": (
                        accuracy_numerator / accuracy_denominator
                        if selected_hit
                        and accuracy_denominator
                        and not accuracy_missing
                        else None
                    ),
                    "matched_accuracy_not_ready_reasons": [
                        f"{query['query_id']}: {field_key}"
                        for field_key in accuracy_missing
                    ],
                    "nonmatched_completeness_values": nonmatched_values,
                    "candidate_confidence": [
                        item["confidence"]
                        for item in all_candidates
                    ],
                    "formal_ready": (
                        not retrieval_reasons
                        and (
                            selected_hit is None
                            or (
                                not completeness_missing
                                and not accuracy_missing
                            )
                        )
                    ),
                    "hit_candidate_ids": [
                        item["candidate_pk"] for item in hit_candidates
                    ],
                    "nonmatched_candidate_ids": [
                        item["candidate_pk"]
                        for item in reviewed_candidates
                        if item["judgement"] in {"NOT_HIT", "SUSPECTED"}
                    ],
                    "candidate_ids": [
                        item["candidate_pk"] for item in successful_candidates
                    ],
                }
            )

        quality = self._metrics_v2_quality(query_rows)
        cost_metrics = {
            field_key: self._metrics_v2_numeric_aggregate(
                query_rows,
                field_key,
            )
            for field_key in (
                "llm_cost",
                "third_party_cost",
                "total_cost",
                "search_duration_ms",
            )
        }
        pdl_metrics = self._metrics_v2_pdl(query_rows)
        confidence_metrics = self._metrics_v2_confidence(candidate_rows)
        result_status_metrics = self._metrics_v2_result_status(query_rows)

        grouped_metrics: list[dict[str, Any]] = []
        for query_stage in sorted(
            {row["query_stage"] or "UNSPECIFIED" for row in query_rows}
        ):
            grouped_queries = [
                row
                for row in query_rows
                if (row["query_stage"] or "UNSPECIFIED") == query_stage
            ]
            grouped_query_ids = {
                row["query_id"] for row in grouped_queries
            }
            grouped_candidates = [
                row
                for row in candidate_rows
                if row["query_id"] in grouped_query_ids
            ]
            grouped_metrics.append(
                {
                    "evaluation_phase": process["evaluation_phase"],
                    "system_version": process["system_version"],
                    "query_stage": query_stage,
                    "query_count": len(grouped_queries),
                    "result_status_metrics": self._metrics_v2_result_status(
                        grouped_queries
                    ),
                    "quality_metrics": self._metrics_v2_quality(
                        grouped_queries
                    ),
                    "cost_metrics": {
                        field_key: self._metrics_v2_numeric_aggregate(
                            grouped_queries,
                            field_key,
                        )
                        for field_key in cost_metrics
                    },
                    "pdl_metrics": self._metrics_v2_pdl(grouped_queries),
                    "confidence_metrics": self._metrics_v2_confidence(
                        grouped_candidates
                    ),
                }
            )

        cost_status_values = [
            item["status"] for item in cost_metrics.values()
        ]
        cost_status = {
            "status": (
                "NOT_CONNECTED"
                if all(status == "NOT_CONNECTED" for status in cost_status_values)
                else "COMPLETE"
                if all(status == "COMPLETE" for status in cost_status_values)
                else "PARTIAL"
            ),
            "task_count": len(query_rows),
            "missing_task_count": sum(
                any(
                    row["task_fields"].get(field_key) in (None, "")
                    for field_key in cost_metrics
                )
                for row in query_rows
            ),
            "llm_cost_total": cost_metrics["llm_cost"]["total"],
            "total_cost_total": cost_metrics["total_cost"]["total"],
            "pdl_called_total": (
                pdl_metrics["true_count"]
                if pdl_metrics["known_count"]
                else None
            ),
        }
        module_metrics, field_metrics = self._process_field_report_metrics(
            process
        )
        return {
            "process_id": process_id,
            "run_id": process["run_id"],
            "evaluation_id": process["evaluation_id"],
            "evaluation_phase": process["evaluation_phase"],
            "system_version": process["system_version"],
            "schema_version": process["schema_version"],
            "baseline_version": process["baseline_version"],
            "rule_version": process["rule_version"],
            "metrics_rule_version": V2_METRICS_RULE_VERSION,
            "formal_ready": quality["formal_ready"],
            "review_status": (
                "REVIEWED" if quality["formal_ready"] else "PENDING_REVIEW"
            ),
            "result_status_metrics": result_status_metrics,
            "quality_metrics": quality,
            "retrieval_success": quality["retrieval_success"],
            "matched_completeness": quality["matched_completeness"],
            "matched_accuracy": quality["matched_accuracy"],
            "nonmatched_completeness": quality[
                "nonmatched_completeness"
            ],
            "cost_metrics": cost_metrics,
            "pdl_metrics": pdl_metrics,
            "confidence_metrics": confidence_metrics,
            "grouped_metrics": grouped_metrics,
            "module_metrics": module_metrics,
            "field_metrics": field_metrics,
            "cost_status": cost_status,
            "query_metrics": query_rows,
            "warnings": sorted(set(warnings)),
        }

    @staticmethod
    def _metrics_v3_value(
        numerator: float,
        denominator: int,
        *,
        status: str | None = None,
        reason_codes: Iterable[str] = (),
        reasons: Iterable[str] = (),
    ) -> dict[str, Any]:
        """构造 metrics-v3 的统一指标状态、正式值和可解释原因。

        ``PARTIAL`` 可以展示已完成部分的正式值；``NOT_READY`` 只展示
        preview，避免把未完成身份归类或人工评分当成最终结论。
        """

        preview_value = numerator / denominator if denominator else None
        resolved_status = status or (
            "READY" if denominator else "NOT_APPLICABLE"
        )
        return {
            "numerator": numerator,
            "denominator": denominator,
            "value": (
                preview_value
                if resolved_status in {"READY", "PARTIAL"} and denominator
                else None
            ),
            "preview_value": preview_value,
            "status": resolved_status,
            "reason_codes": sorted(set(reason_codes)),
            "reasons": sorted(set(reasons)),
            # 兼容旧页面和对比代码读取的 v2 字段名。
            "not_ready_reasons": sorted(set(reasons)),
        }

    def _metrics_v3_field_returns(
        self,
        definitions: list[dict[str, Any]],
        query_rows: list[dict[str, Any]],
        candidate_rows: list[dict[str, Any]],
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        """按 value_scope 分别计算 Query/Candidate 字段返回率。"""

        successful_candidates = [
            row for row in candidate_rows
            if row["detail_status"] == "SUCCESS"
        ]
        field_metrics: dict[str, Any] = {}
        for definition in definitions:
            if not definition["enabled"]:
                continue
            rows = (
                query_rows
                if definition["value_scope"] == "QUERY"
                else successful_candidates
            )
            field_key = definition["field_key"]
            returned_count = sum(
                field_key in row["fields"]
                and not row["empty_fields"].get(field_key, True)
                for row in rows
            )
            entity_count = len(rows)
            status = "READY" if entity_count else "NOT_APPLICABLE"
            field_metrics[field_key] = {
                "field_key": field_key,
                "display_name": definition["display_name"],
                "module": definition["module"],
                "value_scope": definition["value_scope"],
                "scoring_role": definition["scoring_role"],
                "compare_mode": definition["compare_mode"],
                "entity_count": entity_count,
                "returned_count": returned_count,
                "empty_count": entity_count - returned_count,
                "candidate_count": entity_count,
                "return_rate": (
                    returned_count / entity_count if entity_count else None
                ),
                "status": status,
                "reason_codes": (
                    [] if entity_count else ["NO_DENOMINATOR"]
                ),
                "reasons": (
                    []
                    if entity_count
                    else ["当前 Process 没有该作用域的有效实体"]
                ),
            }
        candidate_fields = [
            item for item in field_metrics.values()
            if item["value_scope"] == "CANDIDATE"
        ]
        returned_total = sum(
            item["returned_count"] for item in candidate_fields
        )
        denominator = len(successful_candidates) * len(candidate_fields)
        candidate_overall = self._metrics_v3_value(
            float(returned_total),
            denominator,
            status="READY" if denominator else "NOT_APPLICABLE",
            reason_codes=[] if denominator else ["NO_DENOMINATOR"],
            reasons=(
                []
                if denominator
                else ["没有详情成功候选人或已启用 Candidate 字段"]
            ),
        )
        candidate_overall.update({
            "field_count": len(candidate_fields),
            "candidate_count": len(successful_candidates),
        })
        return field_metrics, candidate_overall

    @staticmethod
    def _metrics_v3_modules(
        definitions: list[dict[str, Any]],
        candidate_rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """按 processing-v3 模块状态和核心内容计算 data/empty/unknown。"""

        successful = [
            row for row in candidate_rows
            if row["detail_status"] == "SUCCESS"
        ]
        modules = ("Insights", "Photos", "Profile", "Social", "Summary")
        definitions_by_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for definition in definitions:
            if (
                definition["enabled"]
                and definition["value_scope"] == "CANDIDATE"
                and definition["module"] in modules
            ):
                definitions_by_module[definition["module"]].append(definition)
        result: dict[str, Any] = {}
        for module in modules:
            module_key = module.lower()
            module_definitions = definitions_by_module.get(module, [])
            status_key = f"{module_key}_status"
            content_keys = [
                item["field_key"]
                for item in module_definitions
                if item["field_key"] != status_key
            ]
            data_count = 0
            empty_count = 0
            unknown_count = 0
            for row in successful:
                status_value = row["fields"].get(status_key)
                status = (
                    status_value.strip().lower()
                    if isinstance(status_value, str)
                    else ""
                )
                content_has_value = any(
                    field_key in row["fields"]
                    and not row["empty_fields"].get(field_key, True)
                    for field_key in content_keys
                )
                if module == "Summary":
                    if content_has_value:
                        data_count += 1
                    else:
                        empty_count += 1
                elif status == "empty":
                    empty_count += 1
                elif status == "data":
                    if content_has_value:
                        data_count += 1
                    else:
                        empty_count += 1
                else:
                    unknown_count += 1
            candidate_count = len(successful)
            result[module] = {
                "module": module,
                "data_count": data_count,
                "empty_count": empty_count,
                "unknown_count": unknown_count,
                "candidate_count": candidate_count,
                "data_rate": (
                    data_count / candidate_count if candidate_count else None
                ),
                # 兼容现有报告模板。
                "returned_candidate_count": data_count,
                "return_rate": (
                    data_count / candidate_count if candidate_count else None
                ),
                "status": (
                    "PARTIAL"
                    if unknown_count
                    else "READY" if candidate_count else "NOT_APPLICABLE"
                ),
                "reason_codes": (
                    ["MODULE_STATUS_UNKNOWN"] if unknown_count else []
                ),
                "reasons": (
                    [f"{module} 有 {unknown_count} 个候选人模块状态未知"]
                    if unknown_count else []
                ),
            }
        return result

    @staticmethod
    def _metrics_v3_confidence(
        candidate_rows: list[dict[str, Any]],
    ) -> dict[str, dict[str, int]]:
        """按正式身份分类拆分 Confidence，建议状态只进入 pending。"""

        def label(row: dict[str, Any]) -> str:
            value = row.get("confidence")
            return (
                value.strip()
                if isinstance(value, str) and value.strip()
                else "UNKNOWN"
            )

        def distribution(rows: Iterable[dict[str, Any]]) -> dict[str, int]:
            values: dict[str, int] = {}
            for row in rows:
                key = label(row)
                values[key] = values.get(key, 0) + 1
            return dict(sorted(values.items()))

        final_rows = [
            row for row in candidate_rows
            if row["reviewed_at"]
            and row["classification_source"] in {"MANUAL", "RULE"}
        ]
        primary = [
            row for row in final_rows
            if row["judgement"] == "HIT" and row["is_primary_hit"]
        ]
        all_hit = [
            row for row in final_rows if row["judgement"] == "HIT"
        ]
        not_hit = [
            row for row in final_rows if row["judgement"] == "NOT_HIT"
        ]
        suspected = [
            row for row in final_rows if row["judgement"] == "SUSPECTED"
        ]
        final_pks = {row["candidate_pk"] for row in final_rows}
        pending = [
            row for row in candidate_rows
            if row["candidate_pk"] not in final_pks
        ]
        nonmatched = [*not_hit, *suspected]
        return {
            "overall": distribution(candidate_rows),
            "all_hit": distribution(all_hit),
            "primary_hit": distribution(primary),
            "not_hit": distribution(not_hit),
            "suspected": distribution(suspected),
            "pending": distribution(pending),
            # 兼容旧模板列名。
            "matched": distribution(primary),
            "nonmatched": distribution(nonmatched),
        }

    def _metrics_v3_quality(
        self,
        query_rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """聚合 Query 级身份与字段质量，并保持各指标独立状态。"""

        pending_rows = [
            row for row in query_rows
            if row["identity_state"] == "PENDING"
        ]
        confirmed_rows = [
            row for row in query_rows
            if row["identity_state"]
            in {"HIT_CONFIRMED", "NO_HIT_CONFIRMED"}
        ]
        hit_rows = [
            row for row in confirmed_rows
            if row["identity_state"] == "HIT_CONFIRMED"
        ]
        executed_rows = [
            row for row in query_rows
            if row.get("result_status") != "EXECUTION_FAILED"
        ]
        candidate_return_rows = [
            row for row in executed_rows
            if int(row.get("candidate_count") or 0) > 0
        ]
        retrieval_codes = (
            ["IDENTITY_PENDING"] if pending_rows else []
        )
        retrieval_reasons = [
            f"{row['query_id']} 有 {row['pending_review_count']} 个候选人待身份归类"
            for row in pending_rows
        ]
        retrieval = self._metrics_v3_value(
            float(len(hit_rows)),
            len(executed_rows),
            status=(
                "NOT_READY"
                if pending_rows
                else "READY" if executed_rows else "NOT_APPLICABLE"
            ),
            reason_codes=(
                retrieval_codes
                if pending_rows
                else ([] if executed_rows else ["NO_DENOMINATOR"])
            ),
            reasons=(
                retrieval_reasons
                if pending_rows
                else ([] if executed_rows else ["没有成功执行的 Query"])
            ),
        )
        candidate_return_rate = self._metrics_v3_value(
            float(len(candidate_return_rows)),
            len(executed_rows),
            status="READY" if executed_rows else "NOT_APPLICABLE",
            reason_codes=[] if executed_rows else ["NO_DENOMINATOR"],
            reasons=[] if executed_rows else ["没有成功执行的 Query"],
        )
        conditional_hit_rate = self._metrics_v3_value(
            float(len(hit_rows)),
            len(candidate_return_rows),
            status=(
                "NOT_READY" if pending_rows
                else "READY" if candidate_return_rows else "NOT_APPLICABLE"
            ),
            reason_codes=(
                ["IDENTITY_PENDING"] if pending_rows
                else ([] if candidate_return_rows else ["NO_CANDIDATE_QUERY"])
            ),
            reasons=(
                retrieval_reasons if pending_rows
                else ([] if candidate_return_rows else ["没有返回候选人的 Query"])
            ),
        )

        completeness_values = [
            row["matched_completeness"]
            for row in hit_rows
            if row["matched_completeness"] is not None
        ]
        completeness_codes = [
            code
            for row in hit_rows
            for code in row["matched_completeness_reason_codes"]
        ]
        completeness_reasons = [
            reason
            for row in hit_rows
            for reason in row["matched_completeness_reasons"]
        ]
        if pending_rows:
            completeness_codes.append("IDENTITY_PENDING")
            completeness_reasons.extend(retrieval_reasons)
            completeness_codes.extend(
                code
                for row in pending_rows
                for code in row["matched_completeness_reason_codes"]
                if code in {
                    "BASELINE_NOT_LINKED",
                    "BASELINE_PERSON_NOT_FOUND",
                    "NO_EFFECTIVE_FIELDS",
                }
            )
            completeness_reasons.extend(
                reason
                for row in pending_rows
                for reason in row["matched_completeness_reasons"]
            )
        if not hit_rows:
            if pending_rows:
                completeness_status = "NOT_READY"
            else:
                completeness_status = "NOT_APPLICABLE"
                completeness_codes.append("NO_HIT_CONFIRMED")
                completeness_reasons.append(
                    "已完成身份归类，但没有主要 HIT 候选人"
                )
        elif completeness_codes:
            completeness_status = (
                "PARTIAL" if completeness_values else "NOT_READY"
            )
        else:
            completeness_status = "READY"
        matched_completeness = self._metrics_v3_value(
            sum(
                float(row["matched_completeness_numerator"])
                for row in hit_rows
            ),
            sum(
                int(row["matched_completeness_denominator"])
                for row in hit_rows
            ),
            status=completeness_status,
            reason_codes=completeness_codes,
            reasons=completeness_reasons,
        )

        accuracy_values = [
            row["matched_accuracy"]
            for row in hit_rows
            if row["matched_accuracy"] is not None
        ]
        accuracy_codes = [
            code
            for row in hit_rows
            for code in row["matched_accuracy_reason_codes"]
        ]
        accuracy_reasons = [
            reason
            for row in hit_rows
            for reason in row["matched_accuracy_reasons"]
        ]
        if pending_rows:
            accuracy_codes.append("IDENTITY_PENDING")
            accuracy_reasons.extend(retrieval_reasons)
            accuracy_codes.extend(
                code
                for row in pending_rows
                for code in row["matched_accuracy_reason_codes"]
                if code in {
                    "BASELINE_NOT_LINKED",
                    "BASELINE_PERSON_NOT_FOUND",
                    "NO_EFFECTIVE_FIELDS",
                }
            )
            accuracy_reasons.extend(
                reason
                for row in pending_rows
                for reason in row["matched_accuracy_reasons"]
            )
        if not hit_rows:
            accuracy_status = (
                "NOT_READY" if pending_rows else "NOT_APPLICABLE"
            )
            if not pending_rows:
                accuracy_codes.append("NO_HIT_CONFIRMED")
                accuracy_reasons.append(
                    "已完成身份归类，但没有主要 HIT 候选人"
                )
        elif accuracy_codes:
            accuracy_status = "PARTIAL" if accuracy_values else "NOT_READY"
        elif not accuracy_values:
            accuracy_status = "NOT_APPLICABLE"
            accuracy_codes.append("NO_DENOMINATOR")
            accuracy_reasons.append("主要 HIT 没有非空且可比较的准确率字段")
        else:
            accuracy_status = "READY"
        matched_accuracy = self._metrics_v3_value(
            sum(float(value) for value in accuracy_values),
            len(accuracy_values),
            status=accuracy_status,
            reason_codes=accuracy_codes,
            reasons=accuracy_reasons,
        )

        nonmatched_data_items = [
            item
            for row in query_rows
            for item in row.get("nonmatched_data_completeness_items", [])
        ]
        nonmatched_data_values = [
            float(item["value"]) for item in nonmatched_data_items
        ] or [
            float(value)
            for row in query_rows
            for value in row["nonmatched_data_completeness_values"]
        ]
        if nonmatched_data_values:
            nonmatched_status = "PARTIAL" if pending_rows else "READY"
            nonmatched_codes = (
                ["IDENTITY_PENDING"] if pending_rows else []
            )
            nonmatched_reasons = retrieval_reasons if pending_rows else []
        elif pending_rows:
            nonmatched_status = "NOT_READY"
            nonmatched_codes = ["IDENTITY_PENDING"]
            nonmatched_reasons = retrieval_reasons
        else:
            nonmatched_status = "NOT_APPLICABLE"
            nonmatched_codes = ["NO_NONMATCHED_CONFIRMED"]
            nonmatched_reasons = ["尚无已确认的 NOT_HIT/SUSPECTED 候选人"]
        nonmatched_data_completeness = self._metrics_v3_value(
            (
                sum(float(item["numerator"]) for item in nonmatched_data_items)
                if nonmatched_data_items else sum(nonmatched_data_values)
            ),
            (
                sum(int(item["denominator"]) for item in nonmatched_data_items)
                if nonmatched_data_items else len(nonmatched_data_values)
            ),
            status=nonmatched_status,
            reason_codes=nonmatched_codes,
            reasons=nonmatched_reasons,
        )
        nonmatched_overlap_items = [
            item
            for row in query_rows
            for item in row.get("nonmatched_baseline_overlap_items", [])
        ]
        nonmatched_overlap_values = [
            float(item["value"]) for item in nonmatched_overlap_items
        ] or [
            float(value)
            for row in query_rows
            for value in row["nonmatched_baseline_overlap_values"]
        ]
        if nonmatched_overlap_values:
            overlap_status = "PARTIAL" if pending_rows else "READY"
            overlap_codes = ["IDENTITY_PENDING"] if pending_rows else []
            overlap_reasons = retrieval_reasons if pending_rows else []
        elif pending_rows:
            overlap_status = "NOT_READY"
            overlap_codes = ["IDENTITY_PENDING"]
            overlap_reasons = retrieval_reasons
        else:
            overlap_status = "NOT_APPLICABLE"
            overlap_codes = ["NO_BASELINE_OVERLAP_FIELDS"]
            overlap_reasons = ["没有可比较的非命中候选人资料字段"]
        nonmatched_baseline_overlap = self._metrics_v3_value(
            (
                sum(float(item["numerator"]) for item in nonmatched_overlap_items)
                if nonmatched_overlap_items else sum(nonmatched_overlap_values)
            ),
            (
                sum(int(item["denominator"]) for item in nonmatched_overlap_items)
                if nonmatched_overlap_items else len(nonmatched_overlap_values)
            ),
            status=overlap_status,
            reason_codes=overlap_codes,
            reasons=overlap_reasons,
        )
        formal_ready = (
            retrieval["status"] == "READY"
            and matched_completeness["status"]
            in {"READY", "NOT_APPLICABLE"}
            and matched_accuracy["status"]
            in {"READY", "NOT_APPLICABLE"}
            and nonmatched_data_completeness["status"]
            in {"READY", "NOT_APPLICABLE"}
            and nonmatched_baseline_overlap["status"]
            in {"READY", "NOT_APPLICABLE"}
        )
        return {
            "formal_ready": formal_ready,
            "candidate_return_rate": candidate_return_rate,
            "conditional_hit_rate": conditional_hit_rate,
            "retrieval_success": retrieval,
            "matched_completeness": matched_completeness,
            "matched_accuracy": matched_accuracy,
            # 旧键保留为资料完整度别名，确保旧报告与 Excel 调用方兼容。
            "nonmatched_completeness": nonmatched_data_completeness,
            "nonmatched_data_completeness": nonmatched_data_completeness,
            "nonmatched_baseline_overlap": nonmatched_baseline_overlap,
        }

    def _calculate_process_metrics_v3(self, process_id: str) -> dict[str, Any]:
        """按 metrics-v3 独立计算执行、数据返回、身份和字段质量指标。"""

        process = self.store.fetch_one(
            """
            SELECT pr.*, r.evaluation_id, r.run_label, r.system_version,
                   r.evaluation_phase, fs.definitions_json
            FROM process_runs AS pr
            JOIN runs AS r ON r.run_id = pr.run_id
            JOIN field_schemas AS fs
              ON fs.schema_version = pr.schema_version
            WHERE pr.process_id = ?
            """,
            (process_id,),
        )
        if process is None:
            raise ReviewValidationError(f"处理结果不存在: {process_id}")
        try:
            definitions = validate_field_definitions(
                json.loads(process["definitions_json"])
            )
        except (TypeError, json.JSONDecodeError, FieldSchemaValidationError) as exc:
            raise ReviewValidationError(f"字段配置快照已损坏: {exc}") from exc
        candidate_definitions = {
            item["field_key"]: item
            for item in definitions
            if item["enabled"] and item["value_scope"] == "CANDIDATE"
        }
        completeness_keys = {
            key for key, item in candidate_definitions.items()
            if item["completeness_enabled"]
        }
        accuracy_keys = {
            key for key, item in candidate_definitions.items()
            # 准确率必须比较 Baseline 与 Candidate；只打开准确度开关但关闭
            # 基准对比的字段不进入准确率分母，也不能被误报为“待人工评分”。
            if "accuracy" in item["scoring_role"]
            and item["baseline_compare_enabled"]
        }

        baseline_people: dict[str, dict[str, Any]] = {}
        if process["baseline_version"]:
            for row in self.store.fetch_all(
                """
                SELECT person_id, fields_json, available_fields_json
                FROM baseline_people WHERE baseline_version = ?
                """,
                (process["baseline_version"],),
            ):
                try:
                    fields = json.loads(row["fields_json"] or "{}")
                    available = json.loads(
                        row["available_fields_json"] or "[]"
                    )
                except (TypeError, json.JSONDecodeError):
                    fields, available = {}, []
                baseline_people[row["person_id"]] = {
                    "fields": fields if isinstance(fields, dict) else {},
                    "available_fields": (
                        available if isinstance(available, list) else []
                    ),
                }

        raw_queries = self.store.fetch_all(
            """
            SELECT rq.query_id, rq.person_id, rq.person_id_source,
                   rq.query_stage, rq.candidate_count_listed,
                   rq.detail_success_count, rq.detail_failure_count,
                   pq.result_status, pq.fields_json,
                   pq.empty_fields_json, pq.processing_errors_json
            FROM run_queries AS rq
            JOIN processed_queries AS pq
              ON pq.run_id = rq.run_id AND pq.query_id = rq.query_id
             AND pq.process_id = ?
            WHERE rq.run_id = ? ORDER BY rq.query_id
            """,
            (process_id, process["run_id"]),
        )
        query_rows: list[dict[str, Any]] = []
        for row in raw_queries:
            item = dict(row)
            try:
                item["fields"] = json.loads(row["fields_json"] or "{}")
                item["empty_fields"] = json.loads(
                    row["empty_fields_json"] or "{}"
                )
                item["task_processing_errors"] = json.loads(
                    row["processing_errors_json"] or "[]"
                )
            except (TypeError, json.JSONDecodeError):
                item["fields"], item["empty_fields"] = {}, {}
                item["task_processing_errors"] = []
            item["task_fields"] = item["fields"]
            query_rows.append(item)

        raw_candidates = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.query_id, c.candidate_id,
                   c.candidate_rank, c.rank_score, c.detail_status,
                   c.detail_error, pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json, rv.judgement, rv.reason,
                   rv.field_scores_json, rv.reviewed_at,
                   rv.classification_source, rv.is_primary_hit
            FROM candidates AS c
            JOIN processed_candidates AS pc
              ON pc.candidate_pk = c.candidate_pk AND pc.process_id = ?
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE c.run_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process_id, process["run_id"]),
        )
        candidate_rows: list[dict[str, Any]] = []
        candidates_by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in raw_candidates:
            item = dict(row)
            for source_name, target_name, default in (
                ("fields_json", "fields", {}),
                ("empty_fields_json", "empty_fields", {}),
                ("processing_errors_json", "processing_errors", []),
                ("field_scores_json", "field_scores", {}),
            ):
                try:
                    value = json.loads(row[source_name] or json_text(default))
                except (TypeError, json.JSONDecodeError):
                    value = default
                item[target_name] = value
            confidence = item["fields"].get("candidate_confidence")
            item["confidence"] = (
                confidence
                if isinstance(confidence, str) and confidence.strip()
                else "UNKNOWN"
            )
            item["is_primary_hit"] = bool(item["is_primary_hit"])
            candidate_rows.append(item)
            candidates_by_query[item["query_id"]].append(item)

        for query in query_rows:
            all_candidates = candidates_by_query.get(query["query_id"], [])
            successful = [
                item for item in all_candidates
                if item["detail_status"] == "SUCCESS"
            ]
            final = [
                item for item in successful
                if item["reviewed_at"]
                and item["classification_source"] in {"MANUAL", "RULE"}
            ]
            primary_hits = [
                item for item in final
                if item["judgement"] == "HIT" and item["is_primary_hit"]
            ]
            selected_hit = primary_hits[0] if primary_hits else None
            # 功能说明：执行失败属于执行层结果，不应被误判成“候选人待身份归类”。
            # 该 Query 仍可进入版本结果对比，但不进入身份指标的有效分母。
            if query["result_status"] == "EXECUTION_FAILED":
                identity_state = "EXECUTION_FAILED"
            elif query["result_status"] == "NO_CANDIDATES":
                identity_state = "NO_HIT_CONFIRMED"
            elif selected_hit is not None:
                identity_state = "HIT_CONFIRMED"
            elif successful and len(final) == len(successful) and all(
                item["judgement"] in {"NOT_HIT", "SUSPECTED"}
                for item in final
            ):
                identity_state = "NO_HIT_CONFIRMED"
            else:
                identity_state = "PENDING"
            pending_count = len(successful) - len(final)
            baseline = baseline_people.get(query["person_id"])
            available_fields = (
                set(baseline["available_fields"]) if baseline else set()
            )
            accuracy_effective = sorted(
                available_fields & accuracy_keys
            )
            completeness_value = None
            completeness_numerator = 0.0
            completeness_denominator = 0
            completeness_codes: list[str] = []
            completeness_reasons: list[str] = []
            accuracy_value = None
            accuracy_numerator = 0.0
            accuracy_denominator = 0
            accuracy_codes: list[str] = []
            accuracy_reasons: list[str] = []
            if not query["person_id"]:
                accuracy_codes.append("BASELINE_NOT_LINKED")
                reason = f"{query['query_id']} 尚未关联 Baseline Person"
                accuracy_reasons.append(reason)
            elif baseline is None:
                accuracy_codes.append("BASELINE_PERSON_NOT_FOUND")
                reason = (
                    f"{query['query_id']} 的 person_id "
                    f"{query['person_id']} 不在所选 Baseline"
                )
                accuracy_reasons.append(reason)
            if selected_hit is not None:
                if not completeness_keys:
                    completeness_codes.append("NO_EFFECTIVE_FIELDS")
                    completeness_reasons.append(
                        f"{query['query_id']} 没有启用的 Candidate 完整度字段"
                    )
                else:
                    # 命中资料完整度只统计所有启用字段是否真实返回，不依赖
                    # Baseline 是否提供该字段；字段正确性由准确度单独衡量。
                    completeness_denominator = len(completeness_keys)
                    completeness_numerator = float(sum(
                        int(
                            field_key in selected_hit["fields"]
                            and not selected_hit["empty_fields"].get(
                                field_key,
                                True,
                            )
                        )
                        for field_key in completeness_keys
                    ))
                    completeness_value = (
                        completeness_numerator / completeness_denominator
                    )
                accuracy_values = []
                if baseline is not None and not accuracy_effective:
                    accuracy_codes.append("NO_EFFECTIVE_FIELDS")
                    accuracy_reasons.append(
                        f"{query['query_id']} 没有 Baseline 与准确率配置交集"
                    )
                for field_key in accuracy_effective:
                    score = selected_hit["field_scores"].get(field_key, {})
                    if not score.get("returned_nonempty"):
                        continue
                    value = score.get("accuracy_score")
                    if value is None:
                        accuracy_codes.append("MANUAL_SCORE_PENDING")
                        accuracy_reasons.append(
                            f"{query['query_id']}.{field_key} 准确率待人工评分"
                        )
                    else:
                        accuracy_values.append(float(value))
                if accuracy_values:
                    accuracy_numerator = sum(accuracy_values)
                    accuracy_denominator = len(accuracy_values)
                    accuracy_value = accuracy_numerator / accuracy_denominator

            # 非命中资料完整度只衡量候选人自身是否实际返回字段；不能把
            # 与 Baseline 的相同/冲突混入该指标。候选人资料相似度另行记录，
            # 用于分析误召回和同名干扰；姓名属于 Query 强相关字段，不参与。
            nonmatched_data_completeness_values: list[float] = []
            nonmatched_baseline_overlap_values: list[float] = []
            nonmatched_data_completeness_items: list[dict[str, Any]] = []
            nonmatched_baseline_overlap_items: list[dict[str, Any]] = []
            for candidate in final:
                if candidate["judgement"] not in {"NOT_HIT", "SUSPECTED"}:
                    continue
                if not completeness_keys:
                    continue
                returned_count = 0
                overlap_values: list[float] = []
                for field_key in completeness_keys:
                    returned_count += int(
                        field_key in candidate["fields"]
                        and not candidate["empty_fields"].get(
                            field_key,
                            True,
                        )
                    )
                    if field_key in NONMATCHED_SIMILARITY_EXCLUDED_FIELDS:
                        continue
                    score = candidate["field_scores"].get(field_key, {})
                    overlap = score.get("completeness_score")
                    if (
                        isinstance(score, dict)
                        and score.get("baseline_available")
                        and overlap is not None
                    ):
                        overlap_values.append(float(overlap))
                data_value = returned_count / len(completeness_keys)
                nonmatched_data_completeness_values.append(data_value)
                nonmatched_data_completeness_items.append({
                    "candidate_pk": candidate["candidate_pk"],
                    "numerator": float(returned_count),
                    "denominator": len(completeness_keys),
                    "value": data_value,
                })
                if overlap_values:
                    overlap_numerator = sum(overlap_values)
                    overlap_denominator = len(overlap_values)
                    overlap_value = overlap_numerator / overlap_denominator
                    nonmatched_baseline_overlap_values.append(overlap_value)
                    nonmatched_baseline_overlap_items.append({
                        "candidate_pk": candidate["candidate_pk"],
                        "numerator": overlap_numerator,
                        "denominator": overlap_denominator,
                        "value": overlap_value,
                    })
            query.update({
                "candidate_count": len(all_candidates),
                "successful_candidate_count": len(successful),
                "pending_review_count": pending_count,
                "identity_state": identity_state,
                "retrieval_success": identity_state == "HIT_CONFIRMED",
                "retrieval_not_ready_reasons": (
                    [
                        f"{query['query_id']}: {pending_count} 个候选人待身份归类"
                    ]
                    if identity_state == "PENDING" else []
                ),
                "baseline_available_fields": sorted(available_fields),
                "selected_hit_candidate_id": (
                    selected_hit["candidate_pk"] if selected_hit else None
                ),
                "matched_completeness": completeness_value,
                "matched_completeness_numerator": completeness_numerator,
                "matched_completeness_denominator": completeness_denominator,
                "matched_completeness_reason_codes": completeness_codes,
                "matched_completeness_reasons": completeness_reasons,
                "matched_completeness_not_ready_reasons": completeness_reasons,
                "matched_accuracy": accuracy_value,
                "matched_accuracy_numerator": accuracy_numerator,
                "matched_accuracy_denominator": accuracy_denominator,
                "matched_accuracy_reason_codes": accuracy_codes,
                "matched_accuracy_reasons": accuracy_reasons,
                "matched_accuracy_not_ready_reasons": accuracy_reasons,
                # 保留旧键，避免历史 Excel/外部脚本读取失败；其语义已固定为
                # “非命中资料完整度”。新代码应使用更明确的新键。
                "nonmatched_completeness_values": (
                    nonmatched_data_completeness_values
                ),
                "nonmatched_data_completeness_values": (
                    nonmatched_data_completeness_values
                ),
                "nonmatched_data_completeness_items": (
                    nonmatched_data_completeness_items
                ),
                "nonmatched_baseline_overlap_values": (
                    nonmatched_baseline_overlap_values
                ),
                "nonmatched_baseline_overlap_items": (
                    nonmatched_baseline_overlap_items
                ),
                "candidate_confidence": [
                    item["confidence"] for item in all_candidates
                ],
                "formal_ready": (
                    identity_state != "PENDING"
                    and (
                        selected_hit is None
                        or not completeness_codes and not accuracy_codes
                    )
                ),
                "hit_candidate_ids": [
                    item["candidate_pk"] for item in primary_hits
                ],
                "nonmatched_candidate_ids": [
                    item["candidate_pk"] for item in final
                    if item["judgement"] in {"NOT_HIT", "SUSPECTED"}
                ],
                "candidate_ids": [
                    item["candidate_pk"] for item in successful
                ],
            })

        quality = self._metrics_v3_quality(query_rows)
        field_metrics, candidate_field_return = (
            self._metrics_v3_field_returns(
                definitions,
                query_rows,
                candidate_rows,
            )
        )
        module_metrics = self._metrics_v3_modules(
            definitions,
            candidate_rows,
        )
        cost_metrics = {
            field_key: self._metrics_v2_numeric_aggregate(
                query_rows,
                field_key,
            )
            for field_key in (
                "llm_cost",
                "third_party_cost",
                "total_cost",
                "search_duration_ms",
            )
        }
        pdl_metrics = self._metrics_v2_pdl(query_rows)
        confidence_metrics = self._metrics_v3_confidence(candidate_rows)
        result_status_metrics = self._metrics_v2_result_status(query_rows)
        identity_summary = {
            "query_count": len(query_rows),
            "classified_query_count": sum(
                row["identity_state"]
                in {"HIT_CONFIRMED", "NO_HIT_CONFIRMED"}
                for row in query_rows
            ),
            "pending_query_count": sum(
                row["identity_state"] == "PENDING" for row in query_rows
            ),
            "execution_failed_query_count": sum(
                row["identity_state"] == "EXECUTION_FAILED"
                for row in query_rows
            ),
            "primary_hit_query_count": sum(
                row["identity_state"] == "HIT_CONFIRMED"
                for row in query_rows
            ),
            "no_hit_query_count": sum(
                row["identity_state"] == "NO_HIT_CONFIRMED"
                for row in query_rows
            ),
            "final_candidate_count": sum(
                bool(row["reviewed_at"])
                and row["classification_source"] in {"MANUAL", "RULE"}
                for row in candidate_rows
            ),
            "pending_candidate_count": sum(
                not row["reviewed_at"]
                or row["classification_source"] not in {"MANUAL", "RULE"}
                for row in candidate_rows
                if row["detail_status"] == "SUCCESS"
            ),
        }
        execution_summary = {
            "query_count": len(query_rows),
            "has_candidates_count": result_status_metrics[
                "has_candidates_count"
            ],
            "no_candidates_count": result_status_metrics[
                "no_candidates_count"
            ],
            "execution_failed_count": result_status_metrics[
                "execution_failed_count"
            ],
            "detail_success_count": sum(
                row["detail_status"] == "SUCCESS" for row in candidate_rows
            ),
            "detail_failure_count": sum(
                row["detail_status"] != "SUCCESS" for row in candidate_rows
            ),
        }
        grouped_metrics = []
        for query_stage in sorted(
            {row["query_stage"] or "UNSPECIFIED" for row in query_rows}
        ):
            stage_queries = [
                row for row in query_rows
                if (row["query_stage"] or "UNSPECIFIED") == query_stage
            ]
            stage_ids = {row["query_id"] for row in stage_queries}
            grouped_metrics.append({
                "evaluation_phase": process["evaluation_phase"],
                "system_version": process["system_version"],
                "query_stage": query_stage,
                "query_count": len(stage_queries),
                "result_status_metrics": self._metrics_v2_result_status(
                    stage_queries
                ),
                "quality_metrics": self._metrics_v3_quality(stage_queries),
                "cost_metrics": {
                    field_key: self._metrics_v2_numeric_aggregate(
                        stage_queries,
                        field_key,
                    )
                    for field_key in cost_metrics
                },
                "pdl_metrics": self._metrics_v2_pdl(stage_queries),
                "confidence_metrics": self._metrics_v3_confidence(
                    [
                        row for row in candidate_rows
                        if row["query_id"] in stage_ids
                    ]
                ),
            })
        cost_status_values = [
            item["status"] for item in cost_metrics.values()
        ]
        cost_status = {
            "status": (
                "NOT_CONNECTED"
                if all(item == "NOT_CONNECTED" for item in cost_status_values)
                else "COMPLETE"
                if all(item == "COMPLETE" for item in cost_status_values)
                else "PARTIAL"
            ),
            "task_count": len(query_rows),
            "missing_task_count": sum(
                any(
                    row["task_fields"].get(field_key) in (None, "")
                    for field_key in cost_metrics
                )
                for row in query_rows
            ),
            "llm_cost_total": cost_metrics["llm_cost"]["total"],
            "total_cost_total": cost_metrics["total_cost"]["total"],
            "pdl_called_total": (
                pdl_metrics["true_count"]
                if pdl_metrics["known_count"] else None
            ),
        }
        warnings = sorted({
            reason
            for metric in quality.values()
            if isinstance(metric, dict)
            for reason in metric.get("reasons", [])
        })
        return {
            "process_id": process_id,
            "run_id": process["run_id"],
            "evaluation_id": process["evaluation_id"],
            "evaluation_phase": process["evaluation_phase"],
            "system_version": process["system_version"],
            "schema_version": process["schema_version"],
            "baseline_version": process["baseline_version"],
            "rule_version": process["rule_version"],
            "metrics_rule_version": V3_METRICS_RULE_VERSION,
            "formal_ready": quality["formal_ready"],
            "review_status": (
                "REVIEWED" if quality["formal_ready"] else "PENDING_REVIEW"
            ),
            "execution_summary": execution_summary,
            "identity_summary": identity_summary,
            "candidate_field_return": candidate_field_return,
            "result_status_metrics": result_status_metrics,
            "quality_metrics": quality,
            "candidate_return_rate": quality["candidate_return_rate"],
            "conditional_hit_rate": quality["conditional_hit_rate"],
            # 兼容旧键；当前展示名称为“目标人物命中率”。
            "retrieval_success": quality["retrieval_success"],
            "matched_completeness": quality["matched_completeness"],
            "matched_accuracy": quality["matched_accuracy"],
            "nonmatched_completeness": quality[
                "nonmatched_completeness"
            ],
            "nonmatched_data_completeness": quality[
                "nonmatched_data_completeness"
            ],
            "nonmatched_baseline_overlap": quality[
                "nonmatched_baseline_overlap"
            ],
            "cost_metrics": cost_metrics,
            "pdl_metrics": pdl_metrics,
            "confidence_metrics": confidence_metrics,
            "grouped_metrics": grouped_metrics,
            "module_metrics": module_metrics,
            "field_metrics": field_metrics,
            "cost_status": cost_status,
            "query_metrics": query_rows,
            "warnings": warnings,
        }

    def _calculate_process_metrics_v4(self, process_id: str) -> dict[str, Any]:
        """计算按资料模块拆分的 metrics-v4，且保留 metrics-v3 兼容键。

        功能说明：仅对 ``field-processing-v5`` 的新 Process 调用。身份与
        通用字段返回率复用已冻结的 v3 计算；新增“主命中基准资料质量”、
        “非命中资料返回率”和“版本回归字段目录”三个独立分区。历史 Process
        不会进入本方法，因此不会被静默改口径。
        """

        result = self._calculate_process_metrics_v3(process_id)
        process = self.store.fetch_one(
            """
            SELECT fs.definitions_json
            FROM process_runs AS pr
            JOIN field_schemas AS fs ON fs.schema_version = pr.schema_version
            WHERE pr.process_id = ?
            """,
            (process_id,),
        )
        if process is None:
            raise ReviewValidationError(f"处理结果不存在: {process_id}")
        definitions = validate_field_definitions(
            json.loads(process["definitions_json"])
        )
        modules = ("Insights", "Photos", "Profile", "Social", "Summary")
        module_definitions: dict[str, list[dict[str, Any]]] = defaultdict(list)
        return_module_definitions: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for definition in definitions:
            if (
                definition["enabled"]
                and definition["value_scope"] == "CANDIDATE"
                and definition["module"] in modules
            ):
                return_module_definitions[definition["module"]].append(definition)
                if definition["baseline_compare_enabled"]:
                    module_definitions[definition["module"]].append(definition)

        score_rows = self.store.fetch_all(
            """
            SELECT candidate_pk, field_scores_json
            FROM reviews WHERE process_id = ?
            """,
            (process_id,),
        )
        score_by_candidate: dict[str, dict[str, Any]] = {}
        for row in score_rows:
            try:
                parsed = json.loads(row["field_scores_json"] or "{}")
            except (TypeError, json.JSONDecodeError):
                parsed = {}
            score_by_candidate[row["candidate_pk"]] = (
                parsed if isinstance(parsed, dict) else {}
            )

        def metric_from_scores(
            scores: dict[str, Any],
            definitions_for_module: list[dict[str, Any]],
            role: str,
        ) -> dict[str, Any]:
            """按单个 Query 的字段分数聚合一个模块指标。"""

            selected = [
                item for item in definitions_for_module
                if item[f"{role}_enabled"]
            ]
            if not selected:
                return self._metrics_v3_value(
                    0.0, 0, status="NOT_APPLICABLE",
                    reason_codes=[f"NO_{role.upper()}_FIELDS"],
                    reasons=[f"该模块没有启用的{role}字段"],
                )
            values: list[float] = []
            unavailable_count = 0
            pending_count = 0
            for definition in selected:
                score = scores.get(definition["field_key"], {})
                if not score.get("baseline_available", False):
                    unavailable_count += 1
                    continue
                value = score.get(
                    "data_completeness_score"
                    if role == "completeness"
                    else f"{role}_score"
                )
                if value is None:
                    pending_count += 1
                    continue
                values.append(float(value))
            if not values:
                reason_codes = []
                reasons = []
                if unavailable_count:
                    reason_codes.append("BASELINE_FIELD_UNAVAILABLE")
                    reasons.append("Baseline 未提供该模块的可用评分字段")
                if pending_count:
                    reason_codes.append("MANUAL_SCORE_PENDING")
                    reasons.append("该模块存在等待人工评分的字段")
                return self._metrics_v3_value(
                    0.0, 0, status="NOT_APPLICABLE",
                    reason_codes=reason_codes or ["NO_DENOMINATOR"],
                    reasons=reasons or ["没有可自动评分的字段"],
                )
            return self._metrics_v3_value(
                sum(values), len(values),
                status="PARTIAL" if pending_count else "READY",
                reason_codes=(
                    ["MANUAL_SCORE_PENDING"] if pending_count else []
                ),
                reasons=(
                    ["部分字段等待人工评分，未进入自动评分分母"]
                    if pending_count else []
                ),
            )

        query_quality: dict[str, dict[str, Any]] = {}
        for query in result["query_metrics"]:
            selected_candidate = query.get("selected_hit_candidate_id")
            per_module: dict[str, dict[str, Any]] = {}
            if not selected_candidate:
                for module in modules:
                    per_module[module] = {
                        "completeness": self._metrics_v3_value(
                            0.0, 0, status="NOT_APPLICABLE",
                            reason_codes=["NO_PRIMARY_HIT"],
                            reasons=["该 Query 没有已确认的主命中候选人"],
                        ),
                        "accuracy": self._metrics_v3_value(
                            0.0, 0, status="NOT_APPLICABLE",
                            reason_codes=["NO_PRIMARY_HIT"],
                            reasons=["该 Query 没有已确认的主命中候选人"],
                        ),
                    }
            else:
                scores = score_by_candidate.get(selected_candidate, {})
                for module in modules:
                    definitions_for_module = module_definitions.get(module, [])
                    per_module[module] = {
                        "completeness": metric_from_scores(
                            scores, definitions_for_module, "completeness"
                        ),
                        "accuracy": metric_from_scores(
                            scores, definitions_for_module, "accuracy"
                        ),
                    }
            query_quality[query["query_id"]] = {
                "query_id": query["query_id"],
                "person_id": query.get("person_id"),
                "primary_hit_candidate_pk": selected_candidate,
                "modules": per_module,
            }

        def aggregate_metrics(items: Iterable[dict[str, Any]]) -> dict[str, Any]:
            """按 Query 等权聚合模块或人物质量，保留 N/A 原因。"""

            rows = list(items)
            ready = [item for item in rows if item.get("value") is not None]
            if not ready:
                return self._metrics_v3_value(
                    0.0, 0, status="NOT_APPLICABLE",
                    reason_codes=["NO_DENOMINATOR"],
                    reasons=["没有可用于该指标的主命中资料字段"],
                )
            partial = any(item.get("status") == "PARTIAL" for item in ready)
            return self._metrics_v3_value(
                sum(float(item["value"]) for item in ready), len(ready),
                status="PARTIAL" if partial else "READY",
                reason_codes=(
                    ["MANUAL_SCORE_PENDING"] if partial else []
                ),
                reasons=(
                    ["部分字段尚待人工评分"] if partial else []
                ),
            )

        for query in result["query_metrics"]:
            quality = query_quality[query["query_id"]]
            quality["overall"] = {
                role: aggregate_metrics(
                    quality["modules"][module][role]
                    for module in modules
                )
                for role in ("completeness", "accuracy")
            }

        module_quality = {
            module: {
                role: aggregate_metrics(
                    quality["modules"][module][role]
                    for quality in query_quality.values()
                    if quality["primary_hit_candidate_pk"]
                )
                for role in ("completeness", "accuracy")
            }
            for module in modules
        }
        overall_quality = {
            role: aggregate_metrics(
                quality["overall"][role]
                for quality in query_quality.values()
                if quality["primary_hit_candidate_pk"]
            )
            for role in ("completeness", "accuracy")
        }

        non_hit_rows = self.store.fetch_all(
            """
            SELECT pc.fields_json, pc.empty_fields_json, rv.judgement
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            JOIN reviews AS rv
              ON rv.process_id = pc.process_id AND rv.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ? AND c.detail_status = 'SUCCESS'
              AND rv.judgement IN ('NOT_HIT', 'SUSPECTED')
              AND rv.reviewed_at IS NOT NULL
            """,
            (process_id,),
        )
        non_hit_fields = [
            item for definitions_for_module in return_module_definitions.values()
            for item in definitions_for_module
            if item["display_enabled"]
        ]
        non_hit_module_metrics: dict[str, Any] = {}
        for module in modules:
            fields = [
                item for item in return_module_definitions.get(module, [])
                if item["display_enabled"]
            ]
            returned = 0
            for row in non_hit_rows:
                try:
                    empty_fields = json.loads(row["empty_fields_json"] or "{}")
                except (TypeError, json.JSONDecodeError):
                    empty_fields = {}
                returned += any(
                    not empty_fields.get(item["field_key"], True)
                    for item in fields
                )
            non_hit_module_metrics[module] = self._metrics_v3_value(
                float(returned), len(non_hit_rows),
                status=("READY" if non_hit_rows and fields else "NOT_APPLICABLE"),
                reason_codes=([] if non_hit_rows and fields else ["NO_DENOMINATOR"]),
                reasons=([] if non_hit_rows and fields else ["没有已确认非命中候选人或启用资料字段"]),
            )
        non_hit_returned = 0
        for row in non_hit_rows:
            try:
                empty_fields = json.loads(row["empty_fields_json"] or "{}")
            except (TypeError, json.JSONDecodeError):
                empty_fields = {}
            non_hit_returned += sum(
                not empty_fields.get(item["field_key"], True)
                for item in non_hit_fields
            )
        non_hit_return = self._metrics_v3_value(
            float(non_hit_returned), len(non_hit_rows) * len(non_hit_fields),
            status=(
                "READY" if non_hit_rows and non_hit_fields else "NOT_APPLICABLE"
            ),
            reason_codes=(
                [] if non_hit_rows and non_hit_fields else ["NO_DENOMINATOR"]
            ),
            reasons=(
                [] if non_hit_rows and non_hit_fields
                else ["没有已确认非命中候选人或启用资料字段"]
            ),
        )
        reason_counts = {
            row["reason"]: row["count"]
            for row in self.store.fetch_all(
                """
                SELECT reason, COUNT(*) AS count FROM reviews
                WHERE process_id = ? AND reviewed_at IS NOT NULL
                GROUP BY reason
                """,
                (process_id,),
            )
        }
        result["metrics_rule_version"] = METRICS_RULE_VERSION
        result["identity_metrics"] = {
            **result["identity_summary"],
            "has_candidates": self._metrics_v3_value(
                float(result["execution_summary"]["has_candidates_count"]),
                result["result_status_metrics"]["total_formal_queries"],
            ),
            "retrieval_success": result["retrieval_success"],
            "reason_counts": reason_counts,
        }
        result["baseline_quality_metrics"] = {
            "modules": module_quality,
            "overall": overall_quality,
            "queries": query_quality,
            "primary_hit_query_count": sum(
                bool(item["primary_hit_candidate_pk"])
                for item in query_quality.values()
            ),
        }
        result["non_hit_data_return"] = {
            "overall": non_hit_return,
            "modules": non_hit_module_metrics,
            "candidate_count": len(non_hit_rows),
        }
        result["regression_metrics"] = {
            "status": "NOT_APPLICABLE",
            "reason_codes": ["COMPARISON_PROCESS_NOT_SELECTED"],
            "reasons": ["创建对比报告后才生成版本回归变化"],
            "eligible_field_keys": [
                item["field_key"] for item in definitions
                if item["enabled"] and item["run_compare_enabled"]
            ],
        }
        return result

    def compare_processes(
        self,
        baseline_process_id: str,
        candidate_process_id: str,
    ) -> dict[str, Any]:
        """生成同条件、新增线索和不可比三类 Process 对比结果。

        正式同条件必须同时满足人物、Query Stage 和 Dataset 输入签名一致。
        缺少回归 Query、签名缺失或签名不一致不会再中断整份报告，而是进入
        ``not_comparable`` 并保留原因。
        """

        if baseline_process_id == candidate_process_id:
            raise ReviewValidationError("baseline 与 candidate 不能是同一 Process")
        processes = []
        for process_id in (baseline_process_id, candidate_process_id):
            row = self.store.fetch_one(
                """
                SELECT pr.*, r.evaluation_id, r.dataset_id,
                       fs.definitions_json
                FROM process_runs AS pr
                JOIN runs AS r ON r.run_id = pr.run_id
                JOIN field_schemas AS fs
                  ON fs.schema_version = pr.schema_version
                WHERE pr.process_id = ?
                """,
                (process_id,),
            )
            if row is None:
                raise ReviewValidationError(f"处理结果不存在: {process_id}")
            processes.append(row)
        baseline_process, candidate_process = processes
        compatibility_fields = (
            "evaluation_id",
            "baseline_version",
            "rule_version",
        )
        differences = [
            field
            for field in compatibility_fields
            if baseline_process[field] != candidate_process[field]
        ]
        if (
            baseline_process["schema_version"]
            != candidate_process["schema_version"]
        ):
            try:
                baseline_definitions = validate_field_definitions(
                    json.loads(baseline_process["definitions_json"])
                )
                candidate_definitions = validate_field_definitions(
                    json.loads(candidate_process["definitions_json"])
                )
            except (
                TypeError,
                json.JSONDecodeError,
                FieldSchemaValidationError,
            ) as exc:
                raise ReviewValidationError(
                    "字段配置快照损坏，无法判断对比兼容性"
                ) from exc
            if json_text(baseline_definitions) != json_text(
                candidate_definitions
            ):
                differences.append("schema_version")
        if differences:
            raise ReviewValidationError(
                "两个 Process 不兼容: " + ", ".join(differences)
            )

        baseline_metrics = self.calculate_process_metrics(
            baseline_process_id
        )
        candidate_metrics = self.calculate_process_metrics(
            candidate_process_id
        )
        metric_maps = {
            baseline_process_id: {
                row["query_id"]: row
                for row in baseline_metrics["query_metrics"]
            },
            candidate_process_id: {
                row["query_id"]: row
                for row in candidate_metrics["query_metrics"]
            },
        }

        def query_map(process: sqlite3.Row) -> dict[tuple[str, str], dict[str, Any]]:
            """读取 Query、稳定输入签名和人物级指标，并检查配对键唯一。"""

            input_signatures: dict[str, str] = {}
            if process["dataset_id"]:
                for input_row in self.store.fetch_all(
                    """
                    SELECT query_id, match_strategy, clues_json,
                           additional_details_json
                    FROM dataset_queries WHERE dataset_id = ?
                    """,
                    (process["dataset_id"],),
                ):
                    try:
                        input_signatures[input_row["query_id"]] = json_text(
                            {
                                "match_strategy": input_row["match_strategy"],
                                "clues": json.loads(input_row["clues_json"]),
                                "additional_details": json.loads(
                                    input_row["additional_details_json"]
                                ),
                            }
                        )
                    except (TypeError, json.JSONDecodeError) as exc:
                        raise ReviewValidationError(
                            f"Dataset Query 输入快照已损坏: {input_row['query_id']}"
                        ) from exc
            rows = self.store.fetch_all(
                """
                SELECT rq.query_id, rq.person_id, rq.query_stage
                FROM run_queries AS rq
                WHERE rq.run_id = ?
                ORDER BY rq.query_id
                """,
                (process["run_id"],),
            )
            result: dict[tuple[str, str], dict[str, Any]] = {}
            for row in rows:
                if not row["person_id"] or not row["query_stage"]:
                    raise ReviewValidationError(
                        f"Process {process['process_id']} 缺少 person_id/query_stage"
                    )
                key = (row["person_id"], row["query_stage"])
                if key in result:
                    raise ReviewValidationError(
                        "同一 Process 中 person_id + query_stage 必须唯一"
                    )
                query_metrics = metric_maps[process["process_id"]].get(
                    row["query_id"],
                    {},
                )
                result[key] = {
                    "query_id": row["query_id"],
                    "has_hit": bool(
                        query_metrics.get("retrieval_success")
                    ),
                    "input_signature": input_signatures.get(row["query_id"]),
                    "metrics": query_metrics,
                }
            return result

        baseline_queries = query_map(baseline_process)
        candidate_queries = query_map(candidate_process)
        category_counts = {
            "持续命中": 0,
            "新增命中": 0,
            "退化未命中": 0,
            "持续未命中": 0,
        }
        pairs: list[dict[str, Any]] = []
        new_clue_queries: list[dict[str, Any]] = []
        not_comparable_queries: list[dict[str, Any]] = []

        def metric_value(
            query: dict[str, Any],
            field_key: str,
        ) -> Any:
            """从人物级指标读取可比较值，缺失保持 None。"""

            return query["metrics"].get(field_key)

        def primary_confidence(query: dict[str, Any]) -> str | None:
            """返回最高排名候选人的原始 Confidence，空值统一 UNKNOWN。"""

            values = query["metrics"].get("candidate_confidence", [])
            if not values:
                return None
            value = values[0]
            return (
                value.strip()
                if isinstance(value, str) and value.strip()
                else "UNKNOWN"
            )

        for person_id, query_stage in sorted(
            set(baseline_queries) | set(candidate_queries)
        ):
            key = (person_id, query_stage)
            baseline_query = baseline_queries.get(key)
            candidate_query = candidate_queries.get(key)
            if baseline_query is None:
                baseline_person_exists = any(
                    item_person_id == person_id
                    for item_person_id, _ in baseline_queries
                )
                if baseline_person_exists:
                    new_clue_queries.append(
                        {
                            "person_id": person_id,
                            "query_stage": query_stage,
                            "candidate_query_id": candidate_query["query_id"],
                            "result_status": candidate_query["metrics"].get(
                                "result_status"
                            ),
                            "retrieval_success": candidate_query[
                                "metrics"
                            ].get("retrieval_success"),
                            "matched_completeness": candidate_query[
                                "metrics"
                            ].get("matched_completeness"),
                            "matched_accuracy": candidate_query[
                                "metrics"
                            ].get("matched_accuracy"),
                            "candidate_confidence": primary_confidence(
                                candidate_query
                            ),
                            "candidate_count": candidate_query[
                                "metrics"
                            ].get("candidate_count", 0),
                            "task_fields": candidate_query["metrics"].get(
                                "task_fields",
                                {},
                            ),
                        }
                    )
                else:
                    not_comparable_queries.append(
                        {
                            "person_id": person_id,
                            "query_stage": query_stage,
                            "baseline_query_id": None,
                            "candidate_query_id": candidate_query["query_id"],
                            "reason": "CANDIDATE_PERSON_NOT_IN_BASELINE",
                        }
                    )
                continue
            if candidate_query is None:
                not_comparable_queries.append(
                    {
                        "person_id": person_id,
                        "query_stage": query_stage,
                        "baseline_query_id": baseline_query["query_id"],
                        "candidate_query_id": None,
                        "reason": "MISSING_REGRESSION_QUERY",
                    }
                )
                continue
            baseline_signature = baseline_query["input_signature"]
            candidate_signature = candidate_query["input_signature"]
            if baseline_signature is None or candidate_signature is None:
                not_comparable_queries.append(
                    {
                        "person_id": person_id,
                        "query_stage": query_stage,
                        "baseline_query_id": baseline_query["query_id"],
                        "candidate_query_id": candidate_query["query_id"],
                        "reason": "INPUT_SIGNATURE_UNAVAILABLE",
                    }
                )
                continue
            if baseline_signature != candidate_signature:
                not_comparable_queries.append(
                    {
                        "person_id": person_id,
                        "query_stage": query_stage,
                        "baseline_query_id": baseline_query["query_id"],
                        "candidate_query_id": candidate_query["query_id"],
                        "reason": "INPUT_SIGNATURE_MISMATCH",
                    }
                )
                continue
            baseline_hit = baseline_query["has_hit"]
            candidate_hit = candidate_query["has_hit"]
            if baseline_hit and candidate_hit:
                category = "持续命中"
            elif not baseline_hit and candidate_hit:
                category = "新增命中"
            elif baseline_hit and not candidate_hit:
                category = "退化未命中"
            else:
                category = "持续未命中"
            category_counts[category] += 1
            baseline_completeness = metric_value(
                baseline_query,
                "matched_completeness",
            )
            candidate_completeness = metric_value(
                candidate_query,
                "matched_completeness",
            )
            baseline_accuracy = metric_value(
                baseline_query,
                "matched_accuracy",
            )
            candidate_accuracy = metric_value(
                candidate_query,
                "matched_accuracy",
            )
            baseline_task = baseline_query["metrics"].get("task_fields", {})
            candidate_task = candidate_query["metrics"].get("task_fields", {})
            pairs.append(
                {
                    "person_id": person_id,
                    "query_stage": query_stage,
                    "baseline_query_id": baseline_query["query_id"],
                    "candidate_query_id": candidate_query["query_id"],
                    "baseline_hit": baseline_hit,
                    "candidate_hit": candidate_hit,
                    "category": category,
                    "baseline_matched_completeness": baseline_completeness,
                    "candidate_matched_completeness": candidate_completeness,
                    "matched_completeness_delta": (
                        candidate_completeness - baseline_completeness
                        if baseline_completeness is not None
                        and candidate_completeness is not None
                        else None
                    ),
                    "baseline_matched_accuracy": baseline_accuracy,
                    "candidate_matched_accuracy": candidate_accuracy,
                    "matched_accuracy_delta": (
                        candidate_accuracy - baseline_accuracy
                        if baseline_accuracy is not None
                        and candidate_accuracy is not None
                        else None
                    ),
                    "baseline_confidence": primary_confidence(
                        baseline_query
                    ),
                    "candidate_confidence": primary_confidence(
                        candidate_query
                    ),
                    "baseline_total_cost": baseline_task.get("total_cost"),
                    "candidate_total_cost": candidate_task.get("total_cost"),
                    "total_cost_delta": (
                        candidate_task["total_cost"]
                        - baseline_task["total_cost"]
                        if isinstance(
                            baseline_task.get("total_cost"),
                            (int, float),
                        )
                        and not isinstance(
                            baseline_task.get("total_cost"),
                            bool,
                        )
                        and isinstance(
                            candidate_task.get("total_cost"),
                            (int, float),
                        )
                        and not isinstance(
                            candidate_task.get("total_cost"),
                            bool,
                        )
                        else None
                    ),
                    "baseline_search_duration_ms": baseline_task.get(
                        "search_duration_ms"
                    ),
                    "candidate_search_duration_ms": candidate_task.get(
                        "search_duration_ms"
                    ),
                    "search_duration_ms_delta": (
                        candidate_task["search_duration_ms"]
                        - baseline_task["search_duration_ms"]
                        if isinstance(
                            baseline_task.get("search_duration_ms"),
                            (int, float),
                        )
                        and not isinstance(
                            baseline_task.get("search_duration_ms"),
                            bool,
                        )
                        and isinstance(
                            candidate_task.get("search_duration_ms"),
                            (int, float),
                        )
                        and not isinstance(
                            candidate_task.get("search_duration_ms"),
                            bool,
                        )
                        else None
                    ),
                    "baseline_pdl_called": baseline_task.get("pdl_called"),
                    "candidate_pdl_called": candidate_task.get("pdl_called"),
                    "baseline_candidate_count": baseline_query[
                        "metrics"
                    ].get("candidate_count", 0),
                    "candidate_candidate_count": candidate_query[
                        "metrics"
                    ].get("candidate_count", 0),
                    "formal_ready": (
                        bool(baseline_query["metrics"].get("formal_ready"))
                        and bool(
                            candidate_query["metrics"].get("formal_ready")
                        )
                    ),
                }
            )

        reason_counts: dict[str, int] = {}
        for item in not_comparable_queries:
            reason = item["reason"]
            reason_counts[reason] = reason_counts.get(reason, 0) + 1
        coverage = {
            "baseline_query_count": len(baseline_queries),
            "candidate_query_count": len(candidate_queries),
            "paired_count": len(pairs),
            "new_clue_count": len(new_clue_queries),
            "baseline_pairable_query_count": len(baseline_queries),
            "candidate_pairable_query_count": len(candidate_queries),
            "successful_pair_count": len(pairs),
            "candidate_new_condition_count": len(new_clue_queries),
            "missing_regression_query_count": reason_counts.get(
                "MISSING_REGRESSION_QUERY",
                0,
            ),
            "input_signature_mismatch_count": reason_counts.get(
                "INPUT_SIGNATURE_MISMATCH",
                0,
            ),
            "input_signature_unavailable_count": reason_counts.get(
                "INPUT_SIGNATURE_UNAVAILABLE",
                0,
            ),
            "not_comparable_count": len(not_comparable_queries),
        }
        same_condition_ready = bool(pairs) and all(
            item["formal_ready"] for item in pairs
        )
        blocking_not_comparable = bool(not_comparable_queries)

        new_clue_stage_metrics: dict[str, Any] = {}
        for query_stage in sorted(
            {item["query_stage"] for item in new_clue_queries}
        ):
            rows = [
                candidate_queries[
                    (item["person_id"], item["query_stage"])
                ]["metrics"]
                for item in new_clue_queries
                if item["query_stage"] == query_stage
            ]
            if candidate_metrics.get("metrics_rule_version") in {
                V3_METRICS_RULE_VERSION,
                METRICS_RULE_VERSION,
            }:
                new_clue_stage_metrics[query_stage] = {
                    "query_count": len(rows),
                    "result_status_metrics": self._metrics_v2_result_status(
                        rows
                    ),
                    "quality_metrics": self._metrics_v3_quality(rows),
                    "cost_metrics": {
                        field_key: self._metrics_v2_numeric_aggregate(
                            rows,
                            field_key,
                        )
                        for field_key in (
                            "llm_cost",
                            "third_party_cost",
                            "total_cost",
                            "search_duration_ms",
                        )
                    },
                    "pdl_metrics": self._metrics_v2_pdl(rows),
                }
            else:
                new_clue_stage_metrics[query_stage] = {
                    "query_count": len(rows),
                    "formal_ready": all(
                        row.get("formal_ready", False) for row in rows
                    ),
                }
        same_condition = {
            "formal_ready": same_condition_ready,
            "pairs": pairs,
            "coverage": coverage,
            "category_counts": category_counts,
        }
        new_clue = {
            "queries": new_clue_queries,
            "query_stage_metrics": new_clue_stage_metrics,
        }
        not_comparable = {
            "queries": not_comparable_queries,
            "reasons": reason_counts,
        }
        if reason_counts.get("INPUT_SIGNATURE_MISMATCH"):
            input_check_status = "MISMATCH"
        elif reason_counts.get("INPUT_SIGNATURE_UNAVAILABLE"):
            input_check_status = "UNAVAILABLE"
        elif not_comparable_queries:
            input_check_status = "PARTIAL"
        else:
            input_check_status = "VERIFIED"
        return {
            "baseline_process_id": baseline_process_id,
            "candidate_process_id": candidate_process_id,
            "formal_ready": (
                same_condition_ready
                and not blocking_not_comparable
            ),
            "input_check_status": input_check_status,
            "same_condition": same_condition,
            "new_clue": new_clue,
            "not_comparable": not_comparable,
            "coverage": coverage,
            # 兼容旧 ReportModel v1 页面和已存在的调用方。
            "category_counts": category_counts,
            "pairs": pairs,
            "baseline_metrics": baseline_metrics,
            "candidate_metrics": candidate_metrics,
        }

    def _report_process(self, process_id: str) -> sqlite3.Row:
        """读取生成报告所需的 Process、Run、Evaluation 和字段快照元数据。"""

        row = self.store.fetch_one(
            """
            SELECT pr.*, r.evaluation_id, r.dataset_id, r.run_label,
                   r.system_version, r.source_type, r.status AS run_status,
                   r.total_queries, r.success_queries, r.failed_queries,
                   r.started_at, r.finished_at, r.evaluation_phase,
                   e.name AS evaluation_name, e.notes AS evaluation_notes,
                   e.thresholds_json, e.threshold_profile_id,
                   tp.name AS threshold_profile_name,
                   tp.version AS threshold_profile_version,
                   fs.name AS schema_name, fs.definitions_json
            FROM process_runs AS pr
            JOIN runs AS r ON r.run_id = pr.run_id
            JOIN evaluations AS e ON e.evaluation_id = r.evaluation_id
            LEFT JOIN threshold_profiles AS tp
              ON tp.profile_id = e.threshold_profile_id
            JOIN field_schemas AS fs
              ON fs.schema_version = pr.schema_version
            WHERE pr.process_id = ?
            """,
            (process_id,),
        )
        if row is None:
            raise ReviewValidationError(f"处理结果不存在: {process_id}")
        if row["status"] != "COMPLETED":
            raise ReviewValidationError(
                f"处理结果尚未完成，不能生成报告: {process_id}"
            )
        return row

    @staticmethod
    def _report_metric(
        values: list[float],
        *,
        formal_ready: bool,
        numerator: float | None = None,
        denominator: int | None = None,
    ) -> dict[str, Any]:
        """构造报告内统一的分子、分母、正式值和预览值。"""

        actual_numerator = sum(values) if numerator is None else numerator
        actual_denominator = len(values) if denominator is None else denominator
        preview_value = (
            actual_numerator / actual_denominator
            if actual_denominator
            else None
        )
        return {
            "numerator": actual_numerator,
            "denominator": actual_denominator,
            "value": preview_value if formal_ready else None,
            "preview_value": preview_value,
        }

    def _query_stage_report_metrics(
        self,
        metrics: dict[str, Any],
    ) -> dict[str, dict[str, Any]]:
        """按 FULL_NAME/FULL_NAME_SOCIAL 聚合阶段5 Query 指标。"""

        result: dict[str, dict[str, Any]] = {}
        for query_stage in sorted(SUPPORTED_QUERY_STAGES):
            rows = [
                row
                for row in metrics["query_metrics"]
                if row["query_stage"] == query_stage
            ]
            if not rows:
                continue
            formal_ready = all(row["formal_ready"] for row in rows)
            successful = [
                row for row in rows if row["retrieval_success"]
            ]
            executed = [
                row for row in rows
                if row.get("result_status") != "EXECUTION_FAILED"
            ]
            candidate_returned = [
                row for row in executed
                if int(row.get("candidate_count") or 0) > 0
            ]
            completeness = [
                float(row["matched_completeness"])
                for row in successful
                if row["matched_completeness"] is not None
            ]
            completeness_numerator = sum(
                float(row["matched_completeness_numerator"])
                for row in successful
            )
            completeness_denominator = sum(
                int(row["matched_completeness_denominator"])
                for row in successful
            )
            accuracy = [
                float(row["matched_accuracy"])
                for row in successful
                if row["matched_accuracy"] is not None
            ]
            nonmatched_data_completeness = [
                float(value)
                for row in rows
                for value in row["nonmatched_data_completeness_values"]
            ]
            nonmatched_data_items = [
                item
                for row in rows
                for item in row.get("nonmatched_data_completeness_items", [])
            ]
            nonmatched_baseline_overlap = [
                float(value)
                for row in rows
                for value in row["nonmatched_baseline_overlap_values"]
            ]
            nonmatched_overlap_items = [
                item
                for row in rows
                for item in row.get("nonmatched_baseline_overlap_items", [])
            ]
            result[query_stage] = {
                "formal_ready": formal_ready,
                "query_count": len(rows),
                "retrieval_success": self._report_metric(
                    [],
                    formal_ready=formal_ready,
                    numerator=float(len(successful)),
                    denominator=len(executed),
                ),
                "candidate_return_rate": self._report_metric(
                    [],
                    formal_ready=True,
                    numerator=float(len(candidate_returned)),
                    denominator=len(executed),
                ),
                "conditional_hit_rate": self._report_metric(
                    [],
                    formal_ready=formal_ready,
                    numerator=float(len(successful)),
                    denominator=len(candidate_returned),
                ),
                "matched_completeness": self._report_metric(
                    completeness,
                    formal_ready=formal_ready,
                    numerator=completeness_numerator,
                    denominator=completeness_denominator,
                ),
                "matched_accuracy": self._report_metric(
                    accuracy,
                    formal_ready=formal_ready,
                ),
                "nonmatched_completeness": self._report_metric(
                    nonmatched_data_completeness,
                    formal_ready=formal_ready,
                    numerator=(
                        sum(
                            float(item["numerator"])
                            for item in nonmatched_data_items
                        )
                        if nonmatched_data_items else None
                    ),
                    denominator=(
                        sum(
                            int(item["denominator"])
                            for item in nonmatched_data_items
                        )
                        if nonmatched_data_items else None
                    ),
                ),
                "nonmatched_data_completeness": self._report_metric(
                    nonmatched_data_completeness,
                    formal_ready=formal_ready,
                    numerator=(
                        sum(
                            float(item["numerator"])
                            for item in nonmatched_data_items
                        )
                        if nonmatched_data_items else None
                    ),
                    denominator=(
                        sum(
                            int(item["denominator"])
                            for item in nonmatched_data_items
                        )
                        if nonmatched_data_items else None
                    ),
                ),
                "nonmatched_baseline_overlap": self._report_metric(
                    nonmatched_baseline_overlap,
                    formal_ready=formal_ready,
                    numerator=(
                        sum(
                            float(item["numerator"])
                            for item in nonmatched_overlap_items
                        )
                        if nonmatched_overlap_items else None
                    ),
                    denominator=(
                        sum(
                            int(item["denominator"])
                            for item in nonmatched_overlap_items
                        )
                        if nonmatched_overlap_items else None
                    ),
                ),
                "query_ids": [row["query_id"] for row in rows],
            }
        return result

    def _process_field_report_metrics(
        self,
        process: sqlite3.Row,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        """计算模块返回率和字段级返回率/复核得分。"""

        try:
            definitions = validate_field_definitions(
                json.loads(process["definitions_json"])
            )
        except (json.JSONDecodeError, FieldSchemaValidationError) as exc:
            raise ReviewValidationError(
                f"字段配置快照已损坏: {process['schema_version']}"
            ) from exc
        rows = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.detail_status,
                   pc.fields_json, pc.empty_fields_json,
                   rv.judgement, rv.field_scores_json, rv.reviewed_at
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ?
            """,
            (process["process_id"],),
        )
        parsed_rows = []
        for row in rows:
            item = dict(row)
            for source_name, target_name in (
                ("fields_json", "fields"),
                ("empty_fields_json", "empty_fields"),
                ("field_scores_json", "field_scores"),
            ):
                try:
                    value = (
                        json.loads(row[source_name])
                        if row[source_name]
                        else {}
                    )
                except (TypeError, json.JSONDecodeError):
                    value = {}
                item[target_name] = value if isinstance(value, dict) else {}
            parsed_rows.append(item)
        successful = [
            row for row in parsed_rows if row["detail_status"] == "SUCCESS"
        ]
        hit_rows = [
            row
            for row in successful
            if row["reviewed_at"] and row["judgement"] == "HIT"
        ]
        nonmatched_rows = [
            row
            for row in successful
            if row["reviewed_at"]
            and row["judgement"] in {"NOT_HIT", "SUSPECTED"}
        ]
        field_metrics: dict[str, Any] = {}
        module_candidates: dict[str, set[str]] = defaultdict(set)
        module_hit_candidates: dict[str, set[str]] = defaultdict(set)
        module_nonmatched_candidates: dict[str, set[str]] = defaultdict(set)
        for definition in definitions:
            if not definition["enabled"]:
                continue
            field_key = definition["field_key"]
            returned_candidates = [
                row
                for row in successful
                if field_key in row["fields"]
                and not row["empty_fields"].get(field_key, True)
            ]
            for row in returned_candidates:
                module_candidates[definition["module"]].add(
                    row["candidate_pk"]
                )
            returned_hit_rows = [
                row
                for row in hit_rows
                if field_key in row["fields"]
                and not row["empty_fields"].get(field_key, True)
            ]
            returned_nonmatched_rows = [
                row
                for row in nonmatched_rows
                if field_key in row["fields"]
                and not row["empty_fields"].get(field_key, True)
            ]
            for row in returned_hit_rows:
                module_hit_candidates[definition["module"]].add(
                    row["candidate_pk"]
                )
            for row in returned_nonmatched_rows:
                module_nonmatched_candidates[definition["module"]].add(
                    row["candidate_pk"]
                )
            completeness_values = [
                float(row["field_scores"][field_key]["completeness_score"])
                for row in hit_rows
                if field_key in row["field_scores"]
                and row["field_scores"][field_key].get(
                    "completeness_score"
                )
                is not None
            ]
            accuracy_values = [
                float(row["field_scores"][field_key]["accuracy_score"])
                for row in hit_rows
                if field_key in row["field_scores"]
                and row["field_scores"][field_key].get("returned_nonempty")
                and row["field_scores"][field_key].get("accuracy_score")
                is not None
            ]
            field_metrics[field_key] = {
                "field_key": field_key,
                "display_name": definition["display_name"],
                "module": definition["module"],
                "scoring_role": definition["scoring_role"],
                "compare_mode": definition["compare_mode"],
                "returned_count": len(returned_candidates),
                "empty_count": len(successful) - len(returned_candidates),
                "candidate_count": len(successful),
                "return_rate": (
                    len(returned_candidates) / len(successful)
                    if successful
                    else None
                ),
                "hit_returned_count": len(returned_hit_rows),
                "hit_candidate_count": len(hit_rows),
                "hit_return_rate": (
                    len(returned_hit_rows) / len(hit_rows)
                    if hit_rows
                    else None
                ),
                "nonmatched_returned_count": len(
                    returned_nonmatched_rows
                ),
                "nonmatched_candidate_count": len(nonmatched_rows),
                "nonmatched_nonempty_rate": (
                    len(returned_nonmatched_rows) / len(nonmatched_rows)
                    if nonmatched_rows
                    else None
                ),
                "hit_completeness": (
                    sum(completeness_values) / len(completeness_values)
                    if completeness_values
                    else None
                ),
                "hit_accuracy": (
                    sum(accuracy_values) / len(accuracy_values)
                    if accuracy_values
                    else None
                ),
            }
        module_metrics = {
            module: {
                "module": module,
                "returned_candidate_count": len(candidate_ids),
                "candidate_count": len(successful),
                "hit_returned_candidate_count": len(
                    module_hit_candidates[module]
                ),
                "hit_candidate_count": len(hit_rows),
                "nonmatched_returned_candidate_count": len(
                    module_nonmatched_candidates[module]
                ),
                "nonmatched_candidate_count": len(nonmatched_rows),
                "return_rate": (
                    len(candidate_ids) / len(successful)
                    if successful
                    else None
                ),
                "hit_return_rate": (
                    len(module_hit_candidates[module]) / len(hit_rows)
                    if hit_rows
                    else None
                ),
                "nonmatched_return_rate": (
                    len(module_nonmatched_candidates[module])
                    / len(nonmatched_rows)
                    if nonmatched_rows
                    else None
                ),
            }
            for module, candidate_ids in sorted(module_candidates.items())
        }
        for module in sorted(
            {
                definition["module"]
                for definition in definitions
                if definition["enabled"]
            }
        ):
            module_metrics.setdefault(
                module,
                {
                    "module": module,
                    "returned_candidate_count": 0,
                    "candidate_count": len(successful),
                    "hit_returned_candidate_count": 0,
                    "hit_candidate_count": len(hit_rows),
                    "nonmatched_returned_candidate_count": 0,
                    "nonmatched_candidate_count": len(nonmatched_rows),
                    "return_rate": 0.0 if successful else None,
                    "hit_return_rate": 0.0 if hit_rows else None,
                    "nonmatched_return_rate": (
                        0.0 if nonmatched_rows else None
                    ),
                },
            )
        return module_metrics, field_metrics

    def _report_candidate_lookup(
        self,
        process_id: str,
    ) -> dict[str, dict[str, Any]]:
        """按 query_id 返回报告配对案例使用的首名候选人。"""

        rows = self.store.fetch_all(
            """
            SELECT c.query_id, c.candidate_pk, c.candidate_id,
                   c.candidate_rank, rv.judgement, rv.reviewed_at
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process_id,),
        )
        result: dict[str, dict[str, Any]] = {}
        for row in rows:
            result.setdefault(row["query_id"], dict(row))
        return result

    def _report_case_groups(
        self,
        process: sqlite3.Row,
        comparison: dict[str, Any] | None,
    ) -> dict[str, list[dict[str, Any]]]:
        """构造可从报告下钻到 Query/Candidate 的人物级案例。"""

        rows = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.candidate_id, c.candidate_rank,
                   c.query_id, c.detail_status, c.detail_error,
                   rq.person_id, rq.person_id_source, rq.query_stage,
                   rq.status AS query_status,
                   bp.display_name AS baseline_display_name,
                   rv.judgement, rv.reason, rv.reviewed_at,
                   rv.classification_source, rv.is_primary_hit,
                   pc.processing_errors_json
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            JOIN run_queries AS rq
              ON rq.run_id = c.run_id AND rq.query_id = c.query_id
            LEFT JOIN baseline_people AS bp
              ON bp.baseline_version = ?
             AND bp.person_id = rq.person_id
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process["baseline_version"], process["process_id"]),
        )
        groups: dict[str, list[dict[str, Any]]] = {
            "命中": [],
            "疑似": [],
            "待复核": [],
            "详情失败": [],
            "处理异常": [],
            "改善最明显": [],
            "持续命中": [],
            "新增命中": [],
            "退化未命中": [],
            "持续未命中": [],
            "新增线索": [],
            "不可比": [],
        }
        for row in rows:
            item = dict(row)
            try:
                processing_errors = json.loads(
                    row["processing_errors_json"]
                )
            except (TypeError, json.JSONDecodeError):
                processing_errors = []
            item["processing_error_count"] = (
                len(processing_errors)
                if isinstance(processing_errors, list)
                else 0
            )
            item.pop("processing_errors_json", None)
            if row["detail_status"] == "FAILED":
                groups["详情失败"].append(item)
            if item["processing_error_count"]:
                groups["处理异常"].append(item)
            if not row["reviewed_at"]:
                groups["待复核"].append(item)
            elif row["judgement"] == "HIT":
                groups["命中"].append(item)
            elif row["judgement"] == "SUSPECTED":
                groups["疑似"].append(item)
        if comparison is not None:
            baseline_lookup = self._report_candidate_lookup(
                comparison["baseline_process_id"]
            )
            candidate_lookup = self._report_candidate_lookup(
                comparison["candidate_process_id"]
            )
            for pair in comparison["pairs"]:
                baseline_candidate = baseline_lookup.get(
                    pair["baseline_query_id"],
                    {},
                )
                candidate_candidate = candidate_lookup.get(
                    pair["candidate_query_id"],
                    {},
                )
                enriched = {
                    **pair,
                    "baseline_candidate_pk": baseline_candidate.get(
                        "candidate_pk"
                    ),
                    "baseline_candidate_id": baseline_candidate.get(
                        "candidate_id"
                    ),
                    "candidate_candidate_pk": candidate_candidate.get(
                        "candidate_pk"
                    ),
                    "candidate_candidate_id": candidate_candidate.get(
                        "candidate_id"
                    ),
                }
                groups[pair["category"]].append(enriched)
            for item in comparison["new_clue"]["queries"]:
                candidate_candidate = candidate_lookup.get(
                    item["candidate_query_id"],
                    {},
                )
                groups["新增线索"].append(
                    {
                        **item,
                        "candidate_candidate_pk": candidate_candidate.get(
                            "candidate_pk"
                        ),
                        "candidate_candidate_id": candidate_candidate.get(
                            "candidate_id"
                        ),
                        "category": "新增线索",
                    }
                )
            groups["不可比"].extend(
                comparison["not_comparable"]["queries"]
            )
            comparable_cases = [
                item
                for category in (
                    "持续命中",
                    "新增命中",
                    "退化未命中",
                    "持续未命中",
                )
                for item in groups[category]
            ]
            improved_cases = [
                item
                for item in comparable_cases
                if item["category"] == "新增命中"
                or (item.get("matched_completeness_delta") or 0) > 0
                or (item.get("matched_accuracy_delta") or 0) > 0
            ]
            improved_cases.sort(
                key=lambda item: (
                    item["category"] == "新增命中",
                    item.get("matched_completeness_delta") is not None,
                    item.get("matched_completeness_delta") or 0,
                    item.get("matched_accuracy_delta") is not None,
                    item.get("matched_accuracy_delta") or 0,
                ),
                reverse=True,
            )
            groups["改善最明显"] = improved_cases[:5]
            groups["新增命中"].sort(
                key=lambda item: (
                    item.get("matched_completeness_delta") is not None,
                    item.get("matched_completeness_delta") or 0,
                    item.get("matched_accuracy_delta") is not None,
                    item.get("matched_accuracy_delta") or 0,
                ),
                reverse=True,
            )
            groups["持续未命中"].sort(
                key=lambda item: (
                    item.get("candidate_candidate_count", 0),
                    str(item.get("candidate_confidence") or ""),
                ),
                reverse=True,
            )
        return groups

    @staticmethod
    def _merge_comparison_metrics(
        baseline: dict[str, Any],
        candidate: dict[str, Any],
    ) -> dict[str, Any]:
        """合并 baseline/candidate 模块或字段指标并计算可用差值。"""

        merged: dict[str, Any] = {}
        for key in sorted(set(baseline) | set(candidate)):
            baseline_item = baseline.get(key, {})
            candidate_item = candidate.get(key, {})
            item = {**baseline_item, **candidate_item}
            for metric_name in (
                "return_rate",
                "hit_return_rate",
                "nonmatched_return_rate",
                "nonmatched_nonempty_rate",
                "hit_completeness",
                "hit_accuracy",
            ):
                baseline_value = baseline_item.get(metric_name)
                candidate_value = candidate_item.get(metric_name)
                item[f"baseline_{metric_name}"] = baseline_value
                item[f"candidate_{metric_name}"] = candidate_value
                item[f"{metric_name}_delta"] = (
                    candidate_value - baseline_value
                    if baseline_value is not None
                    and candidate_value is not None
                    else None
                )
            merged[key] = item
        return merged

    @staticmethod
    def _report_v5_metric(
        value: float | None,
        *,
        numerator: float | None = None,
        denominator: int | None = None,
        status: str | None = None,
        reason_codes: Iterable[str] = (),
        reasons: Iterable[str] = (),
    ) -> dict[str, Any]:
        """构造 v5 候选人或 Query 明细使用的统一指标对象。

        功能说明：报告详情必须区分“值为 0”和“当前不适用”。本方法只负责
        把既有 metrics-v4 的结果转为展示对象，不重新计算指标公式。

        参数说明：
            value: 已由现有指标规则计算出的比例；不可用时为 None。
            numerator: 可选的分子，用于详情审计。
            denominator: 可选的分母，用于详情审计。
            status: 明确的展示状态；为空时根据 value 自动推断。
            reason_codes: 不适用或未就绪的机器可读原因。
            reasons: 面向页面的原因说明。

        返回值：包含 value、分子、分母、状态和原因的 JSON 安全对象。
        """

        resolved_status = status or (
            "READY" if value is not None else "NOT_APPLICABLE"
        )
        return {
            "value": value,
            "numerator": numerator,
            "denominator": denominator,
            "status": resolved_status,
            "reason_codes": sorted(set(reason_codes)),
            "reasons": sorted(set(reasons)),
        }

    @staticmethod
    def _report_v5_safe_json(
        value: Any,
        default: dict[str, Any] | list[Any],
    ) -> dict[str, Any] | list[Any]:
        """安全解析报告快照所需 JSON，损坏数据退化为指定空结构。

        功能说明：历史导入数据中的 JSON 损坏不能让报告在半途中生成；调用方
        根据字段语义传入空对象或空数组，并在上层保留已有处理错误信息。
        """

        try:
            parsed = json.loads(value) if value else default
        except (TypeError, json.JSONDecodeError):
            parsed = default
        return parsed if isinstance(parsed, type(default)) else default

    def _build_report_v5_candidate_snapshot(
        self,
        *,
        row: dict[str, Any],
        query_quality: dict[str, Any],
        definitions: list[dict[str, Any]],
        baseline_fields: dict[str, Any],
        baseline_available_fields: set[str],
        raw_ids: list[str],
    ) -> dict[str, Any]:
        """构造一个候选人的 v5 报告快照，不读取 Raw 完整内容。

        功能说明：将已入库 Candidate、处理字段、复核结论与字段评分转换为
        页面可直接使用的结构。所有评分复用 metrics-v4 使用的字段评分，
        不在报告层另建一套准确度或完整度公式。

        参数说明：
            row: 已批量读取并完成 JSON 解析的 Candidate 行。
            query_quality: 当前 Query 的基准资料模块质量结果。
            definitions: 当前 Process 冻结的 FieldSchema 定义。
            baseline_fields: 当前 Query 关联 Baseline Person 的字段值。
            baseline_available_fields: Baseline 可用于比较的字段集合。
            raw_ids: 对应 Candidate 的 Raw 记录标识，不包含 Raw payload。

        返回值：完整 Candidate 快照，包含指标、模块、字段对比、证据和引用。
        """

        fields = row["fields"]
        empty_fields = row["empty_fields"]
        field_scores = row["field_scores"]
        processing_errors = row["processing_errors"]
        successful = row["detail_status"] == "SUCCESS"
        is_final = bool(row["reviewed_at"]) and row[
            "classification_source"
        ] in {"MANUAL", "RULE"}
        judgement = row["judgement"] if is_final else "UNCLASSIFIED"
        if not successful:
            judgement = "DETAIL_FAILED"
        is_primary_hit = bool(row["is_primary_hit"]) and judgement == "HIT"

        def field_returned(field_key: str) -> bool:
            """判断一个已提取字段是否存在真实非空值。"""

            return (
                field_key in fields
                and not bool(empty_fields.get(field_key, True))
            )

        completeness_definitions = [
            item for item in definitions
            if item["enabled"]
            and item["value_scope"] == "CANDIDATE"
            and item["completeness_enabled"]
        ]
        def matched_metric(
            definitions_for_metric: list[dict[str, Any]],
            role: str,
        ) -> dict[str, Any]:
            """按当前 HIT 候选人的字段评分生成命中指标。

            非主命中同样需要看到自身的资料质量，不能复用主命中的
            Query 汇总值。该函数只读取处理时已冻结的 field_scores，
            不重新请求接口或修改整体汇总口径。
            """

            selected = [
                item for item in definitions_for_metric
                if item["enabled"]
                and item["value_scope"] == "CANDIDATE"
                and item[f"{role}_enabled"]
                and (
                    role == "completeness"
                    or item["baseline_compare_enabled"]
                )
            ]
            if not selected:
                return self._report_v5_metric(
                    None,
                    reason_codes=[f"NO_{role.upper()}_FIELDS"],
                    reasons=[f"当前字段配置没有启用命中{role}字段"],
                )
            values: list[float] = []
            unavailable_count = 0
            pending_count = 0
            for definition in selected:
                if role == "completeness":
                    # 资料完整度只回答字段是否实际返回，全部启用字段都进入
                    # 分母，不因 Baseline 缺字段而缩小统计范围。
                    values.append(float(field_returned(
                        definition["field_key"]
                    )))
                    continue
                score = field_scores.get(definition["field_key"], {})
                if not isinstance(score, dict) or not score.get(
                    "baseline_available", False
                ):
                    unavailable_count += 1
                    continue
                value = score.get("accuracy_score")
                if value is None:
                    pending_count += 1
                    continue
                values.append(float(value))
            if not values:
                reason_codes: list[str] = []
                reasons: list[str] = []
                if unavailable_count:
                    reason_codes.append("BASELINE_FIELD_UNAVAILABLE")
                    reasons.append("Baseline 未提供该模块的可用评分字段")
                if pending_count:
                    reason_codes.append("MANUAL_SCORE_PENDING")
                    reasons.append("该模块存在等待人工评分的字段")
                return self._report_v5_metric(
                    None,
                    reason_codes=reason_codes or ["NO_DENOMINATOR"],
                    reasons=reasons or ["没有可自动评分的字段"],
                )
            return self._report_v5_metric(
                sum(values) / len(values),
                numerator=sum(values),
                denominator=len(values),
                status="PARTIAL" if pending_count else "READY",
                reason_codes=(
                    ["MANUAL_SCORE_PENDING"] if pending_count else []
                ),
                reasons=(
                    ["部分字段等待人工评分，未进入自动评分分母"]
                    if pending_count else []
                ),
            )

        if judgement == "HIT":
            matched_completeness = matched_metric(definitions, "completeness")
            matched_accuracy = matched_metric(definitions, "accuracy")
            nonmatched_completeness = self._report_v5_metric(
                None,
                reason_codes=["NOT_APPLICABLE_TO_HIT"],
                reasons=["命中候选人不计算非命中完整度"],
            )
        elif judgement in {"NOT_HIT", "SUSPECTED"}:
            if completeness_definitions:
                data_completeness_values: list[float] = []
                overlap_values: list[float] = []
                for definition in completeness_definitions:
                    score = field_scores.get(definition["field_key"], {})
                    # 资料完整度只由该字段是否真实返回决定，与基准值无关。
                    data_completeness_values.append(
                        float(field_returned(definition["field_key"]))
                    )
                    if (
                        definition["field_key"]
                        in NONMATCHED_SIMILARITY_EXCLUDED_FIELDS
                    ):
                        continue
                    overlap = (
                        score.get("completeness_score")
                        if isinstance(score, dict) else None
                    )
                    if (
                        isinstance(score, dict)
                        and score.get("baseline_available")
                        and overlap is not None
                    ):
                        overlap_values.append(float(overlap))
                nonmatched_data_completeness = self._report_v5_metric(
                    sum(data_completeness_values) /
                    len(data_completeness_values),
                    numerator=sum(data_completeness_values),
                    denominator=len(data_completeness_values),
                )
                nonmatched_baseline_overlap = (
                    self._report_v5_metric(
                        sum(overlap_values) / len(overlap_values),
                        numerator=sum(overlap_values),
                        denominator=len(overlap_values),
                    )
                    if overlap_values else self._report_v5_metric(
                        None,
                        reason_codes=["NO_BASELINE_OVERLAP_FIELDS"],
                        reasons=["没有可比较的非命中候选人资料字段"],
                    )
                )
            else:
                nonmatched_data_completeness = self._report_v5_metric(
                    None,
                    reason_codes=["NO_COMPLETENESS_FIELDS"],
                    reasons=["当前字段配置没有启用非命中完整度字段"],
                )
                nonmatched_baseline_overlap = self._report_v5_metric(
                    None,
                    reason_codes=["NO_BASELINE_OVERLAP_FIELDS"],
                    reasons=["当前字段配置没有可用于资料相似度的字段"],
                )
            matched_completeness = self._report_v5_metric(
                None,
                reason_codes=["NOT_APPLICABLE_TO_NON_HIT"],
                reasons=["非命中候选人不计算命中完整度"],
            )
            matched_accuracy = self._report_v5_metric(
                None,
                reason_codes=["NOT_APPLICABLE_TO_NON_HIT"],
                reasons=["非命中候选人不计算命中准确度"],
            )
            nonmatched_completeness = nonmatched_data_completeness
        else:
            pending_code = (
                "CANDIDATE_DETAIL_FAILED"
                if not successful else "IDENTITY_UNCLASSIFIED"
            )
            pending_reason = (
                row["detail_error"]
                if not successful and row["detail_error"]
                else "候选人尚无正式身份结论"
            )
            matched_completeness = self._report_v5_metric(
                None, reason_codes=[pending_code], reasons=[pending_reason]
            )
            matched_accuracy = self._report_v5_metric(
                None, reason_codes=[pending_code], reasons=[pending_reason]
            )
            nonmatched_completeness = self._report_v5_metric(
                None, reason_codes=[pending_code], reasons=[pending_reason]
            )
            nonmatched_data_completeness = nonmatched_completeness
            nonmatched_baseline_overlap = self._report_v5_metric(
                None, reason_codes=[pending_code], reasons=[pending_reason]
            )

        if judgement == "HIT":
            nonmatched_data_completeness = nonmatched_completeness
            nonmatched_baseline_overlap = self._report_v5_metric(
                None,
                reason_codes=["NOT_APPLICABLE_TO_HIT"],
                reasons=["命中候选人不计算非命中候选人资料相似度"],
            )

        modules: dict[str, dict[str, Any]] = {}
        field_comparisons: list[dict[str, Any]] = []
        for module in ("Insights", "Photos", "Profile", "Social", "Summary"):
            module_definitions = [
                item for item in definitions
                if item["enabled"]
                and item["display_enabled"]
                and item["value_scope"] == "CANDIDATE"
                and item["module"] == module
            ]
            status_key = f"{module.lower()}_status"
            returned_count = sum(
                field_returned(item["field_key"])
                for item in module_definitions
            )
            compared_scores = [
                field_scores.get(item["field_key"], {})
                for item in module_definitions
                if item["baseline_compare_enabled"]
            ]
            correct_count = sum(
                score.get("accuracy_score") == 1
                for score in compared_scores
                if isinstance(score, dict)
            )
            error_count = sum(
                score.get("comparison_status") == "ERROR"
                for score in compared_scores
                if isinstance(score, dict)
            )
            compared_count = sum(
                bool(score.get("comparison_status"))
                for score in compared_scores
                if isinstance(score, dict)
            )
            if judgement == "HIT":
                module_completeness = matched_metric(
                    module_definitions, "completeness"
                )
                module_accuracy = matched_metric(
                    module_definitions, "accuracy"
                )
            else:
                module_quality = query_quality.get("modules", {}).get(
                    module, {}
                )
                module_completeness = module_quality.get("completeness")
                module_accuracy = module_quality.get("accuracy")
            modules[module] = {
                "module": module,
                "status": fields.get(status_key),
                "has_real_data": bool(returned_count),
                "display_field_count": len(module_definitions),
                "returned_field_count": returned_count,
                "empty_field_count": len(module_definitions) - returned_count,
                "correct_field_count": correct_count,
                "error_field_count": error_count,
                "not_applicable_field_count": (
                    len(compared_scores) - compared_count
                ),
                "completeness": (
                    dict(module_completeness)
                    if isinstance(module_completeness, dict)
                    else self._report_v5_metric(
                        None,
                        reason_codes=["NO_PRIMARY_HIT_QUALITY"],
                    )
                ),
                "accuracy": (
                    dict(module_accuracy)
                    if isinstance(module_accuracy, dict)
                    else self._report_v5_metric(
                        None,
                        reason_codes=["NO_PRIMARY_HIT_QUALITY"],
                    )
                ),
            }
            for definition in module_definitions:
                field_key = definition["field_key"]
                score = field_scores.get(field_key, {})
                if not isinstance(score, dict):
                    score = {}
                baseline_key = definition.get("baseline_field_key") or field_key
                baseline_available = baseline_key in baseline_available_fields
                field_comparisons.append({
                    "field_key": field_key,
                    "display_name": definition["display_name"],
                    "module": module,
                    "candidate_value": (
                        fields.get(field_key)
                        if field_returned(field_key) else None
                    ),
                    "baseline_value": (
                        baseline_fields.get(baseline_key)
                        if baseline_available else None
                    ),
                    "baseline_available": baseline_available,
                    "returned_nonempty": field_returned(field_key),
                    "comparison_status": score.get("comparison_status"),
                    "reason_code": score.get("reason_code"),
                    "completeness_score": score.get("completeness_score"),
                    "data_completeness_score": score.get(
                        "data_completeness_score"
                    ),
                    "baseline_coverage_score": score.get(
                        "baseline_coverage_score"
                    ),
                    "accuracy_score": score.get("accuracy_score"),
                })

        evidence: dict[str, Any] = {}
        if row["evidence"]:
            evidence["review_evidence"] = row["evidence"]
        return {
            "candidate_pk": row["candidate_pk"],
            "candidate_id": row["candidate_id"],
            "candidate_rank": row["candidate_rank"],
            "rank_score": row["rank_score"],
            "display_name": (
                fields.get("profile_full_name")
                or fields.get("summary_display_name")
                or fields.get("display_name")
                or row["candidate_id"]
            ),
            "detail_status": row["detail_status"],
            "detail_error": row["detail_error"] or "",
            "confidence": (
                fields.get("candidate_confidence")
                or fields.get("summary_confidence_level")
                or "UNKNOWN"
            ),
            "identity": {
                "judgement": judgement,
                "classification_source": (
                    row["classification_source"] or "SUGGESTED"
                ),
                "is_primary_hit": is_primary_hit,
                "reason": row["reason"] or "",
                "evidence_summary": row["evidence"] or "",
            },
            "metrics": {
                "matched_completeness": matched_completeness,
                "matched_accuracy": matched_accuracy,
                "nonmatched_completeness": nonmatched_completeness,
                "nonmatched_data_completeness": nonmatched_data_completeness,
                "nonmatched_baseline_overlap": nonmatched_baseline_overlap,
            },
            "modules": modules,
            "field_comparisons": field_comparisons,
            "evidence": evidence,
            "processing_errors": processing_errors,
            "references": {
                "candidate_pk": row["candidate_pk"],
                "raw_ids": raw_ids,
            },
        }

    def _build_report_v5_process_snapshot(
        self,
        process: sqlite3.Row,
        metrics: dict[str, Any],
    ) -> dict[str, Any]:
        """批量读取一个 Process 并生成全部 Query/Candidate 报告快照。

        功能说明：本方法使用固定数量的批量 SQL 查询完成快照组装，避免在
        Query 或 Candidate 循环内访问数据库。它是 Web 与静态 HTML 的共同
        数据来源，且不保存 Candidate Detail Raw payload。
        """

        try:
            definitions = validate_field_definitions(
                json.loads(process["definitions_json"])
            )
        except (TypeError, json.JSONDecodeError, FieldSchemaValidationError) as exc:
            raise ReviewValidationError("报告字段配置快照已损坏") from exc

        query_metrics = {
            item["query_id"]: item
            for item in metrics.get("query_metrics", [])
        }
        query_quality = metrics.get("baseline_quality_metrics", {}).get(
            "queries", {}
        )
        query_rows = self.store.fetch_all(
            """
            SELECT rq.query_id, rq.person_id, rq.person_id_source,
                   rq.query_stage, rq.task_id, rq.status AS query_status,
                   rq.candidate_count_total, rq.candidate_count_listed,
                   rq.detail_success_count, rq.detail_failure_count,
                   pq.result_status, pq.fields_json, pq.empty_fields_json,
                   pq.processing_errors_json,
                   bp.display_name AS baseline_display_name,
                   bp.fields_json AS baseline_fields_json,
                   bp.available_fields_json AS baseline_available_fields_json
            FROM run_queries AS rq
            JOIN processed_queries AS pq
              ON pq.run_id = rq.run_id AND pq.query_id = rq.query_id
             AND pq.process_id = ?
            LEFT JOIN baseline_people AS bp
              ON bp.baseline_version = ? AND bp.person_id = rq.person_id
            WHERE rq.run_id = ?
            ORDER BY rq.query_id
            """,
            (process["process_id"], process["baseline_version"], process["run_id"]),
        )
        candidate_rows = self.store.fetch_all(
            """
            SELECT c.candidate_pk, c.query_id, c.candidate_id,
                   c.candidate_rank, c.rank_score, c.detail_status,
                   c.detail_error, pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json, rv.judgement, rv.reason,
                   rv.evidence, rv.field_scores_json, rv.reviewed_at,
                   rv.classification_source, rv.is_primary_hit
            FROM candidates AS c
            JOIN processed_candidates AS pc
              ON pc.candidate_pk = c.candidate_pk AND pc.process_id = ?
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE c.run_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process["process_id"], process["run_id"]),
        )
        raw_rows = self.store.fetch_all(
            """
            SELECT candidate_pk, query_id, raw_id
            FROM raw_records
            WHERE run_id = ?
            ORDER BY query_id, candidate_pk, sequence_no, collected_at, raw_id
            """,
            (process["run_id"],),
        )
        raw_ids_by_candidate: dict[str, list[str]] = defaultdict(list)
        raw_ids_by_query: dict[str, list[str]] = defaultdict(list)
        for raw_row in raw_rows:
            raw_ids_by_query[raw_row["query_id"]].append(raw_row["raw_id"])
            if raw_row["candidate_pk"]:
                raw_ids_by_candidate[raw_row["candidate_pk"]].append(
                    raw_row["raw_id"]
                )

        candidates_by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for source in candidate_rows:
            row = dict(source)
            row["fields"] = self._report_v5_safe_json(
                row.pop("fields_json"), {}
            )
            row["empty_fields"] = self._report_v5_safe_json(
                row.pop("empty_fields_json"), {}
            )
            row["processing_errors"] = self._report_v5_safe_json(
                row.pop("processing_errors_json"), []
            )
            row["field_scores"] = self._report_v5_safe_json(
                row.pop("field_scores_json"), {}
            )
            candidates_by_query[row["query_id"]].append(row)

        items: list[dict[str, Any]] = []
        total_candidate_count = 0
        for source in query_rows:
            row = dict(source)
            task_fields = self._report_v5_safe_json(row["fields_json"], {})
            baseline_fields = self._report_v5_safe_json(
                row["baseline_fields_json"], {}
            )
            baseline_available = self._report_v5_safe_json(
                row["baseline_available_fields_json"], []
            )
            metric = query_metrics.get(row["query_id"], {})
            candidate_snapshots = [
                self._build_report_v5_candidate_snapshot(
                    row=candidate,
                    query_quality=query_quality.get(row["query_id"], {}),
                    definitions=definitions,
                    baseline_fields=baseline_fields,
                    baseline_available_fields=set(baseline_available),
                    # 早期导入只把 Raw 挂在 Query；优先使用 Candidate Raw，
                    # 不存在时退回同一 Query 的 Raw ID，保证审计可追溯。
                    raw_ids=(
                        raw_ids_by_candidate.get(candidate["candidate_pk"])
                        or raw_ids_by_query.get(row["query_id"], [])
                    ),
                )
                for candidate in candidates_by_query.get(row["query_id"], [])
            ]
            total_candidate_count += len(candidate_snapshots)
            pair_key = (
                f"{row['person_id']}|{row['query_stage']}"
                if row["person_id"] and row["query_stage"]
                else f"query:{row['query_id']}"
            )
            items.append({
                "pair_key": pair_key,
                "query": {
                    "query_id": row["query_id"],
                    "input_id": row["query_id"],
                    "person_id": row["person_id"],
                    "person_id_source": row["person_id_source"],
                    "display_name": (
                        row["baseline_display_name"] or row["person_id"]
                        or row["query_id"]
                    ),
                    "query_stage": row["query_stage"] or "UNSPECIFIED",
                    "query_status": row["query_status"],
                    "result_status": row["result_status"],
                    "candidate_count": len(candidate_snapshots),
                    "candidate_count_reported": row["candidate_count_total"],
                    "candidate_count_listed": row["candidate_count_listed"],
                    "detail_success_count": row["detail_success_count"],
                    "detail_failure_count": row["detail_failure_count"],
                    "retrieval_success": bool(metric.get("retrieval_success")),
                    "identity_state": metric.get("identity_state"),
                    "primary_hit_candidate_id": metric.get(
                        "selected_hit_candidate_id"
                    ),
                    "task_metrics": {
                        field_key: task_fields.get(field_key)
                        for field_key in (
                            "llm_cost",
                            "third_party_cost",
                            "total_cost",
                            "search_duration_ms",
                            "pdl_called",
                        )
                    },
                    "processing_errors": self._report_v5_safe_json(
                        row["processing_errors_json"], []
                    ),
                    "references": {
                        "run_id": process["run_id"],
                        "process_id": process["process_id"],
                        "raw_ids": raw_ids_by_query.get(row["query_id"], []),
                    },
                },
                "candidates": candidate_snapshots,
            })
        return {
            "total_query_count": len(items),
            "total_candidate_count": total_candidate_count,
            "items": items,
        }

    def _build_report_v5_query_explorer(
        self,
        *,
        candidate_process: sqlite3.Row,
        candidate_metrics: dict[str, Any],
        baseline_process: sqlite3.Row | None,
        baseline_metrics: dict[str, Any] | None,
        comparison: dict[str, Any] | None,
    ) -> dict[str, Any]:
        """合并单次或对比报告的全部 Query/Candidate 快照。

        功能说明：单次报告只保存候选运行侧；对比报告按已有 compare_processes
        的配对结果组合两侧快照，不改变其正式可比性判断。
        """

        candidate_snapshot = self._build_report_v5_process_snapshot(
            candidate_process, candidate_metrics
        )
        if baseline_process is None or baseline_metrics is None:
            return {
                "total_query_count": candidate_snapshot["total_query_count"],
                "total_candidate_count": candidate_snapshot[
                    "total_candidate_count"
                ],
                "baseline_total_query_count": 0,
                "baseline_total_candidate_count": 0,
                "initial_query_count": 5,
                "load_more_query_count": 10,
                # 兼容早期 v5 页面读取的单一分段字段。
                "default_chunk_size": 10,
                "items": [
                    {
                        "pair_key": item["pair_key"],
                        "person_id": item["query"]["person_id"],
                        "display_name": item["query"]["display_name"],
                        "query_stage": item["query"]["query_stage"],
                        "change_category": None,
                        "change": None,
                        "candidate_run": item,
                        "baseline_run": None,
                    }
                    for item in candidate_snapshot["items"]
                ],
            }

        baseline_snapshot = self._build_report_v5_process_snapshot(
            baseline_process, baseline_metrics
        )
        baseline_by_query = {
            item["query"]["query_id"]: item
            for item in baseline_snapshot["items"]
        }
        candidate_by_query = {
            item["query"]["query_id"]: item
            for item in candidate_snapshot["items"]
        }
        pair_by_candidate_query = {
            item["candidate_query_id"]: item
            for item in (comparison or {}).get("pairs", [])
            if item.get("candidate_query_id")
        }
        new_clue_by_candidate_query = {
            item["candidate_query_id"]: item
            for item in (comparison or {}).get("new_clue", {}).get(
                "queries", []
            )
            if item.get("candidate_query_id")
        }
        not_comparable_by_candidate_query = {
            item["candidate_query_id"]: item
            for item in (comparison or {}).get("not_comparable", {}).get(
                "queries", []
            )
            if item.get("candidate_query_id")
        }
        items: list[dict[str, Any]] = []
        used_baseline_query_ids: set[str] = set()
        for candidate_item in candidate_snapshot["items"]:
            query_id = candidate_item["query"]["query_id"]
            pair = pair_by_candidate_query.get(query_id)
            new_clue = new_clue_by_candidate_query.get(query_id)
            not_comparable = not_comparable_by_candidate_query.get(query_id)
            baseline_item = None
            change_category = None
            change: dict[str, Any] | None = None
            if pair:
                baseline_item = baseline_by_query.get(pair["baseline_query_id"])
                used_baseline_query_ids.add(pair["baseline_query_id"])
                change_category = pair["category"]
                change = pair
            elif new_clue:
                change_category = "新增线索"
                change = new_clue
            elif not_comparable:
                change_category = "不可比"
                change = not_comparable
            items.append({
                "pair_key": candidate_item["pair_key"],
                "person_id": candidate_item["query"]["person_id"],
                "display_name": candidate_item["query"]["display_name"],
                "query_stage": candidate_item["query"]["query_stage"],
                "change_category": change_category,
                "change": change,
                "candidate_run": candidate_item,
                "baseline_run": baseline_item,
            })
        for baseline_item in baseline_snapshot["items"]:
            query_id = baseline_item["query"]["query_id"]
            if query_id in used_baseline_query_ids:
                continue
            items.append({
                "pair_key": baseline_item["pair_key"],
                "person_id": baseline_item["query"]["person_id"],
                "display_name": baseline_item["query"]["display_name"],
                "query_stage": baseline_item["query"]["query_stage"],
                "change_category": "不可比",
                "change": {
                    "reason": "MISSING_CANDIDATE_QUERY",
                    "baseline_query_id": query_id,
                    "candidate_query_id": None,
                },
                "candidate_run": None,
                "baseline_run": baseline_item,
            })
        return {
            "total_query_count": len(items),
            "total_candidate_count": candidate_snapshot[
                "total_candidate_count"
            ],
            "baseline_total_query_count": baseline_snapshot[
                "total_query_count"
            ],
            "baseline_total_candidate_count": baseline_snapshot[
                "total_candidate_count"
            ],
            "initial_query_count": 5,
            "load_more_query_count": 10,
            "default_chunk_size": 10,
            "items": items,
        }

    @staticmethod
    def _build_report_v5_processing_scope(
        process: sqlite3.Row,
        metrics: dict[str, Any],
    ) -> dict[str, Any]:
        """生成本次处理实际启用字段的报告快照。

        功能说明：字段清单以 Process 绑定的不可变 FieldSchema 为准，并合并
        本次字段返回统计。当前最新配置发生变化时，历史报告仍能准确说明当时
        提取、展示、对比和评分了哪些字段。

        参数说明：
            process: 包含 ``definitions_json`` 和 Schema 标识的 Process 行。
            metrics: 当前 Process 已计算完成的字段统计。

        返回值：
            dict: 汇总计数、模块分组以及全部启用字段的轻量快照。

        异常说明：
            Process 的字段配置损坏时返回空清单；报告主体仍可生成，并由现有
            diagnostics 展示配置问题。
        """

        try:
            definitions = validate_field_definitions(
                json.loads(process["definitions_json"])
            )
        except (TypeError, json.JSONDecodeError, FieldSchemaValidationError):
            definitions = []
        field_metrics = metrics.get("field_metrics", {})
        enabled = [
            definition for definition in definitions
            if definition.get("enabled")
        ]
        fields: list[dict[str, Any]] = []
        for definition in enabled:
            field_key = definition["field_key"]
            field_metric = field_metrics.get(field_key, {})
            fields.append({
                "field_key": field_key,
                "display_name": definition["display_name"],
                "module": definition["module"],
                "value_scope": definition["value_scope"],
                "source_path": definition["source_path"],
                "source_type": definition.get("source_type", "PATH"),
                "data_type": definition.get("data_type", "string"),
                "display_enabled": bool(
                    definition.get("display_enabled", True)
                ),
                "baseline_compare_enabled": bool(
                    definition.get("baseline_compare_enabled")
                ),
                "completeness_enabled": bool(
                    definition.get("completeness_enabled")
                ),
                "accuracy_enabled": bool(
                    definition.get("accuracy_enabled")
                ),
                "identity_enabled": bool(
                    definition.get("identity_enabled")
                ),
                "returned_count": int(
                    field_metric.get("returned_count") or 0
                ),
                "empty_count": int(field_metric.get("empty_count") or 0),
                "entity_count": int(
                    field_metric.get(
                        "entity_count",
                        field_metric.get("candidate_count", 0),
                    ) or 0
                ),
                "return_rate": field_metric.get("return_rate"),
                "status": field_metric.get("status", "NOT_APPLICABLE"),
            })
        module_order = {
            "Task": 0,
            "Candidate": 1,
            "Insights": 2,
            "Photos": 3,
            "Profile": 4,
            "Social": 5,
            "Summary": 6,
        }
        modules: list[dict[str, Any]] = []
        for module in sorted(
            {item["module"] for item in fields},
            key=lambda value: (module_order.get(value, 99), value),
        ):
            module_fields = [
                item for item in fields if item["module"] == module
            ]
            modules.append({
                "module": module,
                "field_count": len(module_fields),
                "actual_returned_field_count": sum(
                    item["returned_count"] > 0 for item in module_fields
                ),
                "fields": module_fields,
            })
        return {
            "schema_version": process["schema_version"],
            "schema_name": process["schema_name"],
            "enabled_field_count": len(fields),
            "actual_returned_field_count": sum(
                item["returned_count"] > 0 for item in fields
            ),
            "baseline_compare_field_count": sum(
                item["baseline_compare_enabled"] for item in fields
            ),
            "completeness_field_count": sum(
                item["completeness_enabled"] for item in fields
            ),
            "accuracy_field_count": sum(
                item["accuracy_enabled"] for item in fields
            ),
            "identity_field_count": sum(
                item["identity_enabled"] for item in fields
            ),
            "modules": modules,
            "fields": fields,
        }

    @staticmethod
    def _build_report_v5_module_return_overview(
        query_explorer: dict[str, Any],
        processing_scope: dict[str, Any],
    ) -> dict[str, Any]:
        """聚合五大资料模块在 HIT 与非命中候选人中的真实返回情况。

        功能说明：
            模块有数据率按“至少一个业务原子字段有值”的候选人数计算；
            字段完整度按实际返回字段槽位除以应有字段槽位计算。状态字段、
            对象容器和重复聚合字段不会进入槽位，Candidate Detail 失败也
            不进入返回率分母，避免把请求失败误判为模块空数据。

        参数说明：
            query_explorer: ReportModel v5 已冻结的全部 Query/Candidate 快照。
            processing_scope: 当前 Process 绑定的 FieldSchema 字段范围快照。

        返回值：
            dict: 包含人群计数、五个模块汇总、字段明细和接口状态分布。

        异常说明：
            历史或不完整快照缺少字段时按空值处理；不影响报告其他章节生成。
        """

        business_fields: dict[str, list[dict[str, Any]]] = {
            module: [] for module in REPORT_V5_PROFILE_MODULES
        }
        for field in processing_scope.get("fields", []):
            module = field.get("module")
            field_key = str(field.get("field_key") or "")
            if (
                module not in business_fields
                or field.get("value_scope") != "CANDIDATE"
                or not field.get("display_enabled", True)
                or field_key in REPORT_V5_MODULE_NON_BUSINESS_FIELDS
                or str(field.get("data_type") or "").lower() == "object"
            ):
                continue
            business_fields[module].append({
                "field_key": field_key,
                "display_name": field.get("display_name") or field_key,
                "source_path": field.get("source_path") or "",
            })

        candidates: list[dict[str, Any]] = []
        for item in query_explorer.get("items", []):
            candidate_run = item.get("candidate_run") or {}
            candidates.extend(candidate_run.get("candidates", []))

        def detail_success(candidate: dict[str, Any]) -> bool:
            """返回候选人详情是否成功，可安全进入模块返回率分母。"""

            return candidate.get("detail_status") == "SUCCESS"

        populations = {
            "all": {
                "label": "全部候选人",
                "candidates": candidates,
            },
            "hit": {
                "label": "全部 HIT",
                "candidates": [
                    item for item in candidates
                    if item.get("identity", {}).get("judgement") == "HIT"
                ],
            },
            "nonmatched": {
                "label": "非命中 / 疑似",
                "candidates": [
                    item for item in candidates
                    if item.get("identity", {}).get("judgement")
                    in {"NOT_HIT", "SUSPECTED"}
                ],
            },
        }
        population_summary: dict[str, dict[str, Any]] = {}
        for key, population in populations.items():
            items = population["candidates"]
            eligible = [item for item in items if detail_success(item)]
            population_summary[key] = {
                "key": key,
                "label": population["label"],
                "candidate_count": len(items),
                "detail_success_count": len(eligible),
                "detail_failure_count": len(items) - len(eligible),
            }

        modules: list[dict[str, Any]] = []
        for module in REPORT_V5_PROFILE_MODULES:
            module_fields = business_fields[module]
            field_keys = [item["field_key"] for item in module_fields]

            def candidate_field_map(
                candidate: dict[str, Any],
            ) -> dict[str, dict[str, Any]]:
                """按 field_key 索引候选人字段快照，缺失字段按未返回处理。"""

                return {
                    str(item.get("field_key")): item
                    for item in candidate.get("field_comparisons", [])
                    if item.get("module") == module
                }

            def population_metric(
                population_key: str,
            ) -> dict[str, Any]:
                """计算一个模块在指定候选人人群中的返回率和字段完整度。"""

                source = populations[population_key]
                eligible = [
                    item for item in source["candidates"]
                    if detail_success(item)
                ]
                returned_candidates = 0
                fully_returned_candidates = 0
                returned_slots = 0
                status_distribution = {
                    "data": 0,
                    "empty": 0,
                    "error": 0,
                    "unknown": 0,
                }
                for candidate in eligible:
                    field_map = candidate_field_map(candidate)
                    returned_count = sum(
                        bool(field_map.get(field_key, {}).get(
                            "returned_nonempty"
                        ))
                        for field_key in field_keys
                    )
                    returned_slots += returned_count
                    returned_candidates += int(returned_count > 0)
                    fully_returned_candidates += int(
                        bool(field_keys)
                        and returned_count == len(field_keys)
                    )
                    raw_status = str(
                        candidate.get("modules", {})
                        .get(module, {})
                        .get("status")
                        or "unknown"
                    ).strip().lower()
                    status = (
                        raw_status
                        if raw_status in status_distribution
                        else "unknown"
                    )
                    status_distribution[status] += 1
                candidate_denominator = len(eligible)
                slot_denominator = candidate_denominator * len(field_keys)
                return {
                    "key": population_key,
                    "label": source["label"],
                    "candidate_count": len(source["candidates"]),
                    "detail_success_count": candidate_denominator,
                    "detail_failure_count": (
                        len(source["candidates"]) - candidate_denominator
                    ),
                    "module_returned_candidate_count": returned_candidates,
                    "module_return_rate": (
                        returned_candidates / candidate_denominator
                        if candidate_denominator and field_keys else None
                    ),
                    "returned_field_slot_count": returned_slots,
                    "field_slot_count": slot_denominator,
                    "field_completeness": (
                        returned_slots / slot_denominator
                        if slot_denominator else None
                    ),
                    "fully_returned_candidate_count": (
                        fully_returned_candidates
                    ),
                    "fully_returned_rate": (
                        fully_returned_candidates / candidate_denominator
                        if candidate_denominator and field_keys else None
                    ),
                    "status_distribution": status_distribution,
                }

            module_populations = {
                key: population_metric(key) for key in populations
            }
            field_rows: list[dict[str, Any]] = []
            for field in module_fields:
                row = dict(field)
                row["populations"] = {}
                for population_key, population in populations.items():
                    eligible = [
                        item for item in population["candidates"]
                        if detail_success(item)
                    ]
                    returned_count = sum(
                        bool(candidate_field_map(candidate).get(
                            field["field_key"], {}
                        ).get("returned_nonempty"))
                        for candidate in eligible
                    )
                    row["populations"][population_key] = {
                        "returned_count": returned_count,
                        "candidate_count": len(eligible),
                        "return_rate": (
                            returned_count / len(eligible)
                            if eligible else None
                        ),
                    }
                hit_rate = row["populations"]["hit"]["return_rate"]
                nonmatched_rate = row["populations"][
                    "nonmatched"
                ]["return_rate"]
                row["hit_nonmatched_delta"] = (
                    hit_rate - nonmatched_rate
                    if hit_rate is not None and nonmatched_rate is not None
                    else None
                )
                field_rows.append(row)
            hit_completeness = module_populations["hit"][
                "field_completeness"
            ]
            nonmatched_completeness = module_populations["nonmatched"][
                "field_completeness"
            ]
            modules.append({
                "module": module,
                "business_field_count": len(module_fields),
                "field_keys": field_keys,
                "populations": module_populations,
                "hit_nonmatched_completeness_delta": (
                    hit_completeness - nonmatched_completeness
                    if (
                        hit_completeness is not None
                        and nonmatched_completeness is not None
                    ) else None
                ),
                "fields": field_rows,
            })
        classified_count = sum(
            item.get("identity", {}).get("judgement")
            in {"HIT", "NOT_HIT", "SUSPECTED"}
            for item in candidates
            if detail_success(item)
        )
        successful_count = sum(detail_success(item) for item in candidates)
        return {
            "module_order": list(REPORT_V5_PROFILE_MODULES),
            "populations": population_summary,
            "unclassified_success_count": successful_count - classified_count,
            "modules": modules,
        }

    @staticmethod
    def _enrich_report_v5_core_metrics(
        core_metrics: dict[str, dict[str, Any]],
        query_explorer: dict[str, Any],
    ) -> dict[str, dict[str, Any]]:
        """为核心指标补充用途、公式和本次计算贡献明细。

        功能说明：计算结果仍完全复用 metrics-v4；本方法只把 Query 或
        Candidate 的贡献项整理进报告快照，供指标卡抽屉解释“本次怎么算”。
        每个 Query 对命中率最多贡献一次，所有质量指标继续遵守主命中或
        非命中候选人的既有聚合口径。

        参数说明：
            core_metrics: 已冻结的七项核心指标。
            query_explorer: 包含全部 Query/Candidate 的 v5 快照。

        返回值：
            dict: 增加 label、purpose、formula、breakdown 的指标对象。

        异常说明：
            历史或不完整候选数据会被跳过，不影响聚合结果本身。
        """

        result = {
            key: dict(value)
            for key, value in core_metrics.items()
        }
        query_items = query_explorer.get("items", [])

        # 将候选人级指标按身份范围重新汇总为报告专用的分组快照。
        # 处理层仍按既有口径计算正式 value；这里仅补齐“全部/主/非主
        # HIT”与“非命中候选人”的人数、字段槽位和辅助计算说明，避免
        # 报告页面把不同统计范围混在同一条算式里。
        candidate_contexts: list[dict[str, Any]] = []
        for item in query_items:
            candidate_run = item.get("candidate_run") or {}
            query = candidate_run.get("query") or {}
            for candidate in candidate_run.get("candidates", []):
                candidate_contexts.append({
                    "query_id": query.get("query_id"),
                    "display_name": item.get("display_name") or query.get(
                        "query_id"
                    ),
                    "candidate": candidate,
                })

        def population_group(
            metric_key: str,
            group_key: str,
            label: str,
            description: str,
            contexts: list[dict[str, Any]],
            *,
            formal: bool = False,
        ) -> dict[str, Any]:
            """汇总一个候选人身份分组，供报告展示而非重算正式指标。"""

            scored: list[tuple[dict[str, Any], dict[str, Any]]] = []
            for context in contexts:
                candidate_metric = (
                    context["candidate"].get("metrics", {}).get(metric_key, {})
                )
                if candidate_metric.get("value") is not None:
                    scored.append((context, candidate_metric))
            group: dict[str, Any] = {
                "key": group_key,
                "label": label,
                "description": description,
                "formal": formal,
                "candidate_count": len(contexts),
                "scored_candidate_count": len(scored),
            }
            if metric_key == "matched_accuracy":
                macro_numerator = sum(
                    float(item[1].get("value") or 0) for item in scored
                )
                macro_denominator = len(scored)
                field_numerator = sum(
                    float(item[1].get("numerator") or 0) for item in scored
                )
                field_denominator = sum(
                    int(item[1].get("denominator") or 0) for item in scored
                )
                group.update({
                    "calculation_label": "候选人宏平均",
                    "numerator": macro_numerator,
                    "denominator": macro_denominator,
                    "value": (
                        macro_numerator / macro_denominator
                        if macro_denominator else None
                    ),
                    "numerator_label": "候选人准确度之和",
                    "denominator_label": "参与准确度计算的候选人",
                    "auxiliary_calculation": {
                        "label": "字段级微平均（辅助观察）",
                        "numerator": field_numerator,
                        "denominator": field_denominator,
                        "value": (
                            field_numerator / field_denominator
                            if field_denominator else None
                        ),
                        "numerator_label": "字段最终得分之和",
                        "denominator_label": "参与准确度计算的字段数",
                    },
                })
            else:
                numerator = sum(
                    float(item[1].get("numerator") or 0) for item in scored
                )
                denominator = sum(
                    int(item[1].get("denominator") or 0) for item in scored
                )
                denominator_label = (
                    "可比较资料字段槽位"
                    if metric_key == "nonmatched_baseline_overlap"
                    else "启用字段槽位"
                )
                group.update({
                    "calculation_label": (
                        "候选人资料相似度"
                        if metric_key == "nonmatched_baseline_overlap"
                        else "资料完整度"
                    ),
                    "numerator": numerator,
                    "denominator": denominator,
                    "value": numerator / denominator if denominator else None,
                    "numerator_label": (
                        "资料相似得分"
                        if metric_key == "nonmatched_baseline_overlap"
                        else "实际返回字段数"
                    ),
                    "denominator_label": denominator_label,
                })
            return group

        hit_contexts = [
            context
            for context in candidate_contexts
            if context["candidate"].get("identity", {}).get("judgement") == "HIT"
        ]
        primary_hit_contexts = [
            context
            for context in hit_contexts
            if context["candidate"].get("identity", {}).get("is_primary_hit")
        ]
        secondary_hit_contexts = [
            context
            for context in hit_contexts
            if not context["candidate"].get("identity", {}).get("is_primary_hit")
        ]
        nonmatched_contexts = [
            context
            for context in candidate_contexts
            if context["candidate"].get("identity", {}).get("judgement")
            in {"NOT_HIT", "SUSPECTED"}
        ]
        population_definitions = {
            "matched_accuracy": [
                ("all_hit", "全部 HIT", "所有规则或人工判定为 HIT 的候选人", hit_contexts, False),
                ("primary_hit", "主 HIT", "每个 Query 仅保留 rank_score 最高的 HIT；正式整体指标口径", primary_hit_contexts, True),
                ("secondary_hit", "非主 HIT", "同一 Query 内其余 HIT，仅用于辅助分析", secondary_hit_contexts, False),
            ],
            "matched_completeness": [
                ("all_hit", "全部 HIT", "所有规则或人工判定为 HIT 的候选人", hit_contexts, False),
                ("primary_hit", "主 HIT", "每个 Query 仅保留 rank_score 最高的 HIT；正式整体指标口径", primary_hit_contexts, True),
                ("secondary_hit", "非主 HIT", "同一 Query 内其余 HIT，仅用于辅助分析", secondary_hit_contexts, False),
            ],
            "nonmatched_data_completeness": [
                ("nonmatched", "非命中 / 疑似", "所有 NOT_HIT 与 SUSPECTED 候选人", nonmatched_contexts, True),
            ],
            "nonmatched_baseline_overlap": [
                ("nonmatched", "非命中 / 疑似", "所有 NOT_HIT 与 SUSPECTED 候选人", nonmatched_contexts, True),
            ],
        }
        for key, definition in REPORT_V5_CORE_METRIC_DEFINITIONS.items():
            metric = result.setdefault(key, {})
            metric.update(definition)
            breakdown: list[dict[str, Any]] = []
            for item in query_items:
                candidate_run = item.get("candidate_run") or {}
                query = candidate_run.get("query") or {}
                query_id = query.get("query_id")
                if not query_id:
                    continue
                candidate_count = int(query.get("candidate_count") or 0)
                executed = query.get("result_status") != "EXECUTION_FAILED"
                hit = bool(query.get("retrieval_success"))
                if key in {
                    "candidate_return_rate",
                    "retrieval_success",
                    "conditional_hit_rate",
                }:
                    if not executed:
                        continue
                    if key == "conditional_hit_rate" and not candidate_count:
                        continue
                    contribution = (
                        candidate_count > 0
                        if key == "candidate_return_rate"
                        else hit
                    )
                    breakdown.append({
                        "query_id": query_id,
                        "display_name": item.get("display_name") or query_id,
                        "candidate_count": candidate_count,
                        "result": "计入分子" if contribution else "仅计入分母",
                        "value": 1.0 if contribution else 0.0,
                    })
                    continue
                for candidate in candidate_run.get("candidates", []):
                    identity = candidate.get("identity", {})
                    judgement = identity.get("judgement")
                    if key in {"matched_accuracy", "matched_completeness"}:
                        if (
                            judgement != "HIT"
                            or not identity.get("is_primary_hit")
                        ):
                            continue
                    elif judgement not in {"NOT_HIT", "SUSPECTED"}:
                        continue
                    candidate_metric = candidate.get("metrics", {}).get(key, {})
                    value = candidate_metric.get("value")
                    if value is None:
                        continue
                    component_score_key = {
                        "matched_accuracy": "accuracy_score",
                        "matched_completeness": "data_completeness_score",
                        "nonmatched_data_completeness": (
                            "data_completeness_score"
                        ),
                        "nonmatched_baseline_overlap": (
                            "baseline_coverage_score"
                        ),
                    }.get(key)
                    field_components = [
                        {
                            "field_key": field.get("field_key"),
                            "display_name": field.get("display_name"),
                            "score": field.get(component_score_key),
                        }
                        for field in candidate.get("field_comparisons", [])
                        if (
                            component_score_key
                            and field.get(component_score_key) is not None
                            and not (
                                key == "nonmatched_baseline_overlap"
                                and field.get("field_key")
                                in NONMATCHED_SIMILARITY_EXCLUDED_FIELDS
                            )
                        )
                    ]
                    breakdown.append({
                        "query_id": query_id,
                        "display_name": item.get("display_name") or query_id,
                        "candidate_id": candidate.get("candidate_id"),
                        "candidate_rank": candidate.get("candidate_rank"),
                        "numerator": candidate_metric.get("numerator"),
                        "denominator": candidate_metric.get("denominator"),
                        "value": value,
                        "field_components": field_components,
                        "result": (
                            "主命中候选人"
                            if key in {
                                "matched_accuracy",
                                "matched_completeness",
                            }
                            else judgement
                        ),
                    })
            metric["breakdown"] = breakdown
            metric["breakdown_count"] = len(breakdown)
            if key in population_definitions:
                metric["population_groups"] = [
                    population_group(
                        key,
                        group_key,
                        label,
                        description,
                        contexts,
                        formal=formal,
                    )
                    for group_key, label, description, contexts, formal
                    in population_definitions[key]
                ]
            if key == "matched_accuracy" and breakdown:
                metric["calculation_expression"] = (
                    "主 HIT "
                    f"{len(breakdown)} 人："
                    + " + ".join(
                        f"{float(item['value']) * 100:.4f}%"
                        for item in breakdown
                    )
                    + f"；再除以 {len(breakdown)} 位主命中候选人"
                )
                field_numerator = sum(
                    float(item.get("numerator") or 0)
                    for item in breakdown
                )
                field_denominator = sum(
                    int(item.get("denominator") or 0)
                    for item in breakdown
                )
                metric["auxiliary_calculation"] = {
                    "label": "字段级微平均（辅助观察）",
                    "numerator": field_numerator,
                    "denominator": field_denominator,
                    "value": (
                        field_numerator / field_denominator
                        if field_denominator else None
                    ),
                }
            elif metric.get("denominator"):
                primary_group = next(
                    (
                        item for item in metric.get("population_groups", [])
                        if item.get("formal")
                    ),
                    None,
                )
                if primary_group:
                    metric["calculation_expression"] = (
                        f"{primary_group['label']} "
                        f"{primary_group['candidate_count']} 人："
                        f"{primary_group['numerator_label']} "
                        f"{float(primary_group['numerator']):.6f} ÷ "
                        f"{primary_group['denominator_label']} "
                        f"{primary_group['denominator']}"
                    )
                else:
                    metric["calculation_expression"] = (
                        f"{float(metric.get('numerator') or 0):.6f} ÷ "
                        f"{int(metric['denominator'])}"
                    )
        return result

    def _build_report_v5_overview(
        self,
        metrics: dict[str, Any],
        query_explorer: dict[str, Any],
    ) -> dict[str, Any]:
        """构造 v5 首屏摘要与可解释明细，不改变 metrics-v4 计算结果。"""

        execution = metrics.get("execution_summary", {})
        identity = metrics.get("identity_summary", {})
        core_metrics = self._enrich_report_v5_core_metrics({
            field_key: dict(metrics.get(field_key, {}))
            for field_key in (
                "candidate_return_rate",
                "conditional_hit_rate",
                "retrieval_success",
                "matched_accuracy",
                "matched_completeness",
                "nonmatched_data_completeness",
                "nonmatched_baseline_overlap",
            )
        }, query_explorer)
        query_count = int(execution.get("query_count") or 0)
        failed_count = int(execution.get("execution_failed_count") or 0)
        if not query_count:
            blocking_alert = {
                "code": "NO_FORMAL_QUERY",
                "message": "没有正式 Query，无法生成核心评测结论。",
            }
        elif failed_count == query_count:
            blocking_alert = {
                "code": "ALL_QUERY_EXECUTION_FAILED",
                "message": "全部 Query 执行失败，核心指标不可用。",
            }
        else:
            blocking_alert = None
        success_metric = core_metrics["retrieval_success"]
        success_value = success_metric.get("value")
        if blocking_alert:
            title = "本次报告存在阻断问题"
        elif success_value is None:
            title = "本次检索已完成，核心质量指标暂不可用"
        else:
            title = "本次检索已完成，可查看核心指标与候选人明细"
        description = (
            f"本次共执行 {query_count} 个 Query，"
            f"其中 {execution.get('has_candidates_count', 0)} 个返回候选人，"
            f"{identity.get('primary_hit_query_count', 0)} 个命中基准人物。"
        )
        return {
            "conclusion": {
                "level": "WARNING" if blocking_alert else "INFO",
                "title": title,
                "description": description,
            },
            "core_metrics": core_metrics,
            "secondary_metrics": {
                "query_count": query_count,
                "has_candidates_count": execution.get(
                    "has_candidates_count", 0
                ),
                "no_candidates_count": execution.get(
                    "no_candidates_count", 0
                ),
                "candidate_count": identity.get("final_candidate_count", 0),
                "primary_hit_query_count": identity.get(
                    "primary_hit_query_count", 0
                ),
                "no_hit_query_count": identity.get("no_hit_query_count", 0),
                "pending_query_count": identity.get("pending_query_count", 0),
                "execution_failed_count": failed_count,
                "detail_success_count": execution.get(
                    "detail_success_count", 0
                ),
                "detail_failure_count": execution.get(
                    "detail_failure_count", 0
                ),
            },
            "blocking_alert": blocking_alert,
        }

    @staticmethod
    def _build_report_v5_optional_sections(
        candidate_process: sqlite3.Row,
        candidate_metrics: dict[str, Any],
        threshold_assessment: dict[str, Any],
        comparison: dict[str, Any] | None,
    ) -> dict[str, bool]:
        """集中计算 v5 可选章节开关，避免模板重复判断业务口径。"""

        has_threshold_actual = any(
            item.get("actual") is not None
            for stage in threshold_assessment.get("stages", {}).values()
            for item in stage.get("items", {}).values()
        )
        show_threshold = bool(
            candidate_process["threshold_profile_id"]
            and threshold_assessment.get("configured_count", 0)
            and has_threshold_actual
        )
        cost_metrics = candidate_metrics.get("cost_metrics", {})
        show_cost = any(
            item.get("total") is not None
            for item in cost_metrics.values()
            if isinstance(item, dict)
        )
        pdl_metrics = candidate_metrics.get("pdl_metrics", {})
        return {
            "show_comparison": comparison is not None,
            "show_cost": show_cost,
            "show_pdl": bool(pdl_metrics.get("known_count", 0)),
            "show_threshold": show_threshold,
            # 当前系统没有独立建议配置，首版只展示满足条件的参考线判断。
            "show_recommendation": False,
            "show_evidence": True,
        }

    @staticmethod
    def _build_report_v5_diagnostics(
        *,
        warnings: list[str],
        not_ready_reasons: list[dict[str, Any]],
        candidate_process: sqlite3.Row,
    ) -> dict[str, Any]:
        """将普通风险和技术信息收敛为页面底部的结构化诊断数据。"""

        items = [
            {
                "severity": "WARNING",
                "code": "REPORT_WARNING",
                "message": warning,
            }
            for warning in warnings
        ]
        items.extend({
            "severity": "WARNING",
            "code": item.get("reason_code", "UNKNOWN"),
            "message": item.get("reason", ""),
            "details": item.get("details", []),
        } for item in not_ready_reasons)
        return {
            "items": items,
            "technical_metadata": {
                "schema_version": candidate_process["schema_version"],
                "rule_version": candidate_process["rule_version"],
                "metrics_rule_version": METRICS_RULE_VERSION,
                "source_type": candidate_process["source_type"],
            },
        }

    @staticmethod
    def assess_evaluation_thresholds(
        metrics: dict[str, Any],
        thresholds: Any,
    ) -> dict[str, Any]:
        """按 Query Stage 判断参考线并给出可解释建议。

        参数说明:
            metrics: ``metrics-v2`` Process 指标模型。
            thresholds: Evaluation 保存的参考线快照。

        返回值:
            每项实际值、参考线、PASS/FAIL/NOT_CONFIGURED/NOT_READY，
            以及“建议上线/继续优化/暂不能判断”的总体建议。

        异常说明:
            ReviewValidationError: 阈值快照损坏或指标规则不是 v2。
        """

        normalized = normalize_evaluation_thresholds(thresholds)
        groups = {
            item["query_stage"]: item
            for item in metrics.get("grouped_metrics", [])
        }
        stage_results: dict[str, Any] = {}
        configured_count = 0
        pass_count = 0
        fail_count = 0
        not_ready_count = 0
        for query_stage in sorted(SUPPORTED_QUERY_STAGES):
            group = groups.get(query_stage)
            actual_values = {
                "min_retrieval_success": (
                    group["quality_metrics"]["retrieval_success"]["value"]
                    if group
                    else None
                ),
                "min_matched_completeness": (
                    group["quality_metrics"]["matched_completeness"]["value"]
                    if group
                    else None
                ),
                "min_matched_accuracy": (
                    group["quality_metrics"]["matched_accuracy"]["value"]
                    if group
                    else None
                ),
                "max_average_total_cost": (
                    group["cost_metrics"]["total_cost"]["average"]
                    if group
                    else None
                ),
                "max_average_search_duration_ms": (
                    group["cost_metrics"]["search_duration_ms"]["average"]
                    if group
                    else None
                ),
            }
            items: dict[str, Any] = {}
            for field_key, rule in EVALUATION_THRESHOLD_FIELDS.items():
                threshold = normalized[query_stage][field_key]
                actual = actual_values[field_key]
                if threshold is None:
                    status = "NOT_CONFIGURED"
                    reason = "未配置参考线"
                else:
                    configured_count += 1
                    if actual is None:
                        status = "NOT_READY"
                        reason = "指标缺失或尚未完成正式复核"
                        not_ready_count += 1
                    else:
                        passed = (
                            actual >= threshold
                            if rule == "minimum_ratio"
                            else actual <= threshold
                        )
                        status = "PASS" if passed else "FAIL"
                        reason = (
                            "达到参考线" if passed else "未达到参考线"
                        )
                        if passed:
                            pass_count += 1
                        else:
                            fail_count += 1
                items[field_key] = {
                    "threshold": threshold,
                    "actual": actual,
                    "direction": (
                        "MINIMUM"
                        if rule == "minimum_ratio"
                        else "MAXIMUM"
                    ),
                    "status": status,
                    "reason": reason,
                }
            stage_results[query_stage] = {
                "query_count": group["query_count"] if group else 0,
                "items": items,
            }
        if fail_count:
            recommendation = "继续优化"
            recommendation_code = "CONTINUE_OPTIMIZATION"
        elif not configured_count or not_ready_count:
            recommendation = "暂不能判断"
            recommendation_code = "NOT_READY"
        else:
            recommendation = "建议上线"
            recommendation_code = "RECOMMEND_RELEASE"
        return {
            "threshold_snapshot": normalized,
            "stages": stage_results,
            "configured_count": configured_count,
            "pass_count": pass_count,
            "fail_count": fail_count,
            "not_ready_count": not_ready_count,
            "recommendation": recommendation,
            "recommendation_code": recommendation_code,
        }

    def build_report_model(
        self,
        *,
        report_id: str,
        candidate_process_id: str,
        baseline_process_id: str | None = None,
        data_marker: str = "REAL_TEST_DATA",
    ) -> dict[str, Any]:
        """构建 Web、静态 HTML 和 Excel 共用的不可变 ReportModel。"""

        if data_marker not in {"REAL_TEST_DATA", "MOCK"}:
            raise ReviewValidationError(
                "data_marker 只支持 REAL_TEST_DATA 或 MOCK"
            )
        candidate_process = self._report_process(candidate_process_id)
        candidate_metrics = self.calculate_process_metrics(
            candidate_process_id
        )
        try:
            threshold_source = json.loads(
                candidate_process["thresholds_json"] or "{}"
            )
        except (TypeError, json.JSONDecodeError) as exc:
            raise ReviewValidationError(
                "Evaluation 参考线快照 JSON 已损坏"
            ) from exc
        threshold_assessment = self.assess_evaluation_thresholds(
            candidate_metrics,
            threshold_source,
        )
        baseline_process = None
        baseline_metrics = None
        comparison = None
        if baseline_process_id:
            baseline_process = self._report_process(baseline_process_id)
            comparison = self.compare_processes(
                baseline_process_id,
                candidate_process_id,
            )
            baseline_metrics = comparison["baseline_metrics"]
        is_v4 = (
            candidate_metrics.get("metrics_rule_version")
            == METRICS_RULE_VERSION
        )
        is_modern = candidate_metrics.get("metrics_rule_version") in {
            V3_METRICS_RULE_VERSION,
            METRICS_RULE_VERSION,
        }
        if is_modern:
            candidate_modules = candidate_metrics.get("module_metrics", {})
            candidate_fields = candidate_metrics.get("field_metrics", {})
        else:
            candidate_modules, candidate_fields = (
                self._process_field_report_metrics(candidate_process)
            )
        if baseline_process is not None:
            if (
                baseline_metrics
                and baseline_metrics.get("metrics_rule_version") in {
                    V3_METRICS_RULE_VERSION,
                    METRICS_RULE_VERSION,
                }
            ):
                baseline_modules = baseline_metrics.get("module_metrics", {})
                baseline_fields = baseline_metrics.get("field_metrics", {})
            else:
                baseline_modules, baseline_fields = (
                    self._process_field_report_metrics(baseline_process)
                )
            module_metrics = self._merge_comparison_metrics(
                baseline_modules,
                candidate_modules,
            )
            field_metrics = self._merge_comparison_metrics(
                baseline_fields,
                candidate_fields,
            )
        else:
            module_metrics = candidate_modules
            field_metrics = candidate_fields
        formal_ready = candidate_metrics["formal_ready"] and (
            comparison is None or comparison["formal_ready"]
        )
        warnings = [
            "系统返回内容仅用于检索能力测试，不代表已经核实的事实。"
        ]
        if not formal_ready:
            warnings.append(
                "部分质量指标尚未就绪；本报告保留可计算的预览值，"
                "具体原因和修复入口见下方说明。"
            )
        if candidate_metrics["cost_status"]["status"] != "COMPLETE":
            warnings.append(
                "成本与 PDL 数据未完整接入，缺失值未按0计算。"
            )
        if comparison is not None:
            coverage = comparison["same_condition"]["coverage"]
            if coverage["input_signature_unavailable_count"]:
                warnings.append(
                    "部分历史结果缺少 Dataset 输入签名，已归入不可比数据，"
                    "未进入正式同条件结论。"
                )
            if coverage["input_signature_mismatch_count"]:
                warnings.append(
                    "部分 Query 输入签名不一致，已归入不可比数据。"
                )
            if coverage["missing_regression_query_count"]:
                warnings.append(
                    f"Candidate Run 缺少 "
                    f"{coverage['missing_regression_query_count']} 条回归 Query。"
                )
            if coverage["new_clue_count"]:
                warnings.append(
                    f"本阶段包含 {coverage['new_clue_count']} 条新增线索结果，"
                    "不表述为系统版本提升。"
                )
        if candidate_process["failed_queries"]:
            warnings.append(
                f"Candidate Run 含 {candidate_process['failed_queries']} 条失败或部分失败 Query。"
            )
        not_ready_reasons: list[dict[str, str]] = []
        if is_modern:
            seen_reasons: set[tuple[str, str]] = set()
            for metric_key in (
                "candidate_return_rate",
                "conditional_hit_rate",
                "retrieval_success",
                "matched_completeness",
                "matched_accuracy",
                "nonmatched_data_completeness",
                "nonmatched_baseline_overlap",
            ):
                metric = candidate_metrics.get(metric_key, {})
                codes = metric.get("reason_codes", [])
                reasons = metric.get("reasons", [])
                for reason_code in codes:
                    reason = {
                        "BASELINE_NOT_LINKED": "Query 尚未关联 Baseline",
                        "BASELINE_PERSON_NOT_FOUND": "关联人物不在所选 Baseline",
                        "IDENTITY_PENDING": "候选人身份归类尚未完成",
                        "NO_HIT_CONFIRMED": "已确认无命中，本指标不适用",
                        "NO_NONMATCHED_CONFIRMED": "尚无已确认非命中候选人",
                        "NO_BASELINE_OVERLAP_FIELDS": "没有可比较的非命中候选人资料字段",
                        "NO_EFFECTIVE_FIELDS": "没有有效评分字段交集",
                        "MANUAL_SCORE_PENDING": "人工准确率评分尚未完成",
                        "FIELD_NOT_CONNECTED": "接口字段尚未接入",
                        "NO_DENOMINATOR": "当前业务没有适用分母",
                    }.get(reason_code, reason_code)
                    key = (str(reason_code), str(reason))
                    if key in seen_reasons:
                        continue
                    seen_reasons.add(key)
                    not_ready_reasons.append({
                        "metric": metric_key,
                        "reason_code": str(reason_code),
                        "reason": str(reason),
                        "details": [str(item) for item in reasons],
                    })
            if candidate_metrics["cost_status"]["status"] != "COMPLETE":
                disconnected_fields = [
                    field_key
                    for field_key, item in candidate_metrics.get(
                        "cost_metrics",
                        {},
                    ).items()
                    if item.get("status") != "COMPLETE"
                ]
                pdl_metrics = candidate_metrics.get("pdl_metrics", {})
                if pdl_metrics.get("unknown_count", 0):
                    disconnected_fields.append("pdl_called")
                not_ready_reasons.append({
                    "metric": "task_public_fields",
                    "reason_code": "FIELD_NOT_CONNECTED",
                    "reason": (
                        "成本、耗时或 PDL 公共字段尚未完整接入"
                        if candidate_metrics["cost_status"]["status"] == "PARTIAL"
                        else "成本、耗时和 PDL 公共字段尚未接入"
                    ),
                    "details": sorted(set(disconnected_fields)),
                })
        field_alignment_summary = None
        field_alignment_matrix = None
        if is_modern and candidate_process["baseline_version"]:
            try:
                matrix = self.build_field_comparison_matrix(
                    candidate_process["schema_version"],
                    candidate_process["baseline_version"],
                    process_id=candidate_process_id,
                )
                field_alignment_summary = {
                    "severity_counts": matrix["severity_counts"],
                    "status_counts": matrix["status_counts"],
                    "field_count": len(matrix["fields"]),
                }
                field_alignment_matrix = matrix
            except FieldSchemaValidationError:
                field_alignment_summary = None
                field_alignment_matrix = None
        if is_v4 and comparison is not None:
            # 版本变化属于报告上下文：只有明确选中另一个可比 Process 时，
            # 才把字段/模块的前后值与 delta 写入不可变 ReportModel。
            candidate_metrics["regression_metrics"] = {
                "status": (
                    "READY" if comparison["formal_ready"] else "PARTIAL"
                ),
                "reason_codes": [],
                "reasons": [],
                "same_condition": comparison["same_condition"],
                "field_changes": field_metrics,
                "module_changes": module_metrics,
            }
        model = {
            "metadata": {
                "report_model_version": (
                    V5_REPORT_MODEL_VERSION
                    if is_v4 else (
                        V3_REPORT_MODEL_VERSION
                        if is_modern else V2_REPORT_MODEL_VERSION
                    )
                ),
                "metrics_rule_version": candidate_metrics.get(
                    "metrics_rule_version",
                    "metrics-v1",
                ),
                "report_id": report_id,
                "report_type": "COMPARE" if comparison else "SINGLE",
                "evaluation_id": candidate_process["evaluation_id"],
                "evaluation_name": candidate_process["evaluation_name"],
                "threshold_profile_id": candidate_process[
                    "threshold_profile_id"
                ],
                "threshold_profile_name": candidate_process[
                    "threshold_profile_name"
                ],
                "threshold_profile_version": candidate_process[
                    "threshold_profile_version"
                ],
                "evaluation_phase": candidate_process["evaluation_phase"],
                "baseline_evaluation_phase": (
                    baseline_process["evaluation_phase"]
                    if baseline_process
                    else None
                ),
                "generated_at": utc_now_text(),
                "data_marker": data_marker,
                "candidate_process_id": candidate_process_id,
                "baseline_process_id": baseline_process_id,
                "candidate_run_id": candidate_process["run_id"],
                "baseline_run_id": (
                    baseline_process["run_id"] if baseline_process else None
                ),
                "candidate_run_label": candidate_process["run_label"],
                "baseline_run_label": (
                    baseline_process["run_label"] if baseline_process else None
                ),
                "candidate_system_version": candidate_process[
                    "system_version"
                ],
                "baseline_system_version": (
                    baseline_process["system_version"]
                    if baseline_process
                    else None
                ),
                "schema_version": candidate_process["schema_version"],
                "schema_name": candidate_process["schema_name"],
                "baseline_version": candidate_process["baseline_version"],
                "rule_version": candidate_process["rule_version"],
                "source_type": candidate_process["source_type"],
            },
            "summary": {
                "formal_ready": formal_ready,
                "review_status": (
                    "REVIEWED" if formal_ready else "PENDING_REVIEW"
                ),
                "candidate": candidate_metrics,
                "baseline": baseline_metrics,
            },
            "execution_summary": candidate_metrics.get(
                "execution_summary",
                {},
            ),
            "data_return_summary": {
                "candidate_field_return": candidate_metrics.get(
                    "candidate_field_return",
                    {},
                ),
                "module_metrics": candidate_modules,
                "cost_status": candidate_metrics["cost_status"],
            },
            "identity_summary": candidate_metrics.get(
                "identity_summary",
                {},
            ),
            "identity_metrics": candidate_metrics.get("identity_metrics", {}),
            "baseline_quality_metrics": candidate_metrics.get(
                "baseline_quality_metrics", {}
            ),
            "non_hit_data_return": candidate_metrics.get(
                "non_hit_data_return", {}
            ),
            "regression_metrics": candidate_metrics.get(
                "regression_metrics", {}
            ),
            "field_alignment_summary": field_alignment_summary,
            "field_alignment_matrix": field_alignment_matrix or {},
            "not_ready_reasons": not_ready_reasons,
            "result_status_metrics": candidate_metrics.get(
                "result_status_metrics",
                {},
            ),
            "quality_metrics": candidate_metrics.get(
                "quality_metrics",
                {
                    # metrics-v4 及更早版本没有新增的两个 Query 漏斗指标。
                    # 旧快照仍可打开，但不把缺失字段伪造成 0%。
                    "retrieval_success": candidate_metrics.get(
                        "retrieval_success", {}
                    ),
                    "candidate_return_rate": candidate_metrics.get(
                        "candidate_return_rate",
                        {"status": "NOT_APPLICABLE", "value": None,
                         "numerator": 0, "denominator": 0,
                         "reasons": ["历史指标未生成"]},
                    ),
                    "conditional_hit_rate": candidate_metrics.get(
                        "conditional_hit_rate",
                        {"status": "NOT_APPLICABLE", "value": None,
                         "numerator": 0, "denominator": 0,
                         "reasons": ["历史指标未生成"]},
                    ),
                    "matched_completeness": candidate_metrics[
                        "matched_completeness"
                    ],
                    "matched_accuracy": candidate_metrics[
                        "matched_accuracy"
                    ],
                    "nonmatched_completeness": candidate_metrics[
                        "nonmatched_completeness"
                    ],
                    "nonmatched_data_completeness": candidate_metrics[
                        "nonmatched_data_completeness"
                    ],
                    "nonmatched_baseline_overlap": candidate_metrics[
                        "nonmatched_baseline_overlap"
                    ],
                },
            ),
            "cost_metrics": candidate_metrics.get("cost_metrics", {}),
            "pdl_metrics": candidate_metrics.get("pdl_metrics", {}),
            "confidence_metrics": candidate_metrics.get(
                "confidence_metrics",
                {},
            ),
            "grouped_metrics": candidate_metrics.get(
                "grouped_metrics",
                [],
            ),
            "comparison": comparison or {
                "same_condition": {},
                "new_clue": {},
                "not_comparable": {},
            },
            "threshold_assessment": threshold_assessment,
            "query_stage_metrics": self._query_stage_report_metrics(
                candidate_metrics
            ),
            "paired_metrics": comparison,
            "module_metrics": module_metrics,
            "field_metrics": field_metrics,
            "case_groups": self._report_case_groups(
                candidate_process,
                comparison,
            ),
            "cost_status": candidate_metrics["cost_status"],
            "warnings": warnings,
        }
        if is_v4:
            # v5 仅用于新 metrics-v4 报告：在生成时一次性冻结完整查询和候选人
            # 明细，页面展示不再依赖临时数据库拼装，也不会把 Raw payload 写入快照。
            query_explorer = self._build_report_v5_query_explorer(
                candidate_process=candidate_process,
                candidate_metrics=candidate_metrics,
                baseline_process=baseline_process,
                baseline_metrics=baseline_metrics,
                comparison=comparison,
            )
            overview = self._build_report_v5_overview(
                candidate_metrics,
                query_explorer,
            )
            overview["secondary_metrics"]["candidate_count"] = (
                query_explorer["total_candidate_count"]
            )
            processing_scope = self._build_report_v5_processing_scope(
                candidate_process,
                candidate_metrics,
            )
            optional_sections = self._build_report_v5_optional_sections(
                candidate_process,
                candidate_metrics,
                threshold_assessment,
                comparison,
            )
            # 未配置参考线时不在 v5 报告中输出空的判断对象，供后续模板直接隐藏。
            if not optional_sections["show_threshold"]:
                model["threshold_assessment"] = None
            model.update({
                "overview": overview,
                "processing_scope": processing_scope,
                "module_return_overview": (
                    self._build_report_v5_module_return_overview(
                        query_explorer,
                        processing_scope,
                    )
                ),
                "query_explorer": query_explorer,
                "diagnostics": self._build_report_v5_diagnostics(
                    warnings=warnings,
                    not_ready_reasons=not_ready_reasons,
                    candidate_process=candidate_process,
                ),
                "optional_sections": optional_sections,
            })
        return model

    def _report_directory(
        self,
        evaluation_id: str,
        report_id: str,
    ) -> Path:
        """返回受控报告目录并验证两个外部标识。"""

        validate_storage_id(evaluation_id, "evaluation_id")
        validate_storage_id(report_id, "report_id")
        target = (self.report_dir / evaluation_id / report_id).resolve()
        try:
            target.relative_to(self.report_dir.resolve())
        except ValueError as exc:
            raise ReviewValidationError("报告目录越出 SEARCH_REPORT_DIR") from exc
        return target

    def _report_export_records(
        self,
        process: sqlite3.Row,
        metrics: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """生成 processed Excel 使用的 Query/Candidate/Failure 统一记录。"""

        query_metric_map = {
            row["query_id"]: row for row in metrics["query_metrics"]
        }
        query_rows = self.store.fetch_all(
            """
            SELECT rq.*, bp.display_name AS baseline_display_name
            FROM run_queries AS rq
            LEFT JOIN baseline_people AS bp
              ON bp.baseline_version = ?
             AND bp.person_id = rq.person_id
            WHERE rq.run_id = ? ORDER BY rq.query_id
            """,
            (process["baseline_version"], process["run_id"]),
        )
        records: list[dict[str, Any]] = []
        for row in query_rows:
            metric = query_metric_map.get(row["query_id"], {})
            task_fields = metric.get("task_fields", {})
            if not isinstance(task_fields, dict):
                task_fields = {}
            task_processing_errors = metric.get(
                "task_processing_errors",
                [],
            )
            if not isinstance(task_processing_errors, list):
                task_processing_errors = []
            records.append(
                {
                    "record_type": "query",
                    "run_id": process["run_id"],
                    "run_label": process["run_label"],
                    "system_version": process["system_version"],
                    "evaluation_phase": process["evaluation_phase"],
                    "query_id": row["query_id"],
                    "person_id": row["person_id"],
                    "person_id_source": row["person_id_source"],
                    "baseline_display_name": row[
                        "baseline_display_name"
                    ],
                    "baseline_match_status": (
                        "UNLINKED"
                        if not row["person_id"]
                        else "MATCHED"
                        if row["baseline_display_name"]
                        else "NOT_CONFIGURED"
                        if not process["baseline_version"]
                        else "NOT_FOUND"
                    ),
                    "query_stage": row["query_stage"],
                    "task_id": row["task_id"],
                    "query_status": row["status"],
                    "result_status": metric.get("result_status"),
                    "candidate_count_total": row["candidate_count_total"],
                    "candidate_count_listed": row[
                        "candidate_count_listed"
                    ],
                    "detail_success_count": row["detail_success_count"],
                    "detail_failure_count": row["detail_failure_count"],
                    "llm_cost": task_fields.get("llm_cost"),
                    "third_party_cost": task_fields.get(
                        "third_party_cost"
                    ),
                    "total_cost": task_fields.get("total_cost"),
                    "pdl_called": task_fields.get("pdl_called"),
                    "search_duration_ms": task_fields.get(
                        "search_duration_ms"
                    ),
                    "retrieval_success": metric.get("retrieval_success"),
                    "matched_completeness": metric.get(
                        "matched_completeness"
                    ),
                    "matched_accuracy": metric.get("matched_accuracy"),
                    "identity_state": metric.get("identity_state"),
                    "formal_ready": metric.get("formal_ready", False),
                    "processing_errors": task_processing_errors,
                }
            )
        candidate_rows = self.store.fetch_all(
            """
            SELECT c.*, rq.person_id, rq.query_stage, rq.task_id,
                   rq.status AS query_status,
                   pc.fields_json, pc.empty_fields_json,
                   pc.processing_errors_json,
                   rv.judgement, rv.reason, rv.evidence,
                   rv.field_scores_json, rv.reviewer, rv.review_note,
                   rv.reviewed_at, rv.classification_source,
                   rv.is_primary_hit
            FROM processed_candidates AS pc
            JOIN candidates AS c ON c.candidate_pk = pc.candidate_pk
            JOIN run_queries AS rq
              ON rq.run_id = c.run_id AND rq.query_id = c.query_id
            LEFT JOIN reviews AS rv
              ON rv.process_id = pc.process_id
             AND rv.candidate_pk = pc.candidate_pk
            WHERE pc.process_id = ?
            ORDER BY c.query_id, c.candidate_rank
            """,
            (process["process_id"],),
        )
        for row in candidate_rows:
            records.append(
                {
                    "record_type": "candidate",
                    "evaluation_id": process["evaluation_id"],
                    "process_id": process["process_id"],
                    "run_id": process["run_id"],
                    "run_label": process["run_label"],
                    "system_version": process["system_version"],
                    "query_id": row["query_id"],
                    "person_id": row["person_id"],
                    "query_stage": row["query_stage"],
                    "task_id": row["task_id"],
                    "query_status": row["query_status"],
                    "candidate_pk": row["candidate_pk"],
                    "candidate_id": row["candidate_id"],
                    "candidate_rank": row["candidate_rank"],
                    "rank_score": row["rank_score"],
                    "detail_status": row["detail_status"],
                    "detail_error": row["detail_error"],
                    "judgement": row["judgement"] or "PENDING_REVIEW",
                    "reason": row["reason"] or "",
                    "classification_source": (
                        row["classification_source"] or "SUGGESTED"
                    ),
                    "is_primary_hit": bool(row["is_primary_hit"]),
                    "evidence": row["evidence"] or "",
                    "reviewer": row["reviewer"] or "",
                    "review_note": row["review_note"] or "",
                    "reviewed_at": row["reviewed_at"],
                    "fields": json.loads(row["fields_json"]),
                    "empty_fields": json.loads(row["empty_fields_json"]),
                    "field_scores": (
                        json.loads(row["field_scores_json"])
                        if row["field_scores_json"]
                        else {}
                    ),
                    "processing_errors": json.loads(
                        row["processing_errors_json"]
                    ),
                }
            )
        for row in self.store.fetch_all(
            """
            SELECT query_id, candidate_id, scope, stage, error, created_at
            FROM failures WHERE run_id = ? ORDER BY created_at
            """,
            (process["run_id"],),
        ):
            records.append({"record_type": "failure", **dict(row)})
        return records

    def create_report(
        self,
        *,
        candidate_process_id: str,
        baseline_process_id: str | None = None,
        data_marker: str = "REAL_TEST_DATA",
        report_id: str | None = None,
    ) -> ReportResult:
        """创建不可变报告模型快照和 processed Excel 输入文件。"""

        object_id = validate_storage_id(
            report_id or f"report_{uuid.uuid4().hex}",
            "report_id",
        )
        candidate_process = self._report_process(candidate_process_id)
        model = self.build_report_model(
            report_id=object_id,
            candidate_process_id=candidate_process_id,
            baseline_process_id=baseline_process_id,
            data_marker=data_marker,
        )
        report_directory = self._report_directory(
            candidate_process["evaluation_id"],
            object_id,
        )
        if report_directory.exists():
            raise ReviewValidationError(
                f"报告目录已经存在，未覆盖: {object_id}"
            )
        report_directory.mkdir(parents=True, exist_ok=False)
        generated_date = model["metadata"]["generated_at"][:10].replace("-", "")
        version_slug = re.sub(
            r"[^A-Za-z0-9_.-]+",
            "-",
            str(candidate_process["system_version"]),
        ).strip("-.")[:48] or "version"
        report_filename = (
            f"{generated_date}_{candidate_process['evaluation_id']}_"
            f"{version_slug}_report.html"
        )
        html_relative = (
            Path(candidate_process["evaluation_id"])
            / object_id
            / report_filename
        ).as_posix()
        try:
            model_path = report_directory / "report_model.json"
            model_path.write_text(
                json.dumps(
                    model,
                    ensure_ascii=False,
                    indent=2,
                    sort_keys=True,
                ),
                encoding="utf-8",
            )
            self._write_jsonl(
                report_directory / "processed_export.jsonl",
                self._report_export_records(
                    candidate_process,
                    model["summary"]["candidate"],
                ),
            )
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    INSERT INTO reports(
                        report_id, evaluation_id, baseline_process_id,
                        candidate_process_id, report_type, status,
                        metrics_json, html_file, excel_file, created_at
                    ) VALUES (?, ?, ?, ?, ?, 'READY', ?, ?, NULL, ?)
                    """,
                    (
                        object_id,
                        candidate_process["evaluation_id"],
                        baseline_process_id,
                        candidate_process_id,
                        model["metadata"]["report_type"],
                        json_text(model),
                        html_relative,
                        model["metadata"]["generated_at"],
                    ),
                )
        except Exception:
            self._cleanup_created(report_directory)
            raise
        return ReportResult(
            report_id=object_id,
            model=model,
            html_file=html_relative,
            excel_file=None,
        )

    def _report_file_from_row(
        self,
        row: sqlite3.Row,
        relative_path: str,
    ) -> Path:
        """把数据库报告相对路径解析到受控报告目录。"""

        target = (self.report_dir / relative_path).resolve()
        try:
            target.relative_to(self.report_dir.resolve())
        except ValueError as exc:
            raise ReviewValidationError("报告文件路径越界") from exc
        expected_directory = self._report_directory(
            row["evaluation_id"],
            row["report_id"],
        )
        try:
            target.relative_to(expected_directory)
        except ValueError as exc:
            raise ReviewValidationError("报告文件不属于对应报告目录") from exc
        return target

    def save_report_html(self, report_id: str, html: str) -> str:
        """原子写入已创建报告的静态 HTML，不修改指标快照。"""

        row = self.store.fetch_one(
            "SELECT * FROM reports WHERE report_id = ?",
            (report_id,),
        )
        if row is None:
            raise ReviewValidationError(f"报告不存在: {report_id}")
        target = self._report_file_from_row(row, row["html_file"])
        target.parent.mkdir(parents=True, exist_ok=True)
        temporary = target.with_suffix(".html.tmp")
        temporary.write_text(html, encoding="utf-8")
        temporary.replace(target)
        return row["html_file"]

    def export_report_excel(
        self,
        report_id: str,
        *,
        exporter_path: Path | str | None = None,
        timeout_seconds: int = 180,
    ) -> str:
        """调用现有 Excel 构建器的 processed 模式并登记导出文件。"""

        row = self.store.fetch_one(
            "SELECT * FROM reports WHERE report_id = ?",
            (report_id,),
        )
        if row is None:
            raise ReviewValidationError(f"报告不存在: {report_id}")
        report_directory = self._report_directory(
            row["evaluation_id"],
            report_id,
        )
        input_path = report_directory / "processed_export.jsonl"
        model_path = report_directory / "report_model.json"
        output_path = report_directory / Path(row["html_file"]).name.replace(
            "_report.html",
            "_report.xlsx",
        )
        script = (
            Path(exporter_path).resolve()
            if exporter_path is not None
            else Path(__file__).resolve().parent / "result_to_excel.py"
        )
        completed = subprocess.run(
            [
                sys.executable,
                str(script),
                "processed",
                "--input-file",
                str(input_path),
                "--report-model",
                str(model_path),
                "--output",
                str(output_path),
            ],
            cwd=script.parent,
            text=True,
            capture_output=True,
            timeout=timeout_seconds,
            check=False,
        )
        if completed.returncode != 0 or not output_path.is_file():
            message = (
                completed.stderr.strip()
                or completed.stdout.strip()
                or "Excel 构建器未生成文件"
            )
            raise RuntimeError(f"Excel 导出不可用: {message[:1000]}")
        relative_path = (
            Path(row["evaluation_id"]) / report_id / output_path.name
        ).as_posix()
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE reports SET excel_file = ?
                WHERE report_id = ?
                """,
                (relative_path, report_id),
            )
        return relative_path

    def resolve_report_artifact(
        self,
        report_id: str,
        artifact: str,
    ) -> Path:
        """解析可下载的 html/excel 报告文件并验证存在性。"""

        columns = {"html": "html_file", "excel": "excel_file"}
        column = columns.get(artifact)
        if column is None:
            raise ReviewValidationError("不支持的报告文件类型")
        row = self.store.fetch_one(
            f"SELECT * FROM reports WHERE report_id = ?",
            (report_id,),
        )
        if row is None or not row[column]:
            raise ReviewValidationError("报告文件不存在")
        target = self._report_file_from_row(row, row[column])
        if not target.is_file():
            raise ReviewValidationError("报告文件不存在")
        return target

    def mark_report_failed(self, report_id: str) -> None:
        """静态报告渲染失败时只标记报告，不删除模型和 Raw。"""

        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE reports SET status = 'FAILED'
                WHERE report_id = ?
                """,
                (report_id,),
            )

    def create_execution_run(
        self,
        *,
        evaluation_id: str,
        dataset_id: str,
        run_label: str,
        system_version: str,
        evaluation_phase: str,
        run_id: str | None = None,
    ) -> str:
        """创建一个等待后台执行的 Run 和全部 PENDING Query。

        功能说明:
            仅创建数据库记录，不在 Web 请求线程中调用真实接口。全局只允许
            一个 ``PENDING`` 或 ``RUNNING`` 的 EXECUTION Run。

        参数说明:
            evaluation_id: 已存在的评测标识。
            dataset_id: 已导入且至少包含一条 Query 的数据集标识。
            run_label: 本次运行标签。
            system_version: 被测系统版本。
            evaluation_phase: 明确的评估阶段；新执行不允许 UNSPECIFIED。
            run_id: 测试或外部编排可显式提供的唯一标识。

        返回值:
            新创建的 Run ID。

        异常说明:
            ImportValidationError: 关联对象不存在或必填值为空。
            ActiveRunError: 已有活动执行任务。
        """

        if not run_label.strip() or not system_version.strip():
            raise ImportValidationError(["run_label 和 system_version 不能为空"])
        if (
            evaluation_phase not in EVALUATION_PHASES
            or evaluation_phase == "UNSPECIFIED"
        ):
            raise ImportValidationError(["新执行必须选择明确的 evaluation_phase"])
        validate_storage_id(evaluation_id, "evaluation_id")
        validate_storage_id(dataset_id, "dataset_id")
        object_id = validate_storage_id(
            run_id or f"run_{uuid.uuid4().hex}",
            "run_id",
        )
        evaluation = self.store.fetch_one(
            "SELECT evaluation_id FROM evaluations WHERE evaluation_id = ?",
            (evaluation_id,),
        )
        dataset = self.store.fetch_one(
            "SELECT dataset_id, query_count FROM datasets WHERE dataset_id = ?",
            (dataset_id,),
        )
        if evaluation is None:
            raise ImportValidationError([f"Evaluation 不存在: {evaluation_id}"])
        if dataset is None:
            raise ImportValidationError([f"Dataset 不存在: {dataset_id}"])
        if int(dataset["query_count"]) <= 0:
            raise ImportValidationError(["Dataset 没有可执行 Query"])
        dataset_queries = self.store.fetch_all(
            """
            SELECT query_id, person_id, query_stage
            FROM dataset_queries
            WHERE dataset_id = ?
            ORDER BY rowid
            """,
            (dataset_id,),
        )
        now = utc_now_text()
        with self._execution_lock:
            active = self.store.fetch_one(
                """
                SELECT run_id FROM runs
                WHERE source_type = 'EXECUTION'
                  AND status IN ('PENDING', 'RUNNING')
                LIMIT 1
                """
            )
            if active is not None:
                raise ActiveRunError(
                    f"已有执行任务 {active['run_id']} 正在等待或运行"
                )
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    INSERT INTO runs(
                        run_id, evaluation_id, dataset_id, run_label,
                        system_version, source_type, status,
                        result_schema_version, total_queries, success_queries,
                        failed_queries, message, evaluation_phase, created_at
                    ) VALUES (?, ?, ?, ?, ?, 'EXECUTION', 'PENDING', ?, ?, 0, 0, ?, ?, ?)
                    """,
                    (
                        object_id,
                        evaluation_id,
                        dataset_id,
                        run_label.strip(),
                        system_version.strip(),
                        RESULT_SCHEMA_VERSION,
                        len(dataset_queries),
                        "等待后台执行",
                        evaluation_phase,
                        now,
                    ),
                )
                for query in dataset_queries:
                    connection.execute(
                        """
                        INSERT INTO run_queries(
                            run_id, query_id, person_id, person_id_source,
                            query_stage, status, current_stage
                        ) VALUES (?, ?, ?, ?, ?, 'PENDING', 'Input')
                        """,
                        (
                            object_id,
                            query["query_id"],
                            query["person_id"],
                            (
                                "DATASET"
                                if query["person_id"]
                                else "UNSPECIFIED"
                            ),
                            query["query_stage"],
                        ),
                    )
        return object_id

    def update_run_evaluation_phase(
        self,
        run_id: str,
        evaluation_phase: str,
    ) -> None:
        """更新 Run 的显式评估阶段并令关联报告过期。

        历史 Run 允许从 ``UNSPECIFIED`` 补录阶段，也允许测试人员纠正误选；
        不根据 run_label 或 system_version 自动猜测。
        """

        if evaluation_phase not in EVALUATION_PHASES:
            raise ImportValidationError(
                [f"evaluation_phase 不受支持: {evaluation_phase}"]
            )
        with self.store.transaction() as connection:
            cursor = connection.execute(
                """
                UPDATE runs SET evaluation_phase = ?
                WHERE run_id = ?
                """,
                (evaluation_phase, run_id),
            )
            if cursor.rowcount != 1:
                raise ImportValidationError([f"Run 不存在: {run_id}"])
            connection.execute(
                """
                UPDATE reports SET status = 'STALE'
                WHERE status = 'READY'
                  AND (
                    candidate_process_id IN (
                        SELECT process_id FROM process_runs WHERE run_id = ?
                    )
                    OR baseline_process_id IN (
                        SELECT process_id FROM process_runs WHERE run_id = ?
                    )
                  )
                """,
                (run_id, run_id),
            )

    @staticmethod
    def _normalize_person_name(value: Any) -> str:
        """把人物显示名转换为仅用于精确建议匹配的稳定文本。"""

        if value is None:
            return ""
        return " ".join(str(value).strip().split()).casefold()

    @staticmethod
    def _query_clue_values(clues: Any) -> tuple[str, str]:
        """从受控 Query clues 中提取 FULL_NAME 和首个 SOCIAL_LINK。

        参数说明:
            clues: Dataset 或归档输入记录中的线索数组。

        返回值:
            ``(full_name, social_link)``；不存在或结构非法时返回空字符串。

        异常说明:
            本方法容忍历史脏结构，不抛异常，避免一条旧 Query 阻断整个
            人物关联工作区。
        """

        if not isinstance(clues, list):
            return "", ""
        full_name = ""
        social_link = ""
        for clue in clues:
            if not isinstance(clue, dict):
                continue
            clue_type = str(clue.get("type") or "")
            value = clue.get("value")
            if clue_type == "FULL_NAME" and value is None:
                full_name_query = clue.get("full_name_query")
                if isinstance(full_name_query, dict):
                    value = full_name_query.get("full_name")
            elif clue_type == "SOCIAL_LINK" and value is None:
                social_link_query = clue.get("social_link_query")
                if isinstance(social_link_query, dict):
                    value = social_link_query.get("url")
            if value is None:
                continue
            text = " ".join(str(value).strip().split())
            if clue_type == "FULL_NAME" and not full_name:
                full_name = text
            elif clue_type == "SOCIAL_LINK" and not social_link:
                social_link = text
        return full_name, social_link

    def _archived_run_query_clues(
        self,
        run: sqlite3.Row,
    ) -> dict[str, list[Any]]:
        """从受控归档 results.jsonl 补充无 Dataset Run 的输入 clues。

        只读取 ``runs.results_file`` 指向的 data_dir 内文件，不访问 Raw
        Candidate，也不从第一名候选人反推目标人物。
        """

        relative_path = str(run["results_file"] or "").strip()
        if not relative_path:
            return {}
        target = (self.data_dir / relative_path).resolve()
        try:
            target.relative_to(self.data_dir)
        except ValueError:
            return {}
        if not target.is_file() or target.suffix.lower() != ".jsonl":
            return {}
        records, _ = read_jsonl(target)
        result: dict[str, list[Any]] = {}
        for record in records:
            if not isinstance(record, dict):
                continue
            query_id = record.get("input_id") or record.get("query_id")
            clues = record.get("clues")
            if (
                isinstance(query_id, str)
                and query_id
                and isinstance(clues, list)
            ):
                result[query_id] = clues
        return result

    def get_run_person_link_context(
        self,
        run_id: str,
        baseline_version: str,
    ) -> dict[str, Any]:
        """返回历史 Run Query 与指定 Baseline 的人物关联工作区。

        参数说明:
            run_id: 已保存的 Run 标识。
            baseline_version: 用于校验和生成姓名建议的 Baseline 版本。

        返回值:
            包含 Run、Baseline、汇总、Baseline 人物选项和逐 Query 建议。
            姓名建议只做规范化后的精确匹配，从不自动写入。

        异常说明:
            PersonLinkValidationError: Run 或 Baseline 不存在。
        """

        run = self.store.fetch_one(
            "SELECT * FROM runs WHERE run_id = ?",
            (run_id,),
        )
        if run is None:
            raise PersonLinkValidationError(
                f"Run {run_id} 不存在，请返回 Run 列表重新选择"
            )
        baseline = self.store.fetch_one(
            """
            SELECT bs.*, COUNT(bp.person_id) AS person_count
            FROM baseline_sets AS bs
            LEFT JOIN baseline_people AS bp
              ON bp.baseline_version = bs.baseline_version
            WHERE bs.baseline_version = ?
            GROUP BY bs.baseline_version
            """,
            (baseline_version,),
        )
        if baseline is None:
            raise PersonLinkValidationError(
                f"Run {run_id} 选择的 Baseline {baseline_version} 不存在，"
                "请重新选择有效版本"
            )
        baseline_rows = self.store.fetch_all(
            """
            SELECT person_id, display_name, fields_json
            FROM baseline_people
            WHERE baseline_version = ?
            ORDER BY COALESCE(display_name, ''), person_id
            """,
            (baseline_version,),
        )
        people: list[dict[str, Any]] = []
        people_by_id: dict[str, dict[str, Any]] = {}
        people_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in baseline_rows:
            display_name = str(row["display_name"] or "").strip()
            if not display_name:
                try:
                    fields = json.loads(row["fields_json"])
                except (TypeError, json.JSONDecodeError):
                    fields = {}
                if isinstance(fields, dict):
                    display_name = str(
                        fields.get("profile_full_name")
                        or fields.get("summary_display_name")
                        or fields.get("display_name")
                        or ""
                    ).strip()
            person = {
                "person_id": row["person_id"],
                "display_name": display_name or row["person_id"],
            }
            people.append(person)
            people_by_id[row["person_id"]] = person
            normalized_name = self._normalize_person_name(display_name)
            if normalized_name:
                people_by_name[normalized_name].append(person)

        dataset_clues: dict[str, list[Any]] = {}
        if run["dataset_id"]:
            for row in self.store.fetch_all(
                """
                SELECT query_id, clues_json FROM dataset_queries
                WHERE dataset_id = ?
                """,
                (run["dataset_id"],),
            ):
                try:
                    clues = json.loads(row["clues_json"])
                except (TypeError, json.JSONDecodeError):
                    clues = []
                dataset_clues[row["query_id"]] = (
                    clues if isinstance(clues, list) else []
                )
        archived_clues = (
            {} if dataset_clues else self._archived_run_query_clues(run)
        )
        history_counts = {
            row["query_id"]: row["history_count"]
            for row in self.store.fetch_all(
                """
                SELECT query_id, COUNT(*) AS history_count
                FROM run_query_person_history
                WHERE run_id = ?
                GROUP BY query_id
                """,
                (run_id,),
            )
        }
        query_rows = self.store.fetch_all(
            """
            SELECT query_id, query_stage, person_id, person_id_source, status
            FROM run_queries
            WHERE run_id = ?
            ORDER BY query_id
            """,
            (run_id,),
        )
        queries: list[dict[str, Any]] = []
        linked_count = 0
        unlinked_count = 0
        invalid_count = 0
        unique_suggestion_count = 0
        for row in query_rows:
            clues = dataset_clues.get(
                row["query_id"],
                archived_clues.get(row["query_id"], []),
            )
            query_name, social_link = self._query_clue_values(clues)
            suggestions = [
                {
                    **person,
                    "match_reason": "NORMALIZED_NAME_EXACT",
                }
                for person in people_by_name.get(
                    self._normalize_person_name(query_name),
                    [],
                )
            ]
            has_unique_suggestion = len(suggestions) == 1
            unique_suggestion_count += int(has_unique_suggestion)
            current_person_id = row["person_id"] or None
            current_baseline_exists = (
                current_person_id in people_by_id
                if current_person_id is not None
                else False
            )
            if current_person_id is None:
                unlinked_count += 1
            elif current_baseline_exists:
                linked_count += 1
            else:
                invalid_count += 1
            queries.append(
                {
                    "query_id": row["query_id"],
                    "query_stage": row["query_stage"],
                    "query_status": row["status"],
                    "query_name": query_name,
                    "social_link": social_link,
                    "current_person_id": current_person_id,
                    "current_source": row["person_id_source"],
                    "current_baseline_exists": current_baseline_exists,
                    "current_baseline": people_by_id.get(current_person_id),
                    "suggestions": suggestions,
                    "has_unique_suggestion": has_unique_suggestion,
                    "history_count": history_counts.get(row["query_id"], 0),
                }
            )
        return {
            "run": dict(run),
            "baseline": dict(baseline),
            "summary": {
                "query_count": len(queries),
                "linked_count": linked_count,
                "unlinked_count": unlinked_count,
                "invalid_count": invalid_count,
                "unique_suggestion_count": unique_suggestion_count,
            },
            "baseline_people": people,
            "queries": queries,
            "editable": run["status"] not in {"PENDING", "RUNNING"},
        }

    def update_run_query_person_links(
        self,
        run_id: str,
        baseline_version: str,
        changes: list[dict[str, Any]],
        *,
        sync_dataset: bool = False,
        note: str = "",
    ) -> dict[str, Any]:
        """原子更新历史 Run Query 人物关联并写入审计。

        参数说明:
            run_id: 已结束的 Run。
            baseline_version: 新 person_id 必须属于该 Baseline。
            changes: 包含 query_id、expected_person_id、person_id 的数组。
            sync_dataset: 是否把相同关联同步到 Run 对应 Dataset。
            note: 本次人工修改说明。

        返回值:
            更新数、Dataset 同步数、报告过期数和审计 ID。

        异常说明:
            PersonLinkValidationError: 关联对象、Run 状态、乐观锁或批量
            输入非法。任一项失败会回滚整个批次。
        """

        if not isinstance(changes, list) or not changes:
            raise PersonLinkValidationError(
                f"Run {run_id} 没有提交人物关联变更"
            )
        if not isinstance(note, str):
            raise PersonLinkValidationError(
                f"Run {run_id} 的修改说明必须是文本"
            )
        normalized_note = note.strip()
        if len(normalized_note) > 1000:
            raise PersonLinkValidationError(
                f"Run {run_id} 的修改说明不能超过1000个字符"
            )
        normalized_changes: list[dict[str, Any]] = []
        seen_query_ids: set[str] = set()
        errors: list[str] = []
        for index, change in enumerate(changes, start=1):
            if not isinstance(change, dict):
                errors.append(f"Run {run_id} 第{index}项变更必须是对象")
                continue
            query_id = change.get("query_id")
            if not isinstance(query_id, str) or not query_id.strip():
                errors.append(f"Run {run_id} 第{index}项缺少 query_id")
                continue
            query_id = query_id.strip()
            if query_id in seen_query_ids:
                errors.append(f"Run {run_id} 的 Query {query_id} 重复提交")
                continue
            seen_query_ids.add(query_id)
            if "expected_person_id" not in change:
                errors.append(
                    f"Run {run_id} 的 Query {query_id} 缺少 expected_person_id"
                )
                continue
            expected = change.get("expected_person_id")
            person_id = change.get("person_id")
            if expected is not None and not isinstance(expected, str):
                errors.append(
                    f"Run {run_id} 的 Query {query_id} expected_person_id 非法"
                )
                continue
            if person_id is not None and not isinstance(person_id, str):
                errors.append(
                    f"Run {run_id} 的 Query {query_id} person_id 非法"
                )
                continue
            normalized_changes.append(
                {
                    "query_id": query_id,
                    "expected_person_id": (
                        expected.strip() if isinstance(expected, str) else None
                    )
                    or None,
                    "person_id": (
                        person_id.strip()
                        if isinstance(person_id, str)
                        else None
                    )
                    or None,
                }
            )
        if errors:
            raise PersonLinkValidationError(errors)

        now = utc_now_text()
        history_ids: list[str] = []
        updated_count = 0
        dataset_synced_count = 0
        stale_report_count = 0
        with self.store.transaction() as connection:
            run = connection.execute(
                "SELECT * FROM runs WHERE run_id = ?",
                (run_id,),
            ).fetchone()
            if run is None:
                raise PersonLinkValidationError(
                    f"Run {run_id} 不存在，请返回 Run 列表重新选择"
                )
            if run["status"] in {"PENDING", "RUNNING"}:
                raise PersonLinkValidationError(
                    f"Run {run_id} 当前状态为 {run['status']}，"
                    "请等待执行结束后再修改人物关联"
                )
            baseline = connection.execute(
                """
                SELECT baseline_version FROM baseline_sets
                WHERE baseline_version = ?
                """,
                (baseline_version,),
            ).fetchone()
            if baseline is None:
                raise PersonLinkValidationError(
                    f"Run {run_id} 的 Baseline {baseline_version} 不存在，"
                    "请重新选择有效版本"
                )
            valid_people = {
                row["person_id"]
                for row in connection.execute(
                    """
                    SELECT person_id FROM baseline_people
                    WHERE baseline_version = ?
                    """,
                    (baseline_version,),
                )
            }
            query_rows = {
                row["query_id"]: row
                for row in connection.execute(
                    """
                    SELECT query_id, person_id, person_id_source
                    FROM run_queries WHERE run_id = ?
                    """,
                    (run_id,),
                )
            }
            for change in normalized_changes:
                query_id = change["query_id"]
                query = query_rows.get(query_id)
                if query is None:
                    raise PersonLinkValidationError(
                        f"Run {run_id} 不包含 Query {query_id}，"
                        "请刷新人物关联页面后重试"
                    )
                if change["person_id"] not in valid_people and change[
                    "person_id"
                ] is not None:
                    raise PersonLinkValidationError(
                        f"Run {run_id} 的 Query {query_id} 选择了不存在的 "
                        f"person_id {change['person_id']}，请重新选择"
                    )
                current_person_id = query["person_id"] or None
                if current_person_id != change["expected_person_id"]:
                    raise PersonLinkValidationError(
                        f"Run {run_id} 的 Query {query_id} 人物关联已被其他"
                        "页面修改，请刷新后重试"
                    )
            if sync_dataset and not run["dataset_id"]:
                raise PersonLinkValidationError(
                    f"Run {run_id} 没有关联 Dataset，不能同步人物关联"
                )
            dataset_rows: dict[str, sqlite3.Row] = {}
            if sync_dataset:
                dataset_rows = {
                    row["query_id"]: row
                    for row in connection.execute(
                        """
                        SELECT query_id, person_id, person_id_source
                        FROM dataset_queries WHERE dataset_id = ?
                        """,
                        (run["dataset_id"],),
                    )
                }
                missing_dataset_queries = [
                    change["query_id"]
                    for change in normalized_changes
                    if change["query_id"] not in dataset_rows
                ]
                if missing_dataset_queries:
                    raise PersonLinkValidationError(
                        f"Run {run_id} 的 Dataset 缺少 Query "
                        + ", ".join(missing_dataset_queries)
                        + "，已取消整个批次"
                    )

            for change in normalized_changes:
                query_id = change["query_id"]
                current_person_id = query_rows[query_id]["person_id"] or None
                new_person_id = change["person_id"]
                run_changed = current_person_id != new_person_id
                dataset_changed = False
                if sync_dataset:
                    dataset_row = dataset_rows[query_id]
                    dataset_changed = (
                        (dataset_row["person_id"] or None) != new_person_id
                        or dataset_row["person_id_source"] != "MANUAL_DATASET"
                    )
                if not run_changed and not dataset_changed:
                    continue
                if run_changed:
                    connection.execute(
                        """
                        UPDATE run_queries
                        SET person_id = ?, person_id_source = ?
                        WHERE run_id = ? AND query_id = ?
                        """,
                        (
                            new_person_id,
                            (
                                "MANUAL_RUN"
                                if new_person_id is not None
                                else "UNSPECIFIED"
                            ),
                            run_id,
                            query_id,
                        ),
                    )
                    updated_count += 1
                if dataset_changed:
                    connection.execute(
                        """
                        UPDATE dataset_queries
                        SET person_id = ?, person_id_source = ?
                        WHERE dataset_id = ? AND query_id = ?
                        """,
                        (
                            new_person_id,
                            (
                                "MANUAL_DATASET"
                                if new_person_id is not None
                                else "UNSPECIFIED"
                            ),
                            run["dataset_id"],
                            query_id,
                        ),
                    )
                    dataset_synced_count += 1
                if new_person_id is None and run_changed:
                    change_source = "CLEAR_LINK"
                elif not run_changed and dataset_changed:
                    change_source = "DATASET_SYNC"
                else:
                    change_source = (
                        "MANUAL_SINGLE"
                        if len(normalized_changes) == 1
                        else "MANUAL_BULK"
                    )
                history_id = f"person_history_{uuid.uuid4().hex}"
                connection.execute(
                    """
                    INSERT INTO run_query_person_history(
                        history_id, run_id, query_id, baseline_version,
                        old_person_id, new_person_id, change_source,
                        sync_dataset, note, changed_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        history_id,
                        run_id,
                        query_id,
                        baseline_version,
                        current_person_id,
                        new_person_id,
                        change_source,
                        int(sync_dataset),
                        normalized_note,
                        now,
                    ),
                )
                history_ids.append(history_id)
            if history_ids:
                cursor = connection.execute(
                    """
                    UPDATE reports SET status = 'STALE'
                    WHERE status = 'READY'
                      AND (
                        candidate_process_id IN (
                            SELECT process_id FROM process_runs
                            WHERE run_id = ?
                        )
                        OR baseline_process_id IN (
                            SELECT process_id FROM process_runs
                            WHERE run_id = ?
                        )
                      )
                    """,
                    (run_id, run_id),
                )
                stale_report_count = cursor.rowcount
        return {
            "run_id": run_id,
            "baseline_version": baseline_version,
            "updated_count": updated_count,
            "dataset_synced_count": dataset_synced_count,
            "stale_report_count": stale_report_count,
            "history_ids": history_ids,
        }

    def _execution_paths(
        self,
        evaluation_id: str,
        run_id: str,
    ) -> tuple[Path, Path]:
        """返回一个执行 Run 的受控 results/failures JSONL 路径。"""

        validate_storage_id(evaluation_id, "evaluation_id")
        validate_storage_id(run_id, "run_id")
        run_dir = self.raw_dir / evaluation_id / run_id
        return run_dir / "results.jsonl", run_dir / "failures.jsonl"

    def _update_execution_progress(
        self,
        run_id: str,
        event: dict[str, Any],
    ) -> None:
        """把采集回调转换为页面可轮询的 Query 与 Run 最新状态。"""

        query_id = str(event.get("input_id") or "")
        stage = str(event.get("stage") or "")
        task_id = str(event.get("task_id") or "")
        message = str(event.get("message") or event.get("status") or stage)
        with self.store.transaction() as connection:
            if query_id:
                connection.execute(
                    """
                    UPDATE run_queries
                    SET current_stage = ?,
                        task_id = CASE WHEN ? = '' THEN task_id ELSE ? END
                    WHERE run_id = ? AND query_id = ?
                    """,
                    (stage, task_id, task_id, run_id, query_id),
                )
            connection.execute(
                "UPDATE runs SET message = ? WHERE run_id = ?",
                (f"{query_id}: {message}" if query_id else message, run_id),
            )

    def _persist_execution_success(
        self,
        *,
        run_id: str,
        result: dict[str, Any],
        raw_records: list[dict[str, Any]],
        candidate_failures: list[dict[str, Any]],
    ) -> None:
        """在一个事务内保存成功或部分成功 Query 的全部数据。"""

        query_id = result["input_id"]
        now = utc_now_text()
        task_fields = (
            result.get("task_fields")
            if isinstance(result.get("task_fields"), dict)
            else {}
        )
        public_fields = {
            key: value
            for key, value in task_fields.items()
            if key not in TASK_FIELD_KEYS
        }
        if isinstance(result.get("public_fields"), dict):
            public_fields.update(result["public_fields"])
        result_status = normalize_result_status(
            result["query_status"],
            result.get("candidate_count_listed"),
        )
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE run_queries
                SET task_id = ?, status = ?, current_stage = 'Completed',
                    candidate_count_total = ?, candidate_count_listed = ?,
                    detail_success_count = ?, detail_failure_count = ?,
                    llm_cost = ?, third_party_cost = ?, total_cost = ?,
                    pdl_called = ?, search_duration_ms = ?,
                    result_status = ?, public_fields_json = ?,
                    error = '', finished_at = ?
                WHERE run_id = ? AND query_id = ?
                """,
                (
                    result.get("task_id"),
                    result["query_status"],
                    result.get("candidate_count_total"),
                    result.get("candidate_count_listed", 0),
                    result.get("detail_success_count", 0),
                    result.get("detail_failure_count", 0),
                    task_fields.get("llm_cost"),
                    task_fields.get("third_party_cost"),
                    task_fields.get("total_cost"),
                    task_fields.get("pdl_called"),
                    task_fields.get("search_duration_ms"),
                    result_status,
                    json_text(public_fields),
                    now,
                    run_id,
                    query_id,
                ),
            )
            candidate_by_identity: dict[tuple[str, int], str] = {}
            for candidate in result.get("results", []):
                candidate_pk = f"candidate_{uuid.uuid4().hex}"
                rank = int(candidate["candidate_rank"])
                candidate_id = str(candidate.get("candidate_id") or "")
                candidate_by_identity[(candidate_id, rank)] = candidate_pk
                connection.execute(
                    """
                    INSERT INTO candidates(
                        candidate_pk, run_id, query_id, candidate_id,
                        candidate_rank, rank_score, detail_status,
                        detail_error, ui_sections_json, detail_data_json,
                        list_item_json, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        candidate_pk,
                        run_id,
                        query_id,
                        candidate_id,
                        rank,
                        candidate.get("rank_score"),
                        candidate.get("detail_status", "FAILED"),
                        str(candidate.get("detail_error") or ""),
                        (
                            json_text(candidate["ui_sections"])
                            if candidate.get("ui_sections") is not None
                            else None
                        ),
                        (
                            json_text(candidate["detail_data_raw"])
                            if candidate.get("detail_data_raw") is not None
                            else None
                        ),
                        json_text(candidate.get("list_item_raw") or {}),
                        now,
                    ),
                )
            for raw in raw_records:
                raw_candidate_id = str(raw.get("candidate_id") or "")
                sequence_no = int(raw.get("sequence_no") or 1)
                candidate_pk = candidate_by_identity.get(
                    (raw_candidate_id, sequence_no)
                )
                if candidate_pk is None and raw_candidate_id:
                    candidate_pk = next(
                        (
                            value
                            for (candidate_id, _), value in candidate_by_identity.items()
                            if candidate_id == raw_candidate_id
                        ),
                        None,
                    )
                self._insert_raw(
                    connection,
                    run_id=run_id,
                    query_id=query_id,
                    candidate_pk=candidate_pk,
                    stage=str(raw.get("stage") or "Import"),
                    sequence_no=sequence_no,
                    payload=raw,
                    collected_at=raw.get("collected_at"),
                )
            for failure in candidate_failures:
                connection.execute(
                    """
                    INSERT INTO failures(
                        failure_id, run_id, query_id, candidate_id,
                        scope, stage, error, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        f"failure_{uuid.uuid4().hex}",
                        run_id,
                        query_id,
                        str(failure.get("candidate_id") or ""),
                        "CANDIDATE",
                        str(failure.get("stage") or "GetTaskCandidateDetail"),
                        str(failure.get("error") or ""),
                        failure.get("created_at") or now,
                    ),
                )

    def _persist_execution_failure(
        self,
        *,
        run_id: str,
        query_id: str,
        error: FlowError,
    ) -> dict[str, Any]:
        """保存 Query 级失败及失败前已获得的 Raw，并返回文件记录。"""

        now = utc_now_text()
        task_fields = error.task_fields if isinstance(error.task_fields, dict) else {}
        public_fields = (
            error.public_fields if isinstance(error.public_fields, dict) else {}
        )
        failure_record = {
            "failure_schema_version": RESULT_SCHEMA_VERSION,
            "run_id": run_id,
            "input_id": query_id,
            "task_id": error.task_id,
            "candidate_id": "",
            "scope": "INPUT" if error.stage == "Input" else "QUERY",
            "stage": error.stage,
            "query_status": "FAILED",
            "result_status": "EXECUTION_FAILED",
            "error": str(error),
            "task_fields": task_fields,
            "public_fields": public_fields,
            "raw": error.raw_records,
            "created_at": now,
        }
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE run_queries
                SET task_id = ?, status = 'FAILED', current_stage = ?,
                    llm_cost = ?, third_party_cost = ?, total_cost = ?,
                    pdl_called = ?, search_duration_ms = ?,
                    public_fields_json = ?,
                    result_status = 'EXECUTION_FAILED',
                    error = ?, finished_at = ?
                WHERE run_id = ? AND query_id = ?
                """,
                (
                    error.task_id,
                    error.stage,
                    task_fields.get("llm_cost"),
                    task_fields.get("third_party_cost"),
                    task_fields.get("total_cost"),
                    task_fields.get("pdl_called"),
                    task_fields.get("search_duration_ms"),
                    json_text(public_fields),
                    str(error),
                    now,
                    run_id,
                    query_id,
                ),
            )
            for raw in error.raw_records:
                self._insert_raw(
                    connection,
                    run_id=run_id,
                    query_id=query_id,
                    candidate_pk=None,
                    stage=str(raw.get("stage") or error.stage),
                    sequence_no=int(raw.get("sequence_no") or 1),
                    payload=raw,
                    collected_at=raw.get("collected_at"),
                )
            connection.execute(
                """
                INSERT INTO failures(
                    failure_id, run_id, query_id, candidate_id,
                    scope, stage, error, created_at
                ) VALUES (?, ?, ?, '', ?, ?, ?, ?)
                """,
                (
                    f"failure_{uuid.uuid4().hex}",
                    run_id,
                    query_id,
                    failure_record["scope"],
                    error.stage,
                    str(error),
                    now,
                ),
            )
        return failure_record

    def validate_query_retry(self, run_id: str, query_id: str) -> None:
        """校验单条 Query 是否可在原 Run 中重新请求。

        仅允许已失败或因应用中断而未执行的 Query 重跑；成功结果不能被此
        入口覆盖，避免意外增加成本或破坏后续处理所依赖的数据快照。
        """

        row = self.store.fetch_one(
            """
            SELECT r.source_type, r.status AS run_status, rq.status AS query_status
            FROM runs AS r
            JOIN run_queries AS rq ON rq.run_id = r.run_id
            WHERE r.run_id = ? AND rq.query_id = ?
            """,
            (run_id, query_id),
        )
        if row is None:
            raise ImportValidationError(["Run 或 Query 不存在"])
        if row["source_type"] != "EXECUTION":
            raise ImportValidationError(["仅真实执行 Run 支持重新请求"])
        if row["run_status"] in {"PENDING", "RUNNING"}:
            raise ActiveRunError("当前 Run 尚未结束，不能单独重跑 Query")
        if row["query_status"] not in {"FAILED", "PENDING"}:
            raise ImportValidationError(
                ["仅 FAILED 或中断后遗留的 PENDING Query 可以重新请求"]
            )

    def _refresh_execution_run_status(self, run_id: str) -> None:
        """根据同一 Run 的全部 Query 状态回写汇总计数和终态。"""

        rows = self.store.fetch_all(
            "SELECT status FROM run_queries WHERE run_id = ?", (run_id,)
        )
        success_count = sum(
            row["status"] in {"SUCCESS", "NO_CANDIDATE"} for row in rows
        )
        failed_count = len(rows) - success_count
        if failed_count == len(rows):
            final_status = "FAILED"
        elif failed_count:
            final_status = "PARTIAL_FAILED"
        else:
            final_status = "COMPLETED"
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE runs
                SET status = ?, success_queries = ?, failed_queries = ?,
                    finished_at = ?, message = ?
                WHERE run_id = ?
                """,
                (
                    final_status,
                    success_count,
                    failed_count,
                    utc_now_text(),
                    f"执行结束：成功 {success_count}，失败或部分失败 {failed_count}",
                    run_id,
                ),
            )

    def execute_query_retry(
        self,
        run_id: str,
        query_id: str,
        client: Any,
        *,
        sleep_fn: Callable[[float], None],
    ) -> None:
        """在原 Run 内重新执行单条失败或中断 Query，并保留其余 Query。"""

        self.validate_query_retry(run_id, query_id)
        query = self.store.fetch_one(
            """
            SELECT dq.query_id, dq.person_id, dq.query_stage, dq.match_strategy,
                   dq.clues_json, dq.additional_details_json
            FROM runs AS r
            JOIN dataset_queries AS dq ON dq.dataset_id = r.dataset_id
            WHERE r.run_id = ? AND dq.query_id = ?
            """,
            (run_id, query_id),
        )
        run = self.store.fetch_one("SELECT * FROM runs WHERE run_id = ?", (run_id,))
        if query is None or run is None:
            raise ImportValidationError(["原始 Dataset Query 不存在，无法重新请求"])
        results_path, failures_path = self._execution_paths(run["evaluation_id"], run_id)
        if not results_path.exists() or not failures_path.exists():
            raise FileNotFoundError("原 Run 的结果文件不存在，无法在原 Run 内追加重跑结果")

        now = utc_now_text()
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE runs SET status = 'RUNNING', finished_at = NULL,
                    message = ? WHERE run_id = ?
                """,
                (f"{query_id}: 正在重新请求", run_id),
            )
            connection.execute(
                """
                UPDATE run_queries
                SET status = 'RUNNING', current_stage = 'Input', error = '',
                    started_at = ?, finished_at = NULL
                WHERE run_id = ? AND query_id = ?
                """,
                (now, run_id, query_id),
            )
        item = {
            "input_id": query_id,
            "query_stage": query["query_stage"],
            "match_strategy": query["match_strategy"],
            "clues": json.loads(query["clues_json"]),
            "additional_details": json.loads(query["additional_details_json"]),
        }
        raw_records: list[dict[str, Any]] = []
        candidate_failures: list[dict[str, Any]] = []
        try:
            result = process_one(
                item,
                client,
                sleep_fn=sleep_fn,
                progress_callback=lambda event: self._update_execution_progress(run_id, event),
                raw_callback=raw_records.append,
                failure_callback=candidate_failures.append,
                run_id=run_id,
            )
            result["person_id"] = query["person_id"]
            self._persist_execution_success(
                run_id=run_id,
                result=result,
                raw_records=raw_records,
                candidate_failures=candidate_failures,
            )
            self._append_jsonl(results_path, result)
            for candidate_failure in candidate_failures:
                self._append_jsonl(failures_path, candidate_failure)
        except FlowError as exc:
            failure = self._persist_execution_failure(
                run_id=run_id, query_id=query_id, error=exc
            )
            self._append_jsonl(failures_path, failure)
        self._refresh_execution_run_status(run_id)

    def mark_query_retry_failed(self, run_id: str, query_id: str, message: str) -> None:
        """记录单条重跑的非业务异常，绝不影响同一 Run 的其他 Query。"""

        safe_message = str(message).replace("\n", " ")[:1000]
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE run_queries
                SET status = 'FAILED', current_stage = 'Execution',
                    result_status = 'EXECUTION_FAILED', error = ?, finished_at = ?
                WHERE run_id = ? AND query_id = ?
                """,
                (safe_message, utc_now_text(), run_id, query_id),
            )
        self._refresh_execution_run_status(run_id)

    def execute_run(
        self,
        run_id: str,
        client: Any,
        *,
        sleep_fn: Callable[[float], None],
    ) -> None:
        """顺序执行一个 PENDING Run，并持续保存页面进度与 Query 结果。

        参数说明:
            run_id: ``create_execution_run`` 创建的 Run。
            client: 与 ``SearchClient`` 相同接口的同步客户端。
            sleep_fn: GetTask 轮询等待函数；生产传入 ``time.sleep``。

        返回值:
            无。最终状态和计数写入 SQLite，结果写入受控 Raw 目录。

        异常说明:
            ValueError: Run 不存在或不是可启动的执行 Run。
            非业务异常向上抛出，由后台协调器标记 Run 失败。
        """

        run = self.store.fetch_one("SELECT * FROM runs WHERE run_id = ?", (run_id,))
        if run is None:
            raise ValueError(f"Run 不存在: {run_id}")
        if run["source_type"] != "EXECUTION" or run["status"] != "PENDING":
            raise ValueError(f"Run {run_id} 当前状态不可执行: {run['status']}")
        results_path, failures_path = self._execution_paths(
            run["evaluation_id"],
            run_id,
        )
        if results_path.exists() or failures_path.exists():
            raise FileExistsError(f"Run {run_id} 的执行文件已存在，禁止覆盖")
        results_path.parent.mkdir(parents=True, exist_ok=True)
        results_path.write_text("", encoding="utf-8")
        failures_path.write_text("", encoding="utf-8")
        now = utc_now_text()
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE runs
                SET status = 'RUNNING', started_at = ?, message = ?,
                    results_file = ?, failures_file = ?
                WHERE run_id = ?
                """,
                (
                    now,
                    "开始执行",
                    results_path.relative_to(self.data_dir).as_posix(),
                    failures_path.relative_to(self.data_dir).as_posix(),
                    run_id,
                ),
            )
        queries = self.store.fetch_all(
            """
            SELECT dq.query_id, dq.person_id, dq.query_stage,
                   dq.match_strategy, dq.clues_json,
                   dq.additional_details_json
            FROM dataset_queries AS dq
            JOIN runs AS r ON r.dataset_id = dq.dataset_id
            WHERE r.run_id = ?
            ORDER BY dq.rowid
            """,
            (run_id,),
        )
        success_count = 0
        failed_count = 0
        for query in queries:
            query_id = query["query_id"]
            started_at = utc_now_text()
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    UPDATE run_queries
                    SET status = 'RUNNING', current_stage = 'Input',
                        started_at = ?
                    WHERE run_id = ? AND query_id = ?
                    """,
                    (started_at, run_id, query_id),
                )
            item = {
                "input_id": query_id,
                "query_stage": query["query_stage"],
                "match_strategy": query["match_strategy"],
                "clues": json.loads(query["clues_json"]),
                "additional_details": json.loads(query["additional_details_json"]),
            }
            raw_records: list[dict[str, Any]] = []
            candidate_failures: list[dict[str, Any]] = []
            try:
                result = process_one(
                    item,
                    client,
                    sleep_fn=sleep_fn,
                    progress_callback=lambda event: self._update_execution_progress(
                        run_id,
                        event,
                    ),
                    raw_callback=raw_records.append,
                    failure_callback=candidate_failures.append,
                    run_id=run_id,
                )
                result["person_id"] = query["person_id"]
                self._persist_execution_success(
                    run_id=run_id,
                    result=result,
                    raw_records=raw_records,
                    candidate_failures=candidate_failures,
                )
                self._append_jsonl(results_path, result)
                for candidate_failure in candidate_failures:
                    self._append_jsonl(failures_path, candidate_failure)
                if result["query_status"] in {"SUCCESS", "NO_CANDIDATE"}:
                    success_count += 1
                else:
                    failed_count += 1
            except FlowError as exc:
                failure = self._persist_execution_failure(
                    run_id=run_id,
                    query_id=query_id,
                    error=exc,
                )
                self._append_jsonl(failures_path, failure)
                failed_count += 1
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    UPDATE runs
                    SET success_queries = ?, failed_queries = ?,
                        message = ?
                    WHERE run_id = ?
                    """,
                    (
                        success_count,
                        failed_count,
                        f"已完成 {success_count + failed_count}/{len(queries)}",
                        run_id,
                    ),
                )
        if failed_count == len(queries):
            final_status = "FAILED"
        elif failed_count:
            final_status = "PARTIAL_FAILED"
        else:
            final_status = "COMPLETED"
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE runs
                SET status = ?, finished_at = ?, message = ?
                WHERE run_id = ?
                """,
                (
                    final_status,
                    utc_now_text(),
                    f"执行结束：成功 {success_count}，失败或部分失败 {failed_count}",
                    run_id,
                ),
            )

    def mark_run_failed(self, run_id: str, message: str) -> None:
        """后台初始化或意外错误时把活动 Run 标记为失败，避免永久卡住。"""

        safe_message = str(message).replace("\n", " ")[:1000]
        with self.store.transaction() as connection:
            connection.execute(
                """
                UPDATE run_queries
                SET status = 'FAILED', current_stage = 'Execution',
                    result_status = 'EXECUTION_FAILED',
                    error = CASE WHEN error = '' THEN ? ELSE error END,
                    finished_at = COALESCE(finished_at, ?)
                WHERE run_id = ? AND status IN ('PENDING', 'RUNNING')
                """,
                (safe_message, utc_now_text(), run_id),
            )
            connection.execute(
                """
                UPDATE runs
                SET status = 'FAILED', finished_at = ?, message = ?
                WHERE run_id = ? AND status IN ('PENDING', 'RUNNING')
                """,
                (utc_now_text(), safe_message, run_id),
            )

    def recover_interrupted_runs(self) -> int:
        """应用启动时标记遗留 RUNNING Run，并保留已落库数据。"""

        now = utc_now_text()
        with self.store.transaction() as connection:
            rows = connection.execute(
                """
                SELECT run_id, total_queries, success_queries
                FROM runs
                WHERE source_type = 'EXECUTION' AND status = 'RUNNING'
                """
            ).fetchall()
            for row in rows:
                connection.execute(
                    """
                    UPDATE run_queries
                    SET status = 'FAILED',
                        result_status = 'EXECUTION_FAILED',
                        error = CASE WHEN error = ''
                            THEN 'Web 应用重启，当前 Query 执行已中断'
                            ELSE error END,
                        finished_at = COALESCE(finished_at, ?)
                    WHERE run_id = ? AND status = 'RUNNING'
                    """,
                    (now, row["run_id"]),
                )
                connection.execute(
                    """
                    UPDATE runs
                    SET status = 'INTERRUPTED', finished_at = ?,
                        failed_queries = MAX(
                            failed_queries,
                            total_queries - success_queries
                        ),
                        message = 'Web 应用重启，执行已中断；已完成数据已保留'
                    WHERE run_id = ?
                    """,
                    (now, row["run_id"]),
                )
        return len(rows)

    def _excel_rows(
        self,
        workbook: Any,
        sheet_name: str,
        required_headers: set[str],
        *,
        optional: bool = False,
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """读取一个 Sheet 为字典行，并验证表头。"""

        if sheet_name not in workbook.sheetnames:
            if optional:
                return [], []
            return [], [f"缺少必需 Sheet: {sheet_name}"]
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration:
            return [], [f"Sheet {sheet_name} 为空"]
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        missing = sorted(required_headers - set(headers))
        if missing:
            return [], [f"Sheet {sheet_name} 缺少列: {', '.join(missing)}"]
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(iterator, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            row = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            row["_row_number"] = row_number
            rows.append(row)
        return rows, []

    def _normalize_dataset_records(
        self,
        records: list[Any],
        initial_errors: list[str],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """校验 Dataset Query 并转换为数据库需要的统一结构。"""

        normalized: list[dict[str, Any]] = []
        errors = list(initial_errors)
        seen_ids: set[str] = set()
        for index, record in enumerate(records, start=1):
            label = f"第 {index} 条"
            if not isinstance(record, dict):
                errors.append(f"{label}必须是对象")
                continue
            query_id = record.get("input_id")
            query_stage = record.get("query_stage")
            clues = record.get("clues")
            additional_details = record.get("additional_details", [])
            if not isinstance(query_id, str) or not query_id.strip():
                errors.append(f"{label}缺少非空 input_id")
                continue
            if query_id in seen_ids:
                errors.append(f"{label} input_id 重复: {query_id}")
                continue
            seen_ids.add(query_id)
            if query_stage not in SUPPORTED_QUERY_STAGES:
                errors.append(
                    f"{label} query_stage 只支持 FULL_NAME/FULL_NAME_SOCIAL"
                )
                continue
            if not isinstance(clues, list) or not clues:
                errors.append(f"{label} clues 必须是非空数组")
                continue
            clue_types = {
                clue.get("type") for clue in clues if isinstance(clue, dict)
            }
            if "FULL_NAME" not in clue_types:
                errors.append(f"{label}缺少 FULL_NAME 线索")
                continue
            if query_stage == "FULL_NAME_SOCIAL" and "SOCIAL_LINK" not in clue_types:
                errors.append(f"{label}缺少 SOCIAL_LINK 线索")
                continue
            if not isinstance(additional_details, list):
                errors.append(f"{label} additional_details 必须是数组")
                continue
            match_strategy = record.get("match_strategy") or "UNION"
            if not isinstance(match_strategy, str):
                errors.append(f"{label} match_strategy 必须是字符串")
                continue
            metadata = {
                key: value
                for key, value in record.items()
                if key
                not in {
                    "input_id",
                    "person_id",
                    "query_stage",
                    "match_strategy",
                    "clues",
                    "additional_details",
                }
            }
            normalized.append(
                {
                    "query_id": query_id,
                    "person_id": record.get("person_id"),
                    "query_stage": query_stage,
                    "match_strategy": match_strategy,
                    "clues": clues,
                    "additional_details": additional_details,
                    "metadata": metadata,
                }
            )
        return normalized, errors

    def preview_dataset_jsonl(self, source: Path | str) -> ImportPreview:
        """预览 Dataset JSONL 的合法数量和全部错误，不写文件或数据库。"""

        path = self._source_path(source, {".jsonl"})
        records, parse_errors = read_jsonl(path)
        normalized, errors = self._normalize_dataset_records(records, parse_errors)
        return ImportPreview(file_checksum([("source", path)]), len(normalized), errors)

    def import_dataset_jsonl(
        self,
        source: Path | str,
        *,
        name: str,
        dataset_id: str | None = None,
    ) -> ImportResult:
        """校验、归档并事务导入 Query Dataset JSONL。"""

        path = self._source_path(source, {".jsonl"})
        records, parse_errors = read_jsonl(path)
        normalized, errors = self._normalize_dataset_records(records, parse_errors)
        return self._import_dataset(
            path,
            normalized,
            errors,
            name=name,
            dataset_id=dataset_id,
            source_type="JSONL",
        )

    def import_dataset_excel(
        self,
        source: Path | str,
        *,
        name: str,
        dataset_id: str | None = None,
    ) -> ImportResult:
        """从固定 Queries Sheet 导入 Query Dataset。"""

        path = self._source_path(source, {".xlsx"})
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            rows, errors = self._excel_rows(
                workbook,
                "Queries",
                {"input_id", "query_stage", "clues"},
            )
        finally:
            workbook.close()
        records = []
        for row in rows:
            record = {
                key: cell_value(value)
                for key, value in row.items()
                if key != "_row_number"
            }
            records.append(record)
        normalized, errors = self._normalize_dataset_records(records, errors)
        return self._import_dataset(
            path,
            normalized,
            errors,
            name=name,
            dataset_id=dataset_id,
            source_type="EXCEL",
        )

    def _import_dataset(
        self,
        source: Path,
        records: list[dict[str, Any]],
        errors: list[str],
        *,
        name: str,
        dataset_id: str | None,
        source_type: str,
    ) -> ImportResult:
        """执行 Dataset JSONL/Excel 共用的归档和事务写入。"""

        if errors:
            raise ImportValidationError(errors)
        if not records:
            raise ImportValidationError(["Dataset 不包含合法 Query"])
        object_id = validate_storage_id(
            dataset_id or f"dataset_{uuid.uuid4().hex}",
            "dataset_id",
        )
        checksum = file_checksum([("source", source)])
        if self.store.fetch_one(
            "SELECT dataset_id FROM datasets WHERE checksum = ?",
            (checksum,),
        ):
            raise DuplicateImportError("相同 Dataset 文件已经导入")
        archive_dir, archived = self._archive_sources(
            object_id,
            [("source", source)],
        )
        now = utc_now_text()
        try:
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    INSERT INTO datasets(
                        dataset_id, name, source_type, source_file,
                        checksum, query_count, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        object_id,
                        name,
                        source_type,
                        archived[0],
                        checksum,
                        len(records),
                        now,
                    ),
                )
                connection.executemany(
                    """
                    INSERT INTO dataset_queries(
                        dataset_id, query_id, person_id, person_id_source,
                        query_stage,
                        match_strategy, clues_json, additional_details_json,
                        metadata_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        (
                            object_id,
                            record["query_id"],
                            record["person_id"],
                            (
                                "INPUT"
                                if record["person_id"]
                                else "UNSPECIFIED"
                            ),
                            record["query_stage"],
                            record["match_strategy"],
                            json_text(record["clues"]),
                            json_text(record["additional_details"]),
                            json_text(record["metadata"]),
                        )
                        for record in records
                    ],
                )
        except sqlite3.IntegrityError as exc:
            self._cleanup_created(archive_dir)
            if "checksum" in str(exc).lower():
                raise DuplicateImportError("相同 Dataset 文件已经导入") from exc
            raise
        except Exception:
            self._cleanup_created(archive_dir)
            raise
        return ImportResult(object_id, len(records), checksum, archived)

    def _normalize_baseline_records(
        self,
        records: list[Any],
        initial_errors: list[str],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """校验 Person 基准并保留字段及证据对象。"""

        normalized: list[dict[str, Any]] = []
        errors = list(initial_errors)
        seen: set[str] = set()
        for index, record in enumerate(records, start=1):
            label = f"第 {index} 条"
            if not isinstance(record, dict):
                errors.append(f"{label}必须是对象")
                continue
            person_id = record.get("person_id")
            fields = record.get("fields", {})
            evidence = record.get("evidence", {})
            if not isinstance(person_id, str) or not person_id.strip():
                errors.append(f"{label}缺少非空 person_id")
                continue
            if person_id in seen:
                errors.append(f"{label} person_id 重复: {person_id}")
                continue
            seen.add(person_id)
            if not isinstance(fields, dict) or not isinstance(evidence, dict):
                errors.append(f"{label} fields 和 evidence 必须是对象")
                continue
            if "baseline_available_fields" in record:
                try:
                    available_fields = normalize_available_fields(
                        record["baseline_available_fields"]
                    )
                except ValueError as exc:
                    errors.append(f"{label} {exc}")
                    continue
                available_fields_source = "IMPORT"
            else:
                available_fields = _derived_available_fields(fields)
                available_fields_source = (
                    "DERIVED_LEGACY"
                    if available_fields
                    else "UNSPECIFIED"
                )
            normalized.append(
                {
                    "person_id": person_id,
                    "display_name": record.get("display_name"),
                    "fields": fields,
                    "evidence": evidence,
                    "available_fields": available_fields,
                    "available_fields_source": available_fields_source,
                }
            )
        return normalized, errors

    def import_baseline_jsonl(
        self,
        source: Path | str,
        *,
        name: str,
        baseline_version: str,
    ) -> ImportResult:
        """校验、归档并事务导入基准 JSONL。"""

        path = self._source_path(source, {".jsonl"})
        records, parse_errors = read_jsonl(path)
        normalized, errors = self._normalize_baseline_records(records, parse_errors)
        return self._import_baseline(
            path,
            normalized,
            errors,
            name=name,
            baseline_version=baseline_version,
            source_type="JSONL",
        )

    def import_baseline_excel(
        self,
        source: Path | str,
        *,
        name: str,
        baseline_version: str,
    ) -> ImportResult:
        """从固定“基准数据”Sheet 导入 Person 基准。"""

        path = self._source_path(source, {".xlsx"})
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            rows, errors = self._excel_rows(
                workbook,
                "基准数据",
                {"person_id"},
            )
        finally:
            workbook.close()
        records = []
        for row in rows:
            available_value = row.get("baseline_available_fields")
            available_present = available_value not in (None, "")
            available_fields: Any = None
            if available_present:
                if isinstance(available_value, str):
                    text = available_value.strip()
                    if text.startswith("["):
                        try:
                            available_fields = json.loads(text)
                        except json.JSONDecodeError:
                            errors.append(
                                f"基准数据第 {row['_row_number']} 行 "
                                "baseline_available_fields JSON 格式错误"
                            )
                            available_fields = []
                    else:
                        available_fields = [
                            item.strip() for item in text.split(",")
                        ]
                else:
                    available_fields = cell_value(available_value)
            fields = {
                key: cell_value(value)
                for key, value in row.items()
                if key
                not in {
                    "_row_number",
                    "person_id",
                    "display_name",
                    "evidence",
                    "baseline_available_fields",
                }
                and value not in (None, "")
            }
            evidence_value = cell_value(row.get("evidence"))
            record = {
                "person_id": row.get("person_id"),
                "display_name": row.get("display_name"),
                "fields": fields,
                "evidence": (
                    evidence_value if isinstance(evidence_value, dict) else {}
                ),
            }
            if available_present:
                record["baseline_available_fields"] = available_fields
            records.append(record)
        normalized, errors = self._normalize_baseline_records(records, errors)
        return self._import_baseline(
            path,
            normalized,
            errors,
            name=name,
            baseline_version=baseline_version,
            source_type="EXCEL",
        )

    def _import_baseline(
        self,
        source: Path,
        records: list[dict[str, Any]],
        errors: list[str],
        *,
        name: str,
        baseline_version: str,
        source_type: str,
    ) -> ImportResult:
        """执行 JSONL/Excel 基准共用的归档和事务写入。"""

        if errors:
            raise ImportValidationError(errors)
        if not records:
            raise ImportValidationError(["基准文件不包含合法人物"])
        object_id = validate_storage_id(baseline_version, "baseline_version")
        checksum = file_checksum([("source", source)])
        if self.store.fetch_one(
            "SELECT baseline_version FROM baseline_sets WHERE checksum = ?",
            (checksum,),
        ):
            raise DuplicateImportError("相同基准文件已经导入")
        archive_dir, archived = self._archive_sources(
            object_id,
            [("source", source)],
        )
        now = utc_now_text()
        try:
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    INSERT INTO baseline_sets(
                        baseline_version, name, source_type,
                        source_file, checksum, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (object_id, name, source_type, archived[0], checksum, now),
                )
                connection.executemany(
                    """
                    INSERT INTO baseline_people(
                        baseline_version, person_id, display_name,
                        fields_json, evidence_json, available_fields_json,
                        available_fields_source
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        (
                            object_id,
                            record["person_id"],
                            record["display_name"],
                            json_text(record["fields"]),
                            json_text(record["evidence"]),
                            json_text(record["available_fields"]),
                            record["available_fields_source"],
                        )
                        for record in records
                    ],
                )
        except sqlite3.IntegrityError as exc:
            self._cleanup_created(archive_dir)
            if "checksum" in str(exc).lower():
                raise DuplicateImportError("相同基准文件已经导入") from exc
            if "baseline_sets.baseline_version" in str(exc):
                raise DuplicateImportError(
                    f"baseline_version 已存在: {object_id}"
                ) from exc
            raise
        except Exception:
            self._cleanup_created(archive_dir)
            raise
        return ImportResult(object_id, len(records), checksum, archived)

    def update_baseline_available_fields(
        self,
        baseline_version: str,
        person_id: str,
        available_fields: Any,
    ) -> None:
        """人工维护人物可评估字段并令关联报告快照过期。

        参数说明:
            baseline_version: 已存在的 Baseline 版本。
            person_id: Baseline 中的人物标识。
            available_fields: 页面提交的字符串数组，允许显式清空。

        返回值:
            无；成功后来源固定标记为 ``MANUAL``。

        异常说明:
            ImportValidationError: 字段数组非法或人物不存在。
        """

        try:
            normalized = normalize_available_fields(available_fields)
        except ValueError as exc:
            raise ImportValidationError([str(exc)]) from exc
        with self.store.transaction() as connection:
            cursor = connection.execute(
                """
                UPDATE baseline_people
                SET available_fields_json = ?,
                    available_fields_source = 'MANUAL'
                WHERE baseline_version = ? AND person_id = ?
                """,
                (json_text(normalized), baseline_version, person_id),
            )
            if cursor.rowcount != 1:
                raise ImportValidationError(
                    [
                        f"Baseline 人物不存在: "
                        f"{baseline_version}/{person_id}"
                    ]
                )
            connection.execute(
                """
                UPDATE reports SET status = 'STALE'
                WHERE status = 'READY'
                  AND (
                    candidate_process_id IN (
                        SELECT process_id FROM process_runs
                        WHERE baseline_version = ?
                    )
                    OR baseline_process_id IN (
                        SELECT process_id FROM process_runs
                        WHERE baseline_version = ?
                    )
                  )
                """,
                (baseline_version, baseline_version),
            )

    def _metadata_records(
        self,
        path: Path | None,
    ) -> tuple[dict[str, dict[str, Any]], list[str]]:
        """读取可选 Query 元数据并按 query_id 建索引。"""

        if path is None:
            return {}, []
        records, errors = read_jsonl(path)
        indexed: dict[str, dict[str, Any]] = {}
        for index, record in enumerate(records, start=1):
            if not isinstance(record, dict):
                errors.append(f"元数据第 {index} 条必须是对象")
                continue
            query_id = record.get("query_id") or record.get("input_id")
            if not isinstance(query_id, str) or not query_id:
                errors.append(f"元数据第 {index} 条缺少 query_id/input_id")
                continue
            if query_id in indexed:
                errors.append(f"元数据 query_id 重复: {query_id}")
                continue
            indexed[query_id] = record
        return indexed, errors

    def _normalize_result_records(
        self,
        records: list[Any],
        metadata: dict[str, dict[str, Any]],
        initial_errors: list[str],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """将旧版和 v1.3 结果转换为同一中间结构。"""

        normalized: list[dict[str, Any]] = []
        errors = list(initial_errors)
        seen: set[str] = set()
        for index, raw_record in enumerate(records, start=1):
            label = f"结果第 {index} 条"
            if not isinstance(raw_record, dict):
                errors.append(f"{label}必须是对象")
                continue
            query_id = raw_record.get("input_id")
            candidates = raw_record.get("results")
            if not isinstance(query_id, str) or not query_id:
                errors.append(f"{label}缺少非空 input_id")
                continue
            if query_id in seen:
                errors.append(f"{label} input_id 重复: {query_id}")
                continue
            seen.add(query_id)
            if not isinstance(candidates, list):
                errors.append(f"{label} results 必须是数组")
                continue
            normalized_candidates = []
            ranks: set[int] = set()
            candidate_error = False
            for candidate_index, candidate in enumerate(candidates, start=1):
                if not isinstance(candidate, dict):
                    errors.append(
                        f"{label} results[{candidate_index - 1}] 必须是对象"
                    )
                    candidate_error = True
                    continue
                rank = candidate.get("candidate_rank", candidate_index)
                if (
                    not isinstance(rank, int)
                    or isinstance(rank, bool)
                    or rank <= 0
                    or rank in ranks
                ):
                    errors.append(f"{label}候选人 rank 无效或重复: {rank!r}")
                    candidate_error = True
                    continue
                ranks.add(rank)
                detail_status = candidate.get("detail_status")
                if detail_status not in {"SUCCESS", "FAILED"}:
                    detail_status = "SUCCESS"
                normalized_candidate = dict(sanitize_raw(candidate))
                normalized_candidate.update(
                    {
                        "candidate_rank": rank,
                        "candidate_id": str(candidate.get("candidate_id") or ""),
                        "rank_score": (
                            candidate.get("rank_score")
                            if isinstance(candidate.get("rank_score"), (int, float))
                            and not isinstance(candidate.get("rank_score"), bool)
                            else None
                        ),
                        "detail_status": detail_status,
                        "detail_error": str(candidate.get("detail_error") or ""),
                        "list_item_raw": candidate.get("list_item_raw")
                        or {
                            "candidate_id": candidate.get("candidate_id"),
                            "rank_score": candidate.get("rank_score"),
                        },
                        "detail_data_raw": candidate.get("detail_data_raw"),
                        "ui_sections": candidate.get("ui_sections"),
                    }
                )
                normalized_candidates.append(normalized_candidate)
            if candidate_error:
                continue
            metadata_record = metadata.get(query_id, {})
            query_stage = raw_record.get("query_stage") or metadata_record.get(
                "query_stage"
            )
            if query_stage not in SUPPORTED_QUERY_STAGES:
                query_stage = None
            failed_details = sum(
                item["detail_status"] == "FAILED"
                for item in normalized_candidates
            )
            success_details = len(normalized_candidates) - failed_details
            query_status = raw_record.get("query_status")
            if query_status not in {
                "SUCCESS",
                "NO_CANDIDATE",
                "PARTIAL_DETAIL_FAILED",
            }:
                if not normalized_candidates:
                    query_status = "NO_CANDIDATE"
                elif failed_details:
                    query_status = "PARTIAL_DETAIL_FAILED"
                else:
                    query_status = "SUCCESS"
            raw_value = raw_record.get("raw")
            has_raw = isinstance(raw_value, dict) and bool(raw_value)
            task_fields = (
                raw_record.get("task_fields")
                if isinstance(raw_record.get("task_fields"), dict)
                else {}
            )
            provided_result_status = raw_record.get("result_status")
            if (
                provided_result_status is not None
                and provided_result_status not in RESULT_STATUSES
            ):
                errors.append(
                    f"{label} result_status 不受支持: "
                    f"{provided_result_status!r}"
                )
                continue
            result_status = provided_result_status or normalize_result_status(
                query_status,
                len(normalized_candidates),
            )
            normalized_task_fields = {
                key: sanitize_raw(value)
                for key, value in task_fields.items()
            }
            public_fields = (
                sanitize_raw(raw_record.get("public_fields"))
                if isinstance(raw_record.get("public_fields"), dict)
                else {}
            )
            for key in TASK_FIELD_KEYS:
                normalized_task_fields.setdefault(key, None)
            normalized_record = dict(sanitize_raw(raw_record))
            normalized_record.update(
                {
                    "result_schema_version": str(
                        raw_record.get("result_schema_version") or "legacy"
                    ),
                    "input_id": query_id,
                    "task_id": str(raw_record.get("task_id") or ""),
                    "person_id": (
                        raw_record.get("person_id")
                        or metadata_record.get("person_id")
                    ),
                    "query_stage": query_stage,
                    "query_status": query_status,
                    "result_status": result_status,
                    "candidate_count_total": (
                        raw_record.get("candidate_count_total")
                        if isinstance(
                            raw_record.get("candidate_count_total"),
                            int,
                        )
                        and not isinstance(
                            raw_record.get("candidate_count_total"),
                            bool,
                        )
                        else None
                    ),
                    "candidate_count_listed": len(normalized_candidates),
                    "detail_success_count": success_details,
                    "detail_failure_count": failed_details,
                    "task_fields": normalized_task_fields,
                    "public_fields": public_fields,
                    "raw": sanitize_raw(raw_value) if has_raw else {},
                    "raw_status": (
                        "COMPLETE_RAW" if has_raw else "LEGACY_PARTIAL_RAW"
                    ),
                    "results": normalized_candidates,
                }
            )
            normalized.append(normalized_record)
        return normalized, errors

    def _normalize_failure_records(
        self,
        records: list[Any],
        initial_errors: list[str],
        metadata: dict[str, dict[str, Any]] | None = None,
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """统一旧版与 v1.3 failures.jsonl，并补全失败 Query 的配对元数据。

        参数说明:
            records: failures.jsonl 解析后的原始记录。
            initial_errors: JSONL 读取阶段已经发现的错误。
            metadata: 可选 Query 元数据索引；失败记录没有 results 行时，用其
                保留 ``person_id`` 和 ``query_stage``。

        返回值:
            规范化失败记录和全部校验错误。
        """

        normalized: list[dict[str, Any]] = []
        errors = list(initial_errors)
        metadata = metadata or {}
        for index, record in enumerate(records, start=1):
            if not isinstance(record, dict):
                errors.append(f"失败记录第 {index} 条必须是对象")
                continue
            query_id = record.get("input_id") or record.get("query_id")
            if not isinstance(query_id, str) or not query_id:
                errors.append(f"失败记录第 {index} 条缺少 input_id/query_id")
                continue
            candidate_id = str(record.get("candidate_id") or "")
            stage = str(record.get("stage") or "Import")
            scope = record.get("scope")
            if scope not in {"INPUT", "QUERY", "CANDIDATE", "IMPORT", "PROCESS"}:
                scope = (
                    "CANDIDATE"
                    if candidate_id
                    else ("INPUT" if stage == "Input" else "QUERY")
                )
            metadata_record = metadata.get(query_id, {})
            person_id = metadata_record.get("person_id")
            query_stage = metadata_record.get("query_stage")
            normalized.append(
                {
                    "input_id": query_id,
                    "person_id": (
                        person_id
                        if isinstance(person_id, str) and person_id
                        else None
                    ),
                    "query_stage": (
                        query_stage
                        if query_stage in SUPPORTED_QUERY_STAGES
                        else None
                    ),
                    "task_id": str(record.get("task_id") or ""),
                    "candidate_id": candidate_id,
                    "scope": scope,
                    "stage": stage,
                    "error": str(record.get("error") or ""),
                    "task_fields": (
                        sanitize_raw(record.get("task_fields"))
                        if isinstance(record.get("task_fields"), dict)
                        else {}
                    ),
                    "public_fields": (
                        sanitize_raw(record.get("public_fields"))
                        if isinstance(record.get("public_fields"), dict)
                        else {}
                    ),
                    "raw": sanitize_raw(record.get("raw")),
                    "created_at": record.get("created_at") or utc_now_text(),
                }
            )
        return normalized, errors

    def import_results_jsonl(
        self,
        results_path: Path | str,
        *,
        evaluation_id: str,
        run_label: str,
        system_version: str,
        evaluation_phase: str = "UNSPECIFIED",
        failures_path: Path | str | None = None,
        metadata_path: Path | str | None = None,
        run_id: str | None = None,
    ) -> ImportResult:
        """导入旧版或 v1.3 results/failures/metadata JSONL 文件包。"""

        results = self._source_path(results_path, {".jsonl"})
        failures = (
            self._source_path(failures_path, {".jsonl"})
            if failures_path is not None
            else None
        )
        metadata_file = (
            self._source_path(metadata_path, {".jsonl"})
            if metadata_path is not None
            else None
        )
        result_values, result_errors = read_jsonl(results)
        failure_values, failure_errors = (
            read_jsonl(failures) if failures is not None else ([], [])
        )
        metadata, metadata_errors = self._metadata_records(metadata_file)
        normalized_results, result_errors = self._normalize_result_records(
            result_values,
            metadata,
            [*result_errors, *metadata_errors],
        )
        normalized_failures, failure_errors = self._normalize_failure_records(
            failure_values,
            failure_errors,
            metadata,
        )
        sources = [("results", results)]
        if failures is not None:
            sources.append(("failures", failures))
        if metadata_file is not None:
            sources.append(("metadata", metadata_file))
        return self._import_run(
            normalized_results,
            normalized_failures,
            [*result_errors, *failure_errors],
            sources=sources,
            evaluation_id=evaluation_id,
            run_label=run_label,
            system_version=system_version,
            evaluation_phase=evaluation_phase,
            source_type="JSONL_IMPORT",
            run_id=run_id,
        )

    def _raw_chunk_values(
        self,
        rows: list[dict[str, Any]],
    ) -> tuple[dict[tuple[str, str, str, str], Any], list[str]]:
        """校验并重组 Raw数据 Sheet 分块。"""

        grouped: dict[
            tuple[str, str, str, str],
            list[tuple[int, int, str]],
        ] = defaultdict(list)
        errors: list[str] = []
        for row in rows:
            key = (
                str(row.get("run_label") or ""),
                str(row.get("query_id") or ""),
                str(row.get("candidate_id") or ""),
                str(row.get("field_name") or ""),
            )
            try:
                index = int(row.get("chunk_index"))
                total = int(row.get("chunk_total"))
            except (TypeError, ValueError):
                errors.append(f"Raw数据第 {row['_row_number']} 行分块序号无效")
                continue
            grouped[key].append((index, total, str(row.get("content") or "")))
        restored: dict[tuple[str, str, str, str], Any] = {}
        for key, chunks in grouped.items():
            totals = {item[1] for item in chunks}
            if len(totals) != 1:
                errors.append(f"Raw数据分块总数不一致: {key}")
                continue
            total = totals.pop()
            ordered = sorted(chunks)
            if [item[0] for item in ordered] != list(range(1, total + 1)):
                errors.append(f"Raw数据分块不完整: {key}")
                continue
            restored[key] = cell_value("".join(item[2] for item in ordered))
        return restored, errors

    def import_results_excel(
        self,
        source: Path | str,
        *,
        evaluation_id: str,
        run_label: str,
        system_version: str,
        evaluation_phase: str = "UNSPECIFIED",
        run_id: str | None = None,
    ) -> ImportResult:
        """导入项目规范化 Excel 的候选、Query、失败和 Raw 分块。"""

        path = self._source_path(source, {".xlsx"})
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            candidates, candidate_errors = self._excel_rows(
                workbook,
                "候选结果",
                {"query_id", "task_id", "candidate_id", "candidate_rank"},
            )
            queries, query_errors = self._excel_rows(
                workbook,
                "Query对比",
                {"query_id"},
                optional=True,
            )
            failures, failure_errors = self._excel_rows(
                workbook,
                "失败记录",
                {"query_id", "task_id", "stage", "error"},
                optional=True,
            )
            raw_rows, raw_errors = self._excel_rows(
                workbook,
                "Raw数据",
                {
                    "run_label",
                    "query_id",
                    "candidate_id",
                    "field_name",
                    "chunk_index",
                    "chunk_total",
                    "content",
                },
                optional=True,
            )
        finally:
            workbook.close()
        restored_raw, chunk_errors = self._raw_chunk_values(raw_rows)
        errors = [
            *candidate_errors,
            *query_errors,
            *failure_errors,
            *raw_errors,
            *chunk_errors,
        ]
        selected = [
            row
            for row in candidates
            if not row.get("run_label") or str(row.get("run_label")) == run_label
        ]
        selected_queries = [
            row
            for row in queries
            if not row.get("run_label")
            or str(row.get("run_label")) == run_label
        ]
        if not selected and not selected_queries:
            errors.append(f"候选结果中没有 run_label={run_label!r} 的记录")
        grouped_candidates: dict[str, list[dict[str, Any]]] = defaultdict(list)
        candidate_ranks: dict[str, set[int]] = defaultdict(set)
        metadata: dict[str, dict[str, Any]] = {}
        task_ids: dict[str, str] = {}
        candidate_totals: dict[str, Any] = {}
        task_fields_by_query: dict[str, dict[str, Any]] = {}
        for row in selected:
            query_id = str(row.get("query_id") or "")
            candidate_id = str(row.get("candidate_id") or "")
            try:
                rank = int(row.get("candidate_rank"))
            except (TypeError, ValueError):
                errors.append(f"候选结果第 {row['_row_number']} 行 candidate_rank 无效")
                continue
            if not query_id or not candidate_id or rank <= 0:
                errors.append(
                    f"候选结果第 {row['_row_number']} 行必需标识为空或 rank 无效"
                )
                continue
            if rank in candidate_ranks[query_id]:
                errors.append(
                    f"候选结果第 {row['_row_number']} 行 candidate_rank 重复"
                )
                continue
            candidate_ranks[query_id].add(rank)
            excel_fields = {
                key: cell_value(value)
                for key, value in row.items()
                if key != "_row_number" and value not in (None, "")
            }
            for key, value in restored_raw.items():
                raw_run, raw_query, raw_candidate, field_name = key
                if (
                    raw_run in {"", run_label}
                    and raw_query == query_id
                    and raw_candidate == candidate_id
                ):
                    excel_fields[field_name] = value
            grouped_candidates[query_id].append(
                {
                    "candidate_rank": rank,
                    "candidate_id": candidate_id,
                    "rank_score": (
                        row.get("rank_score")
                        if isinstance(row.get("rank_score"), (int, float))
                        and not isinstance(row.get("rank_score"), bool)
                        else None
                    ),
                    "detail_status": "SUCCESS",
                    "detail_error": "",
                    "list_item_raw": {
                        "candidate_id": candidate_id,
                        "rank_score": row.get("rank_score"),
                    },
                    "detail_data_raw": {
                        "source_type": "EXCEL_IMPORT",
                        "fields": excel_fields,
                    },
                    "ui_sections": {
                        "_excel_import": {
                            "status": "LEGACY_STRUCTURED_DATA",
                            "data": excel_fields,
                        }
                    },
                }
            )
            task_ids[query_id] = str(row.get("task_id") or "")
            candidate_totals[query_id] = row.get("candidate_count_total")
            task_fields_by_query.setdefault(
                query_id,
                {
                    key: cell_value(row.get(key))
                    for key in TASK_FIELD_KEYS
                    if row.get(key) not in (None, "")
                },
            )
            query_type = row.get("query_type")
            metadata[query_id] = {
                "person_id": row.get("person_id"),
                "query_stage": (
                    query_type if query_type in SUPPORTED_QUERY_STAGES else None
                ),
            }
        query_statuses: dict[str, str] = {}
        result_statuses: dict[str, str] = {}
        for row in selected_queries:
            query_id = str(row.get("query_id") or "")
            status = (
                row.get("current_status")
                or row.get("candidate_status")
                or row.get("baseline_status")
            )
            if status in {
                "SUCCESS",
                "NO_CANDIDATE",
                "PARTIAL_DETAIL_FAILED",
                "FAILED",
            }:
                query_statuses[query_id] = status
            result_status = row.get("result_status")
            if result_status not in (None, ""):
                if result_status not in RESULT_STATUSES:
                    errors.append(
                        f"Query对比第 {row['_row_number']} 行 "
                        f"result_status 不受支持"
                    )
                else:
                    result_statuses[query_id] = str(result_status)
            query_task_fields = task_fields_by_query.setdefault(query_id, {})
            for key in TASK_FIELD_KEYS:
                if row.get(key) not in (None, ""):
                    query_task_fields[key] = cell_value(row.get(key))
            metadata.setdefault(
                query_id,
                {
                    "person_id": row.get("person_id"),
                    "query_stage": (
                        row.get("query_type")
                        if row.get("query_type") in SUPPORTED_QUERY_STAGES
                        else None
                    ),
                },
            )
        result_records = []
        all_query_ids = set(grouped_candidates) | {
            query_id
            for query_id, status in query_statuses.items()
            if status != "FAILED"
        }
        for query_id in sorted(all_query_ids):
            items = sorted(
                grouped_candidates.get(query_id, []),
                key=lambda item: item["candidate_rank"],
            )
            query_status = query_statuses.get(
                query_id,
                "SUCCESS" if items else "NO_CANDIDATE",
            )
            task_fields = {
                key: task_fields_by_query.get(query_id, {}).get(key)
                for key in TASK_FIELD_KEYS
            }
            result_records.append(
                {
                    "result_schema_version": "excel-legacy",
                    "input_id": query_id,
                    "task_id": task_ids.get(query_id, ""),
                    "person_id": metadata.get(query_id, {}).get("person_id"),
                    "query_stage": metadata.get(query_id, {}).get("query_stage"),
                    "query_status": query_status,
                    "result_status": result_statuses.get(
                        query_id,
                        normalize_result_status(query_status, len(items)),
                    ),
                    "candidate_count_total": candidate_totals.get(query_id),
                    "candidate_count_listed": len(items),
                    "detail_success_count": len(items),
                    "detail_failure_count": 0,
                    "task_fields": task_fields,
                    "raw": {},
                    "raw_status": "LEGACY_PARTIAL_RAW",
                    "results": items,
                }
            )
        failure_records = []
        for row in failures:
            if row.get("run_label") and str(row.get("run_label")) != run_label:
                continue
            failure_records.append(
                {
                    "input_id": str(row.get("query_id") or ""),
                    "task_id": str(row.get("task_id") or ""),
                    "candidate_id": str(row.get("candidate_id") or ""),
                    "scope": (
                        "CANDIDATE" if row.get("candidate_id") else "QUERY"
                    ),
                    "stage": str(row.get("stage") or "Import"),
                    "error": str(row.get("error") or ""),
                    "raw": None,
                    "created_at": utc_now_text(),
                }
            )
        if errors:
            raise ImportValidationError(errors)
        return self._import_run(
            result_records,
            failure_records,
            [],
            sources=[("source", path)],
            evaluation_id=evaluation_id,
            run_label=run_label,
            system_version=system_version,
            evaluation_phase=evaluation_phase,
            source_type="EXCEL_IMPORT",
            run_id=run_id,
        )

    def _insert_raw(
        self,
        connection: sqlite3.Connection,
        *,
        run_id: str,
        query_id: str,
        candidate_pk: str | None,
        stage: str,
        sequence_no: int,
        payload: Any,
        collected_at: str | None = None,
    ) -> None:
        """向 append-only Raw 表新增一条脱敏记录。"""

        connection.execute(
            """
            INSERT INTO raw_records(
                raw_id, run_id, query_id, candidate_pk, stage,
                sequence_no, payload_json, collected_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                f"raw_{uuid.uuid4().hex}",
                run_id,
                query_id,
                candidate_pk,
                stage,
                sequence_no,
                json_text(payload),
                collected_at or utc_now_text(),
            ),
        )

    def _insert_result_raw(
        self,
        connection: sqlite3.Connection,
        *,
        run_id: str,
        record: dict[str, Any],
        candidate_pks: list[tuple[str, dict[str, Any]]],
        source_type: str,
    ) -> None:
        """把 v1.3 Raw 拆为可查询记录，旧数据写明确缺失标记。"""

        query_id = record["input_id"]
        if record["raw_status"] != "COMPLETE_RAW":
            self._insert_raw(
                connection,
                run_id=run_id,
                query_id=query_id,
                candidate_pk=None,
                stage="Import",
                sequence_no=1,
                payload={
                    "raw_status": "LEGACY_PARTIAL_RAW",
                    "source_type": source_type,
                },
            )
            return
        raw = record["raw"]
        stage_items = [
            ("CreateIntentTask", raw.get("create_intent_task")),
            *[
                ("GetTask", item)
                for item in raw.get("get_task_history", [])
                if isinstance(item, dict)
            ],
            ("ListTaskCandidates", raw.get("list_task_candidates")),
        ]
        stage_items.extend(
            ("AdminLogin", item)
            for item in raw.get("admin_login", [])
            if isinstance(item, dict)
        )
        debug_history = raw.get("get_search_task_debug_history")
        cost_history = raw.get("get_provider_cost_summary_history")
        if isinstance(debug_history, list):
            stage_items.extend(
                ("GetSearchTaskDebug", item)
                for item in debug_history
                if isinstance(item, dict)
            )
        else:
            stage_items.append(
                ("GetSearchTaskDebug", raw.get("get_search_task_debug"))
            )
        if isinstance(cost_history, list):
            stage_items.extend(
                ("GetProviderCostSummary", item)
                for item in cost_history
                if isinstance(item, dict)
            )
        else:
            stage_items.append(
                (
                    "GetProviderCostSummary",
                    raw.get("get_provider_cost_summary"),
                )
            )
        stage_sequences: dict[str, int] = defaultdict(int)
        for stage, payload in stage_items:
            if not isinstance(payload, dict) or not payload:
                continue
            stage_sequences[stage] += 1
            self._insert_raw(
                connection,
                run_id=run_id,
                query_id=query_id,
                candidate_pk=None,
                stage=stage,
                sequence_no=int(
                    payload.get("sequence_no") or stage_sequences[stage]
                ),
                payload=payload,
                collected_at=payload.get("collected_at"),
            )
        for candidate_pk, candidate in candidate_pks:
            detail_payload = candidate.get("detail_response_raw")
            if detail_payload is None:
                detail_payload = candidate.get("detail_data_raw")
            if detail_payload is None:
                continue
            self._insert_raw(
                connection,
                run_id=run_id,
                query_id=query_id,
                candidate_pk=candidate_pk,
                stage="GetTaskCandidateDetail",
                sequence_no=candidate["candidate_rank"],
                payload=detail_payload,
            )

    def _import_run(
        self,
        result_records: list[dict[str, Any]],
        failure_records: list[dict[str, Any]],
        errors: list[str],
        *,
        sources: list[tuple[str, Path]],
        evaluation_id: str,
        run_label: str,
        system_version: str,
        evaluation_phase: str,
        source_type: str,
        run_id: str | None,
    ) -> ImportResult:
        """执行 JSONL/Excel Run 共用的归档、规范化和单事务写入。"""

        if errors:
            raise ImportValidationError(errors)
        if not result_records and not failure_records:
            raise ImportValidationError(["结果文件没有可导入记录"])
        if evaluation_phase not in EVALUATION_PHASES:
            raise ImportValidationError(
                [f"evaluation_phase 不受支持: {evaluation_phase}"]
            )
        if self.store.fetch_one(
            "SELECT evaluation_id FROM evaluations WHERE evaluation_id = ?",
            (evaluation_id,),
        ) is None:
            raise ImportValidationError([f"Evaluation 不存在: {evaluation_id}"])
        validate_storage_id(evaluation_id, "evaluation_id")
        object_id = validate_storage_id(
            run_id or f"run_{uuid.uuid4().hex}",
            "run_id",
        )
        checksum = file_checksum(sources)
        if self.store.fetch_one(
            "SELECT run_id FROM runs WHERE source_checksum = ?",
            (checksum,),
        ):
            raise DuplicateImportError("相同结果文件包已经导入")
        archive_dir, archived = self._archive_sources(object_id, sources)
        normalized_dir = self.raw_dir / evaluation_id / object_id
        if normalized_dir.exists():
            self._cleanup_created(archive_dir)
            raise DuplicateImportError(f"规范化目录已存在: {normalized_dir}")
        normalized_results_path = normalized_dir / "results.jsonl"
        normalized_failures_path = normalized_dir / "failures.jsonl"
        self._write_jsonl(normalized_results_path, result_records)
        self._write_jsonl(normalized_failures_path, failure_records)
        now = utc_now_text()
        result_by_query = {
            record["input_id"]: record for record in result_records
        }
        failures_by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for failure in failure_records:
            failures_by_query[failure["input_id"]].append(failure)
        query_ids = list(result_by_query)
        query_ids.extend(
            query_id
            for query_id in failures_by_query
            if query_id not in result_by_query
        )
        query_statuses = []
        for query_id in query_ids:
            record = result_by_query.get(query_id)
            if record is None:
                query_statuses.append("FAILED")
            else:
                query_statuses.append(record["query_status"])
        success_count = sum(
            status in {"SUCCESS", "NO_CANDIDATE"} for status in query_statuses
        )
        failed_count = len(query_statuses) - success_count
        if failed_count == len(query_statuses):
            run_status = "FAILED"
        elif failed_count:
            run_status = "PARTIAL_FAILED"
        else:
            run_status = "COMPLETED"
        raw_statuses = {
            record.get("raw_status", "LEGACY_PARTIAL_RAW")
            for record in result_records
        }
        message = (
            "LEGACY_PARTIAL_RAW"
            if "LEGACY_PARTIAL_RAW" in raw_statuses
            else ""
        )
        schema_versions = {
            str(record.get("result_schema_version") or "legacy")
            for record in result_records
        }
        if RESULT_SCHEMA_VERSION in schema_versions:
            schema_version = RESULT_SCHEMA_VERSION
        elif schema_versions:
            schema_version = sorted(schema_versions)[0]
        else:
            schema_version = "legacy"
        try:
            with self.store.transaction() as connection:
                connection.execute(
                    """
                    INSERT INTO runs(
                        run_id, evaluation_id, dataset_id, run_label,
                        system_version, source_type, status,
                        result_schema_version, results_file, failures_file,
                        source_checksum, total_queries, success_queries,
                        failed_queries, started_at, finished_at, message,
                        evaluation_phase, created_at
                    ) VALUES (?, ?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        object_id,
                        evaluation_id,
                        run_label,
                        system_version,
                        source_type,
                        run_status,
                        schema_version,
                        normalized_results_path.relative_to(
                            self.data_dir
                        ).as_posix(),
                        normalized_failures_path.relative_to(
                            self.data_dir
                        ).as_posix(),
                        checksum,
                        len(query_ids),
                        success_count,
                        failed_count,
                        now,
                        now,
                        message,
                        evaluation_phase,
                        now,
                    ),
                )
                for query_id in query_ids:
                    record = result_by_query.get(query_id)
                    query_failures = failures_by_query.get(query_id, [])
                    if record is None:
                        first_failure = (
                            query_failures[0] if query_failures else {}
                        )
                        task_id = (
                            first_failure.get("task_id", "")
                        )
                        status = "FAILED"
                        current_stage = (
                            first_failure.get("stage", "Import")
                        )
                        error = (
                            first_failure.get("error", "缺少结果记录")
                        )
                        person_id = first_failure.get("person_id")
                        query_stage = first_failure.get("query_stage")
                        candidate_total = None
                        candidate_listed = 0
                        detail_success = 0
                        detail_failure = 0
                        task_fields: dict[str, Any] = {}
                        task_fields.update(
                            first_failure.get("task_fields", {})
                            if isinstance(first_failure.get("task_fields"), dict)
                            else {}
                        )
                        public_fields = (
                            first_failure.get("public_fields", {})
                            if isinstance(first_failure.get("public_fields"), dict)
                            else {}
                        )
                        result_status = "EXECUTION_FAILED"
                    else:
                        task_id = record["task_id"]
                        status = record["query_status"]
                        current_stage = (
                            query_failures[0]["stage"]
                            if query_failures
                            else (
                                "Import"
                                if record["raw_status"] == "LEGACY_PARTIAL_RAW"
                                else ""
                            )
                        )
                        error = (
                            query_failures[0]["error"]
                            if status == "FAILED" and query_failures
                            else ""
                        )
                        person_id = record.get("person_id")
                        query_stage = record.get("query_stage")
                        candidate_total = record.get("candidate_count_total")
                        candidate_listed = record["candidate_count_listed"]
                        detail_success = record["detail_success_count"]
                        detail_failure = record["detail_failure_count"]
                        task_fields = record.get("task_fields", {})
                        public_fields = (
                            record.get("public_fields", {})
                            if isinstance(record.get("public_fields"), dict)
                            else {}
                        )
                        result_status = record["result_status"]
                    public_fields = {
                        **public_fields,
                        **{
                            key: value
                            for key, value in task_fields.items()
                            if key not in TASK_FIELD_KEYS
                        },
                    }
                    connection.execute(
                        """
                        INSERT INTO run_queries(
                            run_id, query_id, person_id, person_id_source,
                            query_stage, task_id, status, current_stage,
                            candidate_count_total,
                            candidate_count_listed, detail_success_count,
                            detail_failure_count, llm_cost, third_party_cost,
                            total_cost, pdl_called, search_duration_ms,
                            result_status, public_fields_json, error,
                            started_at, finished_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            object_id,
                            query_id,
                            person_id,
                            (
                                "IMPORT_METADATA"
                                if person_id
                                else "UNSPECIFIED"
                            ),
                            query_stage,
                            task_id,
                            status,
                            current_stage,
                            candidate_total,
                            candidate_listed,
                            detail_success,
                            detail_failure,
                            task_fields.get("llm_cost"),
                            task_fields.get("third_party_cost"),
                            task_fields.get("total_cost"),
                            task_fields.get("pdl_called"),
                            task_fields.get("search_duration_ms"),
                            result_status,
                            json_text(public_fields),
                            error,
                            now,
                            now,
                        ),
                    )
                    candidate_pks: list[tuple[str, dict[str, Any]]] = []
                    if record is not None:
                        for candidate in record["results"]:
                            candidate_pk = f"candidate_{uuid.uuid4().hex}"
                            candidate_pks.append((candidate_pk, candidate))
                            connection.execute(
                                """
                                INSERT INTO candidates(
                                    candidate_pk, run_id, query_id,
                                    candidate_id, candidate_rank, rank_score,
                                    detail_status, detail_error,
                                    ui_sections_json, detail_data_json,
                                    list_item_json, created_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    candidate_pk,
                                    object_id,
                                    query_id,
                                    candidate["candidate_id"],
                                    candidate["candidate_rank"],
                                    candidate["rank_score"],
                                    candidate["detail_status"],
                                    candidate["detail_error"],
                                    (
                                        json_text(candidate["ui_sections"])
                                        if candidate.get("ui_sections") is not None
                                        else None
                                    ),
                                    (
                                        json_text(candidate["detail_data_raw"])
                                        if candidate.get("detail_data_raw") is not None
                                        else None
                                    ),
                                    json_text(candidate["list_item_raw"]),
                                    now,
                                ),
                            )
                        self._insert_result_raw(
                            connection,
                            run_id=object_id,
                            record=record,
                            candidate_pks=candidate_pks,
                            source_type=source_type,
                        )
                    else:
                        failure_raw = first_failure.get("raw")
                        if isinstance(failure_raw, list) and failure_raw:
                            for raw_item in failure_raw:
                                if not isinstance(raw_item, dict):
                                    continue
                                self._insert_raw(
                                    connection,
                                    run_id=object_id,
                                    query_id=query_id,
                                    candidate_pk=None,
                                    stage=str(raw_item.get("stage") or "Import"),
                                    sequence_no=int(raw_item.get("sequence_no") or 1),
                                    payload=raw_item,
                                    collected_at=raw_item.get("collected_at"),
                                )
                        else:
                            self._insert_raw(
                                connection,
                                run_id=object_id,
                                query_id=query_id,
                                candidate_pk=None,
                                stage="Import",
                                sequence_no=1,
                                payload={
                                    "raw_status": "LEGACY_PARTIAL_RAW",
                                    "source_type": source_type,
                                },
                            )
                for failure in failure_records:
                    connection.execute(
                        """
                        INSERT INTO failures(
                            failure_id, run_id, query_id, candidate_id,
                            scope, stage, error, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            f"failure_{uuid.uuid4().hex}",
                            object_id,
                            failure["input_id"],
                            failure["candidate_id"],
                            failure["scope"],
                            failure["stage"],
                            failure["error"],
                            failure["created_at"],
                        ),
                    )
        except sqlite3.IntegrityError as exc:
            self._cleanup_created(archive_dir, normalized_dir)
            if "source_checksum" in str(exc).lower():
                raise DuplicateImportError("相同结果文件包已经导入") from exc
            raise
        except Exception:
            self._cleanup_created(archive_dir, normalized_dir)
            raise
        return ImportResult(object_id, len(query_ids), checksum, archived)
