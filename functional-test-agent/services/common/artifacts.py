"""任务产物原子发布、登记和安全解析。"""

from __future__ import annotations

import hashlib
import json
import os
import secrets
import shutil
from pathlib import Path
from typing import Iterable

from services.common.task_models import ArtifactModel, utc_now
from services.common.task_store import TaskStore
from services.common.redaction import redact_text


def _contained_file(task_dir: Path, relative_path: str) -> Path:
    """解析任务内普通文件，拒绝绝对路径、符号链接和越界。"""

    if not relative_path or Path(relative_path).is_absolute():
        raise ValueError("产物路径不合法")
    target = (task_dir / relative_path).resolve()
    if task_dir.resolve() not in target.parents or not target.is_file() or target.is_symlink():
        raise ValueError("产物路径越界或文件不存在")
    return target


def publish_artifact(
    store: TaskStore,
    task_id: str,
    source: Path,
    *,
    artifact_type: str,
    stage: str,
    destination_group: str,
    review_input: bool = False,
) -> dict:
    """把工作文件复制到 published 并返回可登记元数据。"""

    task_dir = store.task_dir(task_id)
    source = Path(source).resolve()
    if task_dir not in source.parents or not source.is_file() or source.is_symlink():
        raise ValueError("只能发布当前任务内的普通文件")
    destination_dir = task_dir / "published" / destination_group
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / source.name
    temporary = destination.with_name(f".{destination.name}.{secrets.token_hex(6)}.tmp")
    try:
        with source.open("rb") as reader, temporary.open("xb") as writer:
            shutil.copyfileobj(reader, writer)
            writer.flush()
            os.fsync(writer.fileno())
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)
    digest = hashlib.sha256(destination.read_bytes()).hexdigest()
    return ArtifactModel(
        id=f"artifact_{secrets.token_hex(10)}",
        type=artifact_type,
        name=destination.name,
        relative_path=destination.relative_to(task_dir).as_posix(),
        size=destination.stat().st_size,
        sha256=digest,
        stage=stage,
        created_at=utc_now(),
        review_input=review_input,
    ).model_dump()


def save_registry(store: TaskStore, task_id: str, items: Iterable[dict]) -> None:
    """原子保存任务产物索引。"""

    TaskStore.atomic_write_json(store.task_dir(task_id) / "artifacts.json", {"schema_version": 1, "items": list(items)})


def merge_registry(store: TaskStore, task_id: str, items: Iterable[dict]) -> list[dict]:
    """在任务锁内合并阶段产物，避免后续阶段覆盖已成功的上游产物。

    参数说明:
        store: 当前智能体的任务存储。
        task_id: 需要更新产物索引的任务 ID。
        items: 本阶段新产生的产物元数据。

    返回值:
        合并后的完整产物列表。同一产物 ID 会以最新元数据替换，其他条目保留。
    """

    with store.locked():
        merged = {item.get("id"): item for item in load_registry(store, task_id) if item.get("id")}
        for item in items:
            if item.get("id"):
                merged[item["id"]] = item
        result = list(merged.values())
        save_registry(store, task_id, result)
        return result


def load_registry(store: TaskStore, task_id: str) -> list[dict]:
    """读取产物索引；损坏或不存在时返回空列表。"""

    import json

    try:
        payload = json.loads((store.task_dir(task_id) / "artifacts.json").read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return []
    items = payload.get("items", []) if isinstance(payload, dict) else []
    return items if isinstance(items, list) else []


def resolve_artifact(store: TaskStore, task_id: str, artifact_id: str) -> tuple[Path, dict]:
    """按已登记 ID 返回产物，禁止浏览器提供磁盘路径。"""

    for item in load_registry(store, task_id):
        if item.get("id") == artifact_id and not item.get("expired"):
            return _contained_file(store.task_dir(task_id), item["relative_path"]), item
    raise FileNotFoundError("产物不存在")


def preview_artifact(path: Path, *, max_bytes: int = 512 * 1024) -> dict:
    """返回浏览器可安全以纯文本展示的产物摘要。

    JSON 会格式化；文本类文件按 UTF-8 解码。XLSX 只读取首个工作表的
    前 100 行和 30 列，不计算公式，避免预览触发文件内代码或外部连接。
    """

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            lines = ["\t".join("" if value is None else str(value) for value in row) for row in sheet.iter_rows(min_row=1, max_row=100, max_col=30, values_only=True)]
            return {"content": redact_text("\n".join(lines)), "format": "xlsx", "truncated": sheet.max_row > 100 or sheet.max_column > 30}
        finally:
            workbook.close()
    if suffix not in {".json", ".md", ".txt", ".yaml", ".yml", ".csv", ".log"}:
        raise ValueError("该产物类型不支持在线预览")
    payload = path.read_bytes()
    truncated = len(payload) > max_bytes
    content = payload[:max_bytes].decode("utf-8", errors="replace")
    if suffix == ".json" and not truncated:
        try:
            content = json.dumps(json.loads(content), ensure_ascii=False, indent=2)
        except json.JSONDecodeError:
            pass
    return {"content": redact_text(content), "format": suffix.lstrip("."), "truncated": truncated}
