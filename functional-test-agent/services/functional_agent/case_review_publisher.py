"""从确认测试用例 JSON 确定性生成最终 JSON 与 XLSX 产物。"""

from __future__ import annotations

import hashlib
import json
import os
import secrets
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services.common.artifacts import load_registry, merge_registry
from services.common.case_review import CaseReviewService
from services.common.task_models import ArtifactModel, utc_now
from services.common.task_store import TaskStore


COLUMNS = (
    "case_id", "test_point_id", "module", "feature", "scenario", "case_name", "priority",
    "preconditions", "test_steps", "test_data", "expected_result", "actual_result", "其他字段",
)


def _safe_cell(value: Any) -> str:
    """把嵌套值稳定转换为文本并防止表格公式注入。"""

    if isinstance(value, dict):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    elif isinstance(value, list):
        text = "\n".join(
            json.dumps(item, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            if isinstance(item, (dict, list)) else str(item)
            for item in value
        )
    elif value is None:
        text = ""
    else:
        text = str(value)
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _xlsx(path: Path, cases: list[Any]) -> None:
    """以固定列顺序写入 XLSX；异常业务字段也稳定转换为可查看文本。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"
    sheet.append(list(COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D1D1F")
    for item in cases:
        if not isinstance(item, dict):
            raise ValueError("测试用例数组的每个元素必须是对象")
        source = item
        extras = {key: value for key, value in source.items() if key not in COLUMNS}
        row: list[str] = []
        for field in COLUMNS:
            value = extras if field == "其他字段" else source.get(field, "")
            if field == "test_steps" and isinstance(value, list):
                value = "\n".join(f"{index}. {step}" for index, step in enumerate(value, 1))
            elif field == "preconditions" and isinstance(value, list):
                value = "\n".join(str(entry) for entry in value)
            row.append(_safe_cell(value))
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        for cell in column[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def _artifact(task_dir: Path, path: Path, artifact_type: str, version: int) -> dict[str, Any]:
    """根据已完整写入的最终文件构造白名单产物元数据。"""

    payload = path.read_bytes()
    return ArtifactModel(
        id=f"artifact_{hashlib.sha256((artifact_type + str(version)).encode()).hexdigest()[:20]}",
        type=artifact_type,
        name=path.name,
        relative_path=path.relative_to(task_dir).as_posix(),
        size=len(payload), sha256=hashlib.sha256(payload).hexdigest(),
        stage="case_review_published", created_at=utc_now(), review_input=False,
    ).model_dump()


def publish_confirmed_cases(store: TaskStore, task_id: str, metadata: dict[str, Any]) -> list[dict[str, Any]]:
    """发布一个确认版本并幂等登记 JSON/XLSX。

    异常策略：先在临时目录生成并 fsync 两个文件，再原子发布版本目录；任何一步
    失败都不会向 artifact registry 登记半成品，重试可复用已完整发布的目录。
    """

    task_dir = store.task_dir(task_id)
    version = int(metadata["version"])
    cases = CaseReviewService(store).read_confirmed(task_id, metadata)
    destination = task_dir / "published" / "test-cases" / f"v{version}"
    temporary = destination.with_name(f".v{version}.{secrets.token_hex(6)}.tmp")
    if not destination.exists():
        temporary.mkdir(parents=True)
        try:
            json_path = temporary / "test-cases.json"
            json_path.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
            with json_path.open("rb") as handle:
                os.fsync(handle.fileno())
            _xlsx(temporary / "test-cases.xlsx", cases)
            with (temporary / "test-cases.xlsx").open("rb") as handle:
                os.fsync(handle.fileno())
            try:
                os.replace(temporary, destination)
            except OSError:
                if not destination.exists():
                    raise
        finally:
            shutil.rmtree(temporary, ignore_errors=True)
    json_path = destination / "test-cases.json"
    xlsx_path = destination / "test-cases.xlsx"
    if not json_path.is_file() or not xlsx_path.is_file() or json_path.is_symlink() or xlsx_path.is_symlink():
        raise RuntimeError("最终用例产物发布不完整")
    published = [
        _artifact(task_dir, json_path, "test_cases_json", version),
        _artifact(task_dir, xlsx_path, "test_cases_xlsx", version),
    ]
    # 固定 artifact ID 使 registry 写入失败后的重试不会产生重复条目。
    merge_registry(store, task_id, published)
    return published
